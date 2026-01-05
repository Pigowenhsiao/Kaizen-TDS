import openpyxl as px
import logging
import shutil
import glob
import sys
import os

from datetime import datetime, timedelta, date
from time import strftime, localtime


########## 自作関数の定義 ##########
sys.path.append('../../MyModule')
import SQL
import Log
import MOCVD_OldFileSearch
import Check


########## 全体パラメータ定義 ##########
Site = '350'
ProductFamily = 'SAG FAB'
Operation = 'T2-DML'
Teststation = 'T2-DML'
X = '999999'
Y = '999999'


########## Logの設定 ##########
Log_FolderName = str(date.today())

# ----- 格納するLogフォルダがなければ作成する -----
if not os.path.exists("../../Log/" + Log_FolderName):
    os.makedirs("../../Log/" + Log_FolderName)

# ----- ログ書き込み先ファイルパス -----
Log_file = '../../Log/' + Log_FolderName + '/018_T2-DML_F5-10G.log'

# ----- ログ書き込み：プログラム開始 -----
Log.Log_Info(Log_file, 'Program Start')


########## 処理ファイルのあるディレクトリ定義 ##########
Path = 'Z:/MOCVD/MOCVD過去プログラム/F5炉/'
# Path = 'C:/Users/hor78296/Desktop/F5炉/'


########## XML出力先ファイルパス ##########
Output_filepath = '//li.lumentuminc.net/data/SAG/TDS/Data/Files to Insert/XML/'
# Output_filepath = '../../XML/018_T2-DML/F5-10G/'


########## TestStepの定義 ##########
teststep_dict = {
    'TestStep1' : 'Coordinate',
    'TestStep2' : 'XRayDiffraction',
    'TestStep3' : 'CarrierConcentration',
    'TestStep4' : 'Thickness',
    'TestStep5' : 'Particles',
    'TestStep6' : 'Dulation',
    'TestStep7' : 'MO1-TEG',
    'TestStep8' : 'MO9-TMI',
    'TestStep9' : 'AsH3-A-20percent',
    'TestStep10' : 'PH3-50percent',
    'TestStep11' : 'CH3Cl-5percent',
    'TestStep12' : 'DMZn-B-0.1percent',
    'TestStep13' : 'AsH3-5percent',
    'TestStep14' : 'Temperature',
    'TestStep15' : 'Piezocon',
    'TestStep16' : 'BallastN2',
    'TestStep17' : 'CH3Cl-Judge',
    'TestStep18' : 'MO-Temperature',
    'TestStep19' : 'SORTED_DATA',
}


########## HeaderMiscの定義 ##########
HeaderMisc_dict = {
    'HeaderMisc1' : 'RecipeName-Macro',
    'HeaderMisc2' : 'RecipeName-Program',
    'HeaderMisc3' : 'RecipeName-Folder'
}


########## 取得した項目と型の対応表を定義 ##########
key_type = {
    "key_start_date_time" : str,
    "key_serial_number" : str,
    "key_part_number" : str,
    "key_operator" : str,
    "key_batch_number" : str,
    "key_HeaderMisc1" : str,
    "key_HeaderMisc2" : str,
    "key_HeaderMisc3" : str,
    "key_XRayDiffraction_Strain" : float,
    "key_CarrierConcentration_Clad" : float,
    "key_CarrierConcentration_Contact" : float,
    "key_CarrierConcentration_Minimum" : float,
    "key_Thickness_Contact" : float,
    "key_Thickness_Clad" : float,
    "key_Particles_Particles" : float,
    "key_Dulation_Step6" : float,
    "key_Dulation_Step7" : float,
    "key_Dulation_Step10" : float,
    "key_Dulation_Step11" : float,
    "key_Dulation_Step12" : float,
    "key_Dulation_Step13" : float,
    "key_Dulation_Step14" : float,
    "key_Dulation_Step20" : float,
    "key_Dulation_Step26" : float,
    "key_Dulation_Step32" : float,
    "key_Dulation_Step33" : float,
    "key_Dulation_Step34" : float,
    "key_Dulation_Step35" : float,
    "key_Dulation_Step40" : float,
    "key_MO1-TEG_Step20" : float,
    "key_MO1-TEG_Step26" : float,
    "key_MO1-TEG_Step32" : float,
    "key_MO1-TEG_Step33" : float,
    "key_MO1-TEG_Step34" : float,
    "key_MO1-TEG_Step35" : float,
    "key_MO9-TMI_Step10" : float,
    "key_MO9-TMI_Step11" : float,
    "key_MO9-TMI_Step12" : float,
    "key_MO9-TMI_Step13" : float,
    "key_MO9-TMI_Step14" : float,
    "key_MO9-TMI_Step20" : float,
    "key_MO9-TMI_Step26" : float,
    "key_MO9-TMI_Step32" : float,
    "key_MO9-TMI_Step33" : float,
    "key_MO9-TMI_Step34" : float,
    "key_MO9-TMI_Step35" : float,
    "key_MO9-TMI_Step40" : float,
    "key_AsH3-A-20percent_Step26" : float,
    "key_AsH3-A-20percent_Step32" : float,
    "key_AsH3-A-20percent_Step33" : float,
    "key_AsH3-A-20percent_Step34" : float,
    "key_AsH3-A-20percent_Step35" : float,
    "key_PH3-50percent_Step6" : float,
    "key_PH3-50percent_Step7" : float,
    "key_PH3-50percent_Step10" : float,
    "key_PH3-50percent_Step11" : float,
    "key_PH3-50percent_Step12" : float,
    "key_PH3-50percent_Step13" : float,
    "key_PH3-50percent_Step14" : float,
    "key_PH3-50percent_Step20" : float,
    "key_PH3-50percent_Step26" : float,
    "key_PH3-50percent_Step40" : float,
    "key_CH3Cl-5percent_Step7" : float,
    "key_DMZn-B-0.1percent_Step10" : float,
    "key_DMZn-B-0.1percent_Step11" : float,
    "key_DMZn-B-0.1percent_Step12" : float,
    "key_DMZn-B-0.1percent_Step13" : float,
    "key_DMZn-B-0.1percent_Step14" : float,
    "key_DMZn-B-0.1percent_Step20" : float,
    "key_DMZn-B-0.1percent_Step26" : float,
    "key_DMZn-B-0.1percent_Step32" : float,
    "key_DMZn-B-0.1percent_Step33" : float,
    "key_DMZn-B-0.1percent_Step34" : float,
    "key_DMZn-B-0.1percent_Step35" : float,
    "key_DMZn-B-0.1percent_Step40" : float,
    "key_AsH3-5percent_Step20" : float,
    "key_Temperature_Step6" : float,
    "key_Temperature_Step7" : float,
    "key_Temperature_Step10" : float,
    "key_Temperature_Step11" : float,
    "key_Temperature_Step12" : float,
    "key_Temperature_Step13" : float,
    "key_Temperature_Step14" : float,
    "key_Temperature_Step20" : float,
    "key_Temperature_Step26" : float,
    "key_Temperature_Step32" : float,
    "key_Temperature_Step33" : float,
    "key_Temperature_Step34" : float,
    "key_Temperature_Step35" : float,
    "key_Temperature_Step40" : float,
    "key_Piezocon_F1" : float,
    "key_Piezocon_F1_Inverse" : float,
    "key_Piezocon_F5_Start" : float,
    "key_Piezocon_F5_Current" : float,
    "key_BallastN2_BallastN2" : float,
    "key_CH3Cl-Judge_StartTime" : float,
    "key_CH3Cl-Judge_StablWaitTime" : float,
    "key_CH3Cl-Judge_EtchingTime" : float,
    "key_CH3Cl-Judge_EtchingStartTemperature" : float,
    "key_MO-Temperature_MO1-TEG(A)" : float,
    "key_MO-Temperature_MO2-TMA(A)" : float,
    "key_MO-Temperature_MO3-DEZ" : float,
    "key_MO-Temperature_MO4-TMI(A)" : float,
    "key_MO-Temperature_MO5-TMI(B)" : float,
    "key_MO-Temperature_MO6-TEG(B)" : float,
    "key_MO-Temperature_MO7-TMA(B)" : float,
    "key_MO-Temperature_MO8-CBr4" : float,
    "key_MO-Temperature_MO9-TMI" : float,
    "key_STARTTIME_SORTED": float,
    "key_SORTNUMBER" : float,
    "key_LotNumber_9": str
}


########## 対象ロット番号のイニシャルを記載したファイルを取得する ##########
Log.Log_Info(Log_file, 'Get SerialNumber Initial List ')
with open('T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開発/018_T2-DML/SerialNumber_Initial_10G.txt', 'r', encoding='utf-8') as textfile:
    SerialNumber_list = {s.strip() for s in textfile.readlines()}


########## 前回処理を行ったファイル名を取得する ##########
with open('F5_FileName_F5-10G.txt', 'r', encoding='utf-8') as textfile:
    Before_FileName = textfile.readline()


########## 空欄チェック ##########
def Get_Cells_Info(Sheet):

    # ----- ログ書込：空欄判定 -----
    Log.Log_Info(Log_file, "Blank Check")

    # ----- False -> 空欄がない -----
    is_cells_empty = False

    # ----- EpiNo./ 日付 が空欄ならば処理を行わない -----
    if Sheet['I8'].value is None or Sheet['Q7'].value is None:
        is_cells_empty = True

    return is_cells_empty


########## データの取得 ##########
def Open_Data_Sheet(Sheet, filepath, SheetName):

    # ----- ログ書込：データ取得 -----
    Log.Log_Info(Log_file, 'Data Acquisition')

    # ----- データを格納する辞書を作成 -----
    data_dict = dict()

    # ----- ロット番号の取得 -----
    serial_number = Sheet['M8'].value
    conn, cursor = SQL.connSQL()

    # ----- Primeに接続できなかったときはNoneが返ってくる -----
    if conn is None:
        Log.Log_Error(Log_file, serial_number + ' : ' + 'Connection with Prime Failed')
        sys.exit()

    # ----- 品名の取得 -----
    part_number, Nine_Serial_Number = SQL.selectSQL(cursor, serial_number)
    SQL.disconnSQL(conn, cursor)

    # ----- Polaron / SEM / XRD / MOCVD の装置Noを取得する -----
    Polaron, SEM, XRD, MOCVD = '1', '1', '1', '5'
    if '#2' in str(Sheet['J49'].value):
        Polaron = '2'
    if '#2' in str(Sheet['J54'].value):
        SEM = '2'
    if '#2' in str(Sheet['J45'].value):
        XRD = '2'

    # ----- TemperatureのStep11とStep34の末尾に↑↓がついてることがあるので外す -----
    Step11 = str(Sheet['Z23'].value)
    Step33 = str(Sheet['Z30'].value)
    if Step11 is not None and (Step11[-1] == '↑' or Step11[-1] == '↓'):
        Step11 = Step11[:-1]
    if Step33 is not None and (Step33[-1] == '↑' or Step33[-1] == '↓'):
        Step33 = Step33[:-1]


    # ----- 必要なデータは辞書に格納する -----
    data_dict = {
        "key_start_date_time" : str(Sheet['Q7'].value).replace(" ", "T"),
        "key_serial_number" : serial_number,
        "key_part_number" : part_number,
        "key_operator" : Sheet['T9'].value,
        "key_LotNumber_9": Nine_Serial_Number,
        "key_batch_number" : Sheet['I8'].value,
        "key_HeaderMisc1" : Sheet['U3'].value,
        "key_HeaderMisc2" : Sheet['U4'].value,
        "key_HeaderMisc3" : Sheet['U5'].value,
        "key_XRayDiffraction_Strain" : Sheet['M46'].value,
        "key_CarrierConcentration_Clad" : Sheet['M50'].value,
        "key_CarrierConcentration_Contact" : Sheet['M49'].value,
        "key_CarrierConcentration_Minimum" : Sheet['M51'].value,
        "key_Thickness_Contact" : Sheet['M54'].value,
        "key_Thickness_Clad" : Sheet['M55'].value,
        "key_Particles_Particles" : Sheet['M59'].value,
        "key_Dulation_Step6" : Sheet['I20'].value,
        "key_Dulation_Step7" : Sheet['I21'].value,
        "key_Dulation_Step10" : Sheet['I22'].value,
        "key_Dulation_Step11" : Sheet['I23'].value,
        "key_Dulation_Step12" : Sheet['I24'].value,
        "key_Dulation_Step13" : Sheet['I25'].value,
        "key_Dulation_Step14" : Sheet['I26'].value,
        "key_Dulation_Step20" : Sheet['I27'].value,
        "key_Dulation_Step26" : Sheet['I28'].value,
        "key_Dulation_Step32" : Sheet['I29'].value,
        "key_Dulation_Step33" : Sheet['I30'].value,
        "key_Dulation_Step34" : Sheet['I31'].value,
        "key_Dulation_Step35" : Sheet['I32'].value,
        "key_Dulation_Step40" : Sheet['I33'].value,
        "key_MO1-TEG_Step20" : Sheet['K27'].value,
        "key_MO1-TEG_Step26" : Sheet['K28'].value,
        "key_MO1-TEG_Step32" : Sheet['K29'].value,
        "key_MO1-TEG_Step33" : Sheet['K30'].value,
        "key_MO1-TEG_Step34" : Sheet['K31'].value,
        "key_MO1-TEG_Step35" : Sheet['K32'].value,
        "key_MO9-TMI_Step10" : Sheet['S22'].value,
        "key_MO9-TMI_Step11" : Sheet['S23'].value,
        "key_MO9-TMI_Step12" : Sheet['S24'].value,
        "key_MO9-TMI_Step13" : Sheet['S25'].value,
        "key_MO9-TMI_Step14" : Sheet['S26'].value,
        "key_MO9-TMI_Step20" : Sheet['S27'].value,
        "key_MO9-TMI_Step26" : Sheet['S28'].value,
        "key_MO9-TMI_Step32" : Sheet['S29'].value,
        "key_MO9-TMI_Step33" : Sheet['S30'].value,
        "key_MO9-TMI_Step34" : Sheet['S31'].value,
        "key_MO9-TMI_Step35" : Sheet['S32'].value,
        "key_MO9-TMI_Step40" : Sheet['S33'].value,
        "key_AsH3-A-20percent_Step26" : Sheet['T28'].value,
        "key_AsH3-A-20percent_Step32" : Sheet['T29'].value,
        "key_AsH3-A-20percent_Step33" : Sheet['T30'].value,
        "key_AsH3-A-20percent_Step34" : Sheet['T31'].value,
        "key_AsH3-A-20percent_Step35" : Sheet['T32'].value,
        "key_PH3-50percent_Step6" : Sheet['U20'].value,
        "key_PH3-50percent_Step7" : Sheet['U21'].value,
        "key_PH3-50percent_Step10" : Sheet['U22'].value,
        "key_PH3-50percent_Step11" : Sheet['U23'].value,
        "key_PH3-50percent_Step12" : Sheet['U24'].value,
        "key_PH3-50percent_Step13" : Sheet['U25'].value,
        "key_PH3-50percent_Step14" : Sheet['U26'].value,
        "key_PH3-50percent_Step20" : Sheet['U27'].value,
        "key_PH3-50percent_Step26" : Sheet['U28'].value,
        "key_PH3-50percent_Step40" : Sheet['U33'].value,
        "key_CH3Cl-5percent_Step7" : Sheet['V21'].value,
        "key_DMZn-B-0.1percent_Step10" : Sheet['X22'].value,
        "key_DMZn-B-0.1percent_Step11" : Sheet['X23'].value,
        "key_DMZn-B-0.1percent_Step12" : Sheet['X24'].value,
        "key_DMZn-B-0.1percent_Step13" : Sheet['X25'].value,
        "key_DMZn-B-0.1percent_Step14" : Sheet['X26'].value,
        "key_DMZn-B-0.1percent_Step20" : Sheet['X27'].value,
        "key_DMZn-B-0.1percent_Step26" : Sheet['X28'].value,
        "key_DMZn-B-0.1percent_Step32" : Sheet['X29'].value,
        "key_DMZn-B-0.1percent_Step33" : Sheet['X30'].value,
        "key_DMZn-B-0.1percent_Step34" : Sheet['X31'].value,
        "key_DMZn-B-0.1percent_Step35" : Sheet['X32'].value,
        "key_DMZn-B-0.1percent_Step40" : Sheet['X33'].value,
        "key_AsH3-5percent_Step20" : Sheet['Y27'].value,
        "key_Temperature_Step6" : Sheet['Z20'].value,
        "key_Temperature_Step7" : Sheet['Z21'].value,
        "key_Temperature_Step10" : Sheet['Z22'].value,
        "key_Temperature_Step11" : Step11,
        "key_Temperature_Step12" : Sheet['Z24'].value,
        "key_Temperature_Step13" : Sheet['Z25'].value,
        "key_Temperature_Step14" : Sheet['Z26'].value,
        "key_Temperature_Step20" : Sheet['Z27'].value,
        "key_Temperature_Step26" : Sheet['Z28'].value,
        "key_Temperature_Step32" : Sheet['Z29'].value,
        "key_Temperature_Step33" : Step33,
        "key_Temperature_Step34" : Sheet['Z31'].value,
        "key_Temperature_Step35" : Sheet['Z32'].value,
        "key_Temperature_Step40" : Sheet['Z33'].value,
        "key_Piezocon_F1" : Sheet['AE39'].value,
        "key_Piezocon_F1_Inverse" : Sheet['AE40'].value,
        "key_Piezocon_F5_Start" : Sheet['AE45'].value,
        "key_Piezocon_F5_Current" : Sheet['AE46'].value,
        "key_BallastN2_BallastN2" : Sheet['AE42'].value,
        "key_CH3Cl-Judge_StartTime" : Sheet['AE50'].value,
        "key_CH3Cl-Judge_StablWaitTime" : Sheet['AE52'].value,
        "key_CH3Cl-Judge_EtchingTime" : Sheet['AE54'].value,
        "key_CH3Cl-Judge_EtchingStartTemperature" : Sheet['AE56'].value,
        "key_MO-Temperature_MO1-TEG(A)" : Sheet['K42'].value,
        "key_MO-Temperature_MO2-TMA(A)" : Sheet['L42'].value,
        "key_MO-Temperature_MO3-DEZ" : Sheet['M42'].value,
        "key_MO-Temperature_MO4-TMI(A)" : Sheet['N42'].value,
        "key_MO-Temperature_MO5-TMI(B)" : Sheet['O42'].value,
        "key_MO-Temperature_MO6-TEG(B)" : Sheet['P42'].value,
        "key_MO-Temperature_MO7-TMA(B)" : Sheet['Q42'].value,
        "key_MO-Temperature_MO8-CBr4" : Sheet['R42'].value,
        "key_MO-Temperature_MO9-TMI" : Sheet['S42'].value,
        "key_Polaron": Polaron,
        "key_SEM": SEM,
        "key_XRD": XRD,
        "key_MOCVD": MOCVD,
    }

    # ----- 空欄箇所はNoneとして取得される。Noneは文字列に変換できないため、空欄("")に置き換える -----
    for keys in data_dict:
        if data_dict[keys] is None or data_dict[keys] == "None" or (data_dict[keys] == '-' and keys != 'key_operator'):
            data_dict[keys] = ""
        # ----- 指数表記箇所はint型に変換する(ただし負の指数は無理(現状ないため特段の処理は行わない)) -----
        if type(data_dict[keys]) is float and 'e' in str(data_dict[keys]) and keys != "key_start_date_time":
            data_dict[keys] = int(float(data_dict[keys]))

    return data_dict


# XML変換
def Output_XML(xml_file, data_dict):

    # ----- ログ書込：XML変換 -----
    Log.Log_Info(Log_file, 'Excel File To XML File Conversion')

    f = open(Output_filepath + xml_file, 'w', encoding="utf-8")

    f.write('<?xml version="1.0" encoding="utf-8"?>' + '\n' +
            '<Results xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema">' + '\n' +
            '       <Result startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Result="Passed">' + '\n' +
            '               <Header SerialNumber=' + '"' + data_dict["key_serial_number"] + '"' + ' PartNumber=' + '"' + data_dict["key_part_number"] + '"' + ' Operation=' + '"' + Operation + '"' + ' TestStation=' + '"' + Operation + '"' + ' Operator=' + '"' + data_dict["key_operator"] + '"' + ' StartTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Site=' + '"' + Site + '"' + ' BatchNumber=' + '"' + data_dict["key_batch_number"] + '"' + ' LotNumber=' + '"' + data_dict["key_serial_number"] + '"/>' + '\n' +
            '               <HeaderMisc>' + '\n' +
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc1"] + '">' + data_dict["key_HeaderMisc1"] + '</Item>' + '\n'
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc2"] + '">' + data_dict["key_HeaderMisc2"] + '</Item>' + '\n'
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc3"] + '">' + data_dict["key_HeaderMisc3"] + '</Item>' + '\n'
            '               </HeaderMisc>' + '\n' +
            '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep1"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + X + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + Y + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep2"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="XRayDiffraction_Strain" Units="percent" Value=' + '"' + str(data_dict["key_XRayDiffraction_Strain"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep3"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="CarrierConcentration_Clad" Units="cm-3" Value=' + '"' + str(data_dict["key_CarrierConcentration_Clad"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="CarrierConcentration_Contact" Units="cm-3" Value=' + '"' + str(data_dict["key_CarrierConcentration_Contact"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="CarrierConcentration_Minimum" Units="cm-3" Value=' + '"' + str(data_dict["key_CarrierConcentration_Minimum"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep4"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Thickness_Contact" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Contact"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Thickness_Clad" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Clad"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep5"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Particles" Units="pieces" Value=' + '"' + str(data_dict["key_Particles_Particles"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep6"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step6" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step6"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step7" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step7"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step11" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step11"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step12"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step20" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step20"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step26" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step26"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step32" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step32"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step33" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step33"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step34" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step34"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step35" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step35"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step40" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step40"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep7"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step20" Units="sccm" Value=' + '"' + str(data_dict["key_MO1-TEG_Step20"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step26" Units="sccm" Value=' + '"' + str(data_dict["key_MO1-TEG_Step26"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step32" Units="sccm" Value=' + '"' + str(data_dict["key_MO1-TEG_Step32"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step33" Units="sccm" Value=' + '"' + str(data_dict["key_MO1-TEG_Step33"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step34" Units="sccm" Value=' + '"' + str(data_dict["key_MO1-TEG_Step34"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step35" Units="sccm" Value=' + '"' + str(data_dict["key_MO1-TEG_Step35"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep8"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sccm" Value=' + '"' + str(data_dict["key_MO9-TMI_Step10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step11" Units="sccm" Value=' + '"' + str(data_dict["key_MO9-TMI_Step11"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12" Units="sccm" Value=' + '"' + str(data_dict["key_MO9-TMI_Step12"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="sccm" Value=' + '"' + str(data_dict["key_MO9-TMI_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="sccm" Value=' + '"' + str(data_dict["key_MO9-TMI_Step14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step20" Units="sccm" Value=' + '"' + str(data_dict["key_MO9-TMI_Step20"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step26" Units="sccm" Value=' + '"' + str(data_dict["key_MO9-TMI_Step26"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step32" Units="sccm" Value=' + '"' + str(data_dict["key_MO9-TMI_Step32"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step33" Units="sccm" Value=' + '"' + str(data_dict["key_MO9-TMI_Step33"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step34" Units="sccm" Value=' + '"' + str(data_dict["key_MO9-TMI_Step34"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step35" Units="sccm" Value=' + '"' + str(data_dict["key_MO9-TMI_Step35"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step40" Units="sccm" Value=' + '"' + str(data_dict["key_MO9-TMI_Step40"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep9"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step26" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-A-20percent_Step26"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step32" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-A-20percent_Step32"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step33" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-A-20percent_Step33"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step34" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-A-20percent_Step34"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step35" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-A-20percent_Step35"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep10"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step6" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-50percent_Step6"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step7" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-50percent_Step7"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-50percent_Step10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step11" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-50percent_Step11"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-50percent_Step12"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-50percent_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-50percent_Step14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step20" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-50percent_Step20"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step26" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-50percent_Step26"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step40" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-50percent_Step40"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep11"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step7" Units="sccm" Value=' + '"' + str(data_dict["key_CH3Cl-5percent_Step7"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep12"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-B-0.1percent_Step10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step11" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-B-0.1percent_Step11"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-B-0.1percent_Step12"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-B-0.1percent_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-B-0.1percent_Step14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step20" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-B-0.1percent_Step20"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step26" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-B-0.1percent_Step26"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step32" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-B-0.1percent_Step32"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step33" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-B-0.1percent_Step33"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step34" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-B-0.1percent_Step34"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step35" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-B-0.1percent_Step35"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step40" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-B-0.1percent_Step40"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep13"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step20" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-5percent_Step20"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep14"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step6" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step6"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step7" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step7"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step11" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step11"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step12"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step20" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step20"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step26" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step26"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step32" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step32"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step33" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step33"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step34" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step34"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step35" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step35"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step40" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step40"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep15"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="F1" Units="percent" Value=' + '"' + str(data_dict["key_Piezocon_F1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="F1_Inverse" Units="percent" Value=' + '"' + str(data_dict["key_Piezocon_F1_Inverse"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="F5_Start" Units="percent" Value=' + '"' + str(data_dict["key_Piezocon_F5_Start"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="F5_Current" Units="percent" Value=' + '"' + str(data_dict["key_Piezocon_F5_Current"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep16"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="BallastN2" Units="slm" Value=' + '"' + str(data_dict["key_BallastN2_BallastN2"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep17"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="StartTime" Units="sec" Value=' + '"' + str(data_dict["key_CH3Cl-Judge_StartTime"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="StablWaitTime" Units="sec" Value=' + '"' + str(data_dict["key_CH3Cl-Judge_StablWaitTime"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="EtchingTime" Units="sec" Value=' + '"' + str(data_dict["key_CH3Cl-Judge_EtchingTime"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="EtchingStartTemperature" Units="degree" Value=' + '"' + str(data_dict["key_CH3Cl-Judge_EtchingStartTemperature"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep18"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="MO1-TEG(A)" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO1-TEG(A)"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO2-TMA(A)" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO2-TMA(A)"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO3-DEZ" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO3-DEZ"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO4-TMI(A)" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO4-TMI(A)"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO5-TMI(B)" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO5-TMI(B)"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO6-TEG(B)" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO6-TEG(B)"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO7-TMA(B)" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO7-TMA(B)"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO8-CBr4" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO8-CBr4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO9-TMI" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO9-TMI"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep19"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="STARTTIME_SORTED" Units="" Value=' + '"' + str(data_dict["key_STARTTIME_SORTED"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="SORTNUMBER" Units="" Value=' + '"' + str(data_dict["key_SORTNUMBER"]) + '"/>' + '\n' +
            '                   <Data DataType="String" Name="LotNumber_5" Value=' + '"' + str(data_dict["key_serial_number"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="LotNumber_9" Value=' + '"' + str(data_dict["key_LotNumber_9"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '\n'
            '               <TestEquipment>' + '\n' +
            '                   <Item DeviceName="Polaron" DeviceSerialNumber="' + data_dict["key_Polaron"] + '"/>' + '\n' +
            '                   <Item DeviceName="SEM" DeviceSerialNumber="' + data_dict["key_SEM"] + '"/>' + '\n' +
            '                   <Item DeviceName="XRD" DeviceSerialNumber="' + data_dict["key_XRD"] + '"/>' + '\n' +
            '                   <Item DeviceName="MOCVD" DeviceSerialNumber="' + data_dict["key_MOCVD"] + '"/>' + '\n' +
            '               </TestEquipment>' + '\n' +
            '\n'
            '               <ErrorData/>' + '\n' +
            '               <FailureData/>' + '\n' +
            '               <Configuration/>' + '\n' +
            '       </Result>' + '\n' +
            '</Results>'
            )
    f.close()


########## シートの判定からXML変換まで処理 ##########
def Data_Extract(filepath, SheetList, old_check):

    # ----- ログ書込：データ変換処理 -----
    Log.Log_Info(Log_file, 'Sub Start')
    wb = px.load_workbook(filepath, read_only=True, data_only=True)

    # ----- 全シートの処理 -----
    for Sheet_Name in SheetList[::-1]:

        Sheet = wb[Sheet_Name]
        Initial = str(Sheet['M8'].value)[0]

        # ----- シートが処理対象シートかどうか確認 -----
        if '139S-T2.exe' not in str(Sheet['U3'].value) or Initial not in SerialNumber_list:
            Log.Log_Error(Log_file, Sheet_Name + ' : ' + 'Not Covered\n')
            continue

        # ----- 空欄チェック -----
        if Get_Cells_Info(Sheet):
            Log.Log_Error(Log_file, "Blank Error\n")
            continue

        # ----- データの取得 -----
        data_dict = Open_Data_Sheet(Sheet, os.path.basename(filepath), Sheet_Name)

        # ----- oldファイルの実行時のみ、着工者が空欄であれば'-'に置き換える -----
        if data_dict["key_operator"] == "":
            if old_check:
                data_dict["key_operator"] = '-'
            else:
                Log.Log_Error(Log_file, Sheet_Name + ' : ' + 'Operator None\n')
                continue

        # ----- 日付フォーマット変換 -----
        if len(data_dict['key_start_date_time']) != 19 or '年' in data_dict['key_start_date_time']:
            Log.Log_Error(Log_file, data_dict["key_serial_number"] + ' : ' + "Date Error\n")
            continue

        # ----- ロット番号をキーとして品名が得られなかった → 空欄で返ってくる -----
        if len(data_dict["key_part_number"]) == 0:
            Log.Log_Error(Log_file, "Lot No Error\n")
            continue

        # ----- STARTTIME_SORTEDの追加 -----

        # 日付をExcel時間に変換する
        date = datetime.strptime(str(data_dict["key_start_date_time"]).replace('T', ' ').replace('.', ':'), "%Y-%m-%d %H:%M:%S")
        date_excel_number = int(str(date - datetime(1899, 12, 30)).split()[0])

        # エピ番号の数値部だけを取得する
        Epi_Number = 0
        for i in data_dict["key_batch_number"]:
            try:
                if 0<=int(i)<=9:
                    Epi_Number = Epi_Number*10+int(i)
            except:
                pass

        # エピ番号を10^6で割って、excel時間に加算する
        date_excel_number += Epi_Number/10**6

        # data_dictに登録する
        data_dict["key_STARTTIME_SORTED"] = date_excel_number
        data_dict["key_SORTNUMBER"] = Epi_Number

        # ----- データ型の確認 -----
        result = Check.Data_Type(key_type, data_dict)
        if result == False:
            Log.Log_Error(Log_file, data_dict["key_serial_number"] + ' : ' + "Data Error\n")
            continue

        # ----- XMLファイルの作成 -----
        xml_file = 'Site=' + Site + ',ProductFamily=' + ProductFamily + ',Operation=' + Operation + \
                   ',Partnumber=' + data_dict["key_part_number"] + ',Serialnumber=' + data_dict["key_serial_number"] + \
                   ',Testdate=' + data_dict["key_start_date_time"].replace(':', '.') + '.xml'

        Output_XML(xml_file, data_dict)


    wb.close()


########## Main処理 ##########
if __name__ == '__main__':

    # ----- ログ書込：Main処理の開始 -----
    Log.Log_Info(Log_file, 'Main Start')

    # ----- path内のフォルダ、ファイルを全部取得 -----
    all_files = os.listdir(Path)

    # ----- ログ記載：着工ファイル検索 -----
    Log.Log_Info(Log_file, 'File Search')

    # ----- 先頭にFBが付いているファイル名と最終更新日時(sec)を格納 -----
    array = []
    for filename in all_files:
        filepath = os.path.join(Path, filename)
        if "FB" in str(filename) and '$' not in str(filename) and os.path.isfile(filepath):
            dt = strftime("%Y-%m-%d %H:%M:%S", localtime(os.path.getctime(filepath)))
            array.append([filepath, dt])

    # ----- 着工ファイルが見つからなかった -----
    if len(array) == 0:
        Log.Log_Info(Log_file, 'Folder Error')
        sys.exit()

    # ----- 最終更新日時順に並び替える -----
    array = sorted(array, key=lambda x: x[1])
    FileName = os.path.basename(array[0][0])
    Log.Log_Info(Log_file, FileName)

    # ----- 前回処理したエピ番号のNumber部分を取り出す -----
    Number = ""
    for i in Before_FileName:
        if "0" <= i <= "9":
            Number += i

    # ----- ファイルの切り替わりを確認 -----
    if Number[4:] not in FileName:

        # ----- ログ記載：過去フォルダ検索 -----
        Log.Log_Info(Log_file, 'Folder Serach')

        # ----- 対象ファイルパスを返す -----
        Old_File_Path = MOCVD_OldFileSearch.F5(Number)

        if Old_File_Path == -1:
            Log.Log_Info(Log_file, 'Old Folder Error')
            sys.exit()

        # ----- ログ記載：シート名の取得 -----
        Log.Log_Info(Log_file, 'OLD Get SheetName')

        # ----- 上記で指定したファイルのシート一覧を取得 -----
        wb = px.load_workbook(Old_File_Path)
        Old_SheetName = wb.sheetnames
        wb.close()

        # ----- ログ記載：前Excelファイルのデータ取得 -----
        Log.Log_Info(Log_file, 'OLD Excel File Get Data')

        # ----- 過去ファイルの処理 -----
        Data_Extract(Old_File_Path, Old_SheetName, 1)

    # ログ記載：Excelファイルのデータ取得
    Log.Log_Info(Log_file, 'Excel File Get Data')

    # ----- arrayに格納されている全てのファイルの処理を行う -----
    for File_Path, _ in array:

        # ----- シート名の取得 -----
        Log.Log_Info(Log_file, 'Get SheetName')

        # ----- 対象ファイルを開き、シートの一覧を取得する -----
        wb = px.load_workbook(File_Path)
        SheetName = wb.sheetnames
        wb.close()

        # ----- 全シートの処理を行う -----
        Data_Extract(File_Path, SheetName, 0)

    # ----- ログ記載：テキストファイルにシート名を上書きで記載する -----
    Log.Log_Info(Log_file, 'Write SheetName')

    # ----- エピ番号の書き込み -----
    with open('F5_FileName_F5-10G.txt', 'w', encoding='utf-8') as textfile:
        textfile.write(FileName)

    # ----- 最終行を書き込んだファイルをGドライブに転送 -----
    shutil.copy("F5_FileName_F5-10G.txt", 'T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開発/018_T2-DML/F5-10G/13_ProgramUsedFile/')


# ----- ログ書込：プログラムの終了 -----
Log.Log_Info(Log_file, 'Program End')