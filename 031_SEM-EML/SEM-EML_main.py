import openpyxl as px
import logging
import shutil
import glob
import xlrd
import sys
import os

from datetime import datetime, timedelta, date


########## サブプログラムの定義 ##########
from SEM_Func import MESA_Width
from SEM_Func import Contact_Degree
from SEM_Func import Contact_Width
from SEM_Func import MESA_Height
from SEM_Func import Ru_Thickness


########## 自作関数の定義 ##########
sys.path.append('../MyModule')
import SQL
import Log
import Convert_Date


########## Logの設定 ##########

# ----- ログフォルダ名を作成 -----
Log_FolderName = str(date.today())

# ----- 格納するLogフォルダがなければ作成する -----
if not os.path.exists("../Log/" + Log_FolderName):
    os.makedirs("../Log/" + Log_FolderName)

# ----- ログファイルの作成 -----
Log_file = '../Log/' + Log_FolderName + '/031_SEM-EML.log'


########## ログ書き込み：プログラムの開始 ##########
Log.Log_Info(Log_file, 'Program Start')


########## 処理対象フォルダのパスを格納 ##########
Folder_Path = [
    'Z:/SEM_PC/先行SEM/製品/HL13B4/',
    'Z:/SEM_PC/先行SEM/製品/HL13B5/',
    'Z:/SEM_PC/先行SEM/製品/HL13B6/',
    'Z:/SEM_PC/先行SEM/製品/HTL13B2/',
    'Z:/SEM_PC/先行SEM/製品/HTL13D1/',
    'Z:/SEM_PC/先行SEM/製品/HL15B5/'
]


########## 処理対象フォルダの処理済みファイル名を格納したテキストファイルパスを格納 ##########
TextFile_Path = [
    'HL13B4.txt',
    'HL13B5.txt',
    'HL13B6.txt',
    'HTL13B2.txt',
    'HTL13D1.txt',
    'HL15B5.txt'
]


########## 空欄チェック ##########
def Get_Cells_Info(filepath):

    # ----- ログ書込：空欄判定 -----
    Log.Log_Info(Log_file, "Blank Check")

    # ----- False -> 空欄がない -----
    is_cells_empty = False

    # ----- ファイルとシートの定義 -----
    wb = xlrd.open_workbook(filepath, on_demand=True)
    sheet = wb.sheet_by_index(0)

    # ----- 旧式ファイルは42行目までデータがないため、エラーとして処理される -----
    try:

        # ----- 一つでも空欄が存在すれば処理を行わない -----
        if sheet.cell(13, 4).value == "" or \
            sheet.cell(14, 4).value == "" or \
            sheet.cell(18, 4).value == "" or \
            sheet.cell(19, 4).value == "" or \
            sheet.cell(20, 4).value == "" or \
            sheet.cell(21, 4).value == "" or \
            sheet.cell(24, 4).value == "" or \
            sheet.cell(25, 4).value == "" or \
            sheet.cell(26, 4).value == "" or \
            sheet.cell(27, 4).value == "" or \
            sheet.cell(30, 4).value == "" or \
            sheet.cell(31, 4).value == "" or \
            sheet.cell(32, 4).value == "" or \
            sheet.cell(33, 4).value == "" or \
            sheet.cell(42, 4).value == "" or \
            sheet.cell(43, 4).value == "" or \
            sheet.cell(44, 4).value == "" or \
            sheet.cell(45, 4).value == "" or \
            sheet.cell(42, 16).value == "" or \
            sheet.cell(43, 16).value == "" or \
            sheet.cell(44, 16).value == "" or \
            sheet.cell(45, 16).value == "":
            is_cells_empty = True

    except IndexError:
        is_cells_empty = True

    wb.release_resources()

    return is_cells_empty


if __name__ == "__main__":


    ########## 対象フォルダの処理 ##########
    for i in range(len(Folder_Path)):

        # ----- ログ書込：Main処理の開始 -----
        Log.Log_Info(Log_file, 'Main Start')

        # ----- 今日までに処理を行ったファイル名を記載したファイルの読み込み -----
        with open(TextFile_Path[i], 'r', encoding='utf-8') as text_file:
            End_FileName = {s.strip() for s in text_file.readlines()}

        # ----- 対象フォルダのファイルをすべて取得 -----
        Log.Log_Info(Log_file, 'Get All Files')
        All_File_List = set(file for file in os.listdir(Folder_Path[i]) if os.path.isfile(os.path.join(Folder_Path[i], file)) and str(file)[-4:] == '.xls')

        # ----- 処理対象ファイルを差集合で取得 -----
        Process_File = All_File_List - End_FileName


        ########## 未処理フォルダ(Process_File)の処理 ##########
        for file in Process_File:

            # ----- ファイルをパス形式にする -----
            file_path = os.path.join(Folder_Path[i], file)

            # ----- ログ書き込み：対象ファイル名 -----
            Log.Log_Info(Log_file, file_path)

            # ----- 空欄判定 -----
            if Get_Cells_Info(file_path):
                Log.Log_Error(Log_file, "Blank Error\n")
                continue

            # ----- ロット名の抜き出し -----
            wb = xlrd.open_workbook(file_path, on_demand=True)
            sheet = wb.sheet_by_index(0)
            serial_number = str(sheet.cell(13, 4).value)

            # ----- 日付を取得し変換を行う -----
            Log.Log_Info(Log_file, 'Date Format Conversion')
            start_date = Convert_Date.Edit_Date(sheet.cell(14, 4).value)

            wb.release_resources()

            # ----- 品名取得のためにPrimeと接続 -----
            conn, cursor = SQL.connSQL()

            # ----- Primeとの接続に失敗した -----
            if conn is None:
                Log.Log_Error(Log_file, serial_number + ' : ' + 'Connection with Prime Failed')
                sys.exit()

            # ----- 品名取得 -----
            part_number, nine_serial_number = SQL.selectSQL(cursor, str(serial_number[:5]))
            SQL.disconnSQL(conn, cursor)

            # ----- 品名がNoneで返ってきた -----
            if part_number is None:
                Log.Log_Error(Log_file, serial_number + ' : ' + "Part Number Error\n")
                End_FileName.add(file)
                continue

            # ----- LDアレイは書き込めないため処理を行わない -----
            if part_number == 'LDアレイ_':
                End_FileName.add(file)
                continue

            # ----- 変換できなかったときは処理を行わない -----
            if start_date == "":
                Log.Log_Error(Log_file, serial_number + ' : ' + "Date Error\n")
                continue

            # ----- 各値取得モジュールに日付、品名、ロット名を渡す -----
            Log.Log_Info(Log_file, 'Excel -> XML')
            MESA_Width.main(file_path, start_date, part_number, serial_number, nine_serial_number)
            Contact_Degree.main(file_path, start_date, part_number, serial_number, nine_serial_number)
            Contact_Width.main(file_path, start_date, part_number, serial_number, nine_serial_number)
            MESA_Height.main(file_path, start_date, part_number, serial_number, nine_serial_number)
            Ru_Thickness.main(file_path, start_date, part_number, serial_number, nine_serial_number)

            Log.Log_Info(Log_file, serial_number + ' : ' + "OK\n")

            # ----- テキストファイルに処理を行ったファイルパスを書き込む -----
            End_FileName.add(file)

        # ----- テキストファイルに処理が終わったファイル一覧を書き込む -----
        Log.Log_Info(Log_file, 'Write the End Sheet Name')
        End_FileName_list = sorted(list(End_FileName))
        End_FileName_str = '\n'.join(End_FileName_list)
        with open(TextFile_Path[i], 'w', encoding='utf-8') as text_file:
            text_file.write(End_FileName_str)

        # ----- ログ書込：Main処理の終了 -----
        Log.Log_Info(Log_file, 'Main Program End')


    # ----- 最終行を書き込んだファイルをGドライブに転送 -----
    shutil.copy(TextFile_Path[i], 'T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開発/031_SEM-EML/13_ProgramUsedFile/')


# ----- ログ記載：プログラムの終了 -----
Log.Log_Info(Log_file, 'Program End')