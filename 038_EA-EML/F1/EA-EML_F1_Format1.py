import openpyxl as px
import logging
import shutil
import pyodbc
import xlrd
import glob
import sys
import os

from datetime import datetime, timedelta, date
from time import strftime, localtime


########## 自作関数の定義 ##########
sys.path.append('../../MyModule')
import SQL
import Log
import ExpandExp
import Convert_Date
import Row_Number_Func
import MOCVD_OldFileSearch
import Check


########## 全体パラメータの定義 ##########
Site = '350'
ProductFamily = 'SAG FAB'
Operation = 'EA-EML_F1_Format1'
TestStation = 'EA-EML'
X = '999999'
Y = '999999'


########## Logフォルダ名の定義 ##########
Log_FolderName = str(date.today())

# ----- 格納するLogフォルダがなければ作成する -----
if not os.path.exists("../../Log/" + Log_FolderName):
    os.makedirs("../../Log/" + Log_FolderName)

# ----- ログ書き込み先ファイルパス -----
Log_file = '../../Log/' + Log_FolderName + '/038_EA-EML_F1_Format1.log'

# ----- ログ書込：プログラムの開始 -----
Log.Log_Info(Log_file, 'Program Start')


########## 処理ファイルのあるディレクトリ定義 ##########
Path = 'Z:/MOCVD/MOCVD過去プログラム/F1炉/'
# Path = 'C:/Users/hor78296/Desktop/F1炉/'

########## XML出力先ファイルパス ##########
Output_filepath = '//li.lumentuminc.net/data/SAG/TDS/Data/Files to Insert/XML/'
# Output_filepath = '../../XML/038_EA-EML/F1_Format1/'


########## TestStepの定義 ##########
teststep_dict = {
    'TestStep1' : 'Coordinate',
    'TestStep2' : 'TargetWavelength',
    'TestStep3' : 'Thickness',
    'TestStep4' : 'XRayDiffraction',
    'TestStep5' : 'InPBoardLot',
    'TestStep6' : 'Adjust',
    'TestStep7' : 'MeasurementConditions',
    'TestStep8' : 'Dulation',
    'TestStep9' : 'MO1-TMI',
    'TestStep10' : 'MO3-TMI',
    'TestStep11' : 'MO4-TEG',
    'TestStep12' : 'MO5-TMI',
    'TestStep13' : 'MO6-TEG',
    'TestStep14' : 'AsH3-5percent',
    'TestStep15' : 'PH3-A-50percent',
    'TestStep16' : 'Si2H6-10ppm',
    'TestStep17' : 'DMZn-B-01percent',
    'TestStep18' : 'GrowthTemperature',
    'TestStep19' : 'LayerNo',
    'TestStep20' : 'Comment',
    'TestStep21' : 'Thickness_Step',
    'TestStep22' : 'CarrierConcentration',
    'TestStep23' : 'Piezocon',
    'TestStep24' : 'BallastN2',
    'TestStep25' : 'MO-Temperature',
    'TestStep26' : 'SORTED_DATA',
}


########## HeaderMiscの定義 ##########
HeaderMisc_dict = {
    'HeaderMisc1' : 'RecipeName-Macro',
    'HeaderMisc2' : 'RecipeName-Program',
    'HeaderMisc3' : 'RecipeName-Folder'
}


########## 取得した項目と型の対応表を定義 ##########
key_type = {
    "key_start_date_time": str,
    "key_serial_number": str,
    "key_part_number": str,
    "key_operator": str,
    "key_batch_number": str,
    "key_HeaderMisc1": str,
    "key_HeaderMisc2": str,
    "key_HeaderMisc3": str,
    "key_TargetWavelength_TargetWavelength": float,
    "key_Thickness_Thickness_Cap": float,
    "key_Thickness_Thickness_Core": float,
    "key_Thickness_Thickness_Total": float,
    "key_XRayDiffraction_Xray_Thickness": float,
    "key_XRayDiffraction_Xray_Strain": float,
    "key_InPBoardLot_InPBoardLot_No": str,
    "key_InPBoardLot_InPBoardLot_CC": float,
    "key_InPBoardLot_InPBoardLot_EPD": float,
    "key_Adjust_PL_Wavelength": float,
    "key_Adjust_PL_Intensity": float,
    "key_Adjust_PL_FWHM": float,
    "key_Adjust_PL_Adjust": float,
    "key_MeasurementConditions_Templature": float,
    "key_MeasurementConditions_Humidity": float,
    "key_MeasurementConditions_LaserSideFilter1": str,
    "key_MeasurementConditions_LaserSideFilter2": float,
    "key_MeasurementConditions_Zvalue": float,
    "key_MeasurementConditions_PL_IntensityRate_A": float,
    "key_MeasurementConditions_PL_IntensityRate_B": float,
    "key_MeasurementConditions_PL_IntensityRate_C": float,
    "key_MeasurementConditions_PL_IntensityRate_D": float,
    "key_Dulation_Step7": float,
    "key_Dulation_Step9": float,
    "key_Dulation_Step13": float,
    "key_Dulation_Step17": float,
    "key_Dulation_Step21": float,
    "key_Dulation_Step25": float,
    "key_Dulation_Step29": float,
    "key_Dulation_Step33": float,
    "key_Dulation_Step37": float,
    "key_Dulation_Step41": float,
    "key_Dulation_Step45": float,
    "key_Dulation_Step46": float,
    "key_Dulation_Step48": float,
    "key_Dulation_Step49": float,
    "key_Dulation_Step51": float,
    "key_MO1-TMI_Step13": float,
    "key_MO1-TMI_Step17": float,
    "key_MO1-TMI_Step33": float,
    "key_MO1-TMI_Step37": float,
    "key_MO1-TMI_Step41": float,
    "key_MO3-TMI_Step21": float,
    "key_MO3-TMI_Step25": float,
    "key_MO3-TMI_Step29": float,
    "key_MO4-TEG_Step25": float,
    "key_MO4-TEG_Step33": float,
    "key_MO5-TMI_Step9": float,
    "key_MO5-TMI_Step45": float,
    "key_MO5-TMI_Step46": float,
    "key_MO6-TEG_Step13": float,
    "key_MO6-TEG_Step17": float,
    "key_MO6-TEG_Step21": float,
    "key_MO6-TEG_Step29": float,
    "key_MO6-TEG_Step37": float,
    "key_MO6-TEG_Step41": float,
    "key_AsH3-5percent_Step13": float,
    "key_AsH3-5percent_Step17": float,
    "key_AsH3-5percent_Step21": float,
    "key_AsH3-5percent_Step25": float,
    "key_AsH3-5percent_Step29": float,
    "key_AsH3-5percent_Step33": float,
    "key_AsH3-5percent_Step37": float,
    "key_AsH3-5percent_Step41": float,
    "key_PH3-A-50percent_Step7": float,
    "key_PH3-A-50percent_Step9": float,
    "key_PH3-A-50percent_Step13": float,
    "key_PH3-A-50percent_Step17": float,
    "key_PH3-A-50percent_Step21": float,
    "key_PH3-A-50percent_Step25": float,
    "key_PH3-A-50percent_Step29": float,
    "key_PH3-A-50percent_Step33": float,
    "key_PH3-A-50percent_Step37": float,
    "key_PH3-A-50percent_Step41": float,
    "key_PH3-A-50percent_Step45": float,
    "key_PH3-A-50percent_Step46": float,
    "key_PH3-A-50percent_Step48": float,
    "key_Si2H6-10ppm_Step9": float,
    "key_Si2H6-10ppm_Step13": float,
    "key_Si2H6-10ppm_Step17": float,
    "key_DMZn-B-01percent_Step46": float,
    "key_DMZn-B-01percent_Step51": float,
    "key_GrowthTemperature_Step7": float,
    "key_GrowthTemperature_Step9": float,
    "key_GrowthTemperature_Step13": float,
    "key_GrowthTemperature_Step17": float,
    "key_GrowthTemperature_Step21": float,
    "key_GrowthTemperature_Step25": float,
    "key_GrowthTemperature_Step29": float,
    "key_GrowthTemperature_Step33": float,
    "key_GrowthTemperature_Step37": float,
    "key_GrowthTemperature_Step41": float,
    "key_GrowthTemperature_Step45": float,
    "key_GrowthTemperature_Step46": float,
    "key_GrowthTemperature_Step48": float,
    "key_GrowthTemperature_Step49": float,
    "key_LayerNo_Step9": str,
    "key_LayerNo_Step13": str,
    "key_LayerNo_Step17": str,
    "key_LayerNo_Step21": str,
    "key_LayerNo_Step25": str,
    "key_LayerNo_Step29": str,
    "key_LayerNo_Step33": str,
    "key_LayerNo_Step37": str,
    "key_LayerNo_Step41": str,
    "key_LayerNo_Step45": str,
    "key_LayerNo_Step46": str,
    "key_Comment_Step7": str,
    "key_Comment_Step9": str,
    "key_Comment_Step13": str,
    "key_Comment_Step17": str,
    "key_Comment_Step21": str,
    "key_Comment_Step25": str,
    "key_Comment_Step29": str,
    "key_Comment_Step33": str,
    "key_Comment_Step37": str,
    "key_Comment_Step41": str,
    "key_Comment_Step45": str,
    "key_Comment_Step46": str,
    "key_Comment_Step48": str,
    "key_Comment_Step49": str,
    "key_Thickness_Step_Step9": float,
    "key_Thickness_Step_Step13": float,
    "key_Thickness_Step_Step17": float,
    "key_Thickness_Step_Step21": float,
    "key_Thickness_Step_Step25": float,
    "key_Thickness_Step_Step29": float,
    "key_Thickness_Step_Step33": float,
    "key_Thickness_Step_Step37": float,
    "key_Thickness_Step_Step41": float,
    "key_Thickness_Step_Step45": float,
    "key_Thickness_Step_Step46": float,
    "key_CarrierConcentration_Step9": float,
    "key_CarrierConcentration_Step13": float,
    "key_CarrierConcentration_Step17": float,
    "key_CarrierConcentration_Step46": float,
    "key_Piezocon_F1": float,
    "key_Piezocon_F1_Inverse": float,
    "key_BallastN2_BallastN2": float,
    "key_MO-Temperature_MO1-TMI": float,
    "key_MO-Temperature_MO2-TEG": float,
    "key_MO-Temperature_MO3-TMI": float,
    "key_MO-Temperature_MO4-TEG": float,
    "key_MO-Temperature_MO5-TMI": float,
    "key_MO-Temperature_MO6-TEG": float,
    "key_MO-Temperature_Unused": float,
    "key_STARTTIME_SORTED": float,
    "key_SORTNUMBER" : float,
    "key_LotNumber_9": str
}


########## 対象ロット番号のイニシャルを記載したファイルを取得する ##########
Log.Log_Info(Log_file, 'Get SerialNumber Initial List ')
with open('T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開発/004_T2-EML/SerialNumber_Initial.txt', 'r', encoding='utf-8') as textfile:
#with open(r'C:\Users\hsi67063\Downloads\SerialNumber_Initial.txt', 'r', encoding='utf-8') as textfile:    
    SerialNumber_list = {s.strip() for s in textfile.readlines()}


########## 前回処理を行ったファイル名を取得する ##########
with open('F1_FileName_Format1.txt', 'r', encoding='utf-8') as textfile:
    Before_FileName = textfile.readline()


########## 空欄チェック ##########
def Get_Cells_Info(Sheet):

    # ----- ログ書込：空欄判定 -----
    Log.Log_Info(Log_file, "Blank Check")

    # ----- False -> 空欄がない -----
    is_cells_empty = False

    # ----- 日付かエピ番号かバラスト流量が空欄の場合は処理を行わない -----
    if  Sheet['I8'].value is None or Sheet['Q7'].value is None or Sheet['AE40'].value is None:
        is_cells_empty = True

    return is_cells_empty


########## データの取得 ##########
def Open_Data_Sheet(Sheet, filepath, SheetName):

    # ----- ログ書込：データ取得 -----
    Log.Log_Info(Log_file, 'Data Acquisition')

    # ----- 辞書の作成 -----
    data_dict = dict()

    # ----- ロット番号の取得 -----
    Serial_number = Sheet["M8"].value
    conn, cursor = SQL.connSQL()

    # ----- Prime接続できなかったときはNoneが返ってくる -----
    if conn is None:
        Log.Log_Error(Log_file, serial_number + ' : ' + 'Connection with Prime Failed')
        sys.exit()

    # ----- 品名を取得 -----
    Part_number, Nine_Serial_Number = SQL.selectSQL(cursor, Serial_number)
    SQL.disconnSQL(conn, cursor)

    # ----- SEM, XRD, MOCVD の装置Noを取得 -----
    SEM, XRD, MOCVD = '1', '1', '1'
    if '#2' in str(Sheet["J67"].value):
        SEM = '2'
    if '#2' in str(Sheet["J44"].value):
        XRD = '2'

    # ----- 項目の取得 -----
    data_dict = {
        "key_start_date_time": str(Sheet["Q7"].value).replace(" ", "T"),
        "key_serial_number": Serial_number,
        "key_part_number": Part_number,
        "key_LotNumber_9": Nine_Serial_Number,
        "key_operator": Sheet["V8"].value,
        "key_batch_number": Sheet["I8"].value,
        "key_HeaderMisc1": Sheet["U3"].value,
        "key_HeaderMisc2": Sheet["U4"].value,
        "key_HeaderMisc3": Sheet["U5"].value,
        "key_TargetWavelength_TargetWavelength": Sheet['M12'].value,
        "key_Thickness_Thickness_Cap": Sheet['M68'].value,
        "key_Thickness_Thickness_Core": Sheet['M69'].value,
        "key_Thickness_Thickness_Total": Sheet['M70'].value,
        "key_XRayDiffraction_Xray_Thickness": Sheet['M46'].value,
        "key_XRayDiffraction_Xray_Strain": Sheet['M47'].value,
        "key_InPBoardLot_InPBoardLot_No": Sheet['AA8'].value,
        "key_InPBoardLot_InPBoardLot_CC": Sheet['AC9'].value,
        "key_InPBoardLot_InPBoardLot_EPD": Sheet['AF9'].value,
        "key_Adjust_PL_Wavelength": Sheet['L52'].value,
        "key_Adjust_PL_Intensity": Sheet['N52'].value,
        "key_Adjust_PL_FWHM": Sheet['P52'].value,
        "key_Adjust_PL_Adjust": Sheet['L54'].value,
        "key_MeasurementConditions_Templature": Sheet['AA52'].value,
        "key_MeasurementConditions_Humidity": Sheet['AA53'].value,
        "key_MeasurementConditions_LaserSideFilter1": Sheet['AA58'].value,
        "key_MeasurementConditions_LaserSideFilter2": Sheet['AA59'].value,
        "key_MeasurementConditions_Zvalue": Sheet['AA60'].value,
        "key_MeasurementConditions_PL_IntensityRate_A": Sheet['AD58'].value,
        "key_MeasurementConditions_PL_IntensityRate_B": Sheet['AD59'].value,
        "key_MeasurementConditions_PL_IntensityRate_C": Sheet['AD60'].value,
        "key_MeasurementConditions_PL_IntensityRate_D": Sheet['AD61'].value,
        "key_Dulation_Step7": Sheet['I20'].value,
        "key_Dulation_Step9": Sheet['I21'].value,
        "key_Dulation_Step13": Sheet['I22'].value,
        "key_Dulation_Step17": Sheet['I23'].value,
        "key_Dulation_Step21": Sheet['I24'].value,
        "key_Dulation_Step25": Sheet['I25'].value,
        "key_Dulation_Step29": Sheet['I26'].value,
        "key_Dulation_Step33": Sheet['I27'].value,
        "key_Dulation_Step37": Sheet['I28'].value,
        "key_Dulation_Step41": Sheet['I29'].value,
        "key_Dulation_Step45": Sheet['I30'].value,
        "key_Dulation_Step46": Sheet['I31'].value,
        "key_Dulation_Step48": Sheet['I32'].value,
        "key_Dulation_Step49": Sheet['I33'].value,
        "key_Dulation_Step51": Sheet['I34'].value,
        "key_MO1-TMI_Step13": Sheet['K22'].value,
        "key_MO1-TMI_Step17": Sheet['K23'].value,
        "key_MO1-TMI_Step33": Sheet['K27'].value,
        "key_MO1-TMI_Step37": Sheet['K28'].value,
        "key_MO1-TMI_Step41": Sheet['K29'].value,
        "key_MO3-TMI_Step21": Sheet['M24'].value,
        "key_MO3-TMI_Step25": Sheet['M25'].value,
        "key_MO3-TMI_Step29": Sheet['M26'].value,
        "key_MO4-TEG_Step25": Sheet['N25'].value,
        "key_MO4-TEG_Step33": Sheet['N27'].value,
        "key_MO5-TMI_Step9": Sheet['O21'].value,
        "key_MO5-TMI_Step45": Sheet['O30'].value,
        "key_MO5-TMI_Step46": Sheet['O31'].value,
        "key_MO6-TEG_Step13": Sheet['P22'].value,
        "key_MO6-TEG_Step17": Sheet['P23'].value,
        "key_MO6-TEG_Step21": Sheet['P24'].value,
        "key_MO6-TEG_Step29": Sheet['P26'].value,
        "key_MO6-TEG_Step37": Sheet['P28'].value,
        "key_MO6-TEG_Step41": Sheet['P29'].value,
        "key_AsH3-5percent_Step13": Sheet['T22'].value,
        "key_AsH3-5percent_Step17": Sheet['T23'].value,
        "key_AsH3-5percent_Step21": Sheet['T24'].value,
        "key_AsH3-5percent_Step25": Sheet['T25'].value,
        "key_AsH3-5percent_Step29": Sheet['T26'].value,
        "key_AsH3-5percent_Step33": Sheet['T27'].value,
        "key_AsH3-5percent_Step37": Sheet['T28'].value,
        "key_AsH3-5percent_Step41": Sheet['T29'].value,
        "key_PH3-A-50percent_Step7": Sheet['U20'].value,
        "key_PH3-A-50percent_Step9": Sheet['U21'].value,
        "key_PH3-A-50percent_Step13": Sheet['U22'].value,
        "key_PH3-A-50percent_Step17": Sheet['U23'].value,
        "key_PH3-A-50percent_Step21": Sheet['U24'].value,
        "key_PH3-A-50percent_Step25": Sheet['U25'].value,
        "key_PH3-A-50percent_Step29": Sheet['U26'].value,
        "key_PH3-A-50percent_Step33": Sheet['U27'].value,
        "key_PH3-A-50percent_Step37": Sheet['U28'].value,
        "key_PH3-A-50percent_Step41": Sheet['U29'].value,
        "key_PH3-A-50percent_Step45": Sheet['U30'].value,
        "key_PH3-A-50percent_Step46": Sheet['U31'].value,
        "key_PH3-A-50percent_Step48": Sheet['U32'].value,
        "key_Si2H6-10ppm_Step9": Sheet['W21'].value,
        "key_Si2H6-10ppm_Step13": Sheet['W22'].value,
        "key_Si2H6-10ppm_Step17": Sheet['W23'].value,
        "key_DMZn-B-01percent_Step46": Sheet['Y31'].value,
        "key_DMZn-B-01percent_Step51": Sheet['Y34'].value,
        "key_GrowthTemperature_Step7": Sheet['Z20'].value,
        "key_GrowthTemperature_Step9": Sheet['Z21'].value,
        "key_GrowthTemperature_Step13": Sheet['Z22'].value,
        "key_GrowthTemperature_Step17": Sheet['Z23'].value,
        "key_GrowthTemperature_Step21": Sheet['Z24'].value,
        "key_GrowthTemperature_Step25": Sheet['Z25'].value,
        "key_GrowthTemperature_Step29": Sheet['Z26'].value,
        "key_GrowthTemperature_Step33": Sheet['Z27'].value,
        "key_GrowthTemperature_Step37": Sheet['Z28'].value,
        "key_GrowthTemperature_Step41": Sheet['Z29'].value,
        "key_GrowthTemperature_Step45": Sheet['Z30'].value,
        "key_GrowthTemperature_Step46": Sheet['Z31'].value,
        "key_GrowthTemperature_Step48": Sheet['Z32'].value,
        "key_GrowthTemperature_Step49": Sheet['Z33'].value,
        "key_LayerNo_Step9": Sheet['AA21'].value,
        "key_LayerNo_Step13": Sheet['AA22'].value,
        "key_LayerNo_Step17": Sheet['AA23'].value,
        "key_LayerNo_Step21": Sheet['AA24'].value,
        "key_LayerNo_Step25": Sheet['AA25'].value,
        "key_LayerNo_Step29": Sheet['AA26'].value,
        "key_LayerNo_Step33": Sheet['AA27'].value,
        "key_LayerNo_Step37": Sheet['AA28'].value,
        "key_LayerNo_Step41": Sheet['AA29'].value,
        "key_LayerNo_Step45": Sheet['AA30'].value,
        "key_LayerNo_Step46": Sheet['AA31'].value,
        "key_Comment_Step7": Sheet['AB20'].value,
        "key_Comment_Step9": Sheet['AB21'].value,
        "key_Comment_Step13": Sheet['AB22'].value,
        "key_Comment_Step17": Sheet['AB23'].value,
        "key_Comment_Step21": Sheet['AB24'].value,
        "key_Comment_Step25": Sheet['AB25'].value,
        "key_Comment_Step29": Sheet['AB26'].value,
        "key_Comment_Step33": Sheet['AB27'].value,
        "key_Comment_Step37": Sheet['AB28'].value,
        "key_Comment_Step41": Sheet['AB29'].value,
        "key_Comment_Step45": Sheet['AB30'].value,
        "key_Comment_Step46": Sheet['AB31'].value,
        "key_Comment_Step48": Sheet['AB32'].value,
        "key_Comment_Step49": Sheet['AB33'].value,
        "key_Thickness_Step_Step9": Sheet['AF21'].value,
        "key_Thickness_Step_Step13": Sheet['AF22'].value,
        "key_Thickness_Step_Step17": Sheet['AF23'].value,
        "key_Thickness_Step_Step21": Sheet['AF24'].value,
        "key_Thickness_Step_Step25": Sheet['AF25'].value,
        "key_Thickness_Step_Step29": Sheet['AF26'].value,
        "key_Thickness_Step_Step33": Sheet['AF27'].value,
        "key_Thickness_Step_Step37": Sheet['AF28'].value,
        "key_Thickness_Step_Step41": Sheet['AF29'].value,
        "key_Thickness_Step_Step45": Sheet['AF30'].value,
        "key_Thickness_Step_Step46": Sheet['AF31'].value,
        "key_CarrierConcentration_Step9": Sheet['AI21'].value,
        "key_CarrierConcentration_Step13": Sheet['AI22'].value,
        "key_CarrierConcentration_Step17": Sheet['AI23'].value,
        "key_CarrierConcentration_Step46": Sheet['AI31'].value,
        "key_Piezocon_F1": Sheet['AE36'].value,
        "key_Piezocon_F1_Inverse": Sheet['AE37'].value,
        "key_BallastN2_BallastN2": Sheet['AE40'].value,
        "key_MO-Temperature_MO1-TMI": Sheet['K42'].value,
        "key_MO-Temperature_MO2-TEG": Sheet['L42'].value,
        "key_MO-Temperature_MO3-TMI": Sheet['M42'].value,
        "key_MO-Temperature_MO4-TEG": Sheet['N42'].value,
        "key_MO-Temperature_MO5-TMI": Sheet['O42'].value,
        "key_MO-Temperature_MO6-TEG": Sheet['P42'].value,
        "key_MO-Temperature_Unused": Sheet['Q42'].value,
        "key_TestEquipment_SEM": SEM,
        "key_TestEquipment_PLmapper": '1',
        "key_TestEquipment_XRD": XRD,
        "key_TestEquipment_MOCVD": MOCVD
    }

    # ----- 空欄箇所はNoneとして取得される。Noneは文字列に変換できないため、空欄("")に置き換える -----
    for keys in data_dict:
        if data_dict[keys] is None or data_dict[keys] == "None" or (data_dict[keys] == '-' and keys != 'key_operator'):
            data_dict[keys] = ""
        # ----- 指数表記を展開する -----
        if type(data_dict[keys]) is float and 'e' in str(data_dict[keys]) and keys != "key_start_date_time":
            data_dict[keys] = ExpandExp.Expand(data_dict[keys])

    return data_dict


########## XML変換 ##########
def Output_XML(xml_file, data_dict):

    ########## ログ書込：XML変換 ##########
    Log.Log_Info(Log_file, 'Excel File To XML File Conversion')
    
    f = open(Output_filepath + xml_file, 'w', encoding="utf-8")

    f.write('<?xml version="1.0" encoding="utf-8"?>' + '\n' +
            '<Results xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema">' + '\n' +
            '       <Result startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Result="Passed">' + '\n' +
            '               <Header SerialNumber=' + '"' + data_dict["key_serial_number"] + '"' + ' PartNumber=' + '"' + data_dict["key_part_number"] + '"' + ' Operation=' + '"' + Operation + '"' + ' TestStation=' + '"' + TestStation + '"' + ' Operator=' + '"' + data_dict["key_operator"] + '"' + ' StartTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Site=' + '"' + Site + '"' + ' BatchNumber=' + '"' + data_dict["key_batch_number"] + '"' + ' LotNumber=' + '"' + data_dict["key_serial_number"] + '"/>' + '\n' +
            '               <HeaderMisc>' + '\n' +
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc1"] + '">' + data_dict["key_HeaderMisc1"] + '</Item>' + '\n'
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc2"] + '">' + data_dict["key_HeaderMisc2"] + '</Item>' + '\n'
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc3"] + '">' + data_dict["key_HeaderMisc3"] + '</Item>' + '\n'
            '               </HeaderMisc>' + '\n' +
            '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep1"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + X + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + Y + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep2"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="TargetWavelength" Units="nm" Value=' + '"' + str(data_dict["key_TargetWavelength_TargetWavelength"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep3"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Thickness_Cap" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Thickness_Cap"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Thickness_Core" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Thickness_Core"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Thickness_Total" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Thickness_Total"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep4"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Xray_Thickness" Units="nm" Value=' + '"' + str(data_dict["key_XRayDiffraction_Xray_Thickness"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Xray_Strain" Units="percent" Value=' + '"' + str(data_dict["key_XRayDiffraction_Xray_Strain"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep5"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="String" Name="InPBoardLot_No" Value=' + '"' + str(data_dict["key_InPBoardLot_InPBoardLot_No"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="InPBoardLot_CC" Units="nm" Value=' + '"' + str(data_dict["key_InPBoardLot_InPBoardLot_CC"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="InPBoardLot_EPD" Units="nm" Value=' + '"' + str(data_dict["key_InPBoardLot_InPBoardLot_EPD"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep6"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="PL_Wavelength" Units="nm" Value=' + '"' + str(data_dict["key_Adjust_PL_Wavelength"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="PL_Intensity" Units="mV" Value=' + '"' + str(data_dict["key_Adjust_PL_Intensity"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="PL_FWHM" Units="meV" Value=' + '"' + str(data_dict["key_Adjust_PL_FWHM"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="PL_Adjust" Units="nm" Value=' + '"' + str(data_dict["key_Adjust_PL_Adjust"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep7"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Templature" Units="degree" Value=' + '"' + str(data_dict["key_MeasurementConditions_Templature"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Humidity" Units="percent" Value=' + '"' + str(data_dict["key_MeasurementConditions_Humidity"]) + '"/>' + '\n' +
            '                   <Data DataType="String" Name="LaserSideFilter1" Value=' + '"' + str(data_dict["key_MeasurementConditions_LaserSideFilter1"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="LaserSideFilter2" Units="percent" Value=' + '"' + str(data_dict["key_MeasurementConditions_LaserSideFilter2"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Zvalue" Units="um" Value=' + '"' + str(data_dict["key_MeasurementConditions_Zvalue"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="PL_IntensityRate_A" Units="percent" Value=' + '"' + str(data_dict["key_MeasurementConditions_PL_IntensityRate_A"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="PL_IntensityRate_B" Units="percent" Value=' + '"' + str(data_dict["key_MeasurementConditions_PL_IntensityRate_B"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="PL_IntensityRate_C" Units="percent" Value=' + '"' + str(data_dict["key_MeasurementConditions_PL_IntensityRate_C"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="PL_IntensityRate_D" Units="percent" Value=' + '"' + str(data_dict["key_MeasurementConditions_PL_IntensityRate_D"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep8"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step7" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step7"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step9" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step17" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step17"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step21" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step21"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step25" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step25"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step29" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step29"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step33" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step33"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step37" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step37"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step41" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step41"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step45" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step45"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step46" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step46"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step48" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step48"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step49" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step49"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step51" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step51"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep9"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="sccm" Value=' + '"' + str(data_dict["key_MO1-TMI_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step17" Units="sccm" Value=' + '"' + str(data_dict["key_MO1-TMI_Step17"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step33" Units="sccm" Value=' + '"' + str(data_dict["key_MO1-TMI_Step33"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step37" Units="sccm" Value=' + '"' + str(data_dict["key_MO1-TMI_Step37"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step41" Units="sccm" Value=' + '"' + str(data_dict["key_MO1-TMI_Step41"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep10"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step21" Units="sccm" Value=' + '"' + str(data_dict["key_MO3-TMI_Step21"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step25" Units="sccm" Value=' + '"' + str(data_dict["key_MO3-TMI_Step25"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step29" Units="sccm" Value=' + '"' + str(data_dict["key_MO3-TMI_Step29"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep11"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step25" Units="sccm" Value=' + '"' + str(data_dict["key_MO4-TEG_Step25"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step33" Units="sccm" Value=' + '"' + str(data_dict["key_MO4-TEG_Step33"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep12"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step9" Units="sccm" Value=' + '"' + str(data_dict["key_MO5-TMI_Step9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step45" Units="sccm" Value=' + '"' + str(data_dict["key_MO5-TMI_Step45"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step46" Units="sccm" Value=' + '"' + str(data_dict["key_MO5-TMI_Step46"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep13"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="sccm" Value=' + '"' + str(data_dict["key_MO6-TEG_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step17" Units="sccm" Value=' + '"' + str(data_dict["key_MO6-TEG_Step17"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step21" Units="sccm" Value=' + '"' + str(data_dict["key_MO6-TEG_Step21"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step29" Units="sccm" Value=' + '"' + str(data_dict["key_MO6-TEG_Step29"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step37" Units="sccm" Value=' + '"' + str(data_dict["key_MO6-TEG_Step37"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step41" Units="sccm" Value=' + '"' + str(data_dict["key_MO6-TEG_Step41"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep14"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-5percent_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step17" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-5percent_Step17"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step21" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-5percent_Step21"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step25" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-5percent_Step25"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step29" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-5percent_Step29"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step33" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-5percent_Step33"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step37" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-5percent_Step37"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step41" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-5percent_Step41"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep15"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step7" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-A-50percent_Step7"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step9" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-A-50percent_Step9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-A-50percent_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step17" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-A-50percent_Step17"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step21" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-A-50percent_Step21"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step25" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-A-50percent_Step25"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step29" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-A-50percent_Step29"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step33" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-A-50percent_Step33"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step37" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-A-50percent_Step37"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step41" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-A-50percent_Step41"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step45" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-A-50percent_Step45"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step46" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-A-50percent_Step46"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step48" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-A-50percent_Step48"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep16"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step9" Units="sccm" Value=' + '"' + str(data_dict["key_Si2H6-10ppm_Step9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="sccm" Value=' + '"' + str(data_dict["key_Si2H6-10ppm_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step17" Units="sccm" Value=' + '"' + str(data_dict["key_Si2H6-10ppm_Step17"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep17"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step46" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-B-01percent_Step46"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step51" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-B-01percent_Step51"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep18"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step7" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step7"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step9" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step17" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step17"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step21" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step21"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step25" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step25"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step29" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step29"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step33" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step33"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step37" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step37"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step41" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step41"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step45" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step45"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step46" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step46"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step48" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step48"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step49" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step49"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep19"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="String" Name="Step9" Value=' + '"' + str(data_dict["key_LayerNo_Step9"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step13" Value=' + '"' + str(data_dict["key_LayerNo_Step13"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step17" Value=' + '"' + str(data_dict["key_LayerNo_Step17"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step21" Value=' + '"' + str(data_dict["key_LayerNo_Step21"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step25" Value=' + '"' + str(data_dict["key_LayerNo_Step25"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step29" Value=' + '"' + str(data_dict["key_LayerNo_Step29"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step33" Value=' + '"' + str(data_dict["key_LayerNo_Step33"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step37" Value=' + '"' + str(data_dict["key_LayerNo_Step37"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step41" Value=' + '"' + str(data_dict["key_LayerNo_Step41"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step45" Value=' + '"' + str(data_dict["key_LayerNo_Step45"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step46" Value=' + '"' + str(data_dict["key_LayerNo_Step46"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep20"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="String" Name="Step7" Value=' + '"' + str(data_dict["key_Comment_Step7"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step9" Value=' + '"' + str(data_dict["key_Comment_Step9"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step13" Value=' + '"' + str(data_dict["key_Comment_Step13"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step17" Value=' + '"' + str(data_dict["key_Comment_Step17"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step21" Value=' + '"' + str(data_dict["key_Comment_Step21"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step25" Value=' + '"' + str(data_dict["key_Comment_Step25"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step29" Value=' + '"' + str(data_dict["key_Comment_Step29"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step33" Value=' + '"' + str(data_dict["key_Comment_Step33"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step37" Value=' + '"' + str(data_dict["key_Comment_Step37"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step41" Value=' + '"' + str(data_dict["key_Comment_Step41"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step45" Value=' + '"' + str(data_dict["key_Comment_Step45"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step46" Value=' + '"' + str(data_dict["key_Comment_Step46"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step48" Value=' + '"' + str(data_dict["key_Comment_Step48"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step49" Value=' + '"' + str(data_dict["key_Comment_Step49"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep21"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step9" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step17" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step17"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step21" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step21"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step25" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step25"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step29" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step29"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step33" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step33"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step37" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step37"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step41" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step41"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step45" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step45"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step46" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step46"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep22"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step9" Units="cm-3" Value=' + '"' + str(data_dict["key_CarrierConcentration_Step9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="cm-3" Value=' + '"' + str(data_dict["key_CarrierConcentration_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step17" Units="cm-3" Value=' + '"' + str(data_dict["key_CarrierConcentration_Step17"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step46" Units="cm-3" Value=' + '"' + str(data_dict["key_CarrierConcentration_Step46"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep23"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="F1" Units="percent" Value=' + '"' + str(data_dict["key_Piezocon_F1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="F1_Inverse" Units="percent" Value=' + '"' + str(data_dict["key_Piezocon_F1_Inverse"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep24"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="BallastN2" Units="slm" Value=' + '"' + str(data_dict["key_BallastN2_BallastN2"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep25"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="MO1-TMI" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO1-TMI"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO2-TEG" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO2-TEG"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO3-TMI" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO3-TMI"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO4-TEG" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO4-TEG"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO5-TMI" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO5-TMI"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO6-TEG" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO6-TEG"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Unused" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_Unused"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep26"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="STARTTIME_SORTED" Units="" Value=' + '"' + str(data_dict["key_STARTTIME_SORTED"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="SORTNUMBER" Units="" Value=' + '"' + str(data_dict["key_SORTNUMBER"]) + '"/>' + '\n' +
            '                   <Data DataType="String" Name="LotNumber_5" Value=' + '"' + str(data_dict["key_serial_number"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="LotNumber_9" Value=' + '"' + str(data_dict["key_LotNumber_9"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '\n'
            '               <TestEquipment>' + '\n' +
            '                   <Item DeviceName="SEM" DeviceSerialNumber="' + data_dict["key_TestEquipment_SEM"] + '"/>' + '\n' +
            '                   <Item DeviceName="PLmapper" DeviceSerialNumber="' + data_dict["key_TestEquipment_PLmapper"] + '"/>' + '\n' +
            '                   <Item DeviceName="XRD" DeviceSerialNumber="' + data_dict["key_TestEquipment_XRD"] + '"/>' + '\n' +
            '                   <Item DeviceName="MOCVD" DeviceSerialNumber="' + data_dict["key_TestEquipment_MOCVD"] + '"/>' + '\n' +
            '               </TestEquipment>' + '\n' +
            '\n'
            '               <ErrorData/>' + '\n' +
            '               <FailureData/>' + '\n' +
            '               <Configuration/>' + '\n' +
            '       </Result>' + '\n' +
            '</Results>'
            )
    f.close()


########## シートの判定からXML変換までの関数 ##########
def Data_Extract(filepath, SheetList, old_check):

    # ----- ログ書込：データ変換処理 -----
    Log.Log_Info(Log_file, 'Sub Program Main\n')

    wb = px.load_workbook(filepath, read_only=True, data_only=True)

    for Sheet_Name in SheetList[::-1]:

        Sheet = wb[Sheet_Name]
        Initial = str(Sheet['M8'].value)[0]

        if Initial not in SerialNumber_list or '13B2-EA' not in str(Sheet['U3'].value):
            Log.Log_Error(Log_file, Sheet_Name + ' : ' + 'Not Covered')
            continue

        # ----- 空欄チェック -----
        if Get_Cells_Info(Sheet):
            Log.Log_Error(Log_file, "Blank Error")
            continue

        # ----- データの取得 -----
        data_dict = Open_Data_Sheet(Sheet, os.path.basename(filepath), Sheet_Name)

        # ----- 辞書がNone -----
        if data_dict is None:
            Log.Log_Error(Log_file, "Lot No Error")
            continue

        # ----- oldファイルの実行時のみ、着工者が空欄であれば'-'に置き換える -----
        if data_dict["key_operator"] == "":
            if old_check:
                data_dict["key_operator"] = '-'
            else:
                Log.Log_Error(Log_file, Sheet_Name + ' : ' + 'Operator None')
                continue

        # ----- 日付フォーマット変換 -----
        if len(data_dict['key_start_date_time']) != 19 or '年' in data_dict['key_start_date_time']:
            Log.Log_Error(Log_file, data_dict["key_serial_number"] + ' : ' + "Date Error")
            continue

        # ----- ロット番号をキーとして品名が得られなかった -----
        if len(data_dict["key_part_number"]) == 0:
            Log.Log_Error(Log_file, data_dict["key_serial_number"] + ' : ' + "Part Number Error")
            continue

        # ----- STARTTIME_SORTEDの追加 -----

        # 日付をExcel時間に変換する
        date = datetime.strptime(str(data_dict["key_start_date_time"]).replace('T', ' ').replace('.', ':'), "%Y-%m-%d %H:%M:%S")
        date_excel_number = int(str(date - datetime(1899, 12, 30)).split()[0])

        # エピ番号の数値部だけを取得する
        Epi_Number = 0
        for i in data_dict["key_batch_number"]:
            try:
                if 0 <= int(i) <= 9:
                    Epi_Number = Epi_Number * 10 + int(i)
            except:
                pass

        # エピ番号を10^6で割って、excel時間に加算する
        date_excel_number += Epi_Number/10**6

        # data_dictに登録する
        data_dict["key_STARTTIME_SORTED"] = date_excel_number
        data_dict["key_SORTNUMBER"] = Epi_Number

        # ----- データ型の確認 -----
        result = Check.Data_Type(key_type, data_dict)
        if result == False:
            Log.Log_Error(Log_file, data_dict["key_serial_number"] + ' : ' + "Data Error\n")
            continue

        # ----- XMLファイルの作成 -----
        xml_file = 'Site=' + Site + ',ProductFamily=' + ProductFamily + ',Operation=' + Operation + \
                   ',Partnumber=' + data_dict["key_part_number"] + ',Serialnumber=' + data_dict["key_serial_number"] + \
                   ',Testdate=' + data_dict["key_start_date_time"].replace(':','.') + '.xml'

        Output_XML(xml_file, data_dict)

    wb.close()


########## Main処理 ##########
if __name__ == '__main__':

    # ----- ログ書込：Main処理の開始 -----
    Log.Log_Info(Log_file, 'Main Start')

    # ----- path内のフォルダ、ファイルを全部取得 -----
    all_files = os.listdir(Path)

    # ----- ログ書込：着工ファイル検索 -----
    Log.Log_Info(Log_file, 'File Search')

    # ----- ファイルパスの取得 -----
    array = []
    for filename in all_files:
        filepath = os.path.join(Path, filename)
        if "FT" in str(filename) and "$" not in str(filename) and os.path.isfile(filepath):
            dt = strftime("%Y-%m-%d %H:%M:%S", localtime(os.path.getctime(filepath)))
            array.append([filepath, dt])

    # ----- 着工ファイルが見つからなかった -----
    if len(array) == 0:
        Log.Log_Info(Log_file, 'Folder Error')
        sys.exit()

    # ----- 最終更新日時順に並び替える -----
    array = sorted(array, key=lambda x: x[1])
    FileName = os.path.basename(array[0][0])
    Log.Log_Info(Log_file, FileName)

    # ----- 前回処理したエピ番号のNumber部分を取り出す -----
    Number = ""
    for i in Before_FileName:
        if "0" <= i <= "9":
            Number += i

    # ----- ファイルの切り替わりを確認 -----
    if Number not in FileName:

        # ----- ログ記載：過去フォルダ検索 -----
        Log.Log_Info(Log_file, 'Folder Serach')

        Old_File_Path = MOCVD_OldFileSearch.F1(str(Number))
        if Old_File_Path == -1:
            Log.Log_Info(Log_file, 'Old Folder Error')
            sys.exit()

        # ----- ログ書込：シート名の取得 -----
        Log.Log_Info(Log_file, 'OLD Get SheetName')

        # ----- 上記で指定したファイルのシート一覧を取得する -----
        wb = px.load_workbook(Old_File_Path)
        Old_SheetName = wb.sheetnames
        wb.close()

        # ----- ログ書込：前Excelファイルのデータ取得 -----
        Log.Log_Info(Log_file, 'OLD Excel File Get Data')

        # ----- 過去のファイルの全シートの処理を行う -----
        Data_Extract(Old_File_Path, Old_SheetName, 1)

    # ----- ログ書込：Excelファイルのデータ取得 -----
    Log.Log_Info(Log_file, 'Excel File Get Data')

    # ----- arrayに格納されている全てのファイルの処理を行う -----
    for File_Path, _ in array:
        Log.Log_Info(Log_file, os.path.basename(File_Path))

        # ----- 対象ファイルを開き、シートの一覧を取得する -----
        wb = px.load_workbook(File_Path)
        SheetName = wb.sheetnames
        wb.close()

        # ----- 全シートの処理を行う -----
        Data_Extract(File_Path, SheetName, 0)

    # ----- ログ書込：テキストファイルにシート名を上書きで書込する -----
    Log.Log_Info(Log_file, 'Write SheetName')

    # ----- 先ほど処理を行ったファイル名の書き込み -----
    with open('F1_FileName_Format1.txt', 'w', encoding='utf-8') as textfile:
        textfile.write(FileName)

    # ----- 最終行を書き込んだファイルをGドライブに転送 -----
    shutil.copy("F1_FileName_Format1.txt", 'T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開発/038_EA-EML/F1/13_ProgramUsedFile/')
    # shutil.copy("F1_FileName_Format1.txt", '../Test/')


########## ログ書込：プログラムの終了 ##########
Log.Log_Info(Log_file, 'Program End')