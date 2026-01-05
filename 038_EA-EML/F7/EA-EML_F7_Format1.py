import openpyxl as px
import logging
import shutil
import pyodbc
import xlrd
import glob
import sys
import os
import re

from datetime import datetime, timedelta, date
from time import strftime, localtime


########## 自作関数の定義 ##########
sys.path.append('../../MyModule')
import SQL
import Log
import ExpandExp
import Convert_Date
import Row_Number_Func
import MOCVD_OldFileSearch
import Check


########## 全体パラメータの定義 ##########
Site = '350'
ProductFamily = 'SAG FAB'
Operation = 'EA-EML_F7_Format1'
TestStation = 'EA-EML'
X = '999999'
Y = '999999'


########## Logフォルダ名の定義 ##########
Log_FolderName = str(date.today())

# ----- 格納するLogフォルダがなければ作成する -----
if not os.path.exists("../../Log/" + Log_FolderName):
    os.makedirs("../../Log/" + Log_FolderName)

# ----- ログ書き込み先ファイルパス -----
Log_file = '../../Log/' + Log_FolderName + '/038_EA-EML_F7_Format1.log'

# ----- ログ書込：Main処理の開始 -----
Log.Log_Info(Log_file, 'Program Start')


########## 処理ファイルのあるディレクトリ定義 ##########
Path = 'Z:/MOCVD/MOCVD過去プログラム/F7炉/'
# Path = 'C:/Users/hor78296/Desktop/F7炉/'


########## XML出力先ファイルパス ##########
Output_filepath = '//li.lumentuminc.net/data/SAG/TDS/Data/Files to Insert/XML/'
# Output_filepath = '../../XML/038_EA-EML/F7_Format1/'


########## TestStepの定義 ##########
teststep_dict = {
    'TestStep1' : 'Coordinate',
    'TestStep2' : 'TargetWavelength',
    'TestStep3' : 'Thickness',
    'TestStep4' : 'XRayDiffraction',
    'TestStep5' : 'InPBoardLot',
    'TestStep6' : 'Wavelength',
    'TestStep7' : 'EpiTT',
    'TestStep8' : 'GFR',
    'TestStep9' : 'Dulation',
    'TestStep10' : 'TMGa_1',
    'TestStep11' : 'TMGa_3',
    'TestStep12' : 'TMIn_1',
    'TestStep13' : 'TMIn_2',
    'TestStep14' : 'DMZn_1',
    'TestStep15' : 'Si2H6_100ppm',
    'TestStep16' : 'PH3_1_100percent',
    'TestStep17' : 'PH3_2_100percent',
    'TestStep18' : 'AsH3_2_100percent',
    'TestStep19' : 'AsH3_3_100percent',
    'TestStep20' : 'GrowthTemperature',
    'TestStep21' : 'LayerNo',
    'TestStep22' : 'Comment',
    'TestStep23' : 'Thickness_Step',
    'TestStep24' : 'CarrierConcentration',
    'TestStep25' : 'MO-level',
    'TestStep26' : 'GFR-flow',
    'TestStep27' : 'As-Ratio',
    'TestStep28' : 'Ga-Ratio',
    'TestStep29' : 'SORTED_DATA'
}


########## HeaderMiscの定義 ##########
HeaderMisc_dict = {
    'HeaderMisc1': 'RecipeName-Macro',
    'HeaderMisc2': 'RecipeName-Program',
    'HeaderMisc3': 'RecipeName-Folder'
}


########## 取得した項目と型の対応表を定義 ##########
key_type = {
    "key_start_date_time": str,
    "key_serial_number": str,
    "key_part_number": str,
    "key_operator": str,
    "key_batch_number": str,
    "key_HeaderMisc1": str,
    "key_HeaderMisc2": str,
    "key_HeaderMisc3": str,
    "key_TargetWavelength_TargetWavelength": float,
    "key_Thickness_Thickness_Cap": float,
    "key_Thickness_Thickness_Core": float,
    "key_Thickness_Thickness_Total": float,
    "key_XRayDiffraction_Xray_Thickness": float,
    "key_XRayDiffraction_Xray_Strain": float,
    "key_InPBoardLot_InPBoardLot_No": str,
    "key_InPBoardLot_InPBoardLot_CC": float,
    "key_InPBoardLot_InPBoardLot_EPD": float,
    "key_Wavelength_Wavelength_EA": float,
    "key_Wavelength_Wavelength_Intensity": float,
    "key_Wavelength_Wavelength_FWHM": float,
    "key_EpiTT_EpiTT_EA": float,
    "key_GFR_GFR_Setting": float,
    "key_Dulation_Step6": float,
    "key_Dulation_Step8": float,
    "key_Dulation_Step10": float,
    "key_Dulation_Step12-1": float,
    "key_Dulation_Step12-2": float,
    "key_Dulation_Step12-3": float,
    "key_Dulation_Step14": float,
    "key_Dulation_Step16": float,
    "key_Dulation_Step18": float,
    "key_Dulation_Step20": float,
    "key_Dulation_Step21": float,
    "key_TMGa_1_Step12-2": float,
    "key_TMGa_3_Step8": float,
    "key_TMGa_3_Step10": float,
    "key_TMGa_3_Step12-1": float,
    "key_TMGa_3_Step12-3": float,
    "key_TMGa_3_Step14": float,
    "key_TMGa_3_Step16": float,
    "key_TMGa_3_Step18": float,
    "key_TMIn_1_Step6": float,
    "key_TMIn_1_Step8": float,
    "key_TMIn_1_Step10": float,
    "key_TMIn_1_Step12-1": float,
    "key_TMIn_1_Step12-2": float,
    "key_TMIn_1_Step12-3": float,
    "key_TMIn_1_Step14": float,
    "key_TMIn_1_Step16": float,
    "key_TMIn_1_Step18": float,
    "key_TMIn_1_Step20": float,
    "key_TMIn_1_Step21": float,
    "key_TMIn_2_Step6": float,
    "key_TMIn_2_Step8": float,
    "key_TMIn_2_Step10": float,
    "key_TMIn_2_Step12-1": float,
    "key_TMIn_2_Step12-2": float,
    "key_TMIn_2_Step12-3": float,
    "key_TMIn_2_Step14": float,
    "key_TMIn_2_Step16": float,
    "key_TMIn_2_Step18": float,
    "key_TMIn_2_Step20": float,
    "key_TMIn_2_Step21": float,
    "key_DMZn_1_Step21": float,
    "key_Si2H6_100ppm_Step6": float,
    "key_Si2H6_100ppm_Step8": float,
    "key_Si2H6_100ppm_Step10": float,
    "key_PH3_1_100percent_Step6": float,
    "key_PH3_1_100percent_Step8": float,
    "key_PH3_1_100percent_Step10": float,
    "key_PH3_1_100percent_Step12-1": float,
    "key_PH3_1_100percent_Step12-2": float,
    "key_PH3_1_100percent_Step12-3": float,
    "key_PH3_1_100percent_Step14": float,
    "key_PH3_1_100percent_Step16": float,
    "key_PH3_1_100percent_Step18": float,
    "key_PH3_1_100percent_Step20": float,
    "key_PH3_1_100percent_Step21": float,
    "key_PH3_2_100percent_Step6": float,
    "key_PH3_2_100percent_Step8": float,
    "key_PH3_2_100percent_Step10": float,
    "key_PH3_2_100percent_Step12-1": float,
    "key_PH3_2_100percent_Step12-2": float,
    "key_PH3_2_100percent_Step12-3": float,
    "key_PH3_2_100percent_Step14": float,
    "key_PH3_2_100percent_Step16": float,
    "key_PH3_2_100percent_Step18": float,
    "key_PH3_2_100percent_Step20": float,
    "key_PH3_2_100percent_Step21": float,
    "key_AsH3_2_100percent_Step12-2": float,
    "key_AsH3_3_100percent_Step8": float,
    "key_AsH3_3_100percent_Step10": float,
    "key_AsH3_3_100percent_Step12-1": float,
    "key_AsH3_3_100percent_Step12-3": float,
    "key_AsH3_3_100percent_Step14": float,
    "key_AsH3_3_100percent_Step16": float,
    "key_AsH3_3_100percent_Step18": float,
    "key_GrowthTemperature_Step6": float,
    "key_GrowthTemperature_Step8": float,
    "key_GrowthTemperature_Step10": float,
    "key_GrowthTemperature_Step12-1": float,
    "key_GrowthTemperature_Step12-2": float,
    "key_GrowthTemperature_Step12-3": float,
    "key_GrowthTemperature_Step14": float,
    "key_GrowthTemperature_Step16": float,
    "key_GrowthTemperature_Step18": float,
    "key_GrowthTemperature_Step20": float,
    "key_GrowthTemperature_Step21": float,
    "key_LayerNo_Step6": str,
    "key_LayerNo_Step8": str,
    "key_LayerNo_Step10": str,
    "key_LayerNo_Step12-1": str,
    "key_LayerNo_Step12-2": str,
    "key_LayerNo_Step12-3": str,
    "key_LayerNo_Step14": str,
    "key_LayerNo_Step16": str,
    "key_LayerNo_Step18": str,
    "key_LayerNo_Step20": str,
    "key_LayerNo_Step21": str,
    "key_Comment_Step6": str,
    "key_Comment_Step8": str,
    "key_Comment_Step10": str,
    "key_Comment_Step12-1": str,
    "key_Comment_Step12-2": str,
    "key_Comment_Step12-3": str,
    "key_Comment_Step14": str,
    "key_Comment_Step16": str,
    "key_Comment_Step18": str,
    "key_Comment_Step20": str,
    "key_Comment_Step21": str,
    "key_Thickness_Step_Step6": float,
    "key_Thickness_Step_Step8": float,
    "key_Thickness_Step_Step10": float,
    "key_Thickness_Step_Step12-1": float,
    "key_Thickness_Step_Step12-2": float,
    "key_Thickness_Step_Step12-3": float,
    "key_Thickness_Step_Step14": float,
    "key_Thickness_Step_Step16": float,
    "key_Thickness_Step_Step18": float,
    "key_Thickness_Step_Step20": float,
    "key_Thickness_Step_Step21": float,
    "key_CarrierConcentration_Step6": float,
    "key_CarrierConcentration_Step8": float,
    "key_CarrierConcentration_Step10": float,
    "key_CarrierConcentration_Step21": float,
    "key_MO-level_TMGa_1": float,
    "key_MO-level_TMGa_2": float,
    "key_MO-level_TMGa_3": float,
    "key_MO-level_TMIn_1": float,
    "key_MO-level_TMIn_2": float,
    "key_MO-level_TMIn_3": float,
    "key_MO-level_DEZn_1": float,
    "key_MO-level_DMZn_1": float,
    "key_MO-level_TMAl_1": float,
    "key_MO-level_TMAl_2": float,
    "key_MO-level_AMAl_3": float,
    "key_GFR-flow_GFR-flow" : float,
    "key_As-Ratio_Step8" : float,
    "key_As-Ratio_Step10" : float,
    "key_As-Ratio_Step12-1" : float,
    "key_As-Ratio_Step12-2" : float,
    "key_As-Ratio_Step12-3" : float,
    "key_As-Ratio_Step14" : float,
    "key_As-Ratio_Step16" : float,
    "key_As-Ratio_Step18" : float,
    "key_Ga-Ratio_Step8" : float,
    "key_Ga-Ratio_Step10" : float,
    "key_Ga-Ratio_Step12-1" : float,
    "key_Ga-Ratio_Step12-2" : float,
    "key_Ga-Ratio_Step12-3" : float,
    "key_Ga-Ratio_Step14" : float,
    "key_Ga-Ratio_Step16" : float,
    "key_Ga-Ratio_Step18" : float,
    "key_STARTTIME_SORTED" : float,
    "key_SORTNUMBER" : float,
    "key_LotNumber_9": str
}


########## 対象ロット番号のイニシャルを書込したファイルを取得する ##########
Log.Log_Info(Log_file, 'Get SerialNumber Initial List ')
with open('T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開発/004_T2-EML/SerialNumber_Initial.txt', 'r', encoding='utf-8') as textfile:
    SerialNumber_list = {s.strip() for s in textfile.readlines()}


########## 前回処理を行ったファイル名を取得する ##########
with open('F7_FileName_Format1.txt', 'r', encoding='utf-8') as textfile:
    Before_FileName = textfile.readline()


########## 空欄チェック ##########
def Get_Cells_Info(Sheet):
    
    # ----- ログ書込：空欄判定 -----
    Log.Log_Info(Log_file, "Blank Check")

    # ----- False -> 空欄がない -----
    is_cells_empty = False

    # ----- エピ番号かロット番号が空欄あらば処理を行わない -----
    if  Sheet['I8'].value is None or Sheet['Q7'].value is None:
        is_cells_empty = True

    return is_cells_empty


########## データの取得 ##########
def Open_Data_Sheet(Sheet, filepath, SheetName, row_number):
    
    # ----- ログ書込：データ取得 -----
    Log.Log_Info(Log_file, 'Data Acquisition')

    # ----- 各関数に引数として渡すために辞書に格納する -----
    data_dict = dict()

    # ----- Serial_NumberをもとにPrimeから品名を引き出す -----
    serial_number = str(Sheet.cell(row=row_number, column=36).value)
    batch_number_six = str(Sheet.cell(row=row_number, column=34).value)
    conn, cursor = SQL.connSQL()

    # ----- Prime接続できなかったときはNoneが返ってくる -----
    if conn is None:
        Log.Log_Error(Log_file, serial_number + ' : ' + 'Connection with Prime Failed')
        sys.exit()

    # ----- 品名を取得 -----
    part_number, Nine_Serial_Number = SQL.selectSQL(cursor, serial_number)
    SQL.disconnSQL(conn, cursor)

    # ----- SEM / PLmapper / XRD / MOCVD の装置Noを取得 -----
    SEM, PLmapper, XRD, MOCVD = '1', '1', '1', '7'

    if '#2' in str(Sheet["L61"].value):
        SEM = '2'
    if '#2' in str(Sheet['AC63'].value):
        PLmapper = '2'
    if '#2' in str(Sheet['R57'].value):
        XRD = '2'
        
    # ----- GFRの値を格納、空欄であれば150 -----
    if Sheet.cell(row=row_number, column=43).value is None:
        GFR = '150'
    else:
        GFR = Sheet.cell(row=row_number, column=43).value

    # ----- データの取得 -----
    data_dict = {
        "key_start_date_time": str(Sheet["Q7"].value).replace(" ", "T"),
        "key_serial_number": serial_number,
        "key_part_number": part_number,
        "key_operator": Sheet["V8"].value,
        "key_batch_number": Sheet["I8"].value,
        "key_batch_number_six" : batch_number_six,
        "key_LotNumber_9": Nine_Serial_Number,
        "key_HeaderMisc1": Sheet['U3'].value,
        "key_HeaderMisc2": Sheet["U4"].value,
        "key_HeaderMisc3": Sheet["U5"].value,
        "key_TargetWavelength_TargetWavelength": Sheet['M12'].value,
        "key_Thickness_Thickness_Cap": Sheet["L62"].value,
        "key_Thickness_Thickness_Core": Sheet["L63"].value,
        "key_Thickness_Thickness_Total": Sheet["L64"].value,
        "key_XRayDiffraction_Xray_Thickness": Sheet.cell(row=row_number, column=17).value,
        "key_XRayDiffraction_Xray_Strain": Sheet.cell(row=row_number, column=18).value,
        "key_InPBoardLot_InPBoardLot_No": Sheet.cell(row=row_number, column=12).value,
        "key_InPBoardLot_InPBoardLot_CC": Sheet.cell(row=row_number, column=14).value,
        "key_InPBoardLot_InPBoardLot_EPD": Sheet.cell(row=row_number, column=15).value,
        "key_Wavelength_Wavelength_EA": Sheet.cell(row=row_number, column=26).value,
        "key_Wavelength_Wavelength_Intensity": Sheet.cell(row=row_number, column=27).value,
        "key_Wavelength_Wavelength_FWHM": Sheet.cell(row=row_number, column=28).value,
        "key_EpiTT_EpiTT_EA": Sheet.cell(row=row_number, column=11).value,
        "key_GFR_GFR_Setting": GFR,
        "key_Dulation_Step6": Sheet["I20"].value,
        "key_Dulation_Step8": Sheet["I21"].value,
        "key_Dulation_Step10": Sheet["I22"].value,
        "key_Dulation_Step12-1": Sheet["I23"].value,
        "key_Dulation_Step12-2": Sheet["I24"].value,
        "key_Dulation_Step12-3": Sheet["I25"].value,
        "key_Dulation_Step14": Sheet["I26"].value,
        "key_Dulation_Step16": Sheet["I27"].value,
        "key_Dulation_Step18": Sheet["I28"].value,
        "key_Dulation_Step20": Sheet["I29"].value,
        "key_Dulation_Step21": Sheet["I30"].value,
        "key_TMGa_1_Step12-2": Sheet["K24"].value,
        "key_TMGa_3_Step8": Sheet["M21"].value,
        "key_TMGa_3_Step10": Sheet["M22"].value,
        "key_TMGa_3_Step12-1": Sheet["M23"].value,
        "key_TMGa_3_Step12-3": Sheet["M25"].value,
        "key_TMGa_3_Step14": Sheet["M26"].value,
        "key_TMGa_3_Step16": Sheet["M27"].value,
        "key_TMGa_3_Step18": Sheet["M28"].value,
        "key_TMIn_1_Step6": Sheet["N20"].value,
        "key_TMIn_1_Step8": Sheet["N21"].value,
        "key_TMIn_1_Step10": Sheet["N22"].value,
        "key_TMIn_1_Step12-1": Sheet["N23"].value,
        "key_TMIn_1_Step12-2": Sheet["N24"].value,
        "key_TMIn_1_Step12-3": Sheet["N25"].value,
        "key_TMIn_1_Step14": Sheet["N26"].value,
        "key_TMIn_1_Step16": Sheet["N27"].value,
        "key_TMIn_1_Step18": Sheet["N28"].value,
        "key_TMIn_1_Step20": Sheet["N29"].value,
        "key_TMIn_1_Step21": Sheet["N30"].value,
        "key_TMIn_2_Step6": Sheet["O20"].value,
        "key_TMIn_2_Step8": Sheet["O21"].value,
        "key_TMIn_2_Step10": Sheet["O22"].value,
        "key_TMIn_2_Step12-1": Sheet["O23"].value,
        "key_TMIn_2_Step12-2": Sheet["O24"].value,
        "key_TMIn_2_Step12-3": Sheet["O25"].value,
        "key_TMIn_2_Step14": Sheet["O26"].value,
        "key_TMIn_2_Step16": Sheet["O27"].value,
        "key_TMIn_2_Step18": Sheet["O28"].value,
        "key_TMIn_2_Step20": Sheet["O29"].value,
        "key_TMIn_2_Step21": Sheet["O30"].value,
        "key_DMZn_1_Step21": Sheet["R30"].value,
        "key_Si2H6_100ppm_Step6": Sheet["X20"].value,
        "key_Si2H6_100ppm_Step8": Sheet["X21"].value,
        "key_Si2H6_100ppm_Step10": Sheet["X22"].value,
        "key_PH3_1_100percent_Step6": Sheet["Y20"].value,
        "key_PH3_1_100percent_Step8": Sheet["Y21"].value,
        "key_PH3_1_100percent_Step10": Sheet["Y22"].value,
        "key_PH3_1_100percent_Step12-1": Sheet["Y23"].value,
        "key_PH3_1_100percent_Step12-2": Sheet["Y24"].value,
        "key_PH3_1_100percent_Step12-3": Sheet["Y25"].value,
        "key_PH3_1_100percent_Step14": Sheet["Y26"].value,
        "key_PH3_1_100percent_Step16": Sheet["Y27"].value,
        "key_PH3_1_100percent_Step18": Sheet["Y28"].value,
        "key_PH3_1_100percent_Step20": Sheet["Y29"].value,
        "key_PH3_1_100percent_Step21": Sheet["Y30"].value,
        "key_PH3_2_100percent_Step6": Sheet["Z20"].value,
        "key_PH3_2_100percent_Step8": Sheet["Z21"].value,
        "key_PH3_2_100percent_Step10": Sheet["Z22"].value,
        "key_PH3_2_100percent_Step12-1": Sheet["Z23"].value,
        "key_PH3_2_100percent_Step12-2": Sheet["Z24"].value,
        "key_PH3_2_100percent_Step12-3": Sheet["Z25"].value,
        "key_PH3_2_100percent_Step14": Sheet["Z26"].value,
        "key_PH3_2_100percent_Step16": Sheet["Z27"].value,
        "key_PH3_2_100percent_Step18": Sheet["Z28"].value,
        "key_PH3_2_100percent_Step20": Sheet["Z29"].value,
        "key_PH3_2_100percent_Step21": Sheet["Z30"].value,
        "key_AsH3_2_100percent_Step12-2": Sheet["AB24"].value,
        "key_AsH3_3_100percent_Step8": Sheet["AC21"].value,
        "key_AsH3_3_100percent_Step10": Sheet["AC22"].value,
        "key_AsH3_3_100percent_Step12-1": Sheet["AC23"].value,
        "key_AsH3_3_100percent_Step12-3": Sheet["AC25"].value,
        "key_AsH3_3_100percent_Step14": Sheet["AC26"].value,
        "key_AsH3_3_100percent_Step16": Sheet["AC27"].value,
        "key_AsH3_3_100percent_Step18": Sheet["AC28"].value,
        "key_GrowthTemperature_Step6": Sheet['AD20'].value,
        "key_GrowthTemperature_Step8": Sheet['AD21'].value,
        "key_GrowthTemperature_Step10": Sheet['AD22'].value,
        "key_GrowthTemperature_Step12-1": Sheet['AD23'].value,
        "key_GrowthTemperature_Step12-2": Sheet['AD24'].value,
        "key_GrowthTemperature_Step12-3": Sheet['AD25'].value,
        "key_GrowthTemperature_Step14": Sheet['AD26'].value,
        "key_GrowthTemperature_Step16": Sheet['AD27'].value,
        "key_GrowthTemperature_Step18": Sheet['AD28'].value,
        "key_GrowthTemperature_Step20": Sheet['AD29'].value,
        "key_GrowthTemperature_Step21": Sheet['AD30'].value,
        "key_LayerNo_Step6": Sheet['AE20'].value,
        "key_LayerNo_Step8": Sheet['AE21'].value,
        "key_LayerNo_Step10": Sheet['AE22'].value,
        "key_LayerNo_Step12-1": Sheet['AE23'].value,
        "key_LayerNo_Step12-2": Sheet['AE24'].value,
        "key_LayerNo_Step12-3": Sheet['AE25'].value,
        "key_LayerNo_Step14": Sheet['AE26'].value,
        "key_LayerNo_Step16": Sheet['AE27'].value,
        "key_LayerNo_Step18": Sheet['AE28'].value,
        "key_LayerNo_Step20": Sheet['AE29'].value,
        "key_LayerNo_Step21": Sheet['AE30'].value,
        "key_Comment_Step6": Sheet['AF20'].value,
        "key_Comment_Step8": Sheet['AF21'].value,
        "key_Comment_Step10": Sheet['AF22'].value,
        "key_Comment_Step12-1": Sheet['AF23'].value,
        "key_Comment_Step12-2": Sheet['AF24'].value,
        "key_Comment_Step12-3": Sheet['AF25'].value,
        "key_Comment_Step14": Sheet['AF26'].value,
        "key_Comment_Step16": Sheet['AF27'].value,
        "key_Comment_Step18": Sheet['AF28'].value,
        "key_Comment_Step20": Sheet['AF29'].value,
        "key_Comment_Step21": Sheet['AF30'].value,
        "key_Thickness_Step_Step6": Sheet['AI20'].value,
        "key_Thickness_Step_Step8": Sheet['AI21'].value,
        "key_Thickness_Step_Step10": Sheet['AI22'].value,
        "key_Thickness_Step_Step12-1": Sheet['AI23'].value,
        "key_Thickness_Step_Step12-2": Sheet['AI24'].value,
        "key_Thickness_Step_Step12-3": Sheet['AI25'].value,
        "key_Thickness_Step_Step14": Sheet['AI26'].value,
        "key_Thickness_Step_Step16": Sheet['AI27'].value,
        "key_Thickness_Step_Step18": Sheet['AI28'].value,
        "key_Thickness_Step_Step20": Sheet['AI29'].value,
        "key_Thickness_Step_Step21": Sheet['AI30'].value,
        "key_CarrierConcentration_Step6": Sheet['AJ20'].value,
        "key_CarrierConcentration_Step8": Sheet['AJ21'].value,
        "key_CarrierConcentration_Step10": Sheet['AJ22'].value,
        "key_CarrierConcentration_Step21": Sheet['AJ30'].value,
        "key_MO-level_TMGa_1": Sheet['K36'].value,
        "key_MO-level_TMGa_2": Sheet['L36'].value,
        "key_MO-level_TMGa_3": Sheet['M36'].value,
        "key_MO-level_TMIn_1": Sheet['N36'].value,
        "key_MO-level_TMIn_2": Sheet['O36'].value,
        "key_MO-level_TMIn_3": Sheet['P36'].value,
        "key_MO-level_DEZn_1": Sheet['Q36'].value,
        "key_MO-level_DMZn_1": Sheet['R36'].value,
        "key_MO-level_TMAl_1": Sheet['S36'].value,
        "key_MO-level_TMAl_2": Sheet['T36'].value,
        "key_MO-level_AMAl_3": Sheet['U36'].value,
        "key_GFR-flow_GFR-flow" : Sheet.cell(row=row_number, column=43).value,
        "key_As-Ratio_Step8" : Sheet['AQ21'].value,
        "key_As-Ratio_Step10" : Sheet['AQ22'].value,
        "key_As-Ratio_Step12-1" : Sheet['AQ23'].value,
        "key_As-Ratio_Step12-2" : Sheet['AQ24'].value,
        "key_As-Ratio_Step12-3" : Sheet['AQ25'].value,
        "key_As-Ratio_Step14" : Sheet['AQ26'].value,
        "key_As-Ratio_Step16" : Sheet['AQ27'].value,
        "key_As-Ratio_Step18" : Sheet['AQ28'].value,
        "key_Ga-Ratio_Step8" : Sheet['AR21'].value,
        "key_Ga-Ratio_Step10" : Sheet['AR22'].value,
        "key_Ga-Ratio_Step12-1" : Sheet['AR23'].value,
        "key_Ga-Ratio_Step12-2" : Sheet['AR24'].value,
        "key_Ga-Ratio_Step12-3" : Sheet['AR25'].value,
        "key_Ga-Ratio_Step14" : Sheet['AR26'].value,
        "key_Ga-Ratio_Step16" : Sheet['AR27'].value,
        "key_Ga-Ratio_Step18" : Sheet['AR28'].value,
        "key_Equipment_SEM": SEM,
        "key_Equipment_XRD": XRD,
        "key_Equipment_PLmapper": PLmapper,
        "key_Equipment_MOCVD": MOCVD,
    }

    # ----- 空欄箇所はNoneとして取得される。Noneは文字列に変換できないため、空欄("")に置き換える -----
    for keys in data_dict:
        if data_dict[keys] is None or data_dict[keys] == "None" or (data_dict[keys] == '-' and keys != 'key_operator'):
            data_dict[keys] = ""
        # ----- 指数表記を展開する -----
        if type(data_dict[keys]) is float and 'e' in str(data_dict[keys]) and keys != "key_start_date_time":
            data_dict[keys] = ExpandExp.Expand(data_dict[keys])

    return data_dict


########## XML変換 ##########
def Output_XML(xml_file, data_dict):

    ########## ログ書込：XML変換 ##########
    Log.Log_Info(Log_file, 'Excel File To XML File Conversion')    
    
    f = open(Output_filepath + xml_file, 'w', encoding="utf-8")

    f.write('<?xml version="1.0" encoding="utf-8"?>' + '\n' +
            '<Results xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema">' + '\n' +
            '       <Result startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Result="Passed">' + '\n' +
            '               <Header SerialNumber=' + '"' + data_dict["key_serial_number"] + '"' + ' PartNumber=' + '"' + data_dict["key_part_number"] + '"' + ' Operation=' + '"' + Operation + '"' + ' TestStation=' + '"' + TestStation + '"' + ' Operator=' + '"' + data_dict["key_operator"] + '"' + ' StartTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Site=' + '"' + Site + '"' + ' BatchNumber=' + '"' + data_dict["key_batch_number"] + '"' + ' LotNumber=' + '"' + data_dict["key_serial_number"] + '"/>' + '\n' +
            '               <HeaderMisc>' + '\n' +
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc1"] + '">' + data_dict["key_HeaderMisc1"] + '</Item>' + '\n'
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc2"] + '">' + data_dict["key_HeaderMisc2"] + '</Item>' + '\n'
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc3"] + '">' + data_dict["key_HeaderMisc3"] + '</Item>' + '\n'
            '               </HeaderMisc>' + '\n' +
            '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep1"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + X + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + Y + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep2"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="TargetWavelength" Units="nm" Value=' + '"' + str(data_dict["key_TargetWavelength_TargetWavelength"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep3"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Thickness_Cap" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Thickness_Cap"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Thickness_Core" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Thickness_Core"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Thickness_Total" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Thickness_Total"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep4"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Xray_Thickness" Units="nm" Value=' + '"' + str(data_dict["key_XRayDiffraction_Xray_Thickness"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Xray_Strain" Units="percent" Value=' + '"' + str(data_dict["key_XRayDiffraction_Xray_Strain"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep5"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="String" Name="InPBoardLot_No" Value=' + '"' + str(data_dict["key_InPBoardLot_InPBoardLot_No"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="InPBoardLot_CC" Units="nm" Value=' + '"' + str(data_dict["key_InPBoardLot_InPBoardLot_CC"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="InPBoardLot_EPD" Units="nm" Value=' + '"' + str(data_dict["key_InPBoardLot_InPBoardLot_EPD"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep6"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Wavelength_EA" Units="nm" Value=' + '"' + str(data_dict["key_Wavelength_Wavelength_EA"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Wavelength_Intensity" Units="count" Value=' + '"' + str(data_dict["key_Wavelength_Wavelength_Intensity"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Wavelength_FWHM" Units="meV" Value=' + '"' + str(data_dict["key_Wavelength_Wavelength_FWHM"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep7"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="EpiTT_EA" Units="degree" Value=' + '"' + str(data_dict["key_EpiTT_EpiTT_EA"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep8"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="GFR_Setting" Units="sccm" Value=' + '"' + str(data_dict["key_GFR_GFR_Setting"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep9"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step6" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step6"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step8" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step8"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-1" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step12-1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-2" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step12-2"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-3" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step12-3"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step16" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step16"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step18" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step18"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step20" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step20"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step21" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step21"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep10"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-2" Units="sccm" Value=' + '"' + str(data_dict["key_TMGa_1_Step12-2"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep11"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step8" Units="sccm" Value=' + '"' + str(data_dict["key_TMGa_3_Step8"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sccm" Value=' + '"' + str(data_dict["key_TMGa_3_Step10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-1" Units="sccm" Value=' + '"' + str(data_dict["key_TMGa_3_Step12-1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-3" Units="sccm" Value=' + '"' + str(data_dict["key_TMGa_3_Step12-3"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="sccm" Value=' + '"' + str(data_dict["key_TMGa_3_Step14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step16" Units="sccm" Value=' + '"' + str(data_dict["key_TMGa_3_Step16"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step18" Units="sccm" Value=' + '"' + str(data_dict["key_TMGa_3_Step18"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep12"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step6" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_1_Step6"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step8" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_1_Step8"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_1_Step10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-1" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_1_Step12-1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-2" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_1_Step12-2"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-3" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_1_Step12-3"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_1_Step14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step16" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_1_Step16"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step18" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_1_Step18"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step20" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_1_Step20"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step21" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_1_Step21"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n'
            '               <TestStep Name=' + '"' + teststep_dict["TestStep13"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step6" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_2_Step6"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step8" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_2_Step8"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_2_Step10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-1" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_2_Step12-1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-2" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_2_Step12-2"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-3" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_2_Step12-3"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_2_Step14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step16" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_2_Step16"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step18" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_2_Step18"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step20" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_2_Step20"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step21" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_2_Step21"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep14"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step21" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn_1_Step21"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep15"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step6" Units="sccm" Value=' + '"' + str(data_dict["key_Si2H6_100ppm_Step6"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step8" Units="sccm" Value=' + '"' + str(data_dict["key_Si2H6_100ppm_Step8"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sccm" Value=' + '"' + str(data_dict["key_Si2H6_100ppm_Step10"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep16"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step6" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_1_100percent_Step6"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step8" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_1_100percent_Step8"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_1_100percent_Step10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-1" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_1_100percent_Step12-1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-2" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_1_100percent_Step12-2"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-3" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_1_100percent_Step12-3"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_1_100percent_Step14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step16" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_1_100percent_Step16"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step18" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_1_100percent_Step18"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step20" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_1_100percent_Step20"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step21" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_1_100percent_Step21"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep17"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step6" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_2_100percent_Step6"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step8" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_2_100percent_Step8"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_2_100percent_Step10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-1" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_2_100percent_Step12-1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-2" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_2_100percent_Step12-2"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-3" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_2_100percent_Step12-3"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_2_100percent_Step14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step16" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_2_100percent_Step16"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step18" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_2_100percent_Step18"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step20" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_2_100percent_Step20"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step21" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_2_100percent_Step21"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep18"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-2" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3_2_100percent_Step12-2"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep19"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step8" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3_3_100percent_Step8"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3_3_100percent_Step10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-1" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3_3_100percent_Step12-1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-3" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3_3_100percent_Step12-3"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3_3_100percent_Step14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step16" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3_3_100percent_Step16"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step18" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3_3_100percent_Step18"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep20"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step6" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step6"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step8" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step8"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-1" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step12-1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-2" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step12-2"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-3" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step12-3"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step16" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step16"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step18" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step18"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step20" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step20"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step21" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step21"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep21"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="String" Name="Step6" Value=' + '"' + str(data_dict["key_LayerNo_Step6"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step8" Value=' + '"' + str(data_dict["key_LayerNo_Step8"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step10" Value=' + '"' + str(data_dict["key_LayerNo_Step10"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step12-1" Value=' + '"' + str(data_dict["key_LayerNo_Step12-1"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step12-2" Value=' + '"' + str(data_dict["key_LayerNo_Step12-2"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step12-3" Value=' + '"' + str(data_dict["key_LayerNo_Step12-3"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step14" Value=' + '"' + str(data_dict["key_LayerNo_Step14"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step16" Value=' + '"' + str(data_dict["key_LayerNo_Step16"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step18" Value=' + '"' + str(data_dict["key_LayerNo_Step18"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step20" Value=' + '"' + str(data_dict["key_LayerNo_Step20"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step21" Value=' + '"' + str(data_dict["key_LayerNo_Step21"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep22"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="String" Name="Step6" Value=' + '"' + str(data_dict["key_Comment_Step6"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step8" Value=' + '"' + str(data_dict["key_Comment_Step8"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step10" Value=' + '"' + str(data_dict["key_Comment_Step10"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step12-1" Value=' + '"' + str(data_dict["key_Comment_Step12-1"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step12-2" Value=' + '"' + str(data_dict["key_Comment_Step12-2"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step12-3" Value=' + '"' + str(data_dict["key_Comment_Step12-3"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step14" Value=' + '"' + str(data_dict["key_Comment_Step14"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step16" Value=' + '"' + str(data_dict["key_Comment_Step16"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step18" Value=' + '"' + str(data_dict["key_Comment_Step18"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step20" Value=' + '"' + str(data_dict["key_Comment_Step20"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Step21" Value=' + '"' + str(data_dict["key_Comment_Step21"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep23"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step6" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step6"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step8" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step8"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-1" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step12-1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-2" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step12-2"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-3" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step12-3"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step16" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step16"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step18" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step18"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step20" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step20"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step21" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step21"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep24"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step6" Units="cm-3" Value=' + '"' + str(data_dict["key_CarrierConcentration_Step6"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step8" Units="cm-3" Value=' + '"' + str(data_dict["key_CarrierConcentration_Step8"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="cm-3" Value=' + '"' + str(data_dict["key_CarrierConcentration_Step10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step21" Units="cm-3" Value=' + '"' + str(data_dict["key_CarrierConcentration_Step21"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep25"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="TMGa_1" Units="g" Value=' + '"' + str(data_dict["key_MO-level_TMGa_1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="TMGa_2" Units="g" Value=' + '"' + str(data_dict["key_MO-level_TMGa_2"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="TMGa_3" Units="g" Value=' + '"' + str(data_dict["key_MO-level_TMGa_3"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="TMIn_1" Units="g" Value=' + '"' + str(data_dict["key_MO-level_TMIn_1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="TMIn_2" Units="g" Value=' + '"' + str(data_dict["key_MO-level_TMIn_2"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="TMIn_3" Units="g" Value=' + '"' + str(data_dict["key_MO-level_TMIn_3"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="DEZn_1" Units="g" Value=' + '"' + str(data_dict["key_MO-level_DEZn_1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="DMZn_1" Units="g" Value=' + '"' + str(data_dict["key_MO-level_DMZn_1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="TMAl_1" Units="g" Value=' + '"' + str(data_dict["key_MO-level_TMAl_1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="TMAl_2" Units="g" Value=' + '"' + str(data_dict["key_MO-level_TMAl_2"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="AMAl_3" Units="g" Value=' + '"' + str(data_dict["key_MO-level_AMAl_3"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep26"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="GFR-flow" Units="sccm" Value=' + '"' + str(data_dict["key_GFR-flow_GFR-flow"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep27"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step8" Units="a.u." Value=' + '"' + str(data_dict["key_As-Ratio_Step8"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="a.u." Value=' + '"' + str(data_dict["key_As-Ratio_Step10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-1" Units="a.u." Value=' + '"' + str(data_dict["key_As-Ratio_Step12-1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-2" Units="a.u." Value=' + '"' + str(data_dict["key_As-Ratio_Step12-2"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-3" Units="a.u." Value=' + '"' + str(data_dict["key_As-Ratio_Step12-3"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="a.u." Value=' + '"' + str(data_dict["key_As-Ratio_Step14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step16" Units="a.u." Value=' + '"' + str(data_dict["key_As-Ratio_Step16"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step18" Units="a.u." Value=' + '"' + str(data_dict["key_As-Ratio_Step18"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep28"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step8" Units="a.u." Value=' + '"' + str(data_dict["key_Ga-Ratio_Step8"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="a.u." Value=' + '"' + str(data_dict["key_Ga-Ratio_Step10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-1" Units="a.u." Value=' + '"' + str(data_dict["key_Ga-Ratio_Step12-1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-2" Units="a.u." Value=' + '"' + str(data_dict["key_Ga-Ratio_Step12-2"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12-3" Units="a.u." Value=' + '"' + str(data_dict["key_Ga-Ratio_Step12-3"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="a.u." Value=' + '"' + str(data_dict["key_Ga-Ratio_Step14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step16" Units="a.u." Value=' + '"' + str(data_dict["key_Ga-Ratio_Step16"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step18" Units="a.u." Value=' + '"' + str(data_dict["key_Ga-Ratio_Step18"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep29"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="STARTTIME_SORTED" Units="" Value=' + '"' + str(data_dict["key_STARTTIME_SORTED"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="SORTNUMBER" Units="" Value=' + '"' + str(data_dict["key_SORTNUMBER"]) + '"/>' + '\n' +
            '                   <Data DataType="String" Name="BATCHNUMBER_SORTED" Value=' + '"' + str(data_dict["key_batch_number_six"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="LotNumber_5" Value=' + '"' + str(data_dict["key_serial_number"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="LotNumber_9" Value=' + '"' + str(data_dict["key_LotNumber_9"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '\n'
            '               <TestEquipment>' + '\n' +
            '                   <Item DeviceName="SEM" DeviceSerialNumber="' + data_dict["key_Equipment_SEM"] + '"/>' + '\n' +
            '                   <Item DeviceName="PLmapper" DeviceSerialNumber="' + data_dict["key_Equipment_PLmapper"] + '"/>' + '\n' +
            '                   <Item DeviceName="XRD" DeviceSerialNumber="' + data_dict["key_Equipment_XRD"] + '"/>' + '\n' +
            '                   <Item DeviceName="MOCVD" DeviceSerialNumber="' + data_dict["key_Equipment_MOCVD"] + '"/>' + '\n' +
            '               </TestEquipment>' + '\n' +
            '\n'
            '               <ErrorData/>' + '\n' +
            '               <FailureData/>' + '\n' +
            '               <Configuration/>' + '\n' +
            '       </Result>' + '\n' +
            '</Results>'
            )
    f.close()


########## シートのチェックからXML変換関数を呼び出す関数 ##########
def Data_Extract(filepath, SheetList, old_check):
    
    # ----- ログ書込：データ変換処理関数 -----
    Log.Log_Info(Log_file, 'Sub Program Main\n')

    wb = px.load_workbook(filepath, read_only=True, data_only=True)

    for Sheet_Name in SheetList[::-1]:

        Sheet = wb[Sheet_Name]

        Log.Log_Info(Log_file, 'Sheet_Name : ' + Sheet_Name)

        # ----- AJ45 ~ AJ56のループ -----
        for row_number in range(45, 57):

            # ----- 対象ロットか確認する -----
            if Sheet.cell(row=row_number, column=36).value is None:
                Log.Log_Error(Log_file, 'Not LotNumber')
                continue

            Initial = str(Sheet.cell(row=row_number, column=36).value)[0]

            if Initial not in SerialNumber_list or 'EA' not in str(Sheet['U4'].value):
                Log.Log_Error(Log_file, 'Not Covered')
                continue

            # ----- 空欄チェック -----
            if Get_Cells_Info(Sheet):
                Log.Log_Error(Log_file, "Blank Error")
                continue

            # ----- データ取得 -----
            data_dict = Open_Data_Sheet(Sheet, os.path.basename(filepath), Sheet_Name, row_number)

            # ----- oldファイルの実行時のみ、着工者が空欄であれば'-'に置き換える -----
            if data_dict["key_operator"] == "":
                if old_check:
                    data_dict["key_operator"] = '-'
                else:
                    Log.Log_Error(Log_file, Sheet_Name + ' : ' + 'Operator None')
                    break

            # ----- 日付フォーマット変換 -----
            if len(data_dict['key_start_date_time']) != 19 or '年' in data_dict['key_start_date_time']:
                Log.Log_Error(Log_file, data_dict["key_serial_number"] + ' : ' + "Date Error")
                break

            # ----- ロット番号をキーとして品名が得られなかった -----
            if len(data_dict["key_part_number"]) == 0:
                Log.Log_Error(Log_file, data_dict["key_serial_number"] + ' : ' + "Part Number Error")
                continue

            # ----- STARTTIME_SORTEDの追加 -----

            # 日付をExcel時間に変換する
            date = datetime.strptime(str(data_dict["key_start_date_time"]).replace('T', ' ').replace('.', ':'), "%Y-%m-%d %H:%M:%S")
            date_excel_number = int(str(date - datetime(1899, 12, 30)).split()[0])

            # エピ番号の数値部だけを取得する
            Epi_Number = 0
            for i in data_dict["key_batch_number_six"]:
                try:
                    if 0 <= int(i) <= 9:
                        Epi_Number = Epi_Number * 10 + int(i)
                except:
                    pass

            # エピ番号を10^6で割って、excel時間に加算する
            date_excel_number += Epi_Number / 10 ** 6

            # data_dictに登録する
            data_dict["key_STARTTIME_SORTED"] = date_excel_number
            data_dict["key_SORTNUMBER"] = Epi_Number

            # ----- データ型の確認 -----
            result = Check.Data_Type(key_type, data_dict)
            if result == False:
                Log.Log_Error(Log_file, data_dict["key_serial_number"] + ' : ' + "Data Error\n")
                continue

            # ----- XMLファイルの作成 -----
            xml_file = 'Site=' + Site + ',ProductFamily=' + ProductFamily + ',Operation=' + Operation + \
                       ',Partnumber=' + data_dict["key_part_number"] + ',Serialnumber=' + data_dict["key_serial_number"] + \
                       ',Testdate=' + data_dict["key_start_date_time"].replace(':', '.') + '.xml'

            Output_XML(xml_file, data_dict)

        Log.Log_Info(Log_file, 'Next_Sheet\n')

    wb.close()


########## Main処理 ##########
if __name__ == '__main__':

    # ----- path内のフォルダ、ファイルを全部取得 -----
    all_files = os.listdir(Path)

    # ----- ログ書込：着工ファイル検索 -----
    Log.Log_Info(Log_file, 'File Search')

    # ----- ファイルパスの取得(Axxxxの形式を探す) -----
    filepattern = "A[0-9]{4}"
    array = []
    for filename in all_files:
        filepath = os.path.join(Path, filename)
        if re.compile(filepattern).search(filename) and '$' not in filename and os.path.isfile(filepath):
            dt = strftime("%Y-%m-%d %H:%M:%S", localtime(os.path.getctime(filepath)))
            array.append([filepath, dt])

    # ----- 着工ファイルが見つからなかった -----
    if len(array) == 0:
        Log.Log_Info(Log_file, 'Folder Error')
        sys.exit()

    # ----- 最終更新日時順に並び替える -----
    array = sorted(array, key=lambda x: x[1])
    FileName = os.path.basename(array[0][0])
    Log.Log_Info(Log_file, FileName)

    # ----- 前回処理したエピ番号のNumber部分を取り出す -----
    Number = ""
    for i in Before_FileName:
        if "0" <= i <= "9":
            Number += i

    # ----- ファイルの切り替わりを確認 -----
    if Number not in FileName:

        # ----- ログ記載：過去フォルダ検索 -----
        Log.Log_Info(Log_file, 'Folder Search')

        try:
            Old_file_path = MOCVD_OldFileSearch.F7(Number)
            if Old_file_path == -1:
                Log.Log_Info(Log_file, 'Old Folder Error')
                sys.exit()

            # ----- ログ書込：シート名の取得 -----
            Log.Log_Info(Log_file, 'OLD Get SheetName')

            # ----- 上記で指定したファイルのシート一覧を取得する -----
            wb = px.load_workbook(Old_file_path, keep_links=False)
            Old_SheetName = wb.sheetnames
            wb.close()

            # ----- ログ書込：前Excelファイルのデータ取得 -----
            Log.Log_Info(Log_file, 'OLD Excel File Get Data')

            # ----- 過去ファイルの処理 -----
            Data_Extract(Old_file_path, Old_SheetName, 1)
        except Exception as e:
            Log.Log_Error(Log_file, f'Error processing old file: {str(e)}')

    # ----- ログ書込：Excelファイルのデータ取得 -----
    Log.Log_Info(Log_file, 'Excel File Get Data')

    # ----- arrayに格納されている全てのファイルの処理を行う -----
    for File_Path, _ in array:

        Log.Log_Info(Log_file, os.path.basename(File_Path))
        
        try:
            # ----- 対象ファイルを開き、シートの一覧を取得する -----
            wb = px.load_workbook(File_Path)
            SheetName = wb.sheetnames
            wb.close()

            Data_list = Data_Extract(File_Path, SheetName, 0)
        except:
            Log.Log_Error(Log_file, os.path.basename(File_Path) + ':File Open Error\n')

    # ----- ログ書込：テキストファイルにシート名を上書きで書込する -----
    Log.Log_Info(Log_file, 'Write SheetName')

    # ----- 先ほど処理を行ったファイル名の書き込み -----
    with open('F7_FileName_Format1.txt', 'w', encoding='utf-8') as textfile:
        textfile.write(FileName)

    # ----- 最終行を書き込んだファイルをGドライブに転送 -----
    shutil.copy("F7_FileName_Format1.txt", 'T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開発/038_EA-EML/F7/13_ProgramUsedFile/')


# ----- ログ書込：プログラムの終了 -----
Log.Log_Info(Log_file, 'Program End')