import openpyxl as px
import logging
import shutil
import pyodbc
import xlrd
import glob
import sys
import os
import re

from datetime import datetime, timedelta, date
from time import strftime, localtime


########## 自作関数の定義 ##########
sys.path.append('../../MyModule')
import SQL
import Log
import Convert_Date
import Row_Number_Func
import MOCVD_OldFileSearch
import Check


########## 全体パラメータ定義 ##########
Site = '350'
ProductFamily = 'SAG FAB'
Operation = 'WG-EML'
TestStation = 'WG-EML'
X = '999999'
Y = '999999'


########## Logの設定 ##########
Log_FolderName = str(date.today())

# ----- 格納するLogフォルダがなければ作成する -----
if not os.path.exists("../../Log/" + Log_FolderName):
    os.makedirs("../../Log/" + Log_FolderName)

# ----- ログ書き込み先ファイルパス -----
Log_file = '../../Log/' + Log_FolderName + '/008_WG-EML_F7.log'

# ----- ログ書込：プログラムの開始 -----
Log.Log_Info(Log_file, 'Program Start')


########## 処理ファイルのあるディレクトリ定義 ##########
Path = 'Z:/MOCVD/MOCVD過去プログラム/F7炉/'
# Path = 'C:/Users/hor78296/Desktop/F7炉/'


########## XML出力先ファイルパス ##########
Output_filepath = '//li.lumentuminc.net/data/SAG/TDS/Data/Files to Insert/XML/'
# Output_filepath = '../../XML/008_WG-EML/F7/'


########## TestStepの定義 ##########
teststep_dict = {
    'TestStep1' : 'Coordinate',
    'TestStep2' : 'Thickness',
    'TestStep3' : 'Strain',
    'TestStep4' : 'Wavelength',
    'TestStep5' : 'EpiTT',
    'TestStep6' : 'Dulation',
    'TestStep7' : 'TMGa_3',
    'TestStep8' : 'TMIn_1',
    'TestStep9' : 'TMIn_2',
    'TestStep10' : 'PH3_1',
    'TestStep11' : 'PH3_2',
    'TestStep12' : 'AsH3_3_100percent',
    'TestStep13' : 'Temperature',
    'TestStep14' : 'ReactorSetting',
    'TestStep15' : 'Remaining_MO',
    'TestStep16' : 'As-Ratio',
    'TestStep17' : 'Ga-Ratio',
    'TestStep18' : 'SORTED_DATA'
}


########## HeaderMiscの定義 ##########
HeaderMisc_dict = {
    'HeaderMisc1' : 'RecipeName-Macro',
    'HeaderMisc2' : 'RecipeName-Program',
    'HeaderMisc3' : 'RecipeName-Folder'
}


########## 取得した項目と型の対応表を定義 ##########
key_type = {
    "key_start_date_time": str,
    "key_serial_number": str,
    "key_part_number": str,
    "key_operator": str,
    "key_batch_number": str,
    "key_HeaderMisc1": str,
    "key_HeaderMisc2": str,
    "key_HeaderMisc3": str,
    "key_Thickness_Thickness_Cap": float,
    "key_Thickness_Thickness_Core": float,
    "key_Thickness_Thickness_Total": float,
    "key_Strain_Straing_WG": float,
    "key_Wavelength_Wavelength_WG": float,
    "key_Wavelength_Wavelength_Intensity": float,
    "key_Wavelength_Wavelength_FWHM": float,
    "key_EpiTT_EpiTT_WG": float,
    "key_Dulation_Step4": float,
    "key_Dulation_Step6": float,
    "key_Dulation_Step8": float,
    "key_Dulation_Step10": float,
    "key_TMGa_3_Step4": float,
    "key_TMGa_3_Step6": float,
    "key_TMGa_3_Step8": float,
    "key_TMIn_1_Step4": float,
    "key_TMIn_1_Step6": float,
    "key_TMIn_1_Step8": float,
    "key_TMIn_1_Step10": float,
    "key_TMIn_2_Step4": float,
    "key_TMIn_2_Step6": float,
    "key_TMIn_2_Step8": float,
    "key_TMIn_2_Step10": float,
    "key_PH3_1_Step4": float,
    "key_PH3_1_Step6": float,
    "key_PH3_1_Step8": float,
    "key_PH3_1_Step10": float,
    "key_PH3_2_Step4": float,
    "key_PH3_2_Step6": float,
    "key_PH3_2_Step8": float,
    "key_PH3_2_Step10": float,
    "key_AsH3_3_100percent_Step4": float,
    "key_AsH3_3_100percent_Step6": float,
    "key_AsH3_3_100percent_Step8": float,
    "key_Temperature_Step4": float,
    "key_Temperature_Step6": float,
    "key_Temperature_Step8": float,
    "key_Temperature_Step10": float,
    "key_ReactorSetting_CoverStar": str,
    "key_ReactorSetting_Magazine": str,
    "key_ReactorSetting_Deck": str,
    "key_Remaining_MO_TMGa_1": float,
    "key_Remaining_MO_TMGa_2": float,
    "key_Remaining_MO_TMGa_3": float,
    "key_Remaining_MO_TMIn_1": float,
    "key_Remaining_MO_TMIn_2": float,
    "key_Remaining_MO_TMIn_3": float,
    "key_Remaining_MO_DEZn_1": float,
    "key_Remaining_MO_DMZn_1": float,
    "key_Remaining_MO_TMAI_1": float,
    "key_Remaining_MO_TMAI_2": float,
    "key_Remaining_MO_TMAI_3": float,
    "key_As-Ratio_Step4": float,
    "key_As-Ratio_Step6": float,
    "key_As-Ratio_Step8": float,
    "key_Ga-Ratio_Step4": float,
    "key_Ga-Ratio_Step6": float,
    "key_Ga-Ratio_Step8": float,
    "key_STARTTIME_SORTED": float,
    "key_SORTNUMBER" : float,
    "key_LotNumber_9" : str
}


########## 対象ロット番号のイニシャルを書込したファイルを取得する ##########
Log.Log_Info(Log_file, 'Get SerialNumber Initial List ')
#with open('T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開発/004_T2-EML/SerialNumber_Initial.txt', 'r', encoding='utf-8') as textfile:
with open('C:/Users/hsi67063/Downloads/SerialNumber_Initial.txt', 'r', encoding='utf-8') as textfile:    
    SerialNumber_list = {s.strip() for s in textfile.readlines()}


########## 前回処理を行ったファイル名を取得する ##########
with open('F7_FileName.txt', 'r', encoding='utf-8') as textfile:
    Before_FileName = textfile.readline()


########## 空欄チェック ##########
def Get_Cells_Info(Sheet):

    # ----- ログ書込：空欄判定 -----
    Log.Log_Info(Log_file, "Blank Check")

    # ----- False -> 空欄がない -----
    is_cells_empty = False

    # ----- 日付かエピ番号か結果が空欄であれば処理を行わない -----
    if Sheet['I8'].value is None or Sheet['Q7'].value is None or Sheet['L65'].value is None:
        is_cells_empty = True

    return is_cells_empty


########## データの取得 ##########
def Open_Data_Sheet(Sheet, filepath, SheetName, dummy_row, row_number):

    # ----- ログ書込：データ取得 -----
    Log.Log_Info(Log_file, 'Data Acquisition')

    # ----- データを格納する辞書を作成 -----
    data_dict = dict()

    # ----- ロット番号の取得 -----
    serial_number = str(Sheet.cell(row=row_number, column=36).value)
    batch_number_six = str(Sheet.cell(row=row_number, column=34).value)
    conn, cursor = SQL.connSQL()

    # ----- Prime接続できなかったときはNoneが返ってくる -----
    if conn is None:
        Log.Log_Error(Log_file, serial_number + ' : ' + 'Connection with Prime Failed')
        sys.exit()
    part_number, Nine_Serial_Number = SQL.selectSQL(cursor, serial_number)
    SQL.disconnSQL(conn, cursor)

    # ----- SEM / PLmapper / XRD / MOCVD の装置Noを取得 -----
    Equipment1, Equipment2, Equipment3, Equipment4 = '1', '1', '1', '7'
    if '#2' in str(Sheet["L61"].value):
        Equipment1 = '2'
    if '#2' in str(Sheet['AC63'].value):
        Equipment2 = '2'
    if '#2' in str(Sheet['R57'].value):
        Equipment3 = '2'

    # ----- StrainとWavelengthの抜き出し -----
    Strain = str(Sheet.cell(row=dummy_row, column=18).value)
    Wavelength = str(Sheet.cell(row=dummy_row, column=26).value)
    Wavelength_Intensity = str(Sheet.cell(row=dummy_row, column=27).value)
    Wavelength_FWHM = str(Sheet.cell(row=dummy_row, column=28).value)

    if Strain == "None" and Wavelength == "None" and Wavelength_Intensity == "None" and Wavelength_FWHM == "None":
        return None

    # ----- データの取得 -----
    data_dict = {
        "key_start_date_time": str(Sheet["Q7"].value).replace(" ", "T"),
        "key_serial_number": serial_number,
        "key_part_number": part_number,
        "key_operator": Sheet["V8"].value,
        "key_batch_number": Sheet["I8"].value,
        "key_batch_number_six": batch_number_six,
        "key_LotNumber_9": Nine_Serial_Number,
        "key_HeaderMisc1": "",
        "key_HeaderMisc2": Sheet["U4"].value,
        "key_HeaderMisc3": Sheet["U5"].value,
        "key_Thickness_Thickness_Cap": Sheet["L62"].value,
        "key_Thickness_Thickness_Core": Sheet["L63"].value,
        "key_Thickness_Thickness_Total": Sheet["L64"].value,
        "key_Strain_Straing_WG" : Strain,
        "key_Wavelength_Wavelength_WG" : Wavelength,
        "key_Wavelength_Wavelength_Intensity" : Wavelength_Intensity,
        "key_Wavelength_Wavelength_FWHM" : Wavelength_FWHM,
        "key_EpiTT_EpiTT_WG" : str(Sheet.cell(row=row_number, column=11).value),
        "key_Dulation_Step4" : Sheet["I19"].value,
        "key_Dulation_Step6" : Sheet["I20"].value,
        "key_Dulation_Step8" : Sheet["I21"].value,
        "key_Dulation_Step10" : Sheet["I22"].value,
        "key_TMGa_3_Step4" : Sheet["M19"].value,
        "key_TMGa_3_Step6" : Sheet["M20"].value,
        "key_TMGa_3_Step8" : Sheet["M21"].value,
        "key_TMIn_1_Step4" : Sheet["N19"].value,
        "key_TMIn_1_Step6" : Sheet["N20"].value,
        "key_TMIn_1_Step8" : Sheet["N21"].value,
        "key_TMIn_1_Step10" : Sheet["N22"].value,
        "key_TMIn_2_Step4": Sheet["O19"].value,
        "key_TMIn_2_Step6": Sheet["O20"].value,
        "key_TMIn_2_Step8": Sheet["O21"].value,
        "key_TMIn_2_Step10": Sheet["O22"].value,
        "key_PH3_1_Step4" : Sheet["Y19"].value,
        "key_PH3_1_Step6" : Sheet["Y20"].value,
        "key_PH3_1_Step8" : Sheet["Y21"].value,
        "key_PH3_1_Step10" : Sheet["Y22"].value,
        "key_PH3_2_Step4": Sheet["Z19"].value,
        "key_PH3_2_Step6": Sheet["Z20"].value,
        "key_PH3_2_Step8": Sheet["Z21"].value,
        "key_PH3_2_Step10": Sheet["Z22"].value,
        "key_AsH3_3_100percent_Step4" : Sheet["AC19"].value,
        "key_AsH3_3_100percent_Step6" : Sheet["AC20"].value,
        "key_AsH3_3_100percent_Step8" : Sheet["AC21"].value,
        "key_Temperature_Step4" : Sheet["AD19"].value,
        "key_Temperature_Step6" : Sheet["AD20"].value,
        "key_Temperature_Step8" : Sheet["AD21"].value,
        "key_Temperature_Step10" : Sheet["AD22"].value,
        "key_ReactorSetting_CoverStar": Sheet["U12"].value,
        "key_ReactorSetting_Magazine": Sheet["U13"].value,
        "key_ReactorSetting_Deck": Sheet["U14"].value,
        "key_Remaining_MO_TMGa_1" : Sheet["K36"].value,
        "key_Remaining_MO_TMGa_2" : Sheet["L36"].value,
        "key_Remaining_MO_TMGa_3" : Sheet["M36"].value,
        "key_Remaining_MO_TMIn_1" : Sheet["N36"].value,
        "key_Remaining_MO_TMIn_2" : Sheet["O36"].value,
        "key_Remaining_MO_TMIn_3" : Sheet["P36"].value,
        "key_Remaining_MO_DEZn_1" : Sheet["Q36"].value,
        "key_Remaining_MO_DMZn_1" : Sheet["R36"].value,
        "key_Remaining_MO_TMAI_1" : Sheet["S36"].value,
        "key_Remaining_MO_TMAI_2" : Sheet["T36"].value,
        "key_Remaining_MO_TMAI_3" : Sheet["U36"].value,
        "key_As-Ratio_Step4": Sheet["AQ19"].value,
        "key_As-Ratio_Step6": Sheet["AQ20"].value,
        "key_As-Ratio_Step8": Sheet["AQ21"].value,
        "key_Ga-Ratio_Step4": Sheet["AR19"].value,
        "key_Ga-Ratio_Step6": Sheet["AR20"].value,
        "key_Ga-Ratio_Step8": Sheet["AR21"].value,
        "key_SEM" : Equipment1,
        "key_PLmapper" : Equipment2,
        "key_XRD" : Equipment3,
        "key_MOCVD"  : Equipment4
    }

    # ----- 空欄箇所はNoneとして取得される。Noneは文字列に変換できないため、空欄("")に置き換える -----
    for keys in data_dict:
        if data_dict[keys] is None or data_dict[keys] == "None" or (data_dict[keys] == '-' and keys != 'key_operator'):
            data_dict[keys] = ""

    return data_dict


########## XMLファイルの作成 ##########
def Output_XML(xml_file, data_dict):
    print(xml_file)

    # ----- ログ書込：XML変換 -----
    Log.Log_Info(Log_file, 'Excel File To XML File Conversion')

    f = open(Output_filepath + xml_file, 'w', encoding="utf-8")

    f.write('<?xml version="1.0" encoding="utf-8"?>' + '\n' +
            '<Results xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema">' + '\n' +
            '       <Result startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Result="Passed">' + '\n' +
            '               <Header SerialNumber=' + '"' + data_dict["key_serial_number"] + '"' + ' PartNumber=' + '"' + data_dict["key_part_number"] + '"' + ' Operation=' + '"' + Operation + '"' + ' TestStation=' + '"' + Operation + '"' + ' Operator=' + '"' + data_dict["key_operator"] + '"' + ' StartTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Site=' + '"' + Site + '"' + ' BatchNumber=' + '"' + data_dict["key_batch_number"] + '"' + ' LotNumber=' + '"' + data_dict["key_serial_number"] + '"/>' + '\n' +
            '               <HeaderMisc>' + '\n' +
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc1"] + '">' + data_dict["key_HeaderMisc1"] + '</Item>' + '\n'
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc2"] + '">' + data_dict["key_HeaderMisc2"] + '</Item>' + '\n'
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc3"] + '">' + data_dict["key_HeaderMisc3"] + '</Item>' + '\n'
            '               </HeaderMisc>' + '\n' +
            '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep1"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + X + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + Y + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep2"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Thickness_Cap" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Thickness_Cap"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Thickness_Core" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Thickness_Core"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Thickness_Total" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Thickness_Total"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep3"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Straing_WG" Units="percent" Value=' + '"' + str(data_dict["key_Strain_Straing_WG"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep4"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Wavelength_WG" Units="nm" Value=' + '"' + str(data_dict["key_Wavelength_Wavelength_WG"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Wavelength_Intensity" Units="count" Value=' + '"' + str(data_dict["key_Wavelength_Wavelength_Intensity"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Wavelength_FWHM" Units="meV" Value=' + '"' + str(data_dict["key_Wavelength_Wavelength_FWHM"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep5"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="EpiTT_WG" Units="degree" Value=' + '"' + str(data_dict["key_EpiTT_EpiTT_WG"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep6"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step4" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step6" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step6"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step8" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step8"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step10"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep7"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step4" Units="sccm" Value=' + '"' + str(data_dict["key_TMGa_3_Step4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step6" Units="sccm" Value=' + '"' + str(data_dict["key_TMGa_3_Step6"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step8" Units="sccm" Value=' + '"' + str(data_dict["key_TMGa_3_Step8"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep8"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step4" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_1_Step4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step6" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_1_Step6"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step8" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_1_Step8"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_1_Step10"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep9"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step4" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_2_Step4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step6" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_2_Step6"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step8" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_2_Step8"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sccm" Value=' + '"' + str(data_dict["key_TMIn_2_Step10"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep10"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step4" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_1_Step4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step6" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_1_Step6"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step8" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_1_Step8"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_1_Step10"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep11"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step4" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_2_Step4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step6" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_2_Step6"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step8" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_2_Step8"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sccm" Value=' + '"' + str(data_dict["key_PH3_2_Step10"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep12"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step4" Units="slm" Value=' + '"' + str(data_dict["key_AsH3_3_100percent_Step4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step6" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3_3_100percent_Step6"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step8" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3_3_100percent_Step8"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep13"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step4" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step6" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step6"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step8" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step8"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step10"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep14"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="String" Name="CoverStar" Value=' + '"' + str(data_dict["key_ReactorSetting_CoverStar"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Magazine" Value=' + '"' + str(data_dict["key_ReactorSetting_Magazine"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="Deck" Value=' + '"' + str(data_dict["key_ReactorSetting_Deck"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep15"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Remaining_MO_TMGa_1" Units="g" Value=' + '"' + str(data_dict["key_Remaining_MO_TMGa_1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Remaining_MO_TMGa_2" Units="g" Value=' + '"' + str(data_dict["key_Remaining_MO_TMGa_2"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Remaining_MO_TMGa_3" Units="g" Value=' + '"' + str(data_dict["key_Remaining_MO_TMGa_3"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Remaining_MO_TMIn_1" Units="g" Value=' + '"' + str(data_dict["key_Remaining_MO_TMIn_1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Remaining_MO_TMIn_2" Units="g" Value=' + '"' + str(data_dict["key_Remaining_MO_TMIn_2"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Remaining_MO_TMIn_3" Units="g" Value=' + '"' + str(data_dict["key_Remaining_MO_TMIn_3"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Remaining_MO_DEZn_1" Units="g" Value=' + '"' + str(data_dict["key_Remaining_MO_DEZn_1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Remaining_MO_DMZn_1" Units="g" Value=' + '"' + str(data_dict["key_Remaining_MO_DMZn_1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Remaining_MO_TMAl_1" Units="g" Value=' + '"' + str(data_dict["key_Remaining_MO_TMAI_1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Remaining_MO_TMAl_2" Units="g" Value=' + '"' + str(data_dict["key_Remaining_MO_TMAI_2"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Remaining_MO_TMAl_3" Units="g" Value=' + '"' + str(data_dict["key_Remaining_MO_TMAI_3"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep16"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step4" Units="a.u." Value=' + '"' + str(data_dict["key_As-Ratio_Step4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step6" Units="a.u." Value=' + '"' + str(data_dict["key_As-Ratio_Step6"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step8" Units="a.u." Value=' + '"' + str(data_dict["key_As-Ratio_Step8"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep17"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step4" Units="a.u." Value=' + '"' + str(data_dict["key_Ga-Ratio_Step4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step6" Units="a.u." Value=' + '"' + str(data_dict["key_Ga-Ratio_Step6"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step8" Units="a.u." Value=' + '"' + str(data_dict["key_Ga-Ratio_Step8"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep18"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="STARTTIME_SORTED" Units="" Value=' + '"' + str(data_dict["key_STARTTIME_SORTED"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="SORTNUMBER" Units="" Value=' + '"' + str(data_dict["key_SORTNUMBER"]) + '"/>' + '\n' +
            '                   <Data DataType="String" Name="BATCHNUMBER_SORTED" Value=' + '"' + str(data_dict["key_batch_number_six"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="LotNumber_5" Value=' + '"' + str(data_dict["key_serial_number"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="LotNumber_9" Value=' + '"' + str(data_dict["key_LotNumber_9"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '\n'
            '               <TestEquipment>' + '\n' +
            '                   <Item DeviceName="SEM" DeviceSerialNumber="' + data_dict["key_SEM"] + '"/>' + '\n' +
            '                   <Item DeviceName="PLmapper" DeviceSerialNumber="' + data_dict["key_PLmapper"] + '"/>' + '\n' +
            '                   <Item DeviceName="XRD" DeviceSerialNumber="' + data_dict["key_XRD"] + '"/>' + '\n' +
            '                   <Item DeviceName="MOCVD" DeviceSerialNumber="' + data_dict["key_MOCVD"] + '"/>' + '\n' +
            '               </TestEquipment>' + '\n' +
            '\n'
            '               <ErrorData/>' + '\n' +
            '               <FailureData/>' + '\n' +
            '               <Configuration/>' + '\n' +
            '       </Result>' + '\n' +
            '</Results>'
            )
    f.close()


########### シートの判定からXML変換までの処理 ##########
def Data_Extract(filepath, SheetList, old_check):

    # ----- ログ書込：データ変換処理 -----
    Log.Log_Info(Log_file, 'Sub Start')

    try:
        wb = px.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        Log.Log_Error(Log_file, f"Error loading workbook {filepath}: {e}")
        return

    for Sheet_Name in SheetList:

        Sheet = wb[Sheet_Name]
        print(str(Sheet['S11'].value),'WG' not in str(Sheet['S11'].value))

        # ----- シートがXML変換対象シートか確認 -----
        if 'WG' not in str(Sheet['S11'].value):
            Log.Log_Error(Log_file, Sheet_Name + ' : ' + 'Not Covered\n')
            continue

        # ----- ダミー捜索 -----
        for dummy_row in range(45, 57):
            if str(Sheet.cell(row=dummy_row, column=9).value) == "ダミー" or str(Sheet.cell(row=dummy_row, column=9).value) == "ﾀﾞﾐｰ":
                if str(Sheet.cell(row=dummy_row, column=18).value) != "None" and str(Sheet.cell(row=dummy_row, column=26).value) != "None" and str(Sheet.cell(row=dummy_row, column=27).value) != "None" and str(Sheet.cell(row=dummy_row, column=28).value) != "None":
                    break
        else:
            continue

         # ----- AJ45 ~ AJ56のループ -----
        for row_number in range(45, 57):

            # ----- 対象ロットか？ -----
            if Sheet.cell(row=row_number, column=36).value is None:
                continue

            Initial = str(Sheet.cell(row=row_number, column=36).value)[0]
            print(Initial,Initial not in SerialNumber_list)
            if Initial not in SerialNumber_list:
                Log.Log_Error(Log_file, Sheet_Name + ' : ' + 'Not Covered\n')
                continue

            # ----- 空欄チェック -----
            if Get_Cells_Info(Sheet):
                Log.Log_Error(Log_file, "Blank Error\n")
                continue

            # ----- データ取得 -----
            data_dict = Open_Data_Sheet(Sheet, os.path.basename(filepath), Sheet_Name, dummy_row, row_number)

            # ----- QC値が格納されていない場合、Noneで返ってくる -----
            if data_dict is None:
                break

            # ----- oldファイルの実行時のみ、着工者が空欄であれば'-'に置き換える -----
            if data_dict["key_operator"] == "":
                if old_check:
                    data_dict["key_operator"] = '-'
                else:
                    Log.Log_Error(Log_file, Sheet_Name + ' : ' + 'Operator None\n')
                    break

            # ----- 日付フォーマット変換 -----
            if len(data_dict['key_start_date_time']) != 19 or '年' in data_dict['key_start_date_time']:
                Log.Log_Error(Log_file, data_dict["key_serial_number"] + ' : ' + "Date Error\n")
                continue

            # ----- ロット番号をキーとして品名が得られなかった -----
            if len(data_dict["key_part_number"]) == 0:
                Log.Log_Error(Log_file, "Lot No Error\n")
                continue

            # ----- STARTTIME_SORTEDの追加 -----

            # 日付をExcel時間に変換する
            date = datetime.strptime(str(data_dict["key_start_date_time"]).replace('T', ' ').replace('.', ':'), "%Y-%m-%d %H:%M:%S")
            date_excel_number = int(str(date - datetime(1899, 12, 30)).split()[0])

            # エピ番号の数値部だけを取得する
            Epi_Number = 0
            for i in data_dict["key_batch_number_six"]:
                try:
                    if 0 <= int(i) <= 9:
                        Epi_Number = Epi_Number * 10 + int(i)
                except:
                    pass

            # エピ番号を10^6で割って、excel時間に加算する
            date_excel_number += Epi_Number / 10 ** 6

            # data_dictに登録する
            data_dict["key_STARTTIME_SORTED"] = date_excel_number
            data_dict["key_SORTNUMBER"] = Epi_Number

            # ----- データ型の確認 -----
            result = Check.Data_Type(key_type, data_dict)
            if result == False:
                Log.Log_Error(Log_file, data_dict["key_serial_number"] + ' : ' + "Data Error\n")
                continue

            # ----- XMLファイルの作成 -----
            xml_file = 'Site=' + Site + ',ProductFamily=' + ProductFamily + ',Operation=' + Operation + \
                       ',Partnumber=' + data_dict["key_part_number"] + ',Serialnumber=' + data_dict["key_serial_number"] + \
                       ',Testdate=' + data_dict["key_start_date_time"].replace(':', '.') + '.xml'

            Output_XML(xml_file, data_dict)
            Log.Log_Info(Log_file, data_dict["key_serial_number"] + ' : ' + "OK\n")

    wb.close()


########### Main処理 ###########
if __name__ == '__main__':

    # ----- ログ書込：Main処理の開始 -----
    Log.Log_Info(Log_file, 'Main Start')

    # ----- path内のフォルダ、ファイルを全部取得 -----
    all_files = os.listdir(Path)

    # ----- ログ書込：着工ファイル検索 -----
    Log.Log_Info(Log_file, 'File Search')

    # ----- ファイルパスの取得(Axxxxの形式を探す) -----
    filepattern = "A[0-9]{4}"
    array = []
    for filename in all_files:
        filepath = os.path.join(Path, filename)
        if re.compile(filepattern).search(filename) and '$' not in filename and os.path.isfile(filepath):
            dt = strftime("%Y-%m-%d %H:%M:%S", localtime(os.path.getctime(filepath)))
            array.append([filepath, dt])

    # ----- 着工ファイルが見つからなかった -----
    if len(array) == 0:
        Log.Log_Info(Log_file, 'Folder Error')
        sys.exit()

    # ----- 最終更新日時順に並び替える -----
    array = sorted(array, key=lambda x: x[1])
    FileName = os.path.basename(array[0][0])
    Log.Log_Info(Log_file, FileName)

    # ----- 前回処理したエピ番号のNumber部分を取り出す -----
    Number = ""
    for i in Before_FileName:
        if "0" <= i <= "9":
            Number += i

    # ----- ファイルの切り替わりを確認 -----
    if Number not in FileName:

        # ----- ログ書込：フォルダ検索 -----
        Log.Log_Info(Log_file, 'Folder Serach')

        Old_File_Path = MOCVD_OldFileSearch.F7(Number);

        if Old_File_Path == -1:
            Log.Log_Info(Log_file, 'Old Folder Error')
            sys.exit()

        # ----- ログ書込：シート名の取得 -----
        Log.Log_Info(Log_file, 'OLD Get SheetName')

        # ----- 上記で指定したファイルのシート一覧を取得する -----
        wb = px.load_workbook(Old_File_Path, read_only=True, data_only=True, keep_links=False)
        Old_SheetName = wb.sheetnames
        wb.close()

        # ----- ログ書込：前Excelファイルのデータ取得 -----
        Log.Log_Info(Log_file, 'OLD Excel File Get Data')

        # ----- 全シートの処理を再度行う -----
        Data_Extract(Old_File_Path, Old_SheetName, 1)

    # ----- ログ書込：Excelファイルのデータ取得 -----
    Log.Log_Info(Log_file, 'Excel File Get Data')

    # ----- arrayに格納されている全てのファイルの処理を行う -----
    for File_Path, _ in reversed(array):

        Log.Log_Info(Log_file, os.path.basename(File_Path))
        
        try:
            # ----- 対象ファイルを開き、シートの一覧を取得する -----
            wb = px.load_workbook(File_Path)
            SheetName = wb.sheetnames
            wb.close()

            Data_list = Data_Extract(File_Path, SheetName, 0)
        except:
            Log.Log_Error(Log_file, os.path.basename(File_Path) + ':File Open Error\n')

    # ----- ログ書込：テキストファイルにシート名を上書きで書込する -----
    Log.Log_Info(Log_file, 'Write SheetName')

    # ----- 先ほど処理を行ったファイル名の書き込み -----
    with open('F7_FileName.txt', 'w', encoding='utf-8') as textfile:
        textfile.write(FileName)

    # ----- 各テキストファイルをGドライブに転送 -----
    #shutil.copy("F7_FileName.txt", 'T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開発/008_WG-EML/F7/13_ProgramUsedFile/')


# ----- ログ書込：プログラムの終了 -----
Log.Log_Info(Log_file, 'Program End')