import openpyxl as px
import logging
import shutil
import pyodbc
import xlrd
import glob
import sys
import os

from datetime import datetime, timedelta, date
from time import strftime, localtime


########## 自作関数の定義 ##########
sys.path.append('../../MyModule')
import SQL
import Log
import Convert_Date
import Row_Number_Func
import MOCVD_OldFileSearch
import Check


########## 全体パラメータ定義 ##########
Site = '350'
ProductFamily = 'SAG FAB'
Operation = 'WG-EML'
TestStation = 'WG-EML'
X = '999999'
Y = '999999'


########## Logの設定 ##########
Log_FolderName = str(date.today())

# ----- 格納するLogフォルダがなければ作成する -----
if not os.path.exists("../../Log/" + Log_FolderName):
    os.makedirs("../../Log/" + Log_FolderName)

# ----- ログ書き込み先ファイルパス -----
Log_file = '../../Log/' + Log_FolderName + '/008_WG-EML_F6.log'

# ----- ログ書込：プログラムの開始 -----
Log.Log_Info(Log_file, 'Program Start')


########## 処理ファイルのあるディレクトリ定義 ##########
Path = 'Z:/MOCVD/MOCVD過去プログラム/F6炉/'
# Path = 'C:/Users/hor78296/Desktop/F6炉/'


########## XML出力先ファイルパス ##########
Output_filepath = '//li.lumentuminc.net/data/SAG/TDS/Data/Files to Insert/XML/'
# Output_filepath = '../../XML/008_WG-EML/F6/'


########## TestStepの定義 ##########
teststep_dict = {
    'TestStep1' : 'Coordinate',
    'TestStep2' : 'Thickness',
    'TestStep3' : 'Dulation',
    'TestStep4' : 'MO1-TMI',
    'TestStep5' : 'MO2-TEG',
    'TestStep6' : 'MO5-TMI-3',
    'TestStep7' : 'AsH3-21-5percent',
    'TestStep8' : 'PH3-1-50percent',
    'TestStep9' : 'PH3-2-50percent',
    'TestStep10' : 'Temperature',
    'TestStep11' : 'BallastN2',
    'TestStep12' : 'MO-Temperature',
    'TestStep13' : 'Strain',
    'TestStep14' : 'Wavelength',
    'TestStep15' : 'SORTED_DATA',
}


########## HeaderMiscの定義 ##########
HeaderMisc_dict = {
    'HeaderMisc1' : 'RecipeName-Macro',
    'HeaderMisc2' : 'RecipeName-Program',
    'HeaderMisc3' : 'RecipeName-Folder'
}


########## 取得した項目と型の対応表を定義 ##########
key_type = {
    "key_start_date_time": str,
    "key_serial_number": str,
    "key_part_number": str,
    "key_operator": str,
    "key_batch_number": str,
    "key_HeaderMisc1": str,
    "key_HeaderMisc2": str,
    "key_HeaderMisc3": str,
    "key_Thickness_Thickness_Cap": float,
    "key_Thickness_Thickness_Core": float,
    "key_Thickness_Thickness_Total": float,
    "key_Dulation_Step9" : float,
    "key_Dulation_Step13" : float,
    "key_Dulation_Step17" : float,
    "key_Dulation_Step21" : float,
    "key_MO1-TMI_Step9" : float,
    "key_MO1-TMI_Step13" : float,
    "key_MO1-TMI_Step17" : float,
    "key_MO2-TEG_Step9" : float,
    "key_MO2-TEG_Step13" : float,
    "key_MO2-TEG_Step17" : float,
    "key_MO5-TMI-3_Step21" : float,
    "key_AsH3-21-5percent_Step9" : float,
    "key_AsH3-21-5percent_Step13" : float,
    "key_AsH3-21-5percent_Step17" : float,
    "key_AsH3-21-5percent_Cylinder" : float,
    "key_PH3-1-50percent_Step21" : float,
    "key_PH3-2-50percent_Step9" : float,
    "key_PH3-2-50percent_Step13" : float,
    "key_PH3-2-50percent_Step17" : float,
    "key_Temperature_Step9" : float,
    "key_Temperature_Step13" : float,
    "key_Temperature_Step17" : float,
    "key_Temperature_Step21" : float,
    "key_BallastN2_BallastN2" : float,
    "key_MO-Temperature_MO1-TMI" : float,
    "key_MO-Temperature_MO2-TEG" : float,
    "key_MO-Temperature_MO3-TMI" : float,
    "key_MO-Temperature_MO4-TEG" : float,
    "key_MO-Temperature_MO5-TMI" : float,
    "key_MO-Temperature_MO6-TEG" : float,
    "key_MO-Temperature_MO7-Mg" : float,
    "key_MO-Temperature_MO6-TMA1" : float,
    "key_STARTTIME_SORTED": float,
    "key_SORTNUMBER" : float,
    "key_LotNumber_9" : str
}


########## 対象ロット番号のイニシャルを書込したファイルを取得する ##########
Log.Log_Info(Log_file, 'Get SerialNumber Initial List ')
#with open('T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開発/004_T2-EML/SerialNumber_Initial.txt', 'r', encoding='utf-8') as textfile:
with open('C:/Users/hsi67063/Downloads/SerialNumber_Initial.txt', 'r', encoding='utf-8') as textfile: 
    SerialNumber_list = {s.strip() for s in textfile.readlines()}


########## 前回処理を行ったファイル名を取得する ##########
with open('F6_FileName.txt', 'r', encoding='utf-8') as textfile:
    Before_FileName = textfile.readline()


########## テキストファイルから[Straing_WG]と[Wavelength_WG]の値を取得する ##########
with open('F6_Strain_Wavelength.txt','r') as textfile:
    Data_list = [s.strip() for s in textfile.readlines()]


########## 空欄チェック ##########
def Get_Cells_Info(Sheet):

    # ----- ログ書込：空欄判定 -----
    Log.Log_Info(Log_file, 'Empty Judgment')

    # ----- False -> 空欄がない -----
    is_cells_empty = False

    # ----- 日付かエピ番号かバラストN2流量が空欄であれば処理を行わない -----
    if Sheet['I8'].value is None or Sheet['R7'].value is None or Sheet['AG39'].value is None:
        is_cells_empty = True

    return is_cells_empty


########## データの取得 ##########
def Open_Data_Sheet(Sheet, filepath, SheetName):

    # ----- ログ書込：データ取得 -----
    Log.Log_Info(Log_file, 'Data Acquisition')

    # ----- データを格納する辞書を作成 -----
    data_dict = dict()

    # ----- ロット番号の取得 -----
    serial_number = Sheet["M8"].value
    conn, cursor = SQL.connSQL()

    # ----- Prime接続できなかったときはNoneが返ってくる -----
    if conn is None:
        Log.Log_Error(Log_file, serial_number + ' : ' + 'Connection with Prime Failed')
        sys.exit()
    part_number, Nine_Serial_Number = SQL.selectSQL(cursor, serial_number)
    SQL.disconnSQL(conn, cursor)

    # ----- SEM / MOCVD の装置Noを取得 -----
    Equipment1, Equipment2 = '1', '6'
    # Equipment判定
    if '#2' in str(Sheet["J44"].value):
        Equipment1 = '2'

    # ----- データの取得 -----
    data_dict = {
        "key_start_date_time": str(Sheet["R7"].value).replace(" ", "T"),
        "key_serial_number": serial_number,
        "key_part_number": part_number,
        "key_LotNumber_9": Nine_Serial_Number,
        "key_operator": "-",
        "key_batch_number": Sheet["I8"].value,
        "key_HeaderMisc1": Sheet["W3"].value,
        "key_HeaderMisc2": Sheet["W4"].value,
        "key_HeaderMisc3": Sheet["W5"].value,
        "key_Thickness_Thickness_Cap": Sheet["M45"].value,
        "key_Thickness_Thickness_Core": Sheet["M46"].value,
        "key_Thickness_Thickness_Total": Sheet["M47"].value,
        "key_Dulation_Step9" : Sheet["I22"].value,
        "key_Dulation_Step13" : Sheet["I23"].value,
        "key_Dulation_Step17" : Sheet["I24"].value,
        "key_Dulation_Step21" : Sheet["I25"].value,
        "key_MO1-TMI_Step9" : Sheet['K22'].value,
        "key_MO1-TMI_Step13" : Sheet['K23'].value,
        "key_MO1-TMI_Step17" : Sheet['K24'].value,
        "key_MO2-TEG_Step9" : Sheet["L22"].value,
        "key_MO2-TEG_Step13" : Sheet["L23"].value,
        "key_MO2-TEG_Step17" : Sheet["L24"].value,
        "key_MO5-TMI-3_Step21" : Sheet["O25"].value,
        "key_AsH3-21-5percent_Step9" : Sheet["T22"].value,
        "key_AsH3-21-5percent_Step13" : Sheet["T23"].value,
        "key_AsH3-21-5percent_Step17" : Sheet["T24"].value,
        "key_AsH3-21-5percent_Cylinder" : Sheet["AG32"].value,
        "key_PH3-1-50percent_Step21" : Sheet["W25"].value,
        "key_PH3-2-50percent_Step9" : Sheet["X22"].value,
        "key_PH3-2-50percent_Step13" : Sheet["X23"].value,
        "key_PH3-2-50percent_Step17" : Sheet["X24"].value,
        "key_Temperature_Step9" : Sheet["AB22"].value,
        "key_Temperature_Step13" : Sheet["AB23"].value,
        "key_Temperature_Step17" : Sheet["AB24"].value,
        "key_Temperature_Step21" : Sheet["AB25"].value,
        "key_BallastN2_BallastN2" : Sheet["AG39"].value,
        "key_MO-Temperature_MO1-TMI" : Sheet["K41"].value,
        "key_MO-Temperature_MO2-TEG" : Sheet["L41"].value,
        "key_MO-Temperature_MO3-TMI" : Sheet["M41"].value,
        "key_MO-Temperature_MO4-TEG" : Sheet["N41"].value,
        "key_MO-Temperature_MO5-TMI" : Sheet["O41"].value,
        "key_MO-Temperature_MO6-TEG" : Sheet["P41"].value,
        "key_MO-Temperature_MO7-Mg" : Sheet["Q41"].value,
        "key_MO-Temperature_MO6-TMA1" : Sheet["R41"].value,
        "key_SEM" : Equipment1,
        "key_MOCVD" : Equipment2
    }

    # ----- 空欄箇所はNoneとして取得される。Noneは文字列に変換できないため、空欄("")に置き換える -----
    for keys in data_dict:
        if data_dict[keys] is None or data_dict[keys] == "None" or (data_dict[keys] == '-' and keys != 'key_operator'):
            data_dict[keys] = ""


    return data_dict


########## XMLファイルの作成 ##########
def Output_XML(xml_file, data_dict):

    # ----- ログ書込：XML変換 -----
    Log.Log_Info(Log_file, 'Excel File To XML File Conversion')

    f = open(Output_filepath + xml_file, 'w', encoding="utf-8")

    f.write('<?xml version="1.0" encoding="utf-8"?>' + '\n' +
            '<Results xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema">' + '\n' +
            '       <Result startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Result="Passed">' + '\n' +
            '               <Header SerialNumber=' + '"' + data_dict["key_serial_number"] + '"' + ' PartNumber=' + '"' + data_dict["key_part_number"] + '"' + ' Operation=' + '"' + Operation + '"' + ' TestStation=' + '"' + Operation + '"' + ' Operator=' + '"' + data_dict["key_operator"] + '"' + ' StartTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Site=' + '"' + Site + '"' + ' BatchNumber=' + '"' + data_dict["key_batch_number"] + '"' + ' LotNumber=' + '"' + data_dict["key_serial_number"] + '"/>' + '\n' +
            '               <HeaderMisc>' + '\n' +
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc1"] + '">' + data_dict["key_HeaderMisc1"] + '</Item>' + '\n'
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc2"] + '">' + data_dict["key_HeaderMisc2"] + '</Item>' + '\n'
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc3"] + '">' + data_dict["key_HeaderMisc3"] + '</Item>' + '\n'
            '               </HeaderMisc>' + '\n' +
            '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep1"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + X + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + Y + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep2"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Thickness_Cap" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Thickness_Cap"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Thickness_Core" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Thickness_Core"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Thickness_Total" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Thickness_Total"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep3"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step9" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step17" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step17"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step21" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step21"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep4"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step9" Units="sccm" Value=' + '"' + str(data_dict["key_MO1-TMI_Step9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="sccm" Value=' + '"' + str(data_dict["key_MO1-TMI_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step17" Units="sccm" Value=' + '"' + str(data_dict["key_MO1-TMI_Step17"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep5"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step9" Units="sccm" Value=' + '"' + str(data_dict["key_MO2-TEG_Step9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="sccm" Value=' + '"' + str(data_dict["key_MO2-TEG_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step17" Units="sccm" Value=' + '"' + str(data_dict["key_MO2-TEG_Step17"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep6"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step21" Units="sccm" Value=' + '"' + str(data_dict["key_MO5-TMI-3_Step21"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep7"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step9" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-21-5percent_Step9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-21-5percent_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step17" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-21-5percent_Step17"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Cylinder" Units="percent" Value=' + '"' + str(data_dict["key_AsH3-21-5percent_Cylinder"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep8"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step21" Units="slm" Value=' + '"' + str(data_dict["key_PH3-1-50percent_Step21"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep9"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step9" Units="slm" Value=' + '"' + str(data_dict["key_PH3-2-50percent_Step9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-2-50percent_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step17" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-2-50percent_Step17"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep10"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step9" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step17" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step17"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step21" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step21"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep11"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="BallastN2" Units="slm" Value=' + '"' + str(data_dict["key_BallastN2_BallastN2"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep12"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="MO1-TMI" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO1-TMI"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO2-TEG" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO2-TEG"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO3-TMI" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO3-TMI"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO4-TEG" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO4-TEG"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO5-TMI" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO5-TMI"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO6-TEG" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO6-TEG"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO7-Mg" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO7-Mg"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO6-TMA1" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO6-TMA1"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep13"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Straing_WG" Units="percent" Value=' + '"' + str(data_dict["key_Strain_Straing_WG"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep14"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Wavelength_WG" Units="nm" Value=' + '"' + str(data_dict["key_Wavelength_Wavelength_WG"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Wavelength_Intensity" Units="count" Value=' + '"' + str(data_dict["key_Wavelength_Wavelength_Intensity"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Wavelength_FWHM" Units="meV" Value=' + '"' + str(data_dict["key_Wavelength_Wavelength_FWHM"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep15"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="STARTTIME_SORTED" Units="" Value=' + '"' + str(data_dict["key_STARTTIME_SORTED"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="SORTNUMBER" Units="" Value=' + '"' + str(data_dict["key_SORTNUMBER"]) + '"/>' + '\n' +
            '                   <Data DataType="String" Name="LotNumber_5" Value=' + '"' + str(data_dict["key_serial_number"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="LotNumber_9" Value=' + '"' + str(data_dict["key_LotNumber_9"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '\n'
            '               <TestEquipment>' + '\n' +
            '                   <Item DeviceName="SEM" DeviceSerialNumber="' + data_dict["key_SEM"] + '"/>' + '\n' +
            '                   <Item DeviceName="MOCVD" DeviceSerialNumber="' + data_dict["key_MOCVD"] + '"/>' + '\n' +
            '               </TestEquipment>' + '\n' +
            '\n'
            '               <ErrorData/>' + '\n' +
            '               <FailureData/>' + '\n' +
            '               <Configuration/>' + '\n' +
            '       </Result>' + '\n' +
            '</Results>'
            )
    f.close()


########### シートの判定からXML変換までの処理 ##########
def Data_Extract(filepath, SheetList):

    # ----- ログ書込：データ変換処理 -----
    Log.Log_Info(Log_file, 'Sub Start')

    wb = px.load_workbook(filepath, read_only=True, data_only=True)

    # ----- シートを逆順に並び替える -----
    Sheets = SheetList[::-1]

    for Sheet_Name in Sheets:

        Sheet = wb[Sheet_Name]
        Initial = str(Sheet['M8'].value)[0]

        # ----- ログ書込：シート名 -----
        Log.Log_Info(Log_file, Sheet_Name)


        # ----- シートが処理対象シートかどうか確認 -----
        if '13B2-WG.exe' not in str(Sheet['W3'].value):
            Log.Log_Error(Log_file, Sheet_Name + ' : ' + 'Not Covered\n')
            continue

        if 'ﾁｪｯｸ' in str(Sheet['M8'].value) or 'チェック' in str(Sheet['M8'].value):
            Data_list[0] = str(Sheet['M45'].value)
            Data_list[1] = str(Sheet['M48'].value)
            Data_list[2] = str(Sheet['M49'].value)
            Data_list[3] = str(Sheet['M50'].value)

            Log.Log_Error(Log_file, Sheet_Name + ' : ' + 'Multilayer Check\n')

            continue

        if Initial not in SerialNumber_list:
            Log.Log_Error(Log_file, Sheet_Name + ' : ' + 'Not Covered\n')
            continue

        # ----- 空欄チェック -----
        if Get_Cells_Info(Sheet):
            Log.Log_Error(Log_file, "Blank Error\n")
            continue

        # ----- データの取得 -----
        data_dict = Open_Data_Sheet(Sheet, os.path.basename(filepath), Sheet_Name)

        # ----- Straing_WGとWavelength系を辞書に追加する -----
        data_dict["key_Strain_Straing_WG"] = Data_list[0]
        data_dict["key_Wavelength_Wavelength_WG"] = Data_list[1]
        data_dict["key_Wavelength_Wavelength_Intensity"] = Data_list[2]
        data_dict["key_Wavelength_Wavelength_FWHM"] = Data_list[3]

        # ----- 日付フォーマット変換 -----
        if len(data_dict['key_start_date_time']) != 19 or '年' in data_dict['key_start_date_time']:
            Log.Log_Error(Log_file, data_dict["key_serial_number"] + ' : ' + "Date Error\n")
            continue

        # ----- ロット番号をキーとして品名が得られなかった -----
        if len(data_dict["key_part_number"]) == 0:
            Log.Log_Error(Log_file, "Lot No Error\n")
            continue

        # ----- STARTTIME_SORTEDの追加 -----

        # 日付をExcel時間に変換する
        date = datetime.strptime(str(data_dict["key_start_date_time"]).replace('T', ' ').replace('.', ':'), "%Y-%m-%d %H:%M:%S")
        date_excel_number = int(str(date - datetime(1899, 12, 30)).split()[0])

        # エピ番号の数値部だけを取得する
        Epi_Number = 0
        for i in data_dict["key_batch_number"]:
            try:
                if 0<=int(i)<=9:
                    Epi_Number = Epi_Number*10+int(i)
            except:
                pass

        # エピ番号を10^6で割って、excel時間に加算する
        date_excel_number += Epi_Number/10**6

        # data_dictに登録する
        data_dict["key_STARTTIME_SORTED"] = date_excel_number
        data_dict["key_SORTNUMBER"] = Epi_Number

        # ----- データ型の確認 -----
        result = Check.Data_Type(key_type, data_dict)
        if result == False:
            Log.Log_Error(Log_file, data_dict["key_serial_number"] + ' : ' + "Data Error\n")
            continue

        # ----- XMLファイルの作成 -----
        xml_file = 'Site=' + Site + ',ProductFamily=' + ProductFamily + ',Operation=' + Operation + \
                   ',Partnumber=' + data_dict["key_part_number"] + ',Serialnumber=' + data_dict["key_serial_number"] + \
                   ',Testdate=' + data_dict["key_start_date_time"].replace(':', '.') + '.xml'

        Output_XML(xml_file, data_dict)
        Log.Log_Info(Log_file, data_dict["key_serial_number"] + ' : ' + "OK\n")

    wb.close()


########### Main処理 ###########
if __name__ == '__main__':

    # ----- ログ書込：Main処理の開始 -----
    Log.Log_Info(Log_file, 'Main Start')

    # ----- path内のフォルダ、ファイルを全部取得 -----
    all_files = os.listdir(Path)

    # ----- ログ書込：着工ファイル検索 -----
    Log.Log_Info(Log_file, 'File Search')

    # ----- ファイルパスの取得 -----
    array = []
    for filename in all_files:
        filepath = os.path.join(Path, filename)
        if "FM" in str(filename) and '$' not in str(filename) and os.path.isfile(filepath):
            dt = strftime("%Y-%m-%d %H:%M:%S", localtime(os.path.getctime(filepath)))
            array.append([filepath, dt])

    # ----- 着工ファイルが見つからなかった -----
    if len(array) == 0:
        Log.Log_Info(Log_file, 'Folder Error')
        sys.exit()

    # ----- 最終更新日時順に並び替える -----
    array = sorted(array, key=lambda x: x[1])
    FileName = os.path.basename(array[0][0])
    Log.Log_Info(Log_file, FileName)

    # ----- 前回処理したエピ番号のNumber部分を取り出す -----
    Number = ""
    for i in Before_FileName:
        if "0" <= i <= "9":
            Number += i

    # ----- ファイルの切り替わりを確認 -----
    if Number not in FileName:

        # ----- ログ書込：フォルダ検索 -----
        Log.Log_Info(Log_file, 'Folder Serach')

        # ----- 前回行ったファイルの取得 -----
        Old_File_Path = MOCVD_OldFileSearch.F6(Number);
        if Old_File_Path == -1:
            Log.Log_Info(Log_file, 'Folder Error')
            sys.exit()

        # ----- ログ書込：シート名の取得 -----
        Log.Log_Info(Log_file, 'Get Old File Sheet')

        # ----- 上記で指定したファイルのシート一覧を取得する -----
        wb = px.load_workbook(Old_File_Path)
        Old_SheetName = wb.sheetnames
        wb.close()

        # ----- ログ書込：前Excelファイルのデータ取得 -----
        Log.Log_Info(Log_file, 'OLD Excel File Get Data')

        # ----- 全シートの処理を再度行う -----
        Data_Extract(Old_File_Path, Old_SheetName)

        # ----- textファイルに、StrainとWavelength値を書込 -----
        Data_str = "\n".join(Data_list)
        with open('F6_Strain_Wavelength.txt', 'w', encoding='utf-8') as textfile:
            textfile.write(Data_str)

    # ----- ログ書込：Excelファイルのデータ取得 -----
    Log.Log_Info(Log_file, 'Excel File Get Data')

    # ----- arrayに格納されている全てのファイルの処理を行う -----
    for File_Path, _ in reversed(array):
        Log.Log_Info(Log_file, os.path.basename(File_Path))

        # ----- 対象ファイルを開き、シートの一覧を取得する -----
        wb = px.load_workbook(File_Path)
        SheetName = wb.sheetnames
        wb.close()

        Data_Extract(File_Path, SheetName)

    # ----- ログ書込：テキストファイルにシート名を上書きで書込する -----
    Log.Log_Info(Log_file, 'Write SheetName')

    # ----- 先ほど処理を行ったファイル名の書き込み -----
    with open('F6_FileName.txt', 'w', encoding='utf-8') as textfile:
        textfile.write(FileName)

    # ----- 各テキストファイルをGドライブに転送 -----
    shutil.copy("F6_Strain_Wavelength.txt", 'T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開発/008_WG-EML/F6/13_ProgramUsedFile/')
    shutil.copy("F6_FileName.txt", 'T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開発/008_WG-EML/F6/13_ProgramUsedFile/')


# ----- ログ書込：プログラムの終了 -----
Log.Log_Info(Log_file, 'Program End')