import openpyxl as px
import logging
import shutil
import pyodbc
import xlrd
import glob
import sys
import os

from datetime import datetime, timedelta, date
from time import strftime, localtime
from dateutil.relativedelta import relativedelta


########## 自作関数の定義 ##########
from F4_Module import F4_ICP
from F4_Module import F4_RIE


########## 共通関数の定義 ##########
sys.path.append('../MyModule')
import Log
import Row_Number_Func
import MOCVD_OldFileSearch


########## Logフォルダ名の定義 ##########
Log_FolderName = str(date.today())

# ----- 格納するLogフォルダがなければ作成する -----
if not os.path.exists("../Log/" + Log_FolderName):
    os.makedirs("../Log/" + Log_FolderName)

# ----- ログ書き込み先ファイルパス -----
Log_File = '../Log/' + Log_FolderName + '/039_Ru-EML_F4.log'

# ----- ログ書込：プログラムの開始 -----
Log.Log_Info(Log_File, 'Program Start')


########## 処理ファイルのあるディレクトリ定義 ##########
Path = 'Z:/MOCVD/MOCVD過去プログラム/F4炉/'
# Path = 'C:/Users/hor78296/Desktop/F4炉/'

########## 対象ロット番号のイニシャルを書込したファイルを取得する ##########
Log.Log_Info(Log_File, 'Get SerialNumber Initial List ')
#with open('T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開発/004_T2-EML/SerialNumber_Initial.txt', 'r', encoding='utf-8') as textfile:
with open('C:/Users/hsi67063/Downloads/SerialNumber_Initial.txt', 'r', encoding='utf-8') as textfile:    
    SerialNumber_list = {s.strip() for s in textfile.readlines()}


########## 処理が完了したシート名一覧を取得する ##########
with open('F4_Before_FileName.txt', 'r', encoding='utf-8') as textfile:
    Before_FileName = textfile.readline()


########## 多層CHの値を取得 ##########
with open('F4_Multi_CH.txt','r') as textfile:
    Data_list = [s.strip() for s in textfile.readlines()]


########## シートの判定からXML変換までの関数 ##########
def Data_Extract(filepath, SheetName):
    
    # ----- ログ書込：データ変換処理 -----
    Log.Log_Info(Log_File, 'Sub Start')

    wb = px.load_workbook(filepath, read_only=True, data_only=True)

    for Sheet_Name in SheetName[::-1]:
        Sheet = wb[Sheet_Name]
        Initial = str(Sheet['O8'].value)[0]
        Time = Sheet['S7'].value
        one_month_ago = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0) - relativedelta(months=2)
        
        # ----- 今日から1月以内のデータか判定 -----
        if type(Time) == type(one_month_ago):
            if Time <= one_month_ago:
                Log.Log_Error(Log_File, Sheet_Name + " : " + "Time Not Covered")
                continue


        # ----- 多層CHファイルか判定 -----
        if  "13B2-RU_JOU-5DEG" in str(Sheet['W4'].value):

            # ----- ログ書込：多層CHK処理 -----
            Log.Log_Info(Log_File, Sheet_Name + " : " + "Multi-layer check")

            # ----- 空欄判定 -----
            Epi_Number = Sheet['I8'].value
            print("Epi Number:",Epi_Number)

            if Epi_Number[:2] is not None and  Epi_Number[:2] == 'FK' and len(Epi_Number) == 7:
                Data_list[0] = str(Sheet['L51'].value)
                Data_list[1] = str(Sheet['L53'].value)
                Data_list[2] = str(Sheet['L52'].value)

                if "#1" in str(Sheet['K50'].value):
                    Data_list[3] = "1"
                else:
                    Data_list[3] = "2"
            continue

        # ----- 処理対象ロット番号か確認 -----
        if Initial not in SerialNumber_list:
            Log.Log_Info(Log_File, Sheet_Name + " : " + "Lot Type Not Covered")
            continue

        # ----- ICPシートか -----
        if 'HL13B7B8' in str(Sheet['W4'].value) or 'HL13B4B5' in str(Sheet['W4'].value):
            F4_ICP.main(filepath, Sheet_Name, Data_list)
            continue

        # ----- RIEシートか -----
        if '13B2-RU_JOU-5DEG' in str(Sheet['W4'].value):
            F4_RIE.main(filepath, Sheet_Name, Data_list)
            continue
    wb.close()

########## Main処理 ##########
if __name__ == '__main__':

    # ----- ログ書込：Main処理の開始 -----
    Log.Log_Info(Log_File, 'Main Start')

    # ----- path内のフォルダ、ファイルを全部取得 -----
    all_files = os.listdir(Path)

    # ----- ログ書込：着工ファイル検索 -----
    Log.Log_Info(Log_File, 'File Search')

    # ----- ファイルパスの取得 -----
    array = []
    for filename in all_files:
        filepath = os.path.join(Path, filename)
        if "FK" in str(filename) and '$' not in str(filename) and os.path.isfile(filepath):
            dt = strftime("%Y-%m-%d %H:%M:%S", localtime(os.path.getctime(filepath)))
            array.append([filepath, dt])

    # ----- 着工ファイルが見つからなかったときはエラー処理 -----
    if len(array) == 0:
        Log.Log_Info(Log_File, 'Folder Error')
        sys.exit()

    # ----- 最終更新日時順に並び替える -----
    array = sorted(array, key=lambda x: x[1])
    FileName = os.path.basename(array[0][0])
    Log.Log_Info(Log_File, FileName)

    # ----- 前回処理したエピ番号のNumber部分を取り出す -----
    Number = ""
    for i in Before_FileName:
        if "0" <= i <= "9":
            Number += i

    # ----- ファイルの切り替わりを確認 -----
    if Number[:4] not in FileName:

        # ----- ログ書込：フォルダ検索 -----
        Log.Log_Info(Log_File, 'Folder Serach')

        # ----- 対象ファイルパスを返す -----
        Old_File_Path = MOCVD_OldFileSearch.F4(Number)
        if Old_File_Path == -1:
            Log.Log_Info(Log_File, 'Old Folder Error')
            sys.exit()

        # ----- ログにファイルパスを書き込む -----
        Log.Log_Info(Log_File, Old_File_Path)

        # ----- シート一覧の取得 -----
        wb = px.load_workbook(Old_File_Path)
        Old_SheetName = wb.sheetnames
        wb.close()

        # ----- ログ書込：前Excelファイルのデータ取得 -----
        Log.Log_Info(Log_File, 'OLD Excel File Get Data')

        # ----- 過去ファイルの処理 -----
        Data_Extract(Old_File_Path, Old_SheetName)

        # ----- 多層CHの値の書き込み -----
        Multi_CH_str = "\n".join(Data_list)
        with open('F4_Multi_CH.txt', 'w', encoding='utf-8') as textfile:
            textfile.write(Multi_CH_str)


    # ----- ログ書込：Excelファイルのデータ取得 -----
    Log.Log_Info(Log_File, 'Excel File Get Data')

    # ----- arrayに格納されている全てのファイルの処理を行う -----
    for File_Path, _ in array:

        Log.Log_Info(Log_File, os.path.basename(File_Path))

        # ----- 対象ファイルを開き、シートの一覧を取得する -----
        wb = px.load_workbook(File_Path)
        SheetName = wb.sheetnames
        wb.close()

        Data_Extract(File_Path, SheetName)

    # ----- 先ほど処理を行ったファイル名の書き込み -----
    with open('F4_Before_FileName.txt', 'w', encoding='utf-8') as textfile:
        textfile.write(FileName)

    # ----- 最終行を書き込んだファイルをGドライブに転送 -----
    shutil.copy('F4_Before_FileName.txt', 'T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開発/039_Ru-EML/13_ProgramUsedFile/')


########## ログ書込：プログラムの終了 ##########
Log.Log_Info(Log_File, 'Program End')