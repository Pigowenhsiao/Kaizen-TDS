import logging
import openpyxl
import os
from typing import Dict, Any
import sys

sys.path.append('../MyModule')
import SQL

# 設定日誌
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")


########## 全体パラメータ定義 ##########
Site = '350'  # サイトの番号を設定
ProductFamily = 'SAG FAB'  # 製品ファミリーを設定
TestStation = 'N-electrode'  # テストステーション名を設定

class PolishProcessor:
    def __init__(self, output_dir: str):
        """EtchedThickne
        初始化 PolishProcessor
        Args:
            output_dir (str): XML 輸出目錄
        """
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def process_file(self, filepath: str, config: Dict[str, Any]):
        """
        處理單一檔案
        Args:
            filepath (str): Excel 文件路徑
            config (dict): 操作的配置字典
        """
        try:
            data = self._extract_data(filepath, config)
            if data:
                self._generate_xml(data, config)
        except Exception as e:
            logging.error(f"Failed to process file {filepath}: {e}")

    def _extract_data(self, filepath: str, config: Dict[str, Any]) -> Dict[str, Any]:
        """
        從 Excel 提取資料
        Args:
            filepath (str): Excel 文件路徑
            config (dict): 配置字典
        Returns:
            dict: 提取的資料
        """

        logging.info(f"Extracting data from {filepath}")
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb[config["data_sheet_name"]]
        xy_sheet = wb[config["xy_sheet_name"]]

        # 提取基本信息
        start_date_time = sheet.cell(row=config["row_start_date_time"], column=3).value
        serial_number = sheet.cell(row=config["row_serial_number"], column=3).value
        operator = sheet.cell(row=config["row_operator"], column=3).value or "None"

        # 提取拋光厚度數據
        
        # ########## Primeに接続し、品名を取得 ##########

        # ----- Primeへの接続 ------
        conn, cursor = SQL.connSQL()

        # ----- Primeとの接続に失敗した場合、処理を打ち切る -----
        if conn is None:
            logging.error(f"Connection with Prime Failed for Serial Number: {serial_number}")
            return {}
        
        # ----- 品名を取得 -----
        part_number, nine_serial_number = SQL.selectSQL(cursor, serial_number)
        SQL.disconnSQL(conn, cursor)
        
        # ----- 品名がNoneが見つからなかった -----
        if part_number is None:
            logging.error(f"Part Number Error for Serial Number: {serial_number}")
            return {}

        # ----- 品名が LDアレイ_ のときは登録でき
        # ないため、次のファイルへ遷移 -----
        if part_number == 'LDアレイ_':
            logging.info(f"Skipping Serial Number: {serial_number} due to invalid Part Number")
            return {}

        
        
        polish_data = [
            sheet.cell(row=config["row_polish"], column=4 + i).value or 0.0 for i in range(14)
        ]
        #print(polish_data)
        
        polish_avg = sum(polish_data) / len(polish_data) if polish_data else 0.0
 
        # 提取座標數據
        x_coords = [
            xy_sheet.cell(row=i + 1, column=config["col_x"]).value or 0.0 for i in range(1,14)
        ]
        y_coords = [
            xy_sheet.cell(row=i + 1, column=config["col_y"]).value or 0.0 for i in range(1,14)
        ]

        # 構建數據字典
        data_dict = {
            "key_Start_Date_Time": start_date_time,
            "key_Serial_Number": serial_number,
            "key_Operator": operator,
            "key_Part_Number": part_number,  # 假設來自資料庫或其他來源
            "key_LotNumber_9": nine_serial_number,
            "key_Polish1": polish_data[0],
            "key_Polish2": polish_data[1],
            "key_Polish3": polish_data[2],
            "key_Polish4": polish_data[3],
            "key_Polish5": polish_data[4],
            "key_Polish6": polish_data[5],
            "key_Polish7": polish_data[6],
            "key_Polish8": polish_data[7],
            "key_Polish9": polish_data[8],
            "key_Polish10": polish_data[9],
            "key_Polish11": polish_data[10],
            "key_Polish12": polish_data[11],
            "key_Polish13": polish_data[12],
            "key_PolishAVG": polish_avg,
            "key_X1": x_coords[0],
            "key_X2": x_coords[1],
            "key_X3": x_coords[2],
            "key_X4": x_coords[3],
            "key_X5": x_coords[4],
            "key_X6": x_coords[5],
            "key_X7": x_coords[6],
            "key_X8": x_coords[7],
            "key_X9": x_coords[8],
            "key_X10": x_coords[9],
            "key_X11": x_coords[10],
            "key_X12": x_coords[11],
            "key_X13": x_coords[12],
            "key_Y1": y_coords[0],
            "key_Y2": y_coords[1],
            "key_Y3": y_coords[2],
            "key_Y4": y_coords[3],
            "key_Y5": y_coords[4],
            "key_Y6": y_coords[5],
            "key_Y7": y_coords[6],
            "key_Y8": y_coords[7],
            "key_Y9": y_coords[8],
            "key_Y10": y_coords[9],
            "key_Y11": y_coords[10],
            "key_Y12": y_coords[11],
            "key_Y13": y_coords[12],
        }

        wb.close()
        return data_dict

    def _generate_xml(self, data: Dict[str, Any], config: Dict[str, Any]):
        """
        根據資料生成完整 XML
        """
        
        start_date_time = str(data.get("key_Start_Date_Time", ""))

        test_date_str = start_date_time.replace(":", ".").replace(" ", "T")
        xml_file = f'Site={Site},ProductFamily={ProductFamily},Operation=N-electrode_Polish_{config["operation"]},Partnumber={data["key_Part_Number"]},Serialnumber={data["key_Serial_Number"]},Testdate={test_date_str}.xml'

        filepath = os.path.join(self.output_dir, xml_file)

        logging.info(f"Generating XML file: {filepath}")

        with open(filepath, "w", encoding="utf-8") as f:
            # XML 頭部
            f.write('<?xml version="1.0" encoding="utf-8"?>\n')
            f.write('<Results xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" ')
            f.write('xmlns:xsd="http://www.w3.org/2001/XMLSchema">\n')
            f.write(f'  <Result startDateTime="{data["key_Start_Date_Time"]}" Result="Done">\n')

            # Header
            f.write(f'    <Header SerialNumber="{data["key_Serial_Number"]}" ')
            f.write(f'PartNumber="{data["key_Part_Number"]}" ')
            f.write(f'Operation="N-electrode_Polish_{config["operation"]}" ')
            f.write(f'TestStation="N-electrode" ')
            f.write(f'Operator="{data["key_Operator"]}" ')
            f.write(f'StartTime="{data["key_Start_Date_Time"]}" ')
            f.write(f'Site="350" ')
            f.write(f'LotNumber="{data["key_Serial_Number"]}"/>\n')

            # TestStep
            for i in range(1, 14):
                f.write(f'    <TestStep Name="Thickness{i}" startDateTime="{data["key_Start_Date_Time"]}" Status="Done">\n')
                f.write(f'      <Data DataType="Numeric" Name="X" Units="um" Value="{data[f"key_X{i}"]}"/>\n')
                f.write(f'      <Data DataType="Numeric" Name="Y" Units="um" Value="{data[f"key_Y{i}"]}"/>\n')
                f.write(f'      <Data DataType="Numeric" Name="Thickness" Units="um" Value="{data[f"key_Polish{i}"]}"/>\n')
                f.write(f'    </TestStep>\n')
            
            # polish average    
            f.write(f'    <TestStep Name="Thickness_AVG" startDateTime="{data["key_Start_Date_Time"]}" Status="Done">\n')
            f.write(f'      <Data DataType="Numeric" Name="Thickness" Units="um" Value="{data[f"key_PolishAVG"]}"/>\n')
            f.write(f'    </TestStep>\n')    

            # SORTED_DATA
            f.write(f'    <TestStep Name="SORTED_DATA" startDateTime="{data["key_Start_Date_Time"]}" Status="Passed">\n')
            f.write(f'      <Data DataType="String" Name="LotNumber_5" Value="{data["key_Serial_Number"]}" CompOperation="LOG"/>\n')
            f.write(f'      <Data DataType="String" Name="LotNumber_9" Value="{data["key_LotNumber_9"]}" CompOperation="LOG"/>\n')
            f.write(f'    </TestStep>\n')

            # TestEquipment
            f.write(f'    <TestEquipment>\n')
            f.write(f'      <Item DeviceName="Stepmeter" DeviceSerialNumber="1"/>\n')
            f.write(f'    </TestEquipment>\n')

            # Additional Elements
            f.write(f'    <ErrorData/>\n')
            f.write(f'    <FailureData/>\n')
            f.write(f'    <Configuration/>\n')

            f.write(f'  </Result>\n')
            f.write('</Results>\n')
