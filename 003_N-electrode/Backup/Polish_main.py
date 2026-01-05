import logging
import shutil
import glob
import xlrd  # .xls ファイルの処理に使用
import openpyxl  # .xlsx と .xlsm ファイルの処理に使用
import os
import sys
from datetime import datetime, date

########## サブプログラムの定義 ##########
from Polish_Func import InitialWaferThickness
from Polish_Func import WaxThickness
from Polish_Func import RoughPolishedThickness
from Polish_Func import MirrorPolishedThickness
from Polish_Func import EtchedThickness

########## 自作関数の定義 ##########
sys.path.append('../MyModule')
import SQL  # SQL接続のためのモジュール
import Log  # ログ記録のためのモジュール
import Convert_Date  # 日付変換用のモジュール

########## ログの設定 ##########
# ログフォルダ名を今日の日付に設定
Log_FolderName = str(date.today())

# ログフォルダが存在しない場合、新規作成
if not os.path.exists("../Log/" + Log_FolderName):
    os.makedirs("../Log/" + Log_FolderName)

# ログファイルパスを定義
Log_file = '../Log/' + Log_FolderName + '/003_N-electrode.log'
Log.Log_Info(Log_file, 'プログラム開始')

########## 処理シートの定義 ##########
# データシートと座標シートの名前を定義
Data_sheet_name = '3ｲﾝﾁ用'
XY_sheet_name = 'ウェハ座標'

########## エラーの時に取得される番号を定義 ##########
Error_Number = {'empty:\'\'', 'error:7', 'error:15', 'error:23', 'error:29', 'error:36', 'error:42', 'xldate:0.0'}

########## 空欄チェック ##########
def Get_Cells_Info(filepath):
    """
    ファイルを読み込み、特定のセルに空欄やエラーがあるか確認
    """

    # ログに空欄チェックの開始を記録
    Log.Log_Info(Log_file, "空欄チェック")

    is_cells_empty = False  # 空欄があるかどうかのフラグ

    # ファイルの形式を確認し、.xlsm または .xlsx のみを処理
    if not (filepath.lower().endswith('.xlsm') or filepath.lower().endswith('.xlsx')):
        Log.Log_Info(Log_file, f"非対応ファイル形式をスキップ: {filepath}")
        return True  # 非対応ファイルをスキップ

    # .xlsm または .xlsx ファイルを openpyxl で処理
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)  # 実際の値のみを取得
        sheet = wb[Data_sheet_name]  # データシートを選択
    except Exception as e:
        Log.Log_Error(Log_file, f"{Data_sheet_name} : 読み取りエラー ({filepath})\n")
        is_cells_empty = True
        return is_cells_empty

    # 特定のセルが空欄またはエラーコードかを確認
    if isinstance(sheet.cell(3, 3).value, str) and sheet.cell(3, 3).value in Error_Number or \
       isinstance(sheet.cell(4, 3).value, str) and sheet.cell(4, 3).value in Error_Number or \
       isinstance(sheet.cell(21, 3).value, str) and sheet.cell(21, 3).value in Error_Number or \
       isinstance(sheet.cell(46, 3).value, str) and sheet.cell(46, 3).value in Error_Number or \
       isinstance(sheet.cell(71, 3).value, str) and sheet.cell(71, 3).value in Error_Number:
        is_cells_empty = True  # 空欄やエラーが見つかった場合、フラグを立てる

    return is_cells_empty  # 結果を返す

########## 本処理 ##########
if __name__ == '__main__':
    
    # プログラム処理の開始をログに記録
    Log.Log_Info(Log_file, 'メイン処理開始')

    # 今日の日付を取得し、年・月・日で分割
    today = str(date.today()).split('-')

    ########### ディレクトリ検索 ##########
    Log.Log_Info(Log_file, 'ディレクトリ検索')

    # 検索するファイルパスを定義
    FilePath = 'Z:/研磨/2024年/'
    FolderName = '*年*月'
    Folder_List = []  # 対象フォルダのリストを初期化
    for folder in glob.glob(FilePath + FolderName):  # フォルダ検索
        Folder_List.append(folder)

    # 対象フォルダが存在しない場合、エラーログに書き込み、プログラム終了
    if len(Folder_List) == 0:
        Log.Log_Error(Log_file, 'フォルダエラー: フォルダが見つかりません')
        sys.exit()

    # フォルダリストをソートし、最初のフォルダを処理対象に設定
    Folder_List.sort()
    directory = Folder_List[0] + '/'

    ########### 処理済みフォルダがなければ作成する ##########
    # 処理済みフォルダが存在しない場合、新規作成
    if not os.path.exists(directory + '/処理済みフォルダ'):
        os.mkdir(directory + '/処理済みフォルダ')

    ########### 対象ファイルの処理 ##########
    # 対象ディレクトリ内の .xlsm または .xlsx ファイルを処理
    for File_Path in glob.glob(directory + '*.xls*'):

        # ファイル名をログに記録
        Log.Log_Info(Log_file, File_Path)

        # 空欄チェックを実行
        if Get_Cells_Info(File_Path):
            Log.Log_Error(Log_file, "空欄エラー\n")
            continue  # エラーがあればファイル処理をスキップ

        # ロット番号と日付の取得
        try:
            wb = openpyxl.load_workbook(File_Path, data_only=True)  # 実際のデータのみを取得
            sheet = wb[Data_sheet_name]
            Serial_Number = str(sheet.cell(4, 3).value)  # ロット番号の取得
            Date_Array = [sheet.cell(5, 3).value, sheet.cell(25, 3).value, sheet.cell(37, 3).value, sheet.cell(56, 3).value]  # 日付の取得
        except Exception as e:
            Log.Log_Error(Log_file, f"ファイル読み取りエラー: {e}")
            continue  # ファイル読み取りに失敗した場合、処理をスキップ

        ########## Primeに接続し、品名を取得 ##########
        conn, cursor = SQL.connSQL()

        # 接続に失敗した場合、エラーログに書き込み、プログラム終了
        if conn is None:
            Log.Log_Error(Log_file, Serial_Number + ' : ' + 'Primeとの接続に失敗しました\n')
            sys.exit()

        # SQLを使用して品名を取得
        Part_Number, Nine_Serial_Number = SQL.selectSQL(cursor, Serial_Number)
        SQL.disconnSQL(conn, cursor)

        # 品名が見つからない場合、エラーログに記録し、次のファイルへ
        if Part_Number is None:
            Log.Log_Error(Log_file, Serial_Number + ' : ' + "品名エラー\n")
            continue

        # 品名が LDアレイ_ の場合、ファイル処理をスキップ
        if Part_Number == 'LDアレイ_':
            continue

        ########## 日付フォーマットの変換 ##########
        # 日付フォーマットが正しいか確認
        if Convert_Date.Edit_Date(Date_Array[0]) == "" or Convert_Date.Edit_Date(Date_Array[1]) == "" or Convert_Date.Edit_Date(Date_Array[2]) == "" or Convert_Date.Edit_Date(Date_Array[3]) == "":
            Log.Log_Error(Log_file, Serial_Number + ' : ' + "日付フォーマットエラー\n")
            continue  # エラーがあればファイル処理をスキップ

        ########## 各Operationに渡しXML変換を行う ##########
        # 各操作の実行をログに記録
        Log.Log_Info(Log_file, '各操作にデータを渡して処理を開始')
        InitialWaferThickness.main(File_Path, Part_Number, Nine_Serial_Number)
        WaxThickness.main(File_Path, Part_Number, Nine_Serial_Number)
        RoughPolishedThickness.main(File_Path, Part_Number, Nine_Serial_Number)
        MirrorPolishedThickness.main(File_Path, Part_Number, Nine_Serial_Number)
        EtchedThickness.main(File_Path, Part_Number, Nine_Serial_Number)

        ########### 処理済みフォルダに移動 ##########
        # 処理が完了したファイルを処理済みフォルダに移動
        Log.Log_Info(Log_file, 'ファイルを処理済みフォルダに移動')
        try:
            shutil.move(File_Path, directory + '/処理済みフォルダ/')
        except PermissionError:
            Log.Log_Error(Log_file, "ファイル移動エラー: 権限エラー")

########## Main処理の終了 ##########
Log.Log_Info(Log_file, 'プログラム終了')
