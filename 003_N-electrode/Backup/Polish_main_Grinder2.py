import logging
import shutil
import glob
import openpyxl
import os
import sys

from datetime import datetime, timedelta, date


########## サブプログラムの定義 ##########
from Polish_Func_Grinder import InitialWaferThickness
from Polish_Func_Grinder import WaxThickness
from Polish_Func_Grinder import RoughPolishedThickness
from Polish_Func_Grinder import MirrorPolishedThickness
from Polish_Func_Grinder import EtchedThickness


########## 自作関数の定義 ##########
sys.path.append('../MyModule')
import SQL
import Log
import Convert_Date


########## ログの設定 ##########

# ----- Logフォルダ名の定義 -----
Log_FolderName = str(date.today())

# ----- 格納するLogフォルダがなければ作成する -----
if not os.path.exists("../Log/" + Log_FolderName):
    os.makedirs("../Log/" + Log_FolderName)

# ----- ログ書き込み先パスの定義 -----
Log_file = '../Log/' + Log_FolderName + '/003_N-electrode.log'


########## プログラムの開始 ##########
Log.Log_Info(Log_file, 'Program Start')


########## 処理シートの定義 ##########
Data_sheet_name = '3ｲﾝﾁ用'
XY_sheet_name = 'ウェハ座標'


########## エラーの時に取得される番号を定義 ##########
Error_Number = {'empty:\'\'', 'error:7', 'error:15', 'error:23', 'error:29', 'error:36', 'error:42', 'xldate:0.0'}


########## 空欄チェック ##########
def Get_Cells_Info(filepath):

    # ----- ログ書込：空欄判定 -----
    Log.Log_Info(Log_file, "Blank Check")

    # ----- False -> 空欄がない -----
    is_cells_empty = False

    # ----- ファイルとシートの定義 -----
    wb = openpyxl.load_workbook(filepath, data_only=True)
    try:
        sheet = wb[Data_sheet_name]
        sheetx = wb[XY_sheet_name]
    except KeyError:
        Log.Log_Error(Log_file, Data_sheet_name + ' : ' + "Not Exist\n")
        is_cells_empty = True
        return is_cells_empty
        
    # ----- 一つでも空欄が存在すれば処理を行わない -----
    if str(sheet.cell(row=4, column=3).value) in Error_Number or \
        str(sheet.cell(row=5, column=3).value) in Error_Number or \
        str(sheet.cell(row=24, column=3).value) in Error_Number or \
        str(sheet.cell(row=25, column=3).value) in Error_Number or \
        str(sheet.cell(row=36, column=3).value) in Error_Number or \
        str(sheet.cell(row=37, column=3).value) in Error_Number or \
        str(sheet.cell(row=55, column=3).value) in Error_Number or \
        str(sheet.cell(row=56, column=3).value) in Error_Number:
            is_cells_empty = True

    return is_cells_empty


########## 本処理 ##########
if __name__ == '__main__':

    # ----- Main処理の開始 -----
    Log.Log_Info(Log_file, 'Main Start')

    # ----- 処理フォルダの定義 → 今日の年/月/日を取り出す -----
    today = str(date.today()).split('-')  # ['2020', '08', '26']の構造で受け取られる


    ########### ディレクトリ検索 ##########
    Log.Log_Info(Log_file, 'Directory Search')

    FilePath = 'Z:/研磨/2025年/'
    FolderName = '*年*月'
    Folder_List = []
    for folder in glob.glob(FilePath + FolderName):
        Folder_List.append(folder)

    # ----- 対象フォルダが存在しない → エラーログに書き込む -----
    if len(Folder_List) == 0:
        Log.Log_Error(Log_file, 'Folder Error')
        sys.exit()

    Folder_List.sort()
    directory1 = Folder_List[0] + '/'
    directory = 'C:/Users/hsi67063/Downloads/'


    ########### 処理済みフォルダがなければ作成する ##########
    if os.path.exists(directory + '/処理済みフォルダ') == 0:
        os.mkdir(directory + '/処理済みフォルダ')


    ########### 対象ファイルの処理 ##########
    for File_Path in glob.glob(directory1 + '*.xls*'):

        # ----- 処理を行うファイルの出力 -----
        Log.Log_Info(Log_file, File_Path)

        # ----- 空欄チェック -----
        if Get_Cells_Info(File_Path):
            Log.Log_Error(Log_file, "Blank Error\n")
            continue

        # ----- ロット番号と日付の取得 -----
        wb = openpyxl.load_workbook(File_Path, data_only=True)
        sheet = wb[Data_sheet_name]
        Serial_Number = str(sheet.cell(row=4, column=3).value)
        print(Serial_Number)
   
        Date_Array = [sheet.cell(row=5, column=3).value, sheet.cell(row=25, column=3).value, sheet.cell(row=37, column=3).value, sheet.cell(row=56, column=3).value]


        ########## Primeに接続し、品名を取得 ##########

        # ----- Primeへの接続 ------
        conn, cursor = SQL.connSQL()

        # ----- Primeとの接続に失敗した場合、処理を打ち切る -----
        if conn is None:
            Log.Log_Error(Log_file, Serial_Number + ' : ' + 'Connection with Prime Failed\n')
            sys.exit()

        # ----- 品名を取得 -----
        Part_Number, Nine_Serial_Number = SQL.selectSQL(cursor, Serial_Number)
        SQL.disconnSQL(conn, cursor)

        # ----- 品名がNoneが見つからなかった -----
        if Part_Number is None:
            Log.Log_Error(Log_file, Serial_Number + ' : ' + "Part Number Error\n")
            continue

        # ----- 品名が LDアレイ_ のときは登録できないため、次のファイルへ遷移 -----
        if Part_Number == 'LDアレイ_':
            continue


        ########## 日付フォーマットの変換 ##########
        if Convert_Date.Edit_Date(Date_Array[0]) == "" or Convert_Date.Edit_Date(Date_Array[1]) == "" or Convert_Date.Edit_Date(Date_Array[2]) == "" or Convert_Date.Edit_Date(Date_Array[3]) == "":
            Log.Log_Error(Log_file, Serial_Number + ' : ' + "Date Error\n")
            continue


        ########## 各Operationに渡しXML変換を行う ##########
        Log.Log_Info(Log_file, 'Move To Each Function')
        InitialWaferThickness.main(File_Path, Part_Number, Nine_Serial_Number)
        WaxThickness.main(File_Path, Part_Number, Nine_Serial_Number)
        RoughPolishedThickness.main(File_Path, Part_Number, Nine_Serial_Number)
        MirrorPolishedThickness.main(File_Path, Part_Number, Nine_Serial_Number)
        EtchedThickness.main(File_Path, Part_Number, Nine_Serial_Number)


        ########### 処理済みフォルダに移動 ##########
        Log.Log_Info(Log_file, 'Move File')
        try:
            shutil.move(File_Path, directory + '/処理済みフォルダ/')
        except PermissionError:
            Log.Log_Error(Log_file, "Permission Error")


########## Main処理の終了 ##########
Log.Log_Info(Log_file, 'Program End')
