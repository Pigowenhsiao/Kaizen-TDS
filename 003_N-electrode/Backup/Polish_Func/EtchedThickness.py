import logging  # ログ記録モジュールをインポート
import openpyxl  # Excel ファイルを処理するためのモジュールをインポート
import glob  # ファイルパスを検索するためのモジュールをインポート
import os  # OS関連の機能を操作するためのモジュールをインポート
import sys  # システム関連の機能を操作するためのモジュールをインポート

from datetime import datetime, timedelta, date  # 日付と時間を処理するためのクラスをインポート
from time import strftime  # 時間フォーマットを変更するための関数をインポート

########## 自作関数の定義 ##########
sys.path.append('../MyModule')  # 自作モジュールのパスをシステムパスに追加
import Log  # ログ記録モジュールをインポート
import Convert_Date  # 日付フォーマット変換モジュールをインポート
import Check  # データ型を確認するモジュールをインポート

########## 全体パラメータ定義 ##########
Site = '350'  # サイトの番号を設定
ProductFamily = 'SAG FAB'  # 製品ファミリーを設定
Operation = 'N-electrode_Polish_EtchedThickness'  # 操作名を設定
TestStation = 'N-electrode'  # テストステーション名を設定

########## ログファイルの定義 ##########
Log_FolderName = str(date.today())  # 今日の日付をログフォルダ名として使用
Log_file = '../Log/' + Log_FolderName + '/003_N-electrode.log'  # ログファイルのパスを定義

########## シート名の定義 ##########
Data_Sheet_Name = '3ｲﾝﾁ用'  # データが入っているシート名
XY_Sheet_Name = 'ウェハ座標'  # X/Y座標データが入っているシート名

########## XML出力先ファイルパス ##########
Output_filepath = '//li.lumentuminc.net/data/SAG/TDS/Data/Files to Insert/XML/'  # XMLファイルの出力先パス
# Output_filepath = '../XML/003_N-electrode/'

########## 取得するデータの行番号と列番号を定義 ##########
Row_Serial_Number = 4  # シリアル番号が含まれている行（openpyxlは1から始まる）
Row_Start_Date_Time = 72  # 開始日時が含まれている行
Row_Operator = 73  # 操作者が含まれている行
Row_Polish = 83  # 磨きデータが含まれている行
col_x = 2  # X座標の列（openpyxlは1から始まる）
col_y = 3  # Y座標の列

########## 取得した項目と型の対応表を定義 ##########
key_type = {
    "key_Start_Date_Time": str,
    "key_Part_Number": str,
    "key_Serial_Number": str,
    "key_Operator": str,
    "key_Polish1": float,
    "key_Polish2": float,
    "key_Polish3": float,
    "key_Polish4": float,
    "key_Polish5": float,
    "key_Polish6": float,
    "key_Polish7": float,
    "key_Polish8": float,
    "key_Polish9": float,
    "key_Polish10": float,
    "key_Polish11": float,
    "key_Polish12": float,
    "key_Polish13": float,
    "key_X1": float,
    "key_X2": float,
    "key_X3": float,
    "key_X4": float,
    "key_X5": float,
    "key_X6": float,
    "key_X7": float,
    "key_X8": float,
    "key_X9": float,
    "key_X10": float,
    "key_X11": float,
    "key_X12": float,
    "key_X13": float,
    "key_Y1": float,
    "key_Y2": float,
    "key_Y3": float,
    "key_Y4": float,
    "key_Y5": float,
    "key_Y6": float,
    "key_Y7": float,
    "key_Y8": float,
    "key_Y9": float,
    "key_Y10": float,
    "key_Y11": float,
    "key_Y12": float,
    "key_Y13": float,
    "key_LotNumber_9": str  # 新たに追加
}

########## データの格納 ##########
def Open_Data_Sheet(filepath, Part_Number, Nine_Serial_Number):
    """
    指定された Excel ファイルを開き、データを抽出して辞書形式で返す
    """
    # ----- ログ書込：データの取得 -----
    Log.Log_Info(Log_file, 'Data Acquisition')

    # ----- ファイル形式を確認、.xlsx および .xlsm ファイルのみ処理 -----
    if not (filepath.lower().endswith('.xlsm') or filepath.lower().endswith('.xlsx')):  # 対応形式の確認
        Log.Log_Info(Log_file, f"Unsupported file format skipped: {filepath}")  # ログにスキップしたファイルを記録
        return {}  # 空の辞書を返す

    # Excel ファイルを openpyxl で処理
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)  # データだけを取得し、読み取り専用で開く
        sheet = wb[Data_Sheet_Name]  # メインデータシートを選択
    except Exception as e:
        Log.Log_Error(Log_file, f"{Data_Sheet_Name} : Failed to read file ({filepath}) - {e}\n")  # 読み取りエラーのログ記録
        return {}  # 空の辞書を返す

    # シリアル番号、開始日時、操作者のデータを取得
    try:
        Start_Date_Time = sheet.cell(row=Row_Start_Date_Time, column=3).value  # 開始日時（72行3列）
        Serial_Number = sheet.cell(row=Row_Serial_Number, column=3).value  # シリアル番号（4行3列）
        Operator = sheet.cell(row=Row_Operator, column=3).value  # 操作者（73行3列）
    except Exception as e:
        Log.Log_Error(Log_file, f"Main sheet cell read error: {filepath} - {e}\n")  # エラーログの記録
        return {}  # 空の辞書を返す

    # 磨きデータを取得
    Polish = []
    try:
        for i in range(13):  # 13個の磨きデータを取得
            Polish.append(sheet.cell(row=Row_Polish, column=4 + i).value)  # 磨きデータは83行4列から開始
    except Exception as e:
        Log.Log_Error(Log_file, f"Polish data read error: {filepath} - {e}\n")  # エラーログの記録
        return {}  # 空の辞書を返す

    # X/Y座標データを取得
    try:
        sheet_xy = wb[XY_Sheet_Name]  # 座標データのシートを選択
        x, y = [], []
        for i in range(1, 14):  # 13個の座標データを取得
            x.append(sheet_xy.cell(row=i, column=col_x).value)  # X座標（2列）
            y.append(sheet_xy.cell(row=i, column=col_y).value)  # Y座標（3列）
    except Exception as e:
        Log.Log_Error(Log_file, f"XY data read error: {filepath} - {e}\n")  # エラーログの記録
        return {}  # 空の辞書を返す

    wb.close()  # ワークブックを閉じる

    # データを辞書形式に格納
    data_dict = {
        "key_Start_Date_Time": Start_Date_Time,
        "key_Part_Number": Part_Number,
        "key_Serial_Number": Serial_Number,
        "key_LotNumber_9": Nine_Serial_Number,
        "key_Operator": Operator,
        "key_Polish1": Polish[0],
        "key_Polish2": Polish[1],
        "key_Polish3": Polish[2],
        "key_Polish4": Polish[3],
        "key_Polish5": Polish[4],
        "key_Polish6": Polish[5],
        "key_Polish7": Polish[6],
        "key_Polish8": Polish[7],
        "key_Polish9": Polish[8],
        "key_Polish10": Polish[9],
        "key_Polish11": Polish[10],
        "key_Polish12": Polish[11],
        "key_Polish13": Polish[12],
        "key_X1": x[0],
        "key_X2": x[1],
        "key_X3": x[2],
        "key_X4": x[3],
        "key_X5": x[4],
        "key_X6": x[5],
        "key_X7": x[6],
        "key_X8": x[7],
        "key_X9": x[8],
        "key_X10": x[9],
        "key_X11": x[10],
        "key_X12": x[11],
        "key_X13": x[12],
        "key_Y1": y[0],
        "key_Y2": y[1],
        "key_Y3": y[2],
        "key_Y4": y[3],
        "key_Y5": y[4],
        "key_Y6": y[5],
        "key_Y7": y[6],
        "key_Y8": y[7],
        "key_Y9": y[8],
        "key_Y10": y[9],
        "key_Y11": y[10],
        "key_Y12": y[11],
        "key_Y13": y[12],
    }

    # ----- 着工者が空欄であれば'-'を入れる -----
    if data_dict["key_Operator"] in [None, ""]:
        data_dict["key_Operator"] = '-'

    return data_dict  # データ辞書を返す

########## XML変換 ##########
def Output_XML(XML_File_Name, data_dict):
    """
    データを XML 形式に変換してファイルに出力する
    """
    # ----- ログ書込：XML変換 -----
    Log.Log_Info(Log_file, 'Excel File To XML File Conversion')

    try:
        with open(Output_filepath + XML_File_Name, 'w', encoding="utf-8") as f:  # XML ファイルを作成
            # XML ヘッダー
            f.write('<?xml version="1.0" encoding="utf-8"?>\n')
            f.write('<Results xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema">\n')
            # Result タグの開始
            f.write(f'    <Result startDateTime="{data_dict["key_Start_Date_Time"].replace(".", ":")}" Result="Done">\n')
            # Header タグ
            f.write(f'        <Header SerialNumber="{data_dict["key_Serial_Number"]}" PartNumber="{data_dict["key_Part_Number"]}" Operation="{Operation}" TestStation="{TestStation}" Operator="{data_dict["key_Operator"]}" StartTime="{data_dict["key_Start_Date_Time"].replace(".", ":")}" Site="{Site}" LotNumber="{data_dict["key_LotNumber_9"]}"/>\n\n')
            # TestStep タグの生成
            f.write(generate_test_steps(data_dict))
            # TestEquipment タグ
            f.write('        <TestEquipment>\n')
            f.write('            <Item DeviceName="Stepmeter" DeviceSerialNumber="1"/>\n')
            f.write('        </TestEquipment>\n\n')
            # ErrorData, FailureData, Configuration タグ
            f.write('        <ErrorData/>\n')
            f.write('        <FailureData/>\n')
            f.write('        <Configuration/>\n')
            # Result タグの終了
            f.write('    </Result>\n')
            f.write('</Results>')
        Log.Log_Info(Log_file, f'Successfully generated XML file: {Output_filepath + XML_File_Name}')  # 成功ログ
    except Exception as e:
        Log.Log_Error(Log_file, f'XML file generation error: {Output_filepath + XML_File_Name}, Error: {e}')  # エラーログ

########## TestStep の生成 ##########
def generate_test_steps(data_dict):
    """
    TestStep の部分を XML 形式で生成する
    """
    test_steps = ""
    for i in range(1, 14):  # 13個の TestStep を生成
        # startDateTime のフォーマットを修正
        start_time = data_dict["key_Start_Date_Time"].replace(".", ":") if isinstance(data_dict["key_Start_Date_Time"], str) else data_dict["key_Start_Date_Time"]
        test_steps += f'        <TestStep Name="Thickness{i}" startDateTime="{start_time}" Status="Done">\n'
        # X の値が数値でない場合は 0 に設定
        x_value = data_dict.get(f"key_X{i}", 0)
        if not isinstance(x_value, (int, float)):
            x_value = 0
        # Y の値が数値でない場合は 0 に設定
        y_value = data_dict.get(f"key_Y{i}", 0)
        if not isinstance(y_value, (int, float)):
            y_value = 0
        # Polish の値が数値でない場合は 0 に設定
        polish_value = data_dict.get(f"key_Polish{i}", 0)
        if not isinstance(polish_value, (int, float)):
            polish_value = 0
        # Data タグの追加
        test_steps += f'            <Data DataType="Numeric" Name="X" Units="um" Value="{x_value}"/>\n'
        test_steps += f'            <Data DataType="Numeric" Name="Y" Units="um" Value="{y_value}"/>\n'
        test_steps += f'            <Data DataType="Numeric" Name="Thickness" Units="um" Value="{polish_value}"/>\n'
        test_steps += '        </TestStep>\n'
    # SORTED_DATA TestStep の追加
    test_steps += f'        <TestStep Name="SORTED_DATA" startDateTime="{start_time}" Status="Passed">\n'
    test_steps += f'            <Data DataType="String" Name="LotNumber_5" Value="{data_dict["key_Serial_Number"]}" CompOperation="LOG"/>\n'
    test_steps += f'            <Data DataType="String" Name="LotNumber_9" Value="{data_dict["key_LotNumber_9"]}" CompOperation="LOG"/>\n'
    test_steps += '        </TestStep>\n'
    return test_steps  # 生成したテストステップの文字列を返す

########## main処理 ##########
def main(File_Path, Part_Number, Nine_Serial_Number):
    """
    メイン処理：データを取得し、XMLファイルを生成する
    """
    # ----- ログ書込：オペレーション名 -----
    Log.Log_Info(Log_file, Operation)

    ########## データ取得 ##########
    data_dict = Open_Data_Sheet(File_Path, Part_Number, Nine_Serial_Number)

    # データが取得できなかった場合は処理を終了
    if not data_dict:
        Log.Log_Error(Log_file, f"{Part_Number} : Data acquisition failed\n")
        return  # 処理をスキップ

    ########## 日付フォーマット変換 ##########
    try:
        data_dict["key_Start_Date_Time"] = Convert_Date.Edit_Date(data_dict["key_Start_Date_Time"])
    except Exception as e:
        Log.Log_Error(Log_file, f"{data_dict['key_Serial_Number']} : Date format conversion error - {e}\n")
        return  # 処理をスキップ

    ########## 空欄チェック ##########
    for key, val in data_dict.items():
        if val in [None, ""]:
            Log.Log_Error(Log_file, f'{data_dict["key_Serial_Number"]} : Blank Error in {key}\n')
            return  # 処理をスキップ

    ########## データ型の確認 ##########
    result = Check.Data_Type(key_type, data_dict)
    if not result:
        Log.Log_Error(Log_file, f'{data_dict["key_Serial_Number"]} : Data Type Error\n')
        return  # データ型エラーの場合、処理をスキップ

    ########## XMLファイルの作成 ##########
    try:
        xml_file = f'Site={Site},ProductFamily={ProductFamily},Operation={Operation},Partnumber={data_dict["key_Part_Number"]},Serialnumber={data_dict["key_Serial_Number"]},Testdate={data_dict["key_Start_Date_Time"]}.xml'
        Output_XML(xml_file, data_dict)
    except Exception as e:
        Log.Log_Error(Log_file, f'{data_dict["key_Serial_Number"]} : XML file creation error - {e}\n')

# プログラム終了ログ
Log.Log_Info(Log_file, 'Program Ended')
