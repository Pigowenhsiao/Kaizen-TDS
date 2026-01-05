import logging  # ログ記録用のloggingモジュールをインポート
import openpyxl  # Excelファイル操作用のopenpyxlモジュールをインポート
import glob  # ファイルパスのパターンマッチング用のglobモジュールをインポート
import os  # OS操作用のosモジュールをインポート
import sys  # システム関連の機能用のsysモジュールをインポート

from datetime import datetime, timedelta, date  # datetimeモジュールからdatetime, timedelta, dateクラスをインポート
from time import strftime  # 時間のフォーマット用にstrftimeをインポート

########## 自作関数の定義 ##########
sys.path.append('../MyModule')  # カスタムモジュールのパスをシステムパスに追加
import Log  # カスタムのLogモジュールをインポート
import Convert_Date  # カスタムのConvert_Dateモジュールをインポート
import Check  # カスタムのCheckモジュールをインポート

########## 全体パラメータ定義 ##########
Site = '350'  # サイト番号を定義
ProductFamily = 'SAG FAB'  # 製品ファミリーを定義
Operation = 'N-electrode_Polish_InitialWaferThickness'  # 操作名を定義
TestStation = 'N-electrode'  # テストステーションを定義

########## ログファイルの定義 ##########
Log_FolderName = str(date.today())  # 今日の日付をログフォルダ名として設定
Log_file = '../Log/' + Log_FolderName + '/003_N-electrode.log'  # ログファイルのパスを定義

########## シート名の定義 ##########
Data_Sheet_Name = '3ｲﾝﾁ用'  # メインデータのシート名を定義
XY_Sheet_Name = 'ウェハ座標'  # 座標データのシート名を定義

########## XML出力先ファイルパス ##########
Output_filepath = '//li.lumentuminc.net/data/SAG/TDS/Data/Files to Insert/XML/'  # XMLファイルの出力パスを定義
# Output_filepath = '../XML/003_N-electrode/'  # 別の可能なXML出力パス（コメントアウト）

########## 取得するデータの行番号と列番号を定義 ##########
Row_Serial_Number = 3  # シリアル番号がある行番号（4行目）
Row_Start_Date_Time = 36  # 開始日時がある行番号（37行目）
Row_Operator = 38  # 操作者がいる行番号（39行目）
Row_Polish = 46  # ポリッシュデータがある行番号（47行目）
col_x = 1  # X座標がある列番号（2列目）
col_y = 2  # Y座標がある列番号（3列目）

########## 取得した項目と型の対応表を定義 ##########
key_type = {
    "key_Start_Date_Time": str,  # 開始日時の型
    "key_Part_Number": str,  # 部品番号の型
    "key_Serial_Number": str,  # シリアル番号の型
    "key_Operator": str,  # 操作者の型
    "key_Polish1": float,  # ポリッシュ1の型
    "key_Polish2": float,  # ポリッシュ2の型
    "key_Polish3": float,  # ポリッシュ3の型
    "key_Polish4": float,  # ポリッシュ4の型
    "key_Polish5": float,  # ポリッシュ5の型
    "key_Polish6": float,  # ポリッシュ6の型
    "key_Polish7": float,  # ポリッシュ7の型
    "key_Polish8": float,  # ポリッシュ8の型
    "key_Polish9": float,  # ポリッシュ9の型
    "key_Polish10": float,  # ポリッシュ10の型
    "key_Polish11": float,  # ポリッシュ11の型
    "key_Polish12": float,  # ポリッシュ12の型
    "key_Polish13": float,  # ポリッシュ13の型
    "key_X1": float,  # X1座標の型
    "key_X2": float,  # X2座標の型
    "key_X3": float,  # X3座標の型
    "key_X4": float,  # X4座標の型
    "key_X5": float,  # X5座標の型
    "key_X6": float,  # X6座標の型
    "key_X7": float,  # X7座標の型
    "key_X8": float,  # X8座標の型
    "key_X9": float,  # X9座標の型
    "key_X10": float,  # X10座標の型
    "key_X11": float,  # X11座標の型
    "key_X12": float,  # X12座標の型
    "key_X13": float,  # X13座標の型
    "key_Y1": float,  # Y1座標の型
    "key_Y2": float,  # Y2座標の型
    "key_Y3": float,  # Y3座標の型
    "key_Y4": float,  # Y4座標の型
    "key_Y5": float,  # Y5座標の型
    "key_Y6": float,  # Y6座標の型
    "key_Y7": float,  # Y7座標の型
    "key_Y8": float,  # Y8座標の型
    "key_Y9": float,  # Y9座標の型
    "key_Y10": float,  # Y10座標の型
    "key_Y11": float,  # Y11座標の型
    "key_Y12": float,  # Y12座標の型
    "key_Y13": float  # Y13座標の型
}

########## データの格納 ##########
def Open_Data_Sheet(filepath, Part_Number, Nine_Serial_Number):
    # ----- ログ書込：データの取得 -----
    Log.Log_Info(Log_file, 'Data Acquisition')  # 記録開始データ獲取

    # ----- 確認ファイル形式、.xlsm と .xlsx の両方を処理 -----
    if not (filepath.lower().endswith('.xlsm') or filepath.lower().endswith('.xlsx')):  # ファイルが .xlsm または .xlsx でない場合
        Log.Log_Info(Log_file, f"Skipping unsupported file format: {filepath}")  # スキップをログに記録
        return {}  # 空の辞書を返してファイルをスキップ

    # openpyxl を使用して Excel ファイルを処理
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, keep_vba=True)  # Excel ワークブックを読み込む（VBA保持）
        sheet = wb[Data_Sheet_Name]  # 指定されたシートを選択
    except Exception as e:
        Log.Log_Error(Log_file, f"Failed to load sheet {Data_Sheet_Name}. Unsupported file format: {filepath}\n")  # エラーメッセージをログに記録
        return {}  # 空の辞書を返してファイルをスキップ

    # シリアル番号、開始日時、操作者を取得
    try:
        Start_Date_Time = sheet.cell(row=Row_Start_Date_Time + 1, column=3).value  # 開始日時を取得（37行目3列目）
        Serial_Number = sheet.cell(row=Row_Serial_Number + 1, column=3).value  # シリアル番号を取得（4行目3列目）
        Operator = sheet.cell(row=Row_Operator + 1, column=3).value  # 操作者を取得（39行目3列目）
    except Exception as e:
        Log.Log_Error(Log_file, f"Error reading main sheet cells in {filepath}: {e}\n")  # エラーメッセージをログに記録
        return {}  # 空の辞書を返してファイルをスキップ

    # ----- ガラス盤研磨 最終ウェハ厚の格納 -----
    Polish = []
    try:
        for i in range(13):  # 13個のポリッシュデータを取得
            Polish.append(sheet.cell(row=Row_Polish + 1, column=4 + i).value)  # 47行目の4～16列目を取得
    except Exception as e:
        Log.Log_Error(Log_file, f"Error reading Polish data in {filepath}: {e}\n")  # エラーメッセージをログに記録
        return {}  # 空の辞書を返してファイルをスキップ

    # ----- X/Y座標の格納 -----
    try:
        sheet_xy = wb[XY_Sheet_Name]  # 座標シートを選択
        x, y = [], []  # X座標とY座標のリストを初期化
        for i in range(1, 14):  # 1～13行目のX/Y座標を取得
            x.append(sheet_xy.cell(row=i, column=col_x + 1).value)  # X座標を取得（2列目）
            y.append(sheet_xy.cell(row=i, column=col_y + 1).value)  # Y座標を取得（3列目）
    except Exception as e:
        Log.Log_Error(Log_file, f"Error reading XY data in {filepath}: {e}\n")  # エラーメッセージをログに記録
        return {}  # 空の辞書を返してファイルをスキップ

    # ----- 辞書型に格納 ------
    data_dict = {
        "key_Start_Date_Time": Start_Date_Time,  # 開始日時
        "key_Part_Number": Part_Number,  # 部品番号
        "key_Serial_Number": Serial_Number,  # シリアル番号
        "key_LotNumber_9": Nine_Serial_Number,  # ロット番号9
        "key_Operator": Operator,  # 操作者
        "key_Polish1": Polish[0],  # ポリッシュ1
        "key_Polish2": Polish[1],  # ポリッシュ2
        "key_Polish3": Polish[2],  # ポリッシュ3
        "key_Polish4": Polish[3],  # ポリッシュ4
        "key_Polish5": Polish[4],  # ポリッシュ5
        "key_Polish6": Polish[5],  # ポリッシュ6
        "key_Polish7": Polish[6],  # ポリッシュ7
        "key_Polish8": Polish[7],  # ポリッシュ8
        "key_Polish9": Polish[8],  # ポリッシュ9
        "key_Polish10": Polish[9],  # ポリッシュ10
        "key_Polish11": Polish[10],  # ポリッシュ11
        "key_Polish12": Polish[11],  # ポリッシュ12
        "key_Polish13": Polish[12],  # ポリッシュ13
        "key_X1": x[0],  # X1座標
        "key_X2": x[1],  # X2座標
        "key_X3": x[2],  # X3座標
        "key_X4": x[3],  # X4座標
        "key_X5": x[4],  # X5座標
        "key_X6": x[5],  # X6座標
        "key_X7": x[6],  # X7座標
        "key_X8": x[7],  # X8座標
        "key_X9": x[8],  # X9座標
        "key_X10": x[9],  # X10座標
        "key_X11": x[10],  # X11座標
        "key_X12": x[11],  # X12座標
        "key_X13": x[12],  # X13座標
        "key_Y1": y[0],  # Y1座標
        "key_Y2": y[1],  # Y2座標
        "key_Y3": y[2],  # Y3座標
        "key_Y4": y[3],  # Y4座標
        "key_Y5": y[4],  # Y5座標
        "key_Y6": y[5],  # Y6座標
        "key_Y7": y[6],  # Y7座標
        "key_Y8": y[7],  # Y8座標
        "key_Y9": y[8],  # Y9座標
        "key_Y10": y[9],  # Y10座標
        "key_Y11": y[10],  # Y11座標
        "key_Y12": y[11],  # Y12座標
        "key_Y13": y[12],  # Y13座標
    }

    # ----- 操作者が空欄であれば'-'を入れる ------
    if data_dict["key_Operator"] == "":
        data_dict["key_Operator"] = '-'  # 操作者が空の場合、'-'を設定

    return data_dict  # データ辞書を返す

########## XML変換 ##########
def Output_XML(XML_File_Name, data_dict):
    # ----- ログ書込:XML変換 -----
    Log.Log_Info(Log_file, 'Excel File To XML File Conversion')  # 記録開始 XML 轉換

    try:
        with open(Output_filepath + XML_File_Name, 'w', encoding="utf-8") as f:  # XMLファイルを開いて書き込み
            f.write('<?xml version="1.0" encoding="utf-8"?>\n' +
                    '<Results xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema">\n' +
                    f'       <Result startDateTime="{data_dict["key_Start_Date_Time"].replace(".", ":")}" Result="Done">\n' +
                    f'               <Header SerialNumber="{data_dict["key_Serial_Number"]}" PartNumber="{data_dict["key_Part_Number"]}" Operation="{Operation}" TestStation="{TestStation}" Operator="{data_dict["key_Operator"]}" StartTime="{data_dict["key_Start_Date_Time"].replace(".", ":")}" Site="{Site}" LotNumber="{data_dict["key_Serial_Number"]}"/>\n\n' +
                    generate_test_steps(data_dict) +  # 生成 TestStep1 から TestStep13
                    f'               <TestStep Name="SORTED_DATA" startDateTime="{data_dict["key_Start_Date_Time"].replace(".", ":")}" Status="Passed">\n' +
                    f'                   <Data DataType="String" Name="LotNumber_5" Value="{data_dict["key_Serial_Number"]}" CompOperation="LOG"/>\n' +
                    f'                   <Data DataType="String" Name="LotNumber_9" Value="{data_dict["key_LotNumber_9"]}" CompOperation="LOG"/>\n' +
                    f'               </TestStep>\n\n' +
                    f'               <TestEquipment>\n' +
                    f'                   <Item DeviceName="Stepmeter" DeviceSerialNumber="1"/>\n' +
                    f'               </TestEquipment>\n\n' +
                    f'               <ErrorData/>\n' +
                    f'               <FailureData/>\n' +
                    f'               <Configuration/>\n' +
                    f'       </Result>\n' +
                    f'</Results>'
                    )
        Log.Log_Info(Log_file, f'Successfully generated XML file: {Output_filepath + XML_File_Name}')  # 成功メッセージをログに記録
    except Exception as e:
        Log.Log_Error(Log_file, f'Failed to generate XML file: {Output_filepath + XML_File_Name}, Error: {e}')  # エラーメッセージをログに記録

def generate_test_steps(data_dict):
    """生成 TestStep 部分の XML 内容"""
    test_steps = ""
    for i in range(1, 14):  # 1から13までのTestStepを生成
        test_steps += f'               <TestStep Name="Thickness{i}" startDateTime="{data_dict["key_Start_Date_Time"].replace(".", ":")}" Status="Done">\n' + \
                      f'                   <Data DataType="Numeric" Name="X" Units="um" Value="{data_dict[f"key_X{i}"]}"/>\n' + \
                      f'                   <Data DataType="Numeric" Name="Y" Units="um" Value="{data_dict[f"key_Y{i}"]}"/>\n' + \
                      f'                   <Data DataType="Numeric" Name="Thickness" Units="um" Value="{data_dict[f"key_Polish{i}"]}"/>\n' + \
                      f'               </TestStep>\n'
    return test_steps

########## main処理 ##########
def main(File_Path, Part_Number, Nine_Serial_Number):
    # ----- ログ書込：オペレーション名 -----
    Log.Log_Info(Log_file, Operation)  # 記録操作名称

    ########## データ取得 ##########
    data_dict = Open_Data_Sheet(File_Path, Part_Number, Nine_Serial_Number)  # データを取得

    if not data_dict:  # データ辞書が空の場合
        Log.Log_Error(Log_file, f"{Part_Number} : Data Acquisition Failed\n")  # エラーメッセージをログに記録
        return  # 処理を終了

    ########## 日付フォーマット変換 ##########
    data_dict["key_Start_Date_Time"] = Convert_Date.Edit_Date(data_dict["key_Start_Date_Time"])  # 日付フォーマットを変換

    ########## 空欄チェック ##########
    for val in data_dict.values():  # データ辞書内のすべての値をチェック
        if val == "":
            Log.Log_Error(Log_file, f'{data_dict["key_Serial_Number"]} : Blank Error\n')  # 空欄エラーをログに記録
            return  # 処理を終了

    ########## データ型の確認 ##########
    result = Check.Data_Type(key_type, data_dict)  # データ型をチェック
    if result == False:
        Log.Log_Error(Log_file, f'{data_dict["key_Serial_Number"]} : Data Error\n')  # データ型エラーをログに記録
        return  # 処理を終了

    ########## XMLファイルの作成 ##########
    xml_file = f'Site={Site},ProductFamily={ProductFamily},Operation={Operation},Partnumber={data_dict["key_Part_Number"]},Serialnumber={data_dict["key_Serial_Number"]},Testdate={data_dict["key_Start_Date_Time"]}.xml'  # XMLファイル名を定義

    Output_XML(xml_file, data_dict)  # XMLファイルを生成
