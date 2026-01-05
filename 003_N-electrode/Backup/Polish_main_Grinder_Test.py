import logging
import shutil
import glob
import xlrd
import openpyxl  # 新增：導入 openpyxl
import os
import sys

from datetime import datetime, timedelta, date

########## サブプログラムの定義 ##########
from Polish_Func_Grinder import InitialWaferThickness
from Polish_Func_Grinder import WaxThickness
from Polish_Func_Grinder import RoughPolishedThickness
from Polish_Func_Grinder import MirrorPolishedThickness
from Polish_Func_Grinder import EtchedThickness

########## 自作関数の定義 ##########
sys.path.append('../MyModule')
import SQL
import Log
import Convert_Date

########## ログの設定 ##########

# ----- Logフォルダ名の定義 -----
Log_FolderName = str(date.today())

# ----- 格納するLogフォルダがなければ作成する -----
if not os.path.exists("../Log/" + Log_FolderName):
    os.makedirs("../Log/" + Log_FolderName)

# ----- ログ書き込み先パスの定義 -----
Log_file = '../Log/' + Log_FolderName + '/003_N-electrode.log'

########## プログラムの開始 ##########
Log.Log_Info(Log_file, 'Program Start')

########## 処理シートの定義 ##########
Data_sheet_name = '3ｲﾝﾁ用'
XY_sheet_name = 'ウェハ座標'

########## エラーの時に取得される番号を定義 ##########
Error_Number = {'empty:\'\'', 'error:7', 'error:15', 'error:23', 'error:29', 'error:36', 'error:42', 'xldate:0.0'}

########## 空欄チェック ##########
def Get_Cells_Info(filepath):

    # ----- ログ書込：空欄判定 -----
    Log.Log_Info(Log_file, "Blank Check")

    # ----- False -> 空欄がない -----
    is_cells_empty = False

    # ----- ファイル拡張子の判定 -----
    _, ext = os.path.splitext(filepath)

    try:
        if ext.lower() == '.xls':
            # 使用 xlrd 來讀取 .xls 文件
            wb = xlrd.open_workbook(filepath, on_demand=True)
            try:
                sheet = wb.sheet_by_name(Data_sheet_name)
                sheetx = wb.sheet_by_name(XY_sheet_name)
            except:
                Log.Log_Error(Log_file, Data_sheet_name + ' : ' + "Not Exist\n")
                is_cells_empty = True
                return is_cells_empty

            # ----- 一つでも空欄が存在すれば処理を行わない -----
            cells_to_check = [
                sheet.cell(3, 2),
                sheet.cell(4, 2),
                sheet.cell(23, 2),
                sheet.cell(24, 2),
                sheet.cell(35, 2),
                sheet.cell(36, 2),
                sheet.cell(54, 2),
                sheet.cell(55, 2)
            ]

            for cell in cells_to_check:
                if str(cell.value) in Error_Number:
                    is_cells_empty = True
                    break

            wb.release_resources()

        elif ext.lower() in ['.xlsx', '.xlsm']:
            # 使用 openpyxl 來讀取 .xlsx 和 .xlsm 文件
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
            try:
                sheet = wb[Data_sheet_name]
                sheetx = wb[XY_sheet_name]
            except KeyError:
                Log.Log_Error(Log_file, Data_sheet_name + ' : ' + "Not Exist\n")
                is_cells_empty = True
                wb.close()
                return is_cells_empty

            # ----- 一つでも空欄が存在すれば処理を行わない -----
            cells_to_check = [
                sheet.cell(row=4, column=3).value,  # Excelの行は1始まり
                sheet.cell(row=5, column=3).value,
                sheet.cell(row=24, column=3).value,
                sheet.cell(row=25, column=3).value,
                sheet.cell(row=36, column=3).value,
                sheet.cell(row=37, column=3).value,
                sheet.cell(row=55, column=3).value,
                sheet.cell(row=56, column=3).value
            ]

            for value in cells_to_check:
                if str(value) in Error_Number:
                    is_cells_empty = True
                    break

            wb.close()

        else:
            Log.Log_Error(Log_file, f"Unsupported file extension: {ext}")
            is_cells_empty = True

    except Exception as e:
        Log.Log_Error(Log_file, f"Error processing file {filepath}: {e}")
        is_cells_empty = True

    return is_cells_empty

########## 本処理 ##########
if __name__ == '__main__':

    # ----- Main処理の開始 -----
    Log.Log_Info(Log_file, 'Main Start')

    # ----- 処理フォルダの定義 → 今日の年/月/日を取り出す -----
    today = str(date.today()).split('-')  # ['2020', '08', '26']の構造で受け取られる

    ########### ディレクトリ検索 ##########
    Log.Log_Info(Log_file, 'Directory Search')

    FilePath = 'Z:/研磨/2024年/'
    FolderName = '*年*月'
    Folder_List = []
    for folder in glob.glob(FilePath + FolderName):
        Folder_List.append(folder)

    # ----- 対象フォルダが存在しない → エラーログに書き込む -----
    if len(Folder_List) == 0:
        Log.Log_Error(Log_file, 'Folder Error')
        sys.exit()

    Folder_List.sort()
    directory = Folder_List[0] + '/'

    ########### 処理済みフォルダがなければ作成する ##########
    processed_folder = os.path.join(directory, '処理済みフォルダ')
    if not os.path.exists(processed_folder):
        os.mkdir(processed_folder)

    ########### 対象ファイルの処理 ##########
    for File_Path in glob.glob(os.path.join(directory, '*.xls*')):

        # ----- 処理を行うファイルの出力 -----
        Log.Log_Info(Log_file, File_Path)

        # ----- 空欄チェック -----
        if Get_Cells_Info(File_Path):
            Log.Log_Error(Log_file, "Blank Error\n")
            continue

        # ----- ロット番号と日付の取得 -----
        _, ext = os.path.splitext(File_Path)
        Serial_Number = ""
        Date_Array = []

        try:
            if ext.lower() == '.xls':
                # 使用 xlrd 來讀取 .xls 文件
                wb = xlrd.open_workbook(File_Path, on_demand=True)
                sheet = wb.sheet_by_name(Data_sheet_name)
                Serial_Number = str(sheet.cell(3, 2).value)

                Date_Array = [
                    sheet.cell(4, 2).value,
                    sheet.cell(24, 2).value,
                    sheet.cell(36, 2).value,
                    sheet.cell(55, 2).value
                ]

                wb.release_resources()

            elif ext.lower() in ['.xlsx', '.xlsm']:
                # 使用 openpyxl 來讀取 .xlsx 和 .xlsm 文件
                wb = openpyxl.load_workbook(File_Path, data_only=True, read_only=True)
                sheet = wb[Data_sheet_name]
                Serial_Number = str(sheet.cell(row=4, column=3).value)

                Date_Array = [
                    sheet.cell(row=5, column=3).value,
                    sheet.cell(row=25, column=3).value,
                    sheet.cell(row=37, column=3).value,
                    sheet.cell(row=56, column=3).value
                ]

                wb.close()

            else:
                Log.Log_Error(Log_file, f"Unsupported file extension: {ext}")
                continue

        except Exception as e:
            Log.Log_Error(Log_file, f"Error reading file {File_Path}: {e}\n")
            continue

        ########## Primeに接続し、品名を取得 ##########

        # ----- Primeへの接続 ------
        conn, cursor = SQL.connSQL()

        # ----- Primeとの接続に失敗した場合、処理を打ち切る -----
        if conn is None:
            Log.Log_Error(Log_file, Serial_Number + ' : ' + 'Connection with Prime Failed\n')
            sys.exit()

        # ----- 品名を取得 -----
        
        Part_Number, Nine_Serial_Number = SQL.selectSQL(cursor, Serial_Number)
        print('Serial_Number:', Serial_Number, 'Part_Number:', Part_Number, 'Nine_Serial_Number:', Nine_Serial_Number)
        SQL.disconnSQL(conn, cursor)

        # ----- 品名がNoneが見つからなかった -----
        if Part_Number is None:
            Log.Log_Error(Log_file, Serial_Number + ' : ' + "Part Number Error\n")
            continue

        # ----- 品名が LDアレイ_ のときは登録できないため、次のファイルへ遷移 -----
        if Part_Number == 'LDアレイ_':
            continue

        ########## 日付フォーマットの変換 ##########
        try:
            if Convert_Date.Edit_Date(Date_Array[0]) == "" or Convert_Date.Edit_Date(Date_Array[1]) == "" or Convert_Date.Edit_Date(Date_Array[2]) == "" or Convert_Date.Edit_Date(Date_Array[3]) == "":
                Log.Log_Error(Log_file, Serial_Number + ' : ' + "日付フォーマットエラー\n")
                continue
        except Exception as e:
            Log.Log_Error(Log_file, f"{Serial_Number} : Date Conversion Error: {e}\n")
            continue

        ########## 各Operationに渡しXML変換を行う ##########
        Log.Log_Info(Log_file, 'Move To Each Function')
        try:
            InitialWaferThickness.main(File_Path, Part_Number, Nine_Serial_Number)
            WaxThickness.main(File_Path, Part_Number, Nine_Serial_Number)
            RoughPolishedThickness.main(File_Path, Part_Number, Nine_Serial_Number)
            MirrorPolishedThickness.main(File_Path, Part_Number, Nine_Serial_Number)
            EtchedThickness.main(File_Path, Part_Number, Nine_Serial_Number)
        except Exception as e:
            Log.Log_Error(Log_file, f"Error in processing functions: {e}\n")
            continue

        ########### 処理済みフォルダに移動 ##########
        Log.Log_Info(Log_file, 'Move File')
        try:
            shutil.move(File_Path, processed_folder)
            Log.Log_Info(Log_file, f"Successfully moved {File_Path} to {processed_folder}")
        except PermissionError:
            Log.Log_Error(Log_file, "Permission Error")
        except Exception as e:
            Log.Log_Error(Log_file, f"Error moving file: {e}")

    ########## Main処理の終了 ##########
    Log.Log_Info(Log_file, 'Program End')
