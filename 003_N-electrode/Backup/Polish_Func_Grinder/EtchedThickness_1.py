import logging  # ログ記録モジュールをインポート
import openpyxl  # Excel ファイルを処理するためのモジュールをインポート
import glob  # ファイルパスを検索するためのモジュールをインポート
import os  # OS関連の機能を操作するためのモジュールをインポート
import sys  # システム関連の機能を操作するためのモジュールをインポート
from datetime import datetime, timedelta, date  # 日付と時間を処理するためのクラスをインポート
from time import strftime  # 時間フォーマットを変更するための関数をインポート

########## 自作関数の定義 ##########
sys.path.append('../MyModule')  # 自作モジュールのパスをシステムパスに追加
import Log  # ログ記録モジュールをインポート
import Convert_Date  # 日付フォーマット変換モジュールをインポート
import Check  # データ型を確認するモジュールをインポート

########## 全体パラメータ定義 ##########
Site = '350'  # サイトの番号を設定
ProductFamily = 'SAG FAB'  # 製品ファミリーを設定
Operation = 'N-electrode_Polish_EtchedThickness'  # 操作名を設定
TestStation = 'N-electrode'  # テストステーション名を設定

########## ログファイルの定義 ##########
Log_FolderName = str(date.today())  # 今日の日付をログフォルダ名として使用
Log_file = '../Log/' + Log_FolderName + '/003_N-electrode.log'  # ログファイルのパスを定義

########## シート名の定義 ##########
Data_Sheet_Name = '3ｲﾝﾁ用'  # データが入っているシート名
XY_Sheet_Name = 'ウェハ座標'  # X/Y座標データが入っているシート名

########## XML出力先ファイルパス ##########
Output_filepath = '//li.lumentuminc.net/data/SAG/TDS/Data/Files to Insert/XML/'  # XMLファイルの出力先パス

########## 取得すRow_Serial_Number = 3  # シリアル番号が含まれている行
Row_Serial_Number = 54  # シリアル番号がある行番号
Row_Start_Date_Time = 55  # 開始日時が含まれている行
Row_Operator = 56  # 操作者が含まれている行
Row_Polish = 62  # 磨きデータが含まれている行
col_x = 1  # X座標の列
col_y = 2  # Y座標の列るデータの列番号を定義 ##########


########## 取得した項目と型の対応表を定義 ##########
key_type = {
    "key_Start_Date_Time": str,  # 開始日時は文字列型
    "key_Part_Number": str,  # パーツ番号は文字列型
    "key_Serial_Number": str,  # シリアル番号は文字列型
    "key_Operator": str,  # 操作者は文字列型
    "key_Polish1": float,  # 磨きデータは浮動小数点数
    # その他の Polishing データの型定義が続く...
    "key_X1": float,  # X座標データも浮動小数点数
    "key_Y1": float  # Y座標データも浮動小数点数
}

########## データの格納 ##########
def Open_Data_Sheet(filepath, Part_Number, Nine_Serial_Number):
    """
    指定された Excel ファイルを開き、データを抽出して辞書形式で返す
    """
    # ----- ログ書込：データ取得の開始 -----
    Log.Log_Info(Log_file, 'データ取得開始')

    # ----- ファイル形式を確認、.xlsx および .xlsm ファイルのみ処理 -----
    if not (filepath.lower().endswith('.xlsm') or filepath.lower().endswith('.xlsx')):  # 対応形式の確認
        Log.Log_Info(Log_file, f"非対応形式のファイルをスキップ: {filepath}")  # ログにスキップしたファイルを記録
        return {}  # 空の辞書を返す

    # Excel ファイルを openpyxl で処理
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)  # データだけを取得する
        sheet = wb[Data_Sheet_Name]  # メインデータシートを選択
    except Exception as e:
        Log.Log_Error(Log_file, f"{Data_Sheet_Name} : ファイルの読み取りに失敗しました ({filepath})\n")  # 読み取りエラーのログ記録
        return {}  # 空の辞書を返す

    # シリアル番号、開始日時、操作者のデータを取得
    try:
        Start_Date_Time = sheet.cell(row=Row_Start_Date_Time + 1, column=3).value  # 開始日時（56行3列）
        Serial_Number = sheet.cell(row=Row_Serial_Number + 1, column=3).value  # シリアル番号（4行3列）
        Operator = sheet.cell(row=Row_Operator + 1, column=3).value  # 操作者（57行3列）
    except Exception as e:
        Log.Log_Error(Log_file, f"メインシートのセル読み取りエラー: {filepath}: {e}\n")  # エラーログの記録
        return {}  # 空の辞書を返す

    # 磨きデータを取得
    Polish = []
    try:
        for i in range(13):  # 13個の磨きデータを取得
            Polish.append(sheet.cell(row=Row_Polish + 1, column=4 + i).value)  # 磨きデータは63行4列から開始
    except Exception as e:
        Log.Log_Error(Log_file, f"磨きデータの読み取りエラー: {filepath}: {e}\n")  # エラーログの記録
        return {}  # 空の辞書を返す

    # X/Y座標データを取得
    try:
        sheet_xy = wb[XY_Sheet_Name]  # 座標データのシートを選択
        x, y = [], []
        for i in range(1, 14):  # 13個の座標データを取得
            x.append(sheet_xy.cell(row=i, column=col_x + 1).value)  # X座標（第2列）
            y.append(sheet_xy.cell(row=i, column=col_y + 1).value)  # Y座標（第3列）
    except Exception as e:
        Log.Log_Error(Log_file, f"XYデータの読み取りエラー: {filepath}: {e}\n")  # エラーログの記録
        return {}  # 空の辞書を返す

    # データを辞書形式に格納
    data_dict = {
        "key_Start_Date_Time": Start_Date_Time,
        "key_Part_Number": Part_Number,
        "key_Serial_Number": Serial_Number,
        "key_LotNumber_9": Nine_Serial_Number,
        "key_Operator": Operator,
        "key_Polish1": Polish[0],
        "key_Polish2": Polish[1],
        "key_Polish3": Polish[2],
        "key_Polish4": Polish[3],
        "key_Polish5": Polish[4],
        "key_Polish6": Polish[5],
        "key_Polish7": Polish[6],
        "key_Polish8": Polish[7],
        "key_Polish9": Polish[8],
        "key_Polish10": Polish[9],
        "key_Polish11": Polish[10],
        "key_Polish12": Polish[11],
        "key_Polish13": Polish[12],
        "key_X1": x[0],
        "key_X2": x[1],
        "key_X3": x[2],
        "key_X4": x[3],
        "key_X5": x[4],
        "key_X6": x[5],
        "key_X7": x[6],
        "key_X8": x[7],
        "key_X9": x[8],
        "key_X10": x[9],
        "key_X11": x[10],
        "key_X12": x[11],
        "key_X13": x[12],
        "key_Y1": y[0],
        "key_Y2": y[1],
        "key_Y3": y[2],
        "key_Y4": y[3],
        "key_Y5": y[4],
        "key_Y6": y[5],
        "key_Y7": y[6],
        "key_Y8": y[7],
        "key_Y9": y[8],
        "key_Y10": y[9],
        "key_Y11": y[10],
        "key_Y12": y[11],
        "key_Y13": y[12]
    }

    # 操作者欄が空欄の場合 '-' を入力
    if data_dict["key_Operator"] == "":
        data_dict["key_Operator"] = '-'  # 操作者欄が空欄のとき、デフォルト値として '-' を設定

    return data_dict  # データ辞書を返す

########## XML変換 ##########
def Output_XML(XML_File_Name, data_dict):
    """
    データを XML 形式に変換してファイルに出力する
    """
    # ----- ログ書込：XML変換開始 -----
    Log.Log_Info(Log_file, 'Excel ファイルを XML ファイルに変換中')

    try:
        with open(Output_filepath + XML_File_Name, 'w', encoding="utf-8") as f:  # XML ファイルを作成
            f.write('<?xml version="1.0" encoding="utf-8"?>\n')
            f.write('<Results>\n')
            f.write(f'    <Result startDateTime="{data_dict["key_Start_Date_Time"].replace(".", ":")}" Result="Done">\n')
            f.write(f'        <Header SerialNumber="{data_dict["key_Serial_Number"]}" PartNumber="{data_dict["key_Part_Number"]}" Operator="{data_dict["key_Operator"]}" />\n')
            f.write(generate_test_steps(data_dict))  # テストステップの XML 生成
            f.write('    </Result>\n')
            f.write('</Results>')
        Log.Log_Info(Log_file, f'成功生成 XML ファイル: {Output_filepath + XML_File_Name}')  # 成功ログ
    except Exception as e:
        Log.Log_Error(Log_file, f'XML ファイル生成エラー: {Output_filepath + XML_File_Name}, エラー: {e}')  # エラーログ

########## TestStep の生成 ##########
def generate_test_steps(data_dict):
    """
    TestStep の部分を XML 形式で生成する
    """
    test_steps = ""
    for i in range(1, 14):  # 13個の TestStep を生成
        test_steps += f'        <TestStep Name="Thickness{i}" Status="Done">\n'
        test_steps += f'            <Data DataType="Numeric" Name="X" Value="{data_dict[f"key_X{i}"]}"/>\n'
        test_steps += f'            <Data DataType="Numeric" Name="Y" Value="{data_dict[f"key_Y{i}"]}"/>\n'
        test_steps += f'            <Data DataType="Numeric" Name="Thickness" Value="{data_dict[f"key_Polish{i}"]}"/>\n'
        test_steps += '        </TestStep>\n'
    return test_steps  # 生成したテストステップの文字列を返す

########## main処理 ##########
def main(File_Path, Part_Number, Nine_Serial_Number):
    """
    メイン処理：データを取得し、XMLファイルを生成する
    """
    # データを取得
    data_dict = Open_Data_Sheet(File_Path, Part_Number, Nine_Serial_Number)  # Excelファイルからデータを取得

    if not data_dict:  # データが取得できなかった場合
        Log.Log_Error(Log_file, f"{Part_Number} : データ取得に失敗しました\n")
        return  # 処理をスキップ

    # 日付フォーマットを変換
    data_dict["key_Start_Date_Time"] = Convert_Date.Edit_Date(data_dict["key_Start_Date_Time"])  # 日付フォーマット変換

    # 空欄がないかチェック
    for val in data_dict.values():  # データ辞書内の全ての値を確認
        if val == "":
            Log.Log_Error(Log_file, f'{data_dict["key_Serial_Number"]} : 空欄エラー\n')  # 空欄エラーのログ記録
            return  # 処理をスキップ

    # データ型の確認
    result = Check.Data_Type(key_type, data_dict)  # データ型が正しいか確認
    if result == False:
        Log.Log_Error(Log_file, f'{data_dict["key_Serial_Number"]} : データ型エラー\n')  # データ型エラーのログ記録
        return  # データ型エラーの場合、処理をスキップ

    # XMLファイルの作成
    xml_file = f'Site={Site},ProductFamily={ProductFamily},Operation={Operation},Partnumber={data_dict["key_Part_Number"]},Serialnumber={data_dict["key_Serial_Number"]},Testdate={data_dict["key_Start_Date_Time"]}.xml'  # XML ファイル名を定義
    Output_XML(xml_file, data_dict)  # XML ファイルを作成

# プログラム終了ログ
Log.Log_Info(Log_file, 'プログラム終了')
