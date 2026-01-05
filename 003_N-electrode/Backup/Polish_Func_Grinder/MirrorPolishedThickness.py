import logging  # ロギングモジュールをインポート、ログ記録に使用
import openpyxl  # openpyxlをインポート、Excelファイルの操作に使用
import glob  # globモジュールをインポート、ファイルパターンマッチングに使用
import os  # osモジュールをインポート、OS関連の機能に使用
import sys  # sysモジュールをインポート、システム関連の機能に使用

from datetime import datetime, timedelta, date  # datetimeモジュールからdatetime, timedelta, dateクラスをインポート
from time import strftime  # timeモジュールからstrftimeをインポート、時間のフォーマットに使用

########## 自作関数の定義 ##########
sys.path.append('../MyModule')  # カスタムモジュールのパスをシステムパスに追加、カスタムモジュールのインポートを可能にする
import Log  # カスタムのLogモジュールをインポート、ログ記録に使用
import Convert_Date  # カスタムのConvert_Dateモジュールをインポート、日付フォーマット変換に使用
import Check  # カスタムのCheckモジュールをインポート、データ型チェックに使用

########## 全体パラメータ定義 ##########
Site = '350'  # サイト番号を定義
ProductFamily = 'SAG FAB'  # 製品ファミリーを定義
Operation = 'N-electrode_Polish_MirrorPolishedThickness'  # 操作名を定義
TestStation = 'N-electrode'  # テストステーションを定義

########## ログファイルの定義 ##########
Log_FolderName = str(date.today())  # 今日の日付をログフォルダ名として使用
Log_file = '../Log/' + Log_FolderName + '/003_N-electrode.log'  # ログファイルのパスを定義

########## シート名の定義 ##########
Data_Sheet_Name = '3ｲﾝﾁ用'  # 主要データのシート名を定義
XY_Sheet_Name = 'ウェハ座標'  # 座標データのシート名を定義

########## XML出力先ファイルパス ##########
Output_filepath = '//li.lumentuminc.net/data/SAG/TDS/Data/Files to Insert/XML/'  # XMLファイルの出力パスを定義
# Output_filepath = '../XML/003_N-electrode/'  # 別の可能なXML出力パス

########## 取得するデータの行番号および列番号を定義 ##########
Row_Serial_Number = 3  # シリアル番号が存在する行番号（0ベースなのでExcelの4行目）
Row_Start_Date_Time = 36  # 開始日時が存在する行番号（0ベースなのでExcelの37行目）
Row_Operator = 38  # オペレーターが存在する行番号（0ベースなのでExcelの39行目）
Row_Polish = 46  # ポリッシュ厚みデータが存在する行番号（0ベースなのでExcelの47行目）
col_x = 1  # X座標が存在する列番号（0ベースなのでExcelの2列目）
col_y = 2  # Y座標が存在する列番号（0ベースなのでExcelの3列目）

########## 取得した項目と型の対応表を定義 ##########
key_type = {
    "key_Start_Date_Time": str,  # 開始日時は文字列型
    "key_Part_Number": str,  # 部品番号は文字列型
    "key_Serial_Number": str,  # シリアル番号は文字列型
    "key_Operator": str,  # オペレーターは文字列型
    "key_Polish1": float,  # ポリッシュ厚み1は浮動小数点型
    "key_Polish2": float,  # ポリッシュ厚み2は浮動小数点型
    "key_Polish3": float,  # ポリッシュ厚み3は浮動小数点型
    "key_Polish4": float,  # ポリッシュ厚み4は浮動小数点型
    "key_Polish5": float,  # ポリッシュ厚み5は浮動小数点型
    "key_Polish6": float,  # ポリッシュ厚み6は浮動小数点型
    "key_Polish7": float,  # ポリッシュ厚み7は浮動小数点型
    "key_Polish8": float,  # ポリッシュ厚み8は浮動小数点型
    "key_Polish9": float,  # ポリッシュ厚み9は浮動小数点型
    "key_Polish10": float,  # ポリッシュ厚み10は浮動小数点型
    "key_Polish11": float,  # ポリッシュ厚み11は浮動小数点型
    "key_Polish12": float,  # ポリッシュ厚み12は浮動小数点型
    "key_Polish13": float,  # ポリッシュ厚み13は浮動小数点型
    "key_X1": float,  # X座標1は浮動小数点型
    "key_X2": float,  # X座標2は浮動小数点型
    "key_X3": float,  # X座標3は浮動小数点型
    "key_X4": float,  # X座標4は浮動小数点型
    "key_X5": float,  # X座標5は浮動小数点型
    "key_X6": float,  # X座標6は浮動小数点型
    "key_X7": float,  # X座標7は浮動小数点型
    "key_X8": float,  # X座標8は浮動小数点型
    "key_X9": float,  # X座標9は浮動小数点型
    "key_X10": float,  # X座標10は浮動小数点型
    "key_X11": float,  # X座標11は浮動小数点型
    "key_X12": float,  # X座標12は浮動小数点型
    "key_X13": float,  # X座標13は浮動小数点型
    "key_Y1": float,  # Y座標1は浮動小数点型
    "key_Y2": float,  # Y座標2は浮動小数点型
    "key_Y3": float,  # Y座標3は浮動小数点型
    "key_Y4": float,  # Y座標4は浮動小数点型
    "key_Y5": float,  # Y座標5は浮動小数点型
    "key_Y6": float,  # Y座標6は浮動小数点型
    "key_Y7": float,  # Y座標7は浮動小数点型
    "key_Y8": float,  # Y座標8は浮動小数点型
    "key_Y9": float,  # Y座標9は浮動小数点型
    "key_Y10": float,  # Y座標10は浮動小数点型
    "key_Y11": float,  # Y座標11は浮動小数点型
    "key_Y12": float,  # Y座標12は浮動小数点型
    "key_Y13": float  # Y座標13は浮動小数点型
}

########## データの格納 ##########
def Open_Data_Sheet(filepath, Part_Number, Nine_Serial_Number):
    # ----- ログ書込：データの取得 -----
    Log.Log_Info(Log_file, 'Data Acquisition')  # Log the start of data acquisition

    # ----- ファイル形式の確認、.xlsmおよび.xlsxのみ処理 -----
    if not (filepath.lower().endswith('.xlsm') or filepath.lower().endswith('.xlsx')):  # Check if the file is not .xlsm or .xlsx
        Log.Log_Info(Log_file, f"Skipping non .xlsx/.xlsm file: {filepath}")  # Log skipping the file
        return {}  # Return an empty dictionary to skip the file

    # openpyxlを使用して.xlsxおよび.xlsmファイルを処理
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)  # Load the Excel workbook, read-only
        sheet = wb[Data_Sheet_Name]  # Select the specified data sheet
    except Exception as e:
        Log.Log_Error(Log_file, f"{Data_Sheet_Name} : Unsupported file format, skipping ({filepath})\n")  # Log error if unable to load sheet
        return {}  # Return an empty dictionary to skip the file

    # シリアル番号、開始日時、オペレーターを取得
    try:
        Start_Date_Time = sheet.cell(row=Row_Start_Date_Time + 1, column=3).value  # Get start date time from row=37, column=3 (C37)
        Serial_Number = sheet.cell(row=Row_Serial_Number + 1, column=3).value  # Get serial number from row=4, column=3 (C4)
        Operator = sheet.cell(row=Row_Operator + 1, column=3).value  # Get operator from row=39, column=3 (C39)
    except Exception as e:
        Log.Log_Error(Log_file, f"Error reading main sheet cells in {filepath}: {e}\n")  # Log error if unable to read main sheet cells
        return {}  # Return an empty dictionary to skip the file

    # ----- ガラス盤研磨 最終ウェハ厚の格納 -----
    Polish = []
    try:
        for i in range(13):  # Get 13 polish thickness data points
            Polish.append(sheet.cell(row=Row_Polish + 1, column=4 + i).value)  # Get data from row=47, columns=4 to 16 (D47 to P47)
    except Exception as e:
        Log.Log_Error(Log_file, f"Error reading Polish data in {filepath}: {e}\n")  # Log error if unable to read polish data
        return {}  # Return an empty dictionary to skip the file

    # ----- X/Y座標の格納 -----
    try:
        sheet_xy = wb[XY_Sheet_Name]  # Select the coordinate sheet
        x, y = [], []
        for i in range(1, 14):  # Get 13 X/Y coordinates
            x.append(sheet_xy.cell(row=i, column=col_x + 1).value)  # Get X coordinate from row=i, column=2 (B1 to B13)
            y.append(sheet_xy.cell(row=i, column=col_y + 1).value)  # Get Y coordinate from row=i, column=3 (C1 to C13)
    except Exception as e:
        Log.Log_Error(Log_file, f"Error reading XY data in {filepath}: {e}\n")  # Log error if unable to read XY data
        return {}  # Return an empty dictionary to skip the file

    # ----- 辞書型に格納 ------
    data_dict = {
        "key_Start_Date_Time": Start_Date_Time,  # 開始日時
        "key_Part_Number": Part_Number,  # 部品番号（関数引数から取得）
        "key_Serial_Number": Serial_Number,  # シリアル番号
        "key_LotNumber_9": Nine_Serial_Number,  # LotNumber_9（関数引数から取得）
        "key_Operator": Operator,  # オペレーター
        "key_Polish1": Polish[0],  # ポリッシュ厚み1
        "key_Polish2": Polish[1],  # ポリッシュ厚み2
        "key_Polish3": Polish[2],  # ポリッシュ厚み3
        "key_Polish4": Polish[3],  # ポリッシュ厚み4
        "key_Polish5": Polish[4],  # ポリッシュ厚み5
        "key_Polish6": Polish[5],  # ポリッシュ厚み6
        "key_Polish7": Polish[6],  # ポリッシュ厚み7
        "key_Polish8": Polish[7],  # ポリッシュ厚み8
        "key_Polish9": Polish[8],  # ポリッシュ厚み9
        "key_Polish10": Polish[9],  # ポリッシュ厚み10
        "key_Polish11": Polish[10],  # ポリッシュ厚み11
        "key_Polish12": Polish[11],  # ポリッシュ厚み12
        "key_Polish13": Polish[12],  # ポリッシュ厚み13
        "key_X1": x[0],  # X座標1
        "key_X2": x[1],  # X座標2
        "key_X3": x[2],  # X座標3
        "key_X4": x[3],  # X座標4
        "key_X5": x[4],  # X座標5
        "key_X6": x[5],  # X座標6
        "key_X7": x[6],  # X座標7
        "key_X8": x[7],  # X座標8
        "key_X9": x[8],  # X座標9
        "key_X10": x[9],  # X座標10
        "key_X11": x[10],  # X座標11
        "key_X12": x[11],  # X座標12
        "key_X13": x[12],  # X座標13
        "key_Y1": y[0],  # Y座標1
        "key_Y2": y[1],  # Y座標2
        "key_Y3": y[2],  # Y座標3
        "key_Y4": y[3],  # Y座標4
        "key_Y5": y[4],  # Y座標5
        "key_Y6": y[5],  # Y座標6
        "key_Y7": y[6],  # Y座標7
        "key_Y8": y[7],  # Y座標8
        "key_Y9": y[8],  # Y座標9
        "key_Y10": y[9],  # Y座標10
        "key_Y11": y[10],  # Y座標11
        "key_Y12": y[11],  # Y座標12
        "key_Y13": y[12],  # Y座標13
    }

    # ----- オペレーターが空欄であれば'-'を入れる ------
    if data_dict["key_Operator"] == "":
        data_dict["key_Operator"] = '-'  # オペレーターが空の場合、'-'を設定

    return data_dict  # データ辞書を返す

########## XML変換 ##########
def Output_XML(XML_File_Name, data_dict):
    # ----- ログ書込:XML変換 -----
    Log.Log_Info(Log_file, 'Excel File To XML File Conversion')  # Log the start of XML conversion

    try:
        with open(Output_filepath + XML_File_Name, 'w', encoding="utf-8") as f:  # Open the XML file for writing with UTF-8 encoding
            f.write('<?xml version="1.0" encoding="utf-8"?>\n' +
                    '<Results xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema">\n' +
                    f'       <Result startDateTime="{data_dict["key_Start_Date_Time"].replace(".", ":")}" Result="Done">\n' +
                    f'               <Header SerialNumber="{data_dict["key_Serial_Number"]}" PartNumber="{data_dict["key_Part_Number"]}" Operation="{Operation}" TestStation="{TestStation}" Operator="{data_dict["key_Operator"]}" StartTime="{data_dict["key_Start_Date_Time"].replace(".", ":")}" Site="{Site}" LotNumber="{data_dict["key_Serial_Number"]}"/>\n\n' +
                    generate_test_steps(data_dict) +  # Generate TestStep1 to TestStep13
                    f'               <TestStep Name="SORTED_DATA" startDateTime="{data_dict["key_Start_Date_Time"].replace(".", ":")}" Status="Passed">\n' +
                    f'                   <Data DataType="String" Name="LotNumber_5" Value="{data_dict["key_Serial_Number"]}" CompOperation="LOG"/>\n' +
                    f'                   <Data DataType="String" Name="LotNumber_9" Value="{data_dict["key_LotNumber_9"]}" CompOperation="LOG"/>\n' +
                    f'               </TestStep>\n\n' +
                    f'               <TestEquipment>\n' +
                    f'                   <Item DeviceName="Stepmeter" DeviceSerialNumber="1"/>\n' +
                    f'               </TestEquipment>\n\n' +
                    f'               <ErrorData/>\n' +
                    f'               <FailureData/>\n' +
                    f'               <Configuration/>\n' +
                    f'       </Result>\n' +
                    f'</Results>'
                    )
        Log.Log_Info(Log_file, f'Successfully generated XML file: {Output_filepath + XML_File_Name}')  # Log successful XML generation
    except Exception as e:
        Log.Log_Error(Log_file, f'Failed to generate XML file: {Output_filepath + XML_File_Name}, Error: {e}')  # Log error if XML generation fails

def generate_test_steps(data_dict):
    """TestStep部分のXML内容を生成"""
    test_steps = ""
    for i in range(1, 14):  # 13個のTestStepを生成
        test_steps += f'               <TestStep Name="Thickness{i}" startDateTime="{data_dict["key_Start_Date_Time"].replace(".", ":")}" Status="Done">\n' + \
                      f'                   <Data DataType="Numeric" Name="X" Units="um" Value="{data_dict[f"key_X{i}"]}"/>\n' + \
                      f'                   <Data DataType="Numeric" Name="Y" Units="um" Value="{data_dict[f"key_Y{i}"]}"/>\n' + \
                      f'                   <Data DataType="Numeric" Name="Thickness" Units="um" Value="{data_dict[f"key_Polish{i}"]}"/>\n' + \
                      f'               </TestStep>\n'
    return test_steps  # 生成したTestStepを返す

########## main処理 ##########
def main(File_Path, Part_Number, Nine_Serial_Number):
    # ----- ログ書込：オペレーション名 -----
    Log.Log_Info(Log_file, Operation)  # Log the operation name

    ########## データ取得 ##########
    data_dict = Open_Data_Sheet(File_Path, Part_Number, Nine_Serial_Number)  # データを取得

    if not data_dict:  # データ辞書が空の場合、スキップ
        Log.Log_Error(Log_file, f"{Part_Number} : Data Acquisition Failed\n")  # Log data acquisition failure
        return  # スクリプトを終了

    ########## 日付フォーマット変換 ##########
    data_dict["key_Start_Date_Time"] = Convert_Date.Edit_Date(data_dict["key_Start_Date_Time"])  # 日付フォーマットを変換

    ########## 空欄チェック ##########
    for val in data_dict.values():  # データ辞書内の全ての値をチェック
        if val == "":
            Log.Log_Error(Log_file, f'{data_dict["key_Serial_Number"]} : Blank Error\n')  # 空欄エラーをログに記録
            return  # スクリプトを終了

    ########## データ型の確認 ##########
    result = Check.Data_Type(key_type, data_dict)  # データ型を確認
    if result == False:
        Log.Log_Error(Log_file, f'{data_dict["key_Serial_Number"]} : Data Error\n')  # データ型エラーをログに記録
        return  # スクリプトを終了

    ########## XMLファイルの作成 ##########
    xml_file = f'Site={Site},ProductFamily={ProductFamily},Operation={Operation},Partnumber={data_dict["key_Part_Number"]},Serialnumber={data_dict["key_Serial_Number"]},Testdate={data_dict["key_Start_Date_Time"]}.xml'  # XMLファイル名を定義

    Output_XML(xml_file, data_dict)  # XMLファイルを生成
