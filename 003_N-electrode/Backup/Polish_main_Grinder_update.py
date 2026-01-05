import logging
import shutil
import glob
import xlrd
import openpyxl  # 用來處理 xlsx 和 xlsm 檔案
import os
import sys
from datetime import datetime, date

########## 自作関数の定義 ##########
sys.path.append('../MyModule')
import SQL
import Log
import Convert_Date

########## サブプログラムの定義 ##########
from Polish_Func_Grinder import InitialWaferThickness
from Polish_Func_Grinder import WaxThickness
from Polish_Func_Grinder import RoughPolishedThickness
from Polish_Func_Grinder import MirrorPolishedThickness
from Polish_Func_Grinder import EtchedThickness

########## ログの設定 ##########
Log_FolderName = str(date.today())

if not os.path.exists("../Log/" + Log_FolderName):
    os.makedirs("../Log/" + Log_FolderName)

Log_file = '../Log/' + Log_FolderName + '/003_N-electrode.log'
Log.Log_Info(Log_file, 'Program Start')

########## 処理シートの定義 ##########
Data_sheet_name = '3ｲﾝﾁ用'
XY_sheet_name = 'ウェハ座標'

########## エラーの時に取得される番号を定義 ##########
Error_Number = {'empty:\'\'', 'error:7', 'error:15', 'error:23', 'error:29', 'error:36', 'error:42', 'xldate:0.0'}

########## 空欄チェック ##########
def Get_Cells_Info(filepath):

    Log.Log_Info(Log_file, "Blank Check")
    is_cells_empty = False

    # 檔案格式處理
    if filepath.lower().endswith('.xls'):
        # 使用 xlrd 打開 .xls 檔案
        try:
            wb = xlrd.open_workbook(filepath, on_demand=True)
            sheet = wb.sheet_by_name(Data_sheet_name)
        except Exception as e:
            Log.Log_Error(Log_file, f"文件讀取失敗 {filepath}: {e}")
            is_cells_empty = True
            return is_cells_empty
    elif filepath.lower().endswith('.xlsx') or filepath.lower().endswith('.xlsm'):
        # 使用 openpyxl 打開 .xlsx 和 .xlsm 檔案
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheet = wb[Data_sheet_name]
        except Exception as e:
            Log.Log_Error(Log_file, f"文件讀取失敗 {filepath}: {e}")
            is_cells_empty = True
            return is_cells_empty
    else:
        # 非支援的檔案格式
        Log.Log_Info(Log_file, f"跳過不支援的檔案格式: {filepath}")
        return True

    # 檢查特定欄位是否為空值或錯誤代碼
    if str(sheet.cell(3, 2).value) in Error_Number or \
       str(sheet.cell(4, 2).value) in Error_Number or \
       str(sheet.cell(23, 2).value) in Error_Number or \
       str(sheet.cell(24, 2).value) in Error_Number or \
       str(sheet.cell(35, 2).value) in Error_Number or \
       str(sheet.cell(36, 2).value) in Error_Number or \
       str(sheet.cell(54, 2).value) in Error_Number or \
       str(sheet.cell(55, 2).value) in Error_Number:
        is_cells_empty = True

    return is_cells_empty


########## 本処理 ##########
if __name__ == '__main__':

    Log.Log_Info(Log_file, 'Main Start')

    today = str(date.today()).split('-')

    ########### ディレクトリ検索 ##########
    Log.Log_Info(Log_file, 'Directory Search')

    FilePath = 'Z:/研磨/2024年/'
    FolderName = '*年*月'
    Folder_List = []
    for folder in glob.glob(FilePath + FolderName):
        Folder_List.append(folder)

    if len(Folder_List) == 0:
        Log.Log_Error(Log_file, 'Folder Error')
        sys.exit()

    Folder_List.sort()
    directory = Folder_List[0] + '/'

    ########### 処理済みフォルダがなければ作成する ##########
    if not os.path.exists(directory + '/処理済みフォルダ'):
        os.mkdir(directory + '/処理済みフォルダ')

    ########### 対象ファイルの処理 ##########
    for File_Path in glob.glob(directory + '*.xls*'):

        Log.Log_Info(Log_file, File_Path)

        # 檢查檔案並跳過有問題的檔案
        if Get_Cells_Info(File_Path):
            Log.Log_Error(Log_file, "Blank Error\n")
            continue

        # 根據檔案格式獲取數據
        if File_Path.lower().endswith('.xls'):
            wb = xlrd.open_workbook(File_Path, on_demand=True)
            sheet = wb.sheet_by_name(Data_sheet_name)
            Serial_Number = str(sheet.cell(3, 2).value)
            Date_Array = [sheet.cell(5, 3).value, sheet.cell(24, 3).value, sheet.cell(36, 3).value, sheet.cell(55, 3).value]
            wb.release_resources()
            
        elif File_Path.lower().endswith('.xlsx') or File_Path.lower().endswith('.xlsm'):
            wb = openpyxl.load_workbook(File_Path, data_only=True)
            sheet = wb[Data_sheet_name]
            Serial_Number = str(sheet.cell(4, 3).value)
            Date_Array = [sheet.cell(5, 3).value, sheet.cell(25, 3).value, sheet.cell(37, 3).value, sheet.cell(56, 3).value]

        ########## Primeに接続し、品名を取得 ##########
        conn, cursor = SQL.connSQL()

        if conn is None:
            Log.Log_Error(Log_file, Serial_Number + ' : ' + 'Connection with Prime Failed\n')
            sys.exit()

        Part_Number, Nine_Serial_Number = SQL.selectSQL(cursor, Serial_Number)
        SQL.disconnSQL(conn, cursor)

        if Part_Number is None:
            Log.Log_Error(Log_file, Serial_Number + ' : ' + "Part Number Error\n")
            continue

        if Part_Number == 'LDアレイ_':
            continue

        ########## 日付フォーマットの変換 ##########
        if Convert_Date.Edit_Date(Date_Array[0]) == "" or Convert_Date.Edit_Date(Date_Array[1]) == "" or Convert_Date.Edit_Date(Date_Array[2]) == "" or Convert_Date.Edit_Date(Date_Array[3]) == "":
            Log.Log_Error(Log_file, Serial_Number + ' : ' + "Date Error\n")
            continue

        ########## 各Operationに渡しXML変換を行う ##########
        Log.Log_Info(Log_file, 'Move To Each Function')
        InitialWaferThickness.main(File_Path, Part_Number, Nine_Serial_Number)
        WaxThickness.main(File_Path, Part_Number, Nine_Serial_Number)
        RoughPolishedThickness.main(File_Path, Part_Number, Nine_Serial_Number)
        MirrorPolishedThickness.main(File_Path, Part_Number, Nine_Serial_Number)
        EtchedThickness.main(File_Path, Part_Number, Nine_Serial_Number)

        ########### 処理済みフォルダに移動 ##########
        Log.Log_Info(Log_file, 'Move File')
        try:
            shutil.move(File_Path, directory + '/処理済みフォルダ/')
        except PermissionError:
            Log.Log_Error(Log_file, "Permission Error")

########## Main処理の終了 ##########
Log.Log_Info(Log_file, 'Program End')
