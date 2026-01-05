import configparser
import csv
import os
import re
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Optional, List

from openpyxl import load_workbook


class FieldDef:
    def __init__(self, key: str, raw_def: str):
        """
        raw_def 例如: "-1:str_Q7" / "-1:float_M12" / "-1:str" / "-1:float_COL26"
        """
        self.key = key
        parts = raw_def.split(":")
        if len(parts) != 2:
            raise ValueError(f"Field definition format error: {key}={raw_def}")

        # index_str = parts[0]  # 目前沒用到，但先保留結構
        type_cell = parts[1]

        # 處理第三欄：datatype_cell / datatype_COLxx / datatype
        m = re.match(r"^(str|float|int)(?:_(.+))?$", type_cell)
        if not m:
            raise ValueError(f"Unknown type format: {key}={raw_def}")

        self.dtype = m.group(1)          # str / float / int
        self.ref: Optional[str] = m.group(2)  # 例如 Q7 / M12 / COL26 / None

        # 判斷是固定 Cell / Column / 非 Excel 欄位
        self.is_cell = False
        self.is_col = False
        self.col_index: Optional[int] = None

        if self.ref is not None:
            if self.ref.startswith("COL"):
                # 例如 COL26
                self.is_col = True
                try:
                    self.col_index = int(self.ref[3:])
                except ValueError:
                    raise ValueError(f"Bad COL index in {key}: {self.ref}")
            else:
                # 一般 Excel A1 style cell
                self.is_cell = True

    def __repr__(self):
        return f"FieldDef(key={self.key}, dtype={self.dtype}, ref={self.ref}, is_cell={self.is_cell}, is_col={self.is_col})"


def parse_ini(ini_path: str):
    cfg = configparser.ConfigParser()
    cfg.optionxform = str  # 保留大小寫
    cfg.read(ini_path, encoding="utf-8")

    basic = cfg["Basic_info"]
    paths = cfg["Paths"]
    excel = cfg["Excel"]

    # DataFields
    fields_raw = cfg["DataFields"]["fields"].strip().splitlines()
    field_defs: List[FieldDef] = []
    for line in fields_raw:
        line = line.strip()
        if not line or line.startswith(";") or line.startswith("#"):
            continue
        # 例如: key_start_date_time:-1:str_Q7
        if ":" not in line:
            continue
        key, rest = line.split(":", 1)
        key = key.strip()
        rest = rest.strip()
        field_defs.append(FieldDef(key, rest))

    return basic, paths, excel, field_defs


def ensure_dir(path: str):
    p = Path(path)
    p.mkdir(parents=True, exist_ok=True)


def convert_value(raw, dtype: str):
    if raw is None:
        return ""
    if dtype == "str":
        return str(raw)
    if dtype == "float":
        try:
            return float(raw)
        except (ValueError, TypeError):
            return ""
    if dtype == "int":
        try:
            return int(raw)
        except (ValueError, TypeError):
            return ""
    # fallback
    return raw


def guess_row_number(ws):
    """
    對於使用 COLxx 的欄位，需要一個 row_number。
    這裡先給一個暫時的預設：
    - 取有 XRD_Thickness 的那一行 row 當 row_number
    - 若沒有，預設 1 (你可以依 F7/F10 原始邏輯改寫)

    建議你照實際 Format1 程式邏輯，把這個函式改成：
    - 搜尋某個 anchor cell 的 row
    - 或根據特定 header 關鍵字決定 row_number
    """
    # 嘗試找 M46 (常見 XRD 厚度位置)
    try:
        cell = ws["M46"]
        if cell.value not in (None, ""):
            return cell.row
    except Exception:
        pass

    # fallback: 第一列
    return 1


def get_field_value_from_sheet(ws, field: FieldDef, row_number: Optional[int]) -> Any:
    if field.is_cell:
        # 固定 A1 style cell
        try:
            return ws[field.ref].value
        except Exception:
            return None

    if field.is_col:
        # 以 row_number + column index 方式讀
        if row_number is None or field.col_index is None:
            return None
        try:
            return ws.cell(row=row_number, column=field.col_index).value
        except Exception:
            return None

    # 非 Excel 欄位，留給後處理
    return None


def post_process_record(record: Dict[str, Any]) -> Dict[str, Any]:
    """
    給你後續「補資料」用的 Hook。
    例如：
        - 依 key_serial_number 去 DB 查 key_part_number / key_LotNumber_9
        - 用 key_start_date_time 做時間排序填 key_STARTTIME_SORTED
        - 依 batch 判斷 SORTNUMBER

    目前先維持原樣，你可以依實際環境改寫。
    """
    # TODO: 在這裡補 DB or 計算邏輯
    return record


def build_record_from_excel(wb_path: str, excel_cfg, field_defs: List[FieldDef]) -> Dict[str, Any]:
    wb = load_workbook(wb_path, data_only=True)
    sheet_name = excel_cfg.get("sheet_name", "ﾃﾞｰﾀ")
    ws = wb[sheet_name]

    # row_number 給所有 COLxx 欄位共用
    row_number = guess_row_number(ws)

    record: Dict[str, Any] = {}
    for f in field_defs:
        raw = get_field_value_from_sheet(ws, f, row_number)
        value = convert_value(raw, f.dtype)
        record[f.key] = value

    # 讓你在這裡補 DB 或計算欄位
    record = post_process_record(record)
    return record


def find_input_files(patterns: str, base_dir: str):
    """
    patterns: F2プログラムシート*.xlsx 這種字串（可以用 ; 分隔多個）
    """
    base = Path(base_dir)
    if not base.exists():
        return []

    pats = [p.strip() for p in patterns.split(";") if p.strip()]
    files = []
    for pat in pats:
        files.extend(base.glob(pat))
    return sorted(set(files))


def write_csv(records: List[Dict[str, Any]], field_defs: List[FieldDef], csv_path: str):
    ensure_dir(os.path.dirname(csv_path))
    fieldnames = [f.key for f in field_defs]

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for rec in records:
            writer.writerow(rec)


def write_pointer_xml(basic_cfg, csv_filename: str, xml_path: str):
    """
    依你「新版 XML 只是標註 CSV 的位置」的需求，
    這裡做一個非常簡單的 pointer XML：
    <DataFile>
      <Site>350</Site>
      <Operation>LD-EML_F2_Format1</Operation>
      <Tool>MOCVD-F2</Tool>
      <CSV>filename.csv</CSV>
      <Generated>2025-12-01T12:34:56</Generated>
    </DataFile>
    """
    ensure_dir(os.path.dirname(xml_path))

    site = basic_cfg.get("Site", "")
    op = basic_cfg.get("Operation", "")
    tool = basic_cfg.get("Tool_Name", "")
    ts = datetime.now().isoformat(timespec="seconds")

    xml_text = f"""<?xml version="1.0" encoding="UTF-8"?>
<DataFile>
  <Site>{site}</Site>
  <Operation>{op}</Operation>
  <Tool>{tool}</Tool>
  <CSV>{csv_filename}</CSV>
  <Generated>{ts}</Generated>
</DataFile>
"""
    with open(xml_path, "w", encoding="utf-8") as f:
        f.write(xml_text)


def main(ini_path: str):
    basic, paths, excel_cfg, field_defs = parse_ini(ini_path)

    input_base = paths.get("input_paths", "").strip()
    output_base = paths.get("output_path", "").strip()
    csv_base = paths.get("CSV_path", "").strip()
    patterns = basic.get("file_name_patterns", "*.*")

    if not input_base:
        raise RuntimeError("input_paths is empty in INI")

    input_files = find_input_files(patterns, input_base)
    print(f"Found {len(input_files)} files in {input_base} for patterns {patterns}")

    all_records: List[Dict[str, Any]] = []
    csv_filenames: List[str] = []

    for fpath in input_files:
        print(f"Processing: {fpath.name}")
        rec = build_record_from_excel(str(fpath), excel_cfg, field_defs)
        all_records.append(rec)

        # 每個 Excel 各出一個 CSV & XML
        csv_name = fpath.stem + ".csv"
        xml_name = fpath.stem + ".xml"

        csv_full = str(Path(csv_base) / csv_name)
        xml_full = str(Path(output_base) / xml_name)

        write_csv([rec], field_defs, csv_full)
        write_pointer_xml(basic, csv_name, xml_full)
        csv_filenames.append(csv_name)

    print("Done.")
    print("Generated CSV files:")
    for name in csv_filenames:
        print("  ", name)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="LD-EML 共用 INI 驅動資料抽取程式")
    parser.add_argument("ini", help="ini 檔案路徑，例如 F1.ini / F2.ini / F6.ini ...")
    args = parser.parse_args()

    main(args.ini)
