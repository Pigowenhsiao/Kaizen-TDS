from configparser import ConfigParser
def _read_and_parse_ini_config(config_file_path: str) -> ConfigParser:
    """
    INI設定ファイルを読み込んで解析する。
    (Read and parse INI config file)
    """
    config = ConfigParser()
    config.read(config_file_path, encoding='utf-8')
    return config
def setup_logging(log_dir, operation_name):
    """
    ログファイル設定を作成する（未使用）。
    (Setup log file configuration, not used)
    """
    # ログは使用しないのでスキップ (Skip, not used)
    return None
import os
import sys
import glob
import logging
import traceback
from pathlib import Path
from datetime import datetime, date, timedelta
from configparser import ConfigParser
import pandas as pd
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
from xml.dom import minidom

sys.path.append('../../MyModule')
import SQL
import Convert_Date
import Row_Number_Func


# -----------------------------
# Ini 設定物件
# -----------------------------
class IniSettings:
    """
    INI設定を保存する（Formation1共通版）。
    (Store INI settings for Formation1 universal version)
    """

    def __init__(self):
        # 基本情報 / Basic info
        self.site = ""
        self.product_family = ""
        self.operation = ""
        self.test_station = ""
        self.retention_date = 30
        self.file_name_patterns = []
        self.tool_name = ""

        # パス情報 / Path info
        self.input_paths = []
        self.output_path = ""
        self.csv_path = ""
        self.intermediate_data_path = ""
        self.log_path = ""
        self.running_rec = ""
        self.backup_running_rec_path = ""

        # Excel情報 / Excel info
        self.data_columns = ""   # Formation1では必須ではないが保持 (Not always needed, but kept)
        self.skip_rows = 0       # Formation1ではskip_rowsは使わない (Not used, but kept)

        # データフィールド / DataFields
        # key -> {'dtype': str/float/int, 'mode': 'cell'/'col'/'virtual', 'ref': 'Q7'/26/None}
        self.field_map = {}
def _parse_fields_map_from_lines_format1(fields_lines):
    """
    Formation1用 [DataFields].fields を解析する。
    (Parse [DataFields].fields for Formation1)

        例/Example:
        key_start_date_time:-1:str_Q7
        key_TargetWavelength_TargetWavelength:-1:float_M12
        key_Wavelength_Wavelength_LD:-1:float_COL26
        key_serial_number:-1:str

    戻り値/Return:
        { key: {'dtype': 'str'/'float'/'int',
                'mode': 'cell'/'col'/'virtual',
                'ref': 'Q7'/26/None} }
    """
    import re

    fields = {}
    for line in fields_lines:
        raw = line.strip()
        if not raw or raw.startswith(';') or raw.startswith('#'):
            continue
        # 只處理 key:idx:type_cell 三段式
        if ':' not in raw:
            continue
        try:
            key, idx_str, type_cell = map(str.strip, raw.split(':', 2))
        except ValueError:
            continue

        # 解析第三欄 type_cell: e.g. str_Q7 / float_M12 / float_COL26 / str
        m = re.match(r'^(str|float|int)(?:_(.+))?$', type_cell)
        if not m:
            # 格式不符合，直接略過/或記 log
            continue

        dtype = m.group(1)
        ref = m.group(2)  # 可能是 None / "Q7" / "M12" / "COL26"
        mode = 'virtual'
        col_index = None

        if ref is not None:
            if ref.startswith("COL"):
                # 例如 COL26 → 以 row_number, column 26 讀取
                try:
                    col_index = int(ref[3:])
                except ValueError:
                    col_index = None
                mode = 'col'
                ref_val = col_index
            else:
                # 一般 A1 樣式
                mode = 'cell'
                ref_val = ref
        else:
            ref_val = None

        fields[key] = {
            'dtype': dtype,
            'mode': mode,   # 'cell' / 'col' / 'virtual'
            'ref': ref_val  # 'Q7' / 26 / None
        }
    return fields


def _extract_settings_from_config(config: ConfigParser) -> IniSettings:
    """
    INIオブジェクトから全ての設定を抽出する。
    (Extract all settings from INI object)
    """
    s = IniSettings()

    # 基本情報 / Basic_info
    s.site = config.get('Basic_info', 'Site')
    s.product_family = config.get('Basic_info', 'ProductFamily')
    s.operation = config.get('Basic_info', 'Operation')
    s.test_station = config.get('Basic_info', 'TestStation')
    s.retention_date = config.getint('Basic_info', 'retention_date', fallback=30)
    patterns_str = config.get('Basic_info', 'file_name_patterns')
    s.file_name_patterns = [x.strip() for x in patterns_str.split(',') if x.strip()]
    s.tool_name = config.get('Basic_info', 'Tool_Name', fallback="")

    # パス情報 / Paths
    s.input_paths = [x.strip() for x in config.get('Paths', 'input_paths').split(',') if x.strip()]
    s.output_path = config.get('Paths', 'output_path', fallback="")
    s.csv_path = config.get('Paths', 'CSV_path', fallback="")
    s.intermediate_data_path = config.get('Paths', 'intermediate_data_path', fallback="")
    s.log_path = config.get('Paths', 'log_path', fallback="../Log/")
    s.running_rec = config.get('Paths', 'running_rec', fallback="")
    s.backup_running_rec_path = config.get('Paths', 'backup_running_rec_path', fallback="")

    # Excel情報 / Excel
    # **sheet_nameは使用しない**、data_columnsやmain_skip_rowsのみ保持 (sheet_name not used)
    s.data_columns = config.get('Excel', 'data_columns', fallback="")
    s.skip_rows = config.getint('Excel', 'main_skip_rows', fallback=0)

    # データフィールド / DataFields
    fields_lines = config.get('DataFields', 'fields').splitlines()
    s.field_map = _parse_fields_map_from_lines_format1(fields_lines)

    return s


# -----------------------------
# Excel読込・シート判定・フィールド抽出
# (Excel reading / sheet validation / field extraction)
# -----------------------------
def is_format1_sheet(ws, settings) -> bool:
    """
    このシートが有効なFormation1シートか判定します。
    (Check if this sheet is a valid Formation1 sheet)
    新ルール：INI設定に従い必須セルを動的に判定 (Dynamically check required cells by INI)
    """

    if settings is None:
        print("[is_format1_sheet] settingsがありません。シート合格判定不可 / Missing settings, cannot judge sheet.")
        return False

    # 必須チェックセルを動的取得 (Dynamically get required cells)
    must_keys = [k for k in settings.field_map.keys() if k in ('key_start_date_time', 'key_serial_number')]
    must_cells = [settings.field_map[k]['ref'] for k in must_keys if settings.field_map[k]['mode'] == 'cell' and settings.field_map[k]['ref']]

    all_checks = []
    passed = True
    for c in must_cells:
        try:
            v = ws[c].value
            check = {
                'cell': c,
                'value': v,
                'rule': '空であってはならない / Not empty',
                'result': '合格 / Pass' if v not in (None, "") else '不合格 / Fail'
            }
            if v is None or v == "":
                passed = False
        except Exception as e:
            check = {
                'cell': c,
                'value': None,
                'rule': '空であってはならない / Not empty',
                'result': f'不合格 / Fail (エラー: {e})'
            }
            passed = False
        all_checks.append(check)
    # チェック詳細を出力 (Print check details)
    print("    [Sheet Format1 チェック明細 / Check details]")
    for check in all_checks:
        print(f"      セル/cell: {check['cell']}, 値/value: {check['value']}, 規範/rule: {check['rule']}, 判定/result: {check['result']}")
    return passed


def guess_row_number(ws, settings: IniSettings) -> int:
    try:
        cell = ws["M46"]
        if cell.value not in (None, ""):
            return cell.row
    except Exception:
        pass
    return 1


def _get_cell(ws, ref: str):
    try:
        return ws[ref].value
    except Exception:
        return None


def _get_col(ws, row: int, col_index: int):
    try:
        return ws.cell(row=row, column=col_index).value
    except Exception:
        return None


def _convert_value(raw, dtype: str):
    if raw is None:
        return None
    if dtype == "str":
        return str(raw)
    if dtype == "float":
        try:
            return float(raw)
        except (ValueError, TypeError):
            return None
    if dtype == "int":
        try:
            return int(raw)
        except (ValueError, TypeError):
            return None
    return raw


def build_record_from_ws(ws, settings: IniSettings, log_file: str) -> dict:
    """
    INIのfield_map（cell/COL/virtual）に基づき、1シート分のレコードを作成。
    (Build a record for one sheet based on INI field_map)
    """
    row_number = guess_row_number(ws, settings)
    record = {}

    print("[データ抽出 debug / Data extraction debug]")
    for key, meta in settings.field_map.items():
        dtype = meta['dtype']
        mode = meta['mode']
        ref = meta['ref']

        if mode == 'cell':
            raw = _get_cell(ws, ref)
            source = f"cell {ref}"
        elif mode == 'col':
            if ref is None:
                raw = None
                source = f"col None"
            else:
                raw = _get_col(ws, row_number, int(ref))
                source = f"row {row_number}, col {ref}"
        else:
            # virtual項目はExcelから読まず、後処理で補完 (Do not read virtual fields from Excel, fill later)
            raw = None
            source = "virtual"

        value = _convert_value(raw, dtype)
        record[key] = value
        #print(f"  key: {key}, source: {source}, raw: {raw}, converted: {value}")

    return record


# -----------------------------
# DB・日付・ソート等の後処理
# (DB / date / sorting post-processing)
# -----------------------------
def enrich_record_with_db_and_time(record: dict, settings: IniSettings, log_file: str) -> dict | None:
    """
    CVD_Grating_Common.pyの方針に従う：
        - 日付変換 (Convert date)
        - retention_dateチェック (Check retention_date)
        - serialでDB検索しPartNumber/LotNumber_9補完 (DB lookup by serial)
        - STARTTIME_SORTED/SORTNUMBER計算 (Calculate sort fields)
    戻り値/Return:
        - 不正なデータはNone (Return None if invalid)
        - それ以外は更新済みrecord (Else, return updated record)
    """

    # 1) 日付解析 / Parse date
    raw_date = record.get('key_start_date_time') or record.get('key_Start_Date_Time')
    dt = None
    if raw_date:
        try:
            dt = pd.to_datetime(raw_date)
            print(f"[enrich debug] raw_date: {raw_date}, 解析結果: {dt}")
        except Exception as e:
            print(f"[enrich debug] 日期解析失敗: {raw_date}, 錯誤: {e}")
            dt = None

    if dt is None:
        print("  [enrich] start_date_time invalid, drop this sheet.")
        return None

    # retention_dateチェック / Check retention_date
    today = datetime.now()
    if dt < today - timedelta(days=settings.retention_date):
        print(f"  [enrich] date {dt} older than retention {settings.retention_date} days, drop.")
        return None

    # 2) SerialでDB検索しPart/Lot取得 / DB lookup for Part/Lot
    serial = record.get('key_serial_number') or record.get('key_Serial_Number')
    if not serial:
        print("  [enrich] serial_number empty, drop this sheet.")
        return None

    conn, cursor = None, None
    try:
        conn, cursor = SQL.connSQL()
        if conn is None:
            print("  [enrich] DB connection failed.")
            return None

        part, lot9 = SQL.selectSQL(cursor, str(serial))
        record['key_part_number'] = part
        record['key_LotNumber_9'] = lot9

        if not part or part == 'LDアレイ_':
            print(f"  [enrich] invalid PartNumber ({part}), drop this sheet.")
            return None

    except Exception as e:
        print(f"  [enrich] DB error: {e}")
        return None
    finally:
        if conn:
            SQL.disconnSQL(conn, cursor)

    # 3) STARTTIME_SORTED/SORTNUMBER計算 / Calculate sort fields
    base_date = datetime(1899, 12, 30)
    date_excel_number = (dt - base_date).days
    # Formation1 每個 sheet 視為一筆，先以 rowIndex=1 概化
    excel_row = 1
    starttime_sorted = date_excel_number + excel_row / 10**6

    record['key_STARTTIME_SORTED'] = float(starttime_sorted)
    record['key_SORTNUMBER'] = float(excel_row)

    # 4) Operation/TestStation/Site等の共通フィールド補完 / Fill common fields
    record.setdefault('key_Operation', settings.operation)
    record.setdefault('key_TestStation', settings.test_station)
    record.setdefault('key_Site', settings.site)

    # 日付を統一フォーマット文字列に変換 (Format date string for CSV)
    record['key_start_date_time'] = dt.strftime("%Y-%m-%d %H:%M:%S")

    return record


# -----------------------------
# CSV・XML出力
# (CSV / XML output)
# -----------------------------
def record_to_dataframe(record: dict) -> pd.DataFrame:
    """
    1件のrecordをDataFrameに変換し、カラム名をリネーム。
    (Convert a record to DataFrame and rename columns)
        key_xxx → xxx
        Formation1要件に応じて特定カラムをマッピング (Map special columns as needed)
    """
    df = pd.DataFrame([record])

    special_renames = {
        'key_start_date_time': 'Start_Date_Time',
        'key_serial_number': 'Serial_Number',
        'key_part_number': 'Part_Number',
        'key_batch_number': 'BatchNumber',
        'key_Operation': 'Operation',
        'key_TestStation': 'TestStation',
        'key_Site': 'Site',
        'key_STARTTIME_SORTED': 'STARTTIME_SORTED',
        'key_SORTNUMBER': 'SORTNUMBER',
        'key_LotNumber_9': 'LotNumber_9',
    }

    new_cols = {}
    for col in df.columns:
        if col in special_renames:
            new_cols[col] = special_renames[col]
        elif col.startswith("key_"):
            new_cols[col] = col[4:]  # 去掉 key_
        else:
            new_cols[col] = col

    df = df.rename(columns=new_cols)

    # カラム順序：主要カラムを先に、それ以外を後ろに (Column order: preferred first)
    preferred_order = [
        'Start_Date_Time', 'Serial_Number', 'Part_Number',
        'Operation', 'TestStation', 'Site',
        'LotNumber_9', 'STARTTIME_SORTED', 'SORTNUMBER'
    ]
    cols = list(df.columns)
    ordered_cols = []

    for c in preferred_order:
        if c in cols:
            ordered_cols.append(c)
            cols.remove(c)
    ordered_cols.extend(cols)
    df = df[ordered_cols]

    return df


def write_to_csv(csv_filepath: str, df: pd.DataFrame, log_file: str) -> bool:
    """
    DataFrameをCSVに追記保存（存在すればappend、なければheader付き）。
    (Append DataFrame to CSV, add header if not exists)
    """
    print(f"  [CSV] writing to {csv_filepath}")
    try:
        os.makedirs(os.path.dirname(csv_filepath), exist_ok=True)
        file_exists = os.path.isfile(csv_filepath)
        df.to_csv(csv_filepath, mode='a', header=not file_exists,
                  index=False, encoding='utf-8-sig')
        return True
    except Exception as e:
        print(f"  [CSV] write failed: {e}")
        return False


def generate_pointer_xml(output_path: str, csv_path: str, settings: IniSettings, log_file: str):
    """
    pointer XMLを生成（CVD_Grating_Common.py形式）。
    (Generate pointer XML, CVD_Grating_Common.py style)
    <Results>
        <Result ...>
            <Header ... Operation=... TestStation=... Site=... />
            <TestStep Name=Operation ...>
                <Data DataType="Table"
                            Name="tbl_OPERATION"
                            Value="CSV full path"
                            CompOperation="LOG" />
            </TestStep>
        </Result>
    </Results>
    """
    print("  [XML] generating pointer xml...")
    try:
        os.makedirs(output_path, exist_ok=True)
        now_iso = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
        serial_no = Path(csv_path).stem

        xml_file_name = (
            f"Site={settings.site},"
            f"ProductFamily={settings.product_family},"
            f"Operation={settings.operation},"
            f"Partnumber=UNKNOWPN,"
            f"Serialnumber={serial_no},"
            f"Testdate={now_iso}.xml"
        ).replace(":", ".")

        xml_file_path = os.path.join(output_path, xml_file_name)

        results = ET.Element(
            "Results",
            {
                "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance",
                "xmlns:xsd": "http://www.w3.org/2001/XMLSchema",
            },
        )

        result = ET.SubElement(
            results,
            "Result",
            startDateTime=now_iso,
            endDateTime=now_iso,
            Result="Passed",
        )

        ET.SubElement(
            result,
            "Header",
            SerialNumber=serial_no,
            PartNumber="UNKNOWPN",
            Operation=settings.operation,
            TestStation=settings.test_station,
            Operator="NA",
            StartTime=now_iso,
            Site=settings.site,
            LotNumber=""
        )

        test_step = ET.SubElement(
            result,
            "TestStep",
            Name=settings.operation,
            startDateTime=now_iso,
            endDateTime=now_iso,
            Status="Passed",
        )

        ET.SubElement(
            test_step,
            "Data",
            DataType="Table",
            Name=f"tbl_{settings.operation.upper()}",
            Value=str(csv_path),
            CompOperation="LOG",
        )

        xml_str = minidom.parseString(
            ET.tostring(results)
        ).toprettyxml(indent="  ", encoding="utf-8")

        with open(xml_file_path, "wb") as f:
            f.write(xml_str)

        print(f"  [XML] pointer XML generated at: {xml_file_path}")
    except Exception as e:
        print(f"  [XML] generate_pointer_xml failed: {e}")


# -----------------------------
# Excelファイル処理メインフロー（複数シート）
# (Main flow for processing Excel files, multi-sheet)
# -----------------------------
def process_excel_file(filepath_str: str, settings: IniSettings, log_file: str, csv_filepath: str):
    """
    1つのFormation1 Excelファイルを処理：
    (Process one Formation1 Excel file)
        1) 全シートを順にスキャン (Scan all sheets)
        2) is_format1_sheet()でFormat1判定 (Check Format1)
        3) 各Format1シートで:
                - build_record_from_ws()
                - enrich_record_with_db_and_time()
        4) 全有効recordをDataFrame化しCSVにappend (Collect valid records, append to CSV)
    """
    filepath = Path(filepath_str)
    print(f"--- ファイル処理開始 / Start processing file: {filepath.name} ---")

    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  [処理] load_workbook 失敗 / load_workbook failed: {e}")
        return

    all_rows = []

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        print(f"  [シート] チェック中 / Checking sheet: {sheetname}")
        passed = is_format1_sheet(ws, settings)
        if not passed:
            print(f"    -> 判定/result: 不合格 / Fail, このシートをスキップ / skip this sheet")
            continue
        print(f"    -> 判定/result: 合格 / Pass, データ抽出開始 / extracting data")
        try:
            record = build_record_from_ws(ws, settings, log_file)
            #print(f"      [raw record] {record}")
            record = enrich_record_with_db_and_time(record, settings, log_file)
            if record is not None:
                df_one = record_to_dataframe(record)
                all_rows.append(df_one.iloc[0])
        except Exception as e:
            print(f"    [sheet] error processing sheet {sheetname}: {e}")
            continue

    if not all_rows:
        print("  [process] no valid Format1 sheets in this file.")
        return

    df_all = pd.DataFrame(all_rows)
    if not write_to_csv(csv_filepath, df_all, log_file):
        print("  [process] writing CSV failed.")
        return

    print(f"--- Finished processing {len(all_rows)} sheet(s) in {filepath.name} ---")
    return all_rows


# -----------------------------
# メイン処理：iniを検索し順次処理
# (Main routine: search and process all ini)
# -----------------------------
def main():
    """
    現在のディレクトリ内の全iniを検索し、Formation1 Excelを順次処理。
    (Search all ini in current directory and process Formation1 Excel files)
    """
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

    init_log = setup_logging('../Log/', 'Formation1_Universal_Init')
    print("===== Formation1 Universal Script Start =====")

    ini_files = [f for f in os.listdir('.') if f.endswith('.ini')]
    if not ini_files:
        Log.Log_Info(init_log, "No .ini config files found in current directory.")
        print("No .ini config files found.")
        return

    print(f"Found {len(ini_files)} ini file(s): {', '.join(ini_files)}")

    for ini_path in ini_files:
        try:
            print(f"--- Processing config: {ini_path} ---")
            config = _read_and_parse_ini_config(ini_path)
            settings = _extract_settings_from_config(config)

            # 為每個 operation 建立獨立 log
            log_file = setup_logging(settings.log_path, settings.operation)
            print(f"===== Start operation: {settings.operation} ({ini_path}) =====")

            # 決定 CSV 檔名：以 operation + 日期組成（yyyy_mm_ddThh.mm.ss）
            now_str = datetime.now().strftime("%Y_%m_%dT%H.%M.%S")
            csv_filename = f"{settings.operation}_{now_str}.csv"
            csv_filepath = os.path.join(settings.csv_path, csv_filename)

            # 依 INI 的 input_paths & file_name_patterns 找 Excel 檔
            excel_files = []
            for base in settings.input_paths:
                base = base.strip()
                if not base:
                    continue
                for pat in settings.file_name_patterns:
                    pat = pat.strip()
                    if not pat:
                        continue
                    full_pattern = os.path.join(base, pat)
                    excel_files.extend(glob.glob(full_pattern))

            if not excel_files:
                print(f"No Excel files found for patterns {settings.file_name_patterns}.")
                continue

            excel_files = sorted(set(excel_files))
            print(f"Found {len(excel_files)} Excel file(s) for this INI.")

            # 逐一處理 Excel，收集所有合格資料
            all_rows = []
            for excel_path in excel_files:
                rows = process_excel_file(excel_path, settings, log_file, csv_filepath)
                if rows:
                    all_rows.extend(rows)

            if not all_rows:
                print(f"No valid Format1 sheets found for {settings.operation}, skip CSV/XML.")
                continue

            # 寫入一個 CSV（疊加所有合格 sheet 資料）
            df_all = pd.DataFrame(all_rows)
            if not write_to_csv(csv_filepath, df_all, log_file):
                print("  [process] writing CSV failed.")
                continue

            # 產生一個 XML（對應此 INI）
            generate_pointer_xml(settings.output_path, csv_filepath, settings, log_file)

            print(f"===== Finished operation: {settings.operation} ({ini_path}) =====")

        except Exception:
            error_message = f"FATAL error with INI {ini_path}: {traceback.format_exc()}"
            print(error_message)
            print(error_message)

    print("===== Formation1 Universal Script End =====")
    print("✅ All Formation1 INI configurations have been processed.")


if __name__ == '__main__':
    main()
