import openpyxl as px
import logging
import shutil
import pyodbc
import xlrd
import glob
import sys
import os

from datetime import datetime, timedelta, date
from time import strftime, localtime


########## 自作関数の定義 ##########
sys.path.append('../../MyModule')
import SQL
import Log
import ExpandExp
import Convert_Date
import Row_Number_Func
import MOCVD_OldFileSearch
import Check


########## 全体パラメータの定義 ##########
Site = '350'
ProductFamily = 'SAG FAB'
Operation = 'LD-EML_F6_Format1'
TestStation = 'LD-EML'
X = '999999'
Y = '999999'


########## Logフォルダ名の定義 ##########
Log_FolderName = str(date.today())


# ----- 格納するLogフォルダがなければ作成する -----
if not os.path.exists("../../Log/" + Log_FolderName):
    os.makedirs("../../Log/" + Log_FolderName)

# ----- ログ書き込み先ファイルパス -----
Log_File = '../../Log/' + Log_FolderName + '/040_LD-EML_F6_Format1.log'

# ----- ログ書込：プログラムの開始 -----
Log.Log_Info(Log_File, 'Program Start')


########## 処理ファイルのあるディレクトリ定義 ##########
Path = 'Z:/MOCVD/MOCVD過去プログラム/F6炉/'
# Path = '../../InputFolder/'


########## XML出力先ファイルパス ##########
Output_filepath = '//li.lumentuminc.net/data/SAG/TDS/Data/Files to Insert/XML/'
# Output_filepath = '../../XML/'


########## TestStepの定義　##########
teststep_dict = {
    'TestStep1' : 'Coordinate',
    'TestStep2' : 'TargetWavelength',
    'TestStep3' : 'Thickness',
    'TestStep4' : 'XRayDiffraction',
    'TestStep5' : 'Adjust',
    'TestStep6' : 'MeasurementConditions',
    # 'TestStep7' : 'Dulation',
    # 'TestStep8' : 'MO1-TEG-1',
    # 'TestStep9' : 'MO2-TMI-1',
    # 'TestStep10' : 'MO3-TEG-2',
    # 'TestStep11' : 'MO4-TMI-2',
    # 'TestStep12' : 'MO5-TMI-3',
    # 'TestStep13' : 'AsH3-1-20percent',
    # 'TestStep14' : 'AsH3-21-5percent',
    # 'TestStep15' : 'PH3-2-50percent',
    # 'TestStep16': 'DMZn-2-01percent',
    # 'TestStep17' : 'Si2H6-10ppm',
    # 'TestStep18' : 'GrowthTemperature',
    # 'TestStep19' : 'LayerNo',
    # 'TestStep20' : 'Comment',
    # 'TestStep21' : 'Thickness_Step',
    # 'TestStep22' : 'CarrierConcentration',
    'TestStep23' : 'Piezocon',
    'TestStep24' : 'AsH3_BonbeConc',
    'TestStep25' : 'BallastN2',
    'TestStep26' : 'DMZn_Conc',
    'TestStep27' : 'MO-Temperature',
    'TestStep28' : 'SORTED_DATA'
}


########## HeaderMiscの定義　##########
HeaderMisc_dict = {
    'HeaderMisc1' : 'RecipeName-Macro',
    'HeaderMisc2' : 'RecipeName-Program',
    'HeaderMisc3' : 'RecipeName-Folder'
}


########## 取得した項目と型の対応表を定義 ##########
key_type = {
    "key_start_date_time": str,
    "key_serial_number": str,
    "key_part_number": str,
    "key_operator": str,
    "key_batch_number": str,
    "key_HeaderMisc1": str,
    "key_HeaderMisc2": str,
    "key_HeaderMisc3": str,
    "key_TargetWavelength_TargetWavelength": float,
    "key_Thickness_Thickness_Cap": float,
    "key_Thickness_Thickness_p-Q_115": float,
    "key_Thickness_Thickness_InP_Spacer1": float,
    "key_Thickness_Thickness_Core": float,
    "key_XRayDiffraction_Xray_Thickness": float,
    "key_XRayDiffraction_Xray_Strain": float,
    "key_Adjust_PL_Wavelength": float,
    "key_Adjust_PL_Intensity": float,
    "key_Adjust_PL_FWHM": float,
    "key_Adjust_PL_Adjust": float,
    "key_MeasurementConditions_Templature": float,
    "key_MeasurementConditions_Humidity": float,
    "key_MeasurementConditions_LaserSideFilter1": str,
    "key_MeasurementConditions_LaserSideFilter2": float,
    "key_MeasurementConditions_Zvalue": float,
    "key_MeasurementConditions_PL_IntensityRate_A": float,
    "key_MeasurementConditions_PL_IntensityRate_B": float,
    "key_MeasurementConditions_PL_IntensityRate_C": float,
    "key_MeasurementConditions_PL_IntensityRate_D": float,
    # "key_Dulation_Step9": float,
    # "key_Dulation_Step13": float,
    # "key_Dulation_Step17": float,
    # "key_Dulation_Step21": float,
    # "key_Dulation_Step25": float,
    # "key_Dulation_Step26": float,
    # "key_Dulation_Step31": float,
    # "key_Dulation_Step35": float,
    # "key_Dulation_Step39": float,
    # "key_Dulation_Step43": float,
    # "key_Dulation_Step47": float,
    # "key_MO1-TEG-1_Step9": float,
    # "key_MO1-TEG-1_Step13": float,
    # "key_MO1-TEG-1_Step21": float,
    # "key_MO1-TEG-1_Step25": float,
    # "key_MO1-TEG-1_Step26": float,
    # "key_MO1-TEG-1_Step35": float,
    # "key_MO2-TMI-1_Step9": float,
    # "key_MO2-TMI-1_Step13": float,
    # "key_MO2-TMI-1_Step21": float,
    # "key_MO2-TMI-1_Step25": float,
    # "key_MO2-TMI-1_Step26": float,
    # "key_MO2-TMI-1_Step35": float,
    # "key_MO3-TEG-2_Step17": float,
    # "key_MO3-TEG-2_Step43": float,
    # "key_MO4-TMI-2_Step17": float,
    # "key_MO4-TMI-2_Step43": float,
    # "key_MO5-TMI-3_Step31": float,
    # "key_MO5-TMI-3_Step39": float,
    # "key_MO5-TMI-3_Step47": float,
    # "key_AsH3-1-20percent_Step43": float,
    # "key_AsH3-21-5percent_Step9": float,
    # "key_AsH3-21-5percent_Step13": float,
    # "key_AsH3-21-5percent_Step17": float,
    # "key_AsH3-21-5percent_Step21": float,
    # "key_AsH3-21-5percent_Step25": float,
    # "key_AsH3-21-5percent_Step26": float,
    # "key_AsH3-21-5percent_Step35": float,
    # "key_PH3-2-50percent_Step9": float,
    # "key_PH3-2-50percent_Step13": float,
    # "key_PH3-2-50percent_Step17": float,
    # "key_PH3-2-50percent_Step21": float,
    # "key_PH3-2-50percent_Step25": float,
    # "key_PH3-2-50percent_Step26": float,
    # "key_PH3-2-50percent_Step31": float,
    # "key_PH3-2-50percent_Step35": float,
    # "key_PH3-2-50percent_Step39": float,
    # "key_PH3-2-50percent_Step47": float,
    # "key_DMZn-2-01percent_Step26": float,
    # "key_DMZn-2-01percent_Step31": float,
    # "key_DMZn-2-01percent_Step35": float,
    # "key_DMZn-2-01percent_Step39": float,
    # "key_DMZn-2-01percent_Step43": float,
    # "key_DMZn-2-01percent_Step47": float,
    # "key_Si2H6-10ppm_Step9": float,
    # "key_GrowthTemperature_Step9": float,
    # "key_GrowthTemperature_Step13": float,
    # "key_GrowthTemperature_Step17": float,
    # "key_GrowthTemperature_Step21": float,
    # "key_GrowthTemperature_Step25": float,
    # "key_GrowthTemperature_Step26": float,
    # "key_GrowthTemperature_Step31": float,
    # "key_GrowthTemperature_Step35": float,
    # "key_GrowthTemperature_Step39": float,
    # "key_GrowthTemperature_Step43": float,
    # "key_GrowthTemperature_Step47": float,
    # "key_LayerNo_Step9": str,
    # "key_LayerNo_Step13": str,
    # "key_LayerNo_Step17": str,
    # "key_LayerNo_Step21": str,
    # "key_LayerNo_Step25": str,
    # "key_LayerNo_Step26": str,
    # "key_LayerNo_Step31": str,
    # "key_LayerNo_Step35": str,
    # "key_LayerNo_Step39": str,
    # "key_LayerNo_Step43": str,
    # "key_LayerNo_Step47": str,
    # "key_Comment_Step9": str,
    # "key_Comment_Step13": str,
    # "key_Comment_Step17": str,
    # "key_Comment_Step21": str,
    # "key_Comment_Step25": str,
    # "key_Comment_Step26": str,
    # "key_Comment_Step31": str,
    # "key_Comment_Step35": str,
    # "key_Comment_Step39": str,
    # "key_Comment_Step43": str,
    # "key_Comment_Step47": str,
    # "key_Thickness_Step_Step9": float,
    # "key_Thickness_Step_Step13": float,
    # "key_Thickness_Step_Step17": float,
    # "key_Thickness_Step_Step21": float,
    # "key_Thickness_Step_Step25": float,
    # "key_Thickness_Step_Step26": float,
    # "key_Thickness_Step_Step31": float,
    # "key_Thickness_Step_Step35": float,
    # "key_Thickness_Step_Step39": float,
    # "key_Thickness_Step_Step43": float,
    # "key_Thickness_Step_Step47": float,
    # "key_CarrierConcentration_Step9": float,
    # "key_CarrierConcentration_Step26": float,
    # "key_CarrierConcentration_Step31": float,
    # "key_CarrierConcentration_Step35": float,
    # "key_CarrierConcentration_Step39": float,
    # "key_CarrierConcentration_Step43": float,
    # "key_CarrierConcentration_Step47": float,
    "key_Piezocon_F1": float,
    "key_Piezocon_F1_Inverse": float,
    "key_AsH3_BonbeConc_AsH3_5percent_BonbeDMZn": float,
    "key_BallastN2_BallastN2": float,
    "key_DMZn_Conc_DMZn_ConcAfterReplacing": float,
    "key_DMZn_Conc_DMZn_Conc": float,
    "key_DMZn_Conc_Volatility_5percentOrMore": float,
    "key_MO-Temperature_MO1-TEG": float,
    "key_MO-Temperature_MO2-TMI": float,
    "key_MO-Temperature_MO3-TEG": float,
    "key_MO-Temperature_MO4-TMI": float,
    "key_MO-Temperature_MO5-TMI": float,
    "key_MO-Temperature_MO6-CBr4": float,
    "key_MO-Temperature_MO7-Cp2Mg": float,
    "key_MO-Temperature_MO8-TMA": float,
    "key_STARTTIME_SORTED": float,
    'key_SORTNUMBER' : float,
    "key_LotNumber_9": str
}


########## 対象ロット番号のイニシャルを書込したファイルを取得する ##########
Log.Log_Info(Log_File, 'Get SerialNumber Initial List ')
with open('../../SerialNumber_Initial.txt', 'r', encoding='utf-8') as textfile:
# with open('T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開発/004_T2-EML/SerialNumber_Initial.txt', 'r', encoding='utf-8') as textfile:
#with open('C:/Users/hsi67063/Downloads/SerialNumber_Initial.txt', 'r', encoding='utf-8') as textfile:    
    SerialNumber_list = {s.strip() for s in textfile.readlines()}


########## 前回処理を行ったファイル名を取得する ##########
with open('F6_FileName_Format1.txt', 'r', encoding='utf-8') as textfile:
    Before_FileName = textfile.readline()


########## 空欄チェック ##########
def Get_Cells_Info(Sheet):

    # ----- ログ書込：空欄判定 -----
    Log.Log_Info(Log_File, "Blank Check")

    # ----- False -> 空欄がない -----
    is_cells_empty = False

    # ----- 日付かエピ番号が空欄であれば処理を行わない -----
    if  Sheet['I8'].value is None or Sheet['R7'].value is None:
        is_cells_empty = True

    return is_cells_empty


########## データの取得 ##########
def Open_Data_Sheet(Sheet, filepath, SheetName):

    # ----- ログ書込：データ取得 -----
    Log.Log_Info(Log_File, 'Data Acquisition')

    # ----- 辞書の作成 -----
    data_dict = dict()

    # ----- Serial_NumberをもとにPrimeから品名を引き出す -----
    serial_number = Sheet["M8"].value
    conn, cursor = SQL.connSQL()

    # ----- Prime接続できなかったときはNoneが返ってくる -----
    if conn is None:
        Log.Log_Error(Log_File, serial_number + ' : ' + 'Connection with Prime Failed')
        sys.exit()

    # ----- 品名を取得 -----
    part_number, Nine_Serial_Number = SQL.selectSQL(cursor, serial_number)
    SQL.disconnSQL(conn, cursor)

    # ----- SEM, XRD, MOCVD の装置Noを取得 -----
    SEM, XRD, MOCVD = '1', '1', '6'
    # Equipment判定
    if '#2' in str(Sheet["J72"].value):
        SEM = '2'
    if '#2' in str(Sheet["J43"].value):
        XRD = '2'

    # 品種によってプログラムシートのフォーマット(セル位置)が異なるので場合分けが必要
    if part_number is None:
        part_number_group = 'unknown'
    elif '13B5' in part_number:
        part_number_group = 'B5'
    elif '13B8' in part_number:
        part_number_group = 'B8'
    else:
        part_number_group = 'unknown'
    # ----- データの格納 -----
    if part_number_group == 'B5':
        data_dict = {
            "key_start_date_time": str(Sheet["R7"].value).replace(" ", "T"),
            "key_serial_number": serial_number,
            "key_part_number": part_number,
            "key_operator": '-',
            "key_LotNumber_9": Nine_Serial_Number,
            "key_batch_number": Sheet["I8"].value,
            "key_HeaderMisc1": Sheet["W3"].value,
            "key_HeaderMisc2": Sheet["W4"].value,
            "key_HeaderMisc3": Sheet["W5"].value,
            "key_TargetWavelength_TargetWavelength": Sheet['M13'].value,
            "key_Thickness_Thickness_Cap": Sheet['M74'].value,
            "key_Thickness_Thickness_p-Q_115": Sheet['M75'].value,
            "key_Thickness_Thickness_InP_Spacer1": Sheet['M76'].value,
            "key_Thickness_Thickness_Core": Sheet['M77'].value,
            "key_XRayDiffraction_Xray_Thickness": Sheet['M46'].value,
            "key_XRayDiffraction_Xray_Strain": Sheet['M47'].value,
            "key_Adjust_PL_Wavelength": Sheet['L52'].value,
            "key_Adjust_PL_Intensity": Sheet['N52'].value,
            "key_Adjust_PL_FWHM": Sheet['P52'].value,
            "key_Adjust_PL_Adjust": Sheet['L54'].value,
            "key_MeasurementConditions_Templature": Sheet['AA52'].value,
            "key_MeasurementConditions_Humidity": Sheet['AA53'].value,
            "key_MeasurementConditions_LaserSideFilter1": Sheet['AA58'].value,
            "key_MeasurementConditions_LaserSideFilter2": Sheet['AA59'].value,
            "key_MeasurementConditions_Zvalue": Sheet['AA60'].value,
            "key_MeasurementConditions_PL_IntensityRate_A": Sheet['AD58'].value,
            "key_MeasurementConditions_PL_IntensityRate_B": Sheet['AD59'].value,
            "key_MeasurementConditions_PL_IntensityRate_C": Sheet['AD60'].value,
            "key_MeasurementConditions_PL_IntensityRate_D": Sheet['AD61'].value,
#            "key_Dulation_Step9": Sheet['I22'].value,
#            "key_Dulation_Step13": Sheet['I23'].value,
#            "key_Dulation_Step17": Sheet['I24'].value,
#            "key_Dulation_Step21": Sheet['I25'].value,
#            "key_Dulation_Step25": Sheet['I26'].value,
#            "key_Dulation_Step26": Sheet['I27'].value,
#            "key_Dulation_Step31": Sheet['I28'].value,
#            "key_Dulation_Step35": Sheet['I29'].value,
#            "key_Dulation_Step39": Sheet['I30'].value,
#            "key_Dulation_Step43": Sheet['I31'].value,
#            "key_Dulation_Step47": Sheet['I32'].value,
#            "key_MO1-TEG-1_Step9": Sheet['K22'].value,
#            "key_MO1-TEG-1_Step13": Sheet['K23'].value,
#            "key_MO1-TEG-1_Step21": Sheet['K25'].value,
#            "key_MO1-TEG-1_Step25": Sheet['K26'].value,
#            "key_MO1-TEG-1_Step26": Sheet['K27'].value,
#            "key_MO1-TEG-1_Step35": Sheet['K29'].value,
#            "key_MO2-TMI-1_Step9": Sheet['L22'].value,
#            "key_MO2-TMI-1_Step13": Sheet['L23'].value,
#            "key_MO2-TMI-1_Step21": Sheet['L25'].value,
#            "key_MO2-TMI-1_Step25": Sheet['L26'].value,
#            "key_MO2-TMI-1_Step26": Sheet['L27'].value,
#            "key_MO2-TMI-1_Step35": Sheet['L29'].value,
#            "key_MO3-TEG-2_Step17": Sheet['M24'].value,
#            "key_MO3-TEG-2_Step43": Sheet['M31'].value,
#            "key_MO4-TMI-2_Step17": Sheet['N24'].value,
#            "key_MO4-TMI-2_Step43": Sheet['N31'].value,
#            "key_MO5-TMI-3_Step31": Sheet['O28'].value,
#            "key_MO5-TMI-3_Step39": Sheet['O30'].value,
#            "key_MO5-TMI-3_Step47": Sheet['O32'].value,
#            "key_AsH3-1-20percent_Step43": Sheet['S31'].value,
#            "key_AsH3-21-5percent_Step9": Sheet['T22'].value,
#            "key_AsH3-21-5percent_Step13": Sheet['T23'].value,
#            "key_AsH3-21-5percent_Step17": Sheet['T24'].value,
#            "key_AsH3-21-5percent_Step21": Sheet['T25'].value,
#            "key_AsH3-21-5percent_Step25": Sheet['T26'].value,
#            "key_AsH3-21-5percent_Step26": Sheet['T27'].value,
#            "key_AsH3-21-5percent_Step35": Sheet['T29'].value,
#            "key_PH3-2-50percent_Step9": Sheet['X22'].value,
#            "key_PH3-2-50percent_Step13": Sheet['X23'].value,
#            "key_PH3-2-50percent_Step17": Sheet['X24'].value,
#            "key_PH3-2-50percent_Step21": Sheet['X25'].value,
#            "key_PH3-2-50percent_Step25": Sheet['X26'].value,
#            "key_PH3-2-50percent_Step26": Sheet['X27'].value,
#            "key_PH3-2-50percent_Step31": Sheet['X28'].value,
#            "key_PH3-2-50percent_Step35": Sheet['X29'].value,
#            "key_PH3-2-50percent_Step39": Sheet['X30'].value,
#            "key_PH3-2-50percent_Step47": Sheet['X32'].value,
#            "key_DMZn-2-01percent_Step26": Sheet['Z27'].value,
#            "key_DMZn-2-01percent_Step31": Sheet['Z28'].value,
#            "key_DMZn-2-01percent_Step35": Sheet['Z29'].value,
#            "key_DMZn-2-01percent_Step39": Sheet['Z30'].value,
#            "key_DMZn-2-01percent_Step43": Sheet['Z31'].value,
#            "key_DMZn-2-01percent_Step47": Sheet['Z32'].value,
#            "key_Si2H6-10ppm_Step9": Sheet['AA22'].value,
#            "key_GrowthTemperature_Step9": Sheet['AB22'].value,
#            "key_GrowthTemperature_Step13": Sheet['AB23'].value,
#            "key_GrowthTemperature_Step17": Sheet['AB24'].value,
#            "key_GrowthTemperature_Step21": Sheet['AB25'].value,
#            "key_GrowthTemperature_Step25": Sheet['AB26'].value,
#            "key_GrowthTemperature_Step26": Sheet['AB27'].value,
#            "key_GrowthTemperature_Step31": Sheet['AB28'].value,
#            "key_GrowthTemperature_Step35": Sheet['AB29'].value,
#            "key_GrowthTemperature_Step39": Sheet['AB30'].value,
#            "key_GrowthTemperature_Step43": Sheet['AB31'].value,
#            "key_GrowthTemperature_Step47": Sheet['AB32'].value,
#            "key_LayerNo_Step9": Sheet['AC22'].value,
#            "key_LayerNo_Step13": Sheet['AC23'].value,
#            "key_LayerNo_Step17": Sheet['AC24'].value,
#            "key_LayerNo_Step21": Sheet['AC25'].value,
#            "key_LayerNo_Step25": Sheet['AC26'].value,
#            "key_LayerNo_Step26": Sheet['AC27'].value,
#            "key_LayerNo_Step31": Sheet['AC28'].value,
#            "key_LayerNo_Step35": Sheet['AC29'].value,
#            "key_LayerNo_Step39": Sheet['AC30'].value,
#            "key_LayerNo_Step43": Sheet['AC31'].value,
#            "key_LayerNo_Step47": Sheet['AC32'].value,
#            "key_Comment_Step9": Sheet['AD22'].value,
#            "key_Comment_Step13": Sheet['AD23'].value,
#            "key_Comment_Step17": Sheet['AD24'].value,
#            "key_Comment_Step21": Sheet['AD25'].value,
#            "key_Comment_Step25": Sheet['AD26'].value,
#            "key_Comment_Step26": Sheet['AD27'].value,
#            "key_Comment_Step31": Sheet['AD28'].value,
#            "key_Comment_Step35": Sheet['AD29'].value,
#            "key_Comment_Step39": Sheet['AD30'].value,
#            "key_Comment_Step43": Sheet['AD31'].value,
#            "key_Comment_Step47": Sheet['AD32'].value,
#            "key_Thickness_Step_Step9": Sheet['AH22'].value,
#            "key_Thickness_Step_Step13": Sheet['AH23'].value,
#            "key_Thickness_Step_Step17": Sheet['AH24'].value,
#            "key_Thickness_Step_Step21": Sheet['AH25'].value,
#            "key_Thickness_Step_Step25": Sheet['AH26'].value,
#            "key_Thickness_Step_Step26": Sheet['AH27'].value,
#            "key_Thickness_Step_Step31": Sheet['AH28'].value,
#            "key_Thickness_Step_Step35": Sheet['AH29'].value,
#            "key_Thickness_Step_Step39": Sheet['AH30'].value,
#            "key_Thickness_Step_Step43": Sheet['AH31'].value,
#            "key_Thickness_Step_Step47": Sheet['AH32'].value,
#            "key_CarrierConcentration_Step9": Sheet['AI22'].value,
#            "key_CarrierConcentration_Step26": Sheet['AI27'].value,
#            "key_CarrierConcentration_Step31": Sheet['AI28'].value,
#            "key_CarrierConcentration_Step35": Sheet['AI29'].value,
#            "key_CarrierConcentration_Step39": Sheet['AI30'].value,
#            "key_CarrierConcentration_Step43": Sheet['AI31'].value,
#            "key_CarrierConcentration_Step47": Sheet['AI32'].value,
            "key_Piezocon_F1": Sheet['AG37'].value,
            "key_Piezocon_F1_Inverse": Sheet['AG38'].value,
            "key_AsH3_BonbeConc_AsH3_5percent_BonbeDMZn": Sheet['AG39'].value,
            "key_BallastN2_BallastN2": Sheet['AG40'].value,
            "key_DMZn_Conc_DMZn_ConcAfterReplacing": Sheet['AG43'].value,
            "key_DMZn_Conc_DMZn_Conc": Sheet['AG44'].value,
            "key_DMZn_Conc_Volatility_5percentOrMore": Sheet['AG45'].value,
            "key_MO-Temperature_MO1-TEG": Sheet['K42'].value,
            "key_MO-Temperature_MO2-TMI": Sheet['L42'].value,
            "key_MO-Temperature_MO3-TEG": Sheet['M42'].value,
            "key_MO-Temperature_MO4-TMI": Sheet['N42'].value,
            "key_MO-Temperature_MO5-TMI": Sheet['O42'].value,
            "key_MO-Temperature_MO6-CBr4": Sheet['P42'].value,
            "key_MO-Temperature_MO7-Cp2Mg": Sheet['Q42'].value,
            "key_MO-Temperature_MO8-TMA": Sheet['R42'].value,
            "key_TestEquipment_SEM": SEM,
            "key_TestEquipment_PLmapper": '1',
            "key_TestEquipment_XRD": XRD,
            "key_TestEquipment_MOCVD": MOCVD
        }
    elif part_number_group == 'B8':
        data_dict = {
            "key_start_date_time": str(Sheet["R7"].value).replace(" ", "T"),
            "key_serial_number": serial_number,
            "key_part_number": part_number,
            "key_operator": '-',
            "key_LotNumber_9": Nine_Serial_Number,
            "key_batch_number": Sheet["I8"].value,
            "key_HeaderMisc1": Sheet["W3"].value,
            "key_HeaderMisc2": Sheet["W4"].value,
            "key_HeaderMisc3": Sheet["W5"].value,
            "key_TargetWavelength_TargetWavelength": Sheet['M13'].value,
            "key_Thickness_Thickness_Cap": Sheet['M71'].value,
            "key_Thickness_Thickness_p-Q_115": Sheet['M72'].value,
            "key_Thickness_Thickness_InP_Spacer1": Sheet['M73'].value,
            "key_Thickness_Thickness_Core": Sheet['M74'].value,
            "key_XRayDiffraction_Xray_Thickness": Sheet['M45'].value,
            "key_XRayDiffraction_Xray_Strain": Sheet['M46'].value,
            "key_Adjust_PL_Wavelength": Sheet['L51'].value,
            "key_Adjust_PL_Intensity": Sheet['N51'].value,
            "key_Adjust_PL_FWHM": Sheet['P51'].value,
            "key_Adjust_PL_Adjust": Sheet['L53'].value,
            "key_MeasurementConditions_Templature": Sheet['AA51'].value,
            "key_MeasurementConditions_Humidity": Sheet['AA52'].value,
            "key_MeasurementConditions_LaserSideFilter1": Sheet['AA57'].value,
            "key_MeasurementConditions_LaserSideFilter2": Sheet['AA58'].value,
            "key_MeasurementConditions_Zvalue": Sheet['AA59'].value,
            "key_MeasurementConditions_PL_IntensityRate_A": Sheet['AD57'].value,
            "key_MeasurementConditions_PL_IntensityRate_B": Sheet['AD58'].value,
            "key_MeasurementConditions_PL_IntensityRate_C": Sheet['AD59'].value,
            "key_MeasurementConditions_PL_IntensityRate_D": Sheet['AD60'].value,
        # "key_Dulation_Step9": Sheet['I22'].value,
        # "key_Dulation_Step13": Sheet['I23'].value,
        # "key_Dulation_Step17": Sheet['I24'].value,
        # "key_Dulation_Step21": Sheet['I25'].value,
        # "key_Dulation_Step25": Sheet['I26'].value,
        # "key_Dulation_Step26": Sheet['I27'].value,
        # "key_Dulation_Step31": Sheet['I28'].value,
        # "key_Dulation_Step35": Sheet['I29'].value,
        # "key_Dulation_Step39": Sheet['I30'].value,
        # "key_Dulation_Step43": Sheet['I31'].value,
        # "key_Dulation_Step47": Sheet['I32'].value,
        # "key_MO1-TEG-1_Step9": Sheet['K22'].value,
        # "key_MO1-TEG-1_Step13": Sheet['K23'].value,
        # "key_MO1-TEG-1_Step21": Sheet['K25'].value,
        # "key_MO1-TEG-1_Step25": Sheet['K26'].value,
        # "key_MO1-TEG-1_Step26": Sheet['K27'].value,
        # "key_MO1-TEG-1_Step35": Sheet['K29'].value,
        # "key_MO2-TMI-1_Step9": Sheet['L22'].value,
        # "key_MO2-TMI-1_Step13": Sheet['L23'].value,
        # "key_MO2-TMI-1_Step21": Sheet['L25'].value,
        # "key_MO2-TMI-1_Step25": Sheet['L26'].value,
        # "key_MO2-TMI-1_Step26": Sheet['L27'].value,
        # "key_MO2-TMI-1_Step35": Sheet['L29'].value,
        # "key_MO3-TEG-2_Step17": Sheet['M24'].value,
        # "key_MO3-TEG-2_Step43": Sheet['M31'].value,
        # "key_MO4-TMI-2_Step17": Sheet['N24'].value,
        # "key_MO4-TMI-2_Step43": Sheet['N31'].value,
        # "key_MO5-TMI-3_Step31": Sheet['O28'].value,
        # "key_MO5-TMI-3_Step39": Sheet['O30'].value,
        # "key_MO5-TMI-3_Step47": Sheet['O32'].value,
        # "key_AsH3-1-20percent_Step43": Sheet['S31'].value,
        # "key_AsH3-21-5percent_Step9": Sheet['T22'].value,
        # "key_AsH3-21-5percent_Step13": Sheet['T23'].value,
        # "key_AsH3-21-5percent_Step17": Sheet['T24'].value,
        # "key_AsH3-21-5percent_Step21": Sheet['T25'].value,
        # "key_AsH3-21-5percent_Step25": Sheet['T26'].value,
        # "key_AsH3-21-5percent_Step26": Sheet['T27'].value,
        # "key_AsH3-21-5percent_Step35": Sheet['T29'].value,
        # "key_PH3-2-50percent_Step9": Sheet['X22'].value,
        # "key_PH3-2-50percent_Step13": Sheet['X23'].value,
        # "key_PH3-2-50percent_Step17": Sheet['X24'].value,
        # "key_PH3-2-50percent_Step21": Sheet['X25'].value,
        # "key_PH3-2-50percent_Step25": Sheet['X26'].value,
        # "key_PH3-2-50percent_Step26": Sheet['X27'].value,
        # "key_PH3-2-50percent_Step31": Sheet['X28'].value,
        # "key_PH3-2-50percent_Step35": Sheet['X29'].value,
        # "key_PH3-2-50percent_Step39": Sheet['X30'].value,
        # "key_PH3-2-50percent_Step47": Sheet['X32'].value,
        # "key_DMZn-2-01percent_Step26": Sheet['Z27'].value,
        # "key_DMZn-2-01percent_Step31": Sheet['Z28'].value,
        # "key_DMZn-2-01percent_Step35": Sheet['Z29'].value,
        # "key_DMZn-2-01percent_Step39": Sheet['Z30'].value,
        # "key_DMZn-2-01percent_Step43": Sheet['Z31'].value,
        # "key_DMZn-2-01percent_Step47": Sheet['Z32'].value,
        # "key_Si2H6-10ppm_Step9": Sheet['AA22'].value,
        # "key_GrowthTemperature_Step9": Sheet['AB22'].value,
        # "key_GrowthTemperature_Step13": Sheet['AB23'].value,
        # "key_GrowthTemperature_Step17": Sheet['AB24'].value,
        # "key_GrowthTemperature_Step21": Sheet['AB25'].value,
        # "key_GrowthTemperature_Step25": Sheet['AB26'].value,
        # "key_GrowthTemperature_Step26": Sheet['AB27'].value,
        # "key_GrowthTemperature_Step31": Sheet['AB28'].value,
        # "key_GrowthTemperature_Step35": Sheet['AB29'].value,
        # "key_GrowthTemperature_Step39": Sheet['AB30'].value,
        # "key_GrowthTemperature_Step43": Sheet['AB31'].value,
        # "key_GrowthTemperature_Step47": Sheet['AB32'].value,
        # "key_LayerNo_Step9": Sheet['AC22'].value,
        # "key_LayerNo_Step13": Sheet['AC23'].value,
        # "key_LayerNo_Step17": Sheet['AC24'].value,
        # "key_LayerNo_Step21": Sheet['AC25'].value,
        # "key_LayerNo_Step25": Sheet['AC26'].value,
        # "key_LayerNo_Step26": Sheet['AC27'].value,
        # "key_LayerNo_Step31": Sheet['AC28'].value,
        # "key_LayerNo_Step35": Sheet['AC29'].value,
        # "key_LayerNo_Step39": Sheet['AC30'].value,
        # "key_LayerNo_Step43": Sheet['AC31'].value,
        # "key_LayerNo_Step47": Sheet['AC32'].value,
        # "key_Comment_Step9": Sheet['AD22'].value,
        # "key_Comment_Step13": Sheet['AD23'].value,
        # "key_Comment_Step17": Sheet['AD24'].value,
        # "key_Comment_Step21": Sheet['AD25'].value,
        # "key_Comment_Step25": Sheet['AD26'].value,
        # "key_Comment_Step26": Sheet['AD27'].value,
        # "key_Comment_Step31": Sheet['AD28'].value,
        # "key_Comment_Step35": Sheet['AD29'].value,
        # "key_Comment_Step39": Sheet['AD30'].value,
        # "key_Comment_Step43": Sheet['AD31'].value,
        # "key_Comment_Step47": Sheet['AD32'].value,
        # "key_Thickness_Step_Step9": Sheet['AH22'].value,
        # "key_Thickness_Step_Step13": Sheet['AH23'].value,
        # "key_Thickness_Step_Step17": Sheet['AH24'].value,
        # "key_Thickness_Step_Step21": Sheet['AH25'].value,
        # "key_Thickness_Step_Step25": Sheet['AH26'].value,
        # "key_Thickness_Step_Step26": Sheet['AH27'].value,
        # "key_Thickness_Step_Step31": Sheet['AH28'].value,
        # "key_Thickness_Step_Step35": Sheet['AH29'].value,
        # "key_Thickness_Step_Step39": Sheet['AH30'].value,
        # "key_Thickness_Step_Step43": Sheet['AH31'].value,
        # "key_Thickness_Step_Step47": Sheet['AH32'].value,
        # "key_CarrierConcentration_Step9": Sheet['AI22'].value,
        # "key_CarrierConcentration_Step26": Sheet['AI27'].value,
        # "key_CarrierConcentration_Step31": Sheet['AI28'].value,
        # "key_CarrierConcentration_Step35": Sheet['AI29'].value,
        # "key_CarrierConcentration_Step39": Sheet['AI30'].value,
        # "key_CarrierConcentration_Step43": Sheet['AI31'].value,
        # "key_CarrierConcentration_Step47": Sheet['AI32'].value,
            "key_Piezocon_F1": Sheet['AG36'].value,
            "key_Piezocon_F1_Inverse": Sheet['AG37'].value,
            "key_AsH3_BonbeConc_AsH3_5percent_BonbeDMZn": Sheet['AG38'].value,
            "key_BallastN2_BallastN2": Sheet['AG39'].value,
            "key_DMZn_Conc_DMZn_ConcAfterReplacing": Sheet['AG42'].value,
            "key_DMZn_Conc_DMZn_Conc": Sheet['AG43'].value,
            "key_DMZn_Conc_Volatility_5percentOrMore": Sheet['AG44'].value,
            "key_MO-Temperature_MO1-TEG": Sheet['K41'].value,
            "key_MO-Temperature_MO2-TMI": Sheet['L41'].value,
            "key_MO-Temperature_MO3-TEG": Sheet['M41'].value,
            "key_MO-Temperature_MO4-TMI": Sheet['N41'].value,
            "key_MO-Temperature_MO5-TMI": Sheet['O41'].value,
            "key_MO-Temperature_MO6-CBr4": Sheet['P41'].value,
            "key_MO-Temperature_MO7-Cp2Mg": Sheet['Q41'].value,
            "key_MO-Temperature_MO8-TMA": Sheet['R41'].value,
            "key_TestEquipment_SEM": SEM,
            "key_TestEquipment_PLmapper": '1',
            "key_TestEquipment_XRD": XRD,
            "key_TestEquipment_MOCVD": MOCVD
        }
    else:
        pass # data_dict remains empty.

    # ----- 空欄箇所はNoneとして取得される。Noneは文字列に変換できないため、空欄("")に置き換える -----
    for keys in data_dict:
        if data_dict[keys] is None or str(data_dict[keys]) == "#VALUE!" or (data_dict[keys] == '-' and keys != "key_operator"):
            data_dict[keys] = ""
        # ----- 指数表記を展開する -----
        if type(data_dict[keys]) is float and 'e' in str(data_dict[keys]) and keys != "key_start_date_time":
            data_dict[keys] = ExpandExp.Expand(data_dict[keys])

    return data_dict


########## XML変換 ##########
def Output_XML(xml_file, data_dict):

    # ----- ログ書込：XML変換 -----
    Log.Log_Info(Log_File, 'Excel File To XML File Conversion')
    
    f = open(Output_filepath + xml_file, 'w', encoding="utf-8")

    f.write('<?xml version="1.0" encoding="utf-8"?>' + '\n' +
            '<Results xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema">' + '\n' +
            '       <Result startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Result="Passed">' + '\n' +
            '               <Header SerialNumber=' + '"' + data_dict["key_serial_number"] + '"' + ' PartNumber=' + '"' + data_dict["key_part_number"] + '"' + ' Operation=' + '"' + Operation + '"' + ' TestStation=' + '"' + TestStation + '"' + ' Operator=' + '"' + data_dict["key_operator"] + '"' + ' StartTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Site=' + '"' + Site + '"' + ' BatchNumber=' + '"' + data_dict["key_batch_number"] + '"' + ' LotNumber=' + '"' + data_dict["key_serial_number"] + '"/>' + '\n' +
            '               <HeaderMisc>' + '\n' +
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc1"] + '">' + data_dict["key_HeaderMisc1"] + '</Item>' + '\n'
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc2"] + '">' + data_dict["key_HeaderMisc2"] + '</Item>' + '\n'
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc3"] + '">' + data_dict["key_HeaderMisc3"] + '</Item>' + '\n'
            '               </HeaderMisc>' + '\n' +
            '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep1"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + X + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + Y + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep2"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="TargetWavelength" Units="nm" Value=' + '"' + str(data_dict["key_TargetWavelength_TargetWavelength"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep3"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Thickness_Cap" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Thickness_Cap"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Thickness_p-Q_115" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Thickness_p-Q_115"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Thickness_InP_Spacer1" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Thickness_InP_Spacer1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Thickness_Core" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Thickness_Core"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep4"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Xray_Thickness" Units="nm" Value=' + '"' + str(data_dict["key_XRayDiffraction_Xray_Thickness"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Xray_Strain" Units="percent" Value=' + '"' + str(data_dict["key_XRayDiffraction_Xray_Strain"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep5"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="PL_Wavelength" Units="nm" Value=' + '"' + str(data_dict["key_Adjust_PL_Wavelength"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="PL_Intensity" Units="mV" Value=' + '"' + str(data_dict["key_Adjust_PL_Intensity"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="PL_FWHM" Units="meV" Value=' + '"' + str(data_dict["key_Adjust_PL_FWHM"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="PL_Adjust" Units="nm" Value=' + '"' + str(data_dict["key_Adjust_PL_Adjust"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep6"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Templature" Units="degree" Value=' + '"' + str(data_dict["key_MeasurementConditions_Templature"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Humidity" Units="percent" Value=' + '"' + str(data_dict["key_MeasurementConditions_Humidity"]) + '"/>' + '\n' +
            '                   <Data DataType="String" Name="LaserSideFilter1" Value=' + '"' + str(data_dict["key_MeasurementConditions_LaserSideFilter1"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="LaserSideFilter2" Units="percent" Value=' + '"' + str(data_dict["key_MeasurementConditions_LaserSideFilter2"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Zvalue" Units="um" Value=' + '"' + str(data_dict["key_MeasurementConditions_Zvalue"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="PL_IntensityRate_A" Units="percent" Value=' + '"' + str(data_dict["key_MeasurementConditions_PL_IntensityRate_A"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="PL_IntensityRate_B" Units="percent" Value=' + '"' + str(data_dict["key_MeasurementConditions_PL_IntensityRate_B"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="PL_IntensityRate_C" Units="percent" Value=' + '"' + str(data_dict["key_MeasurementConditions_PL_IntensityRate_C"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="PL_IntensityRate_D" Units="percent" Value=' + '"' + str(data_dict["key_MeasurementConditions_PL_IntensityRate_D"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
#            '               <TestStep Name=' + '"' + teststep_dict["TestStep7"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step9" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step9"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step13" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step13"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step17" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step17"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step21" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step21"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step25" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step25"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step26" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step26"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step31" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step31"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step35" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step35"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step39" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step39"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step43" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step43"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step47" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step47"]) + '"/>' + '\n' +
#            '               </TestStep>' + '\n' +
#            '               <TestStep Name=' + '"' + teststep_dict["TestStep8"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step9" Units="sccm" Value=' + '"' + str(data_dict["key_MO1-TEG-1_Step9"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step13" Units="sccm" Value=' + '"' + str(data_dict["key_MO1-TEG-1_Step13"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step21" Units="sccm" Value=' + '"' + str(data_dict["key_MO1-TEG-1_Step21"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step25" Units="sccm" Value=' + '"' + str(data_dict["key_MO1-TEG-1_Step25"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step26" Units="sccm" Value=' + '"' + str(data_dict["key_MO1-TEG-1_Step26"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step35" Units="sccm" Value=' + '"' + str(data_dict["key_MO1-TEG-1_Step35"]) + '"/>' + '\n' +
#            '               </TestStep>' + '\n' +
#            '               <TestStep Name=' + '"' + teststep_dict["TestStep9"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step9" Units="sccm" Value=' + '"' + str(data_dict["key_MO2-TMI-1_Step9"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step13" Units="sccm" Value=' + '"' + str(data_dict["key_MO2-TMI-1_Step13"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step21" Units="sccm" Value=' + '"' + str(data_dict["key_MO2-TMI-1_Step21"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step25" Units="sccm" Value=' + '"' + str(data_dict["key_MO2-TMI-1_Step25"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step26" Units="sccm" Value=' + '"' + str(data_dict["key_MO2-TMI-1_Step26"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step35" Units="sccm" Value=' + '"' + str(data_dict["key_MO2-TMI-1_Step35"]) + '"/>' + '\n' +
#            '               </TestStep>' + '\n' +
#            '               <TestStep Name=' + '"' + teststep_dict["TestStep10"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step17" Units="sccm" Value=' + '"' + str(data_dict["key_MO3-TEG-2_Step17"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step43" Units="sccm" Value=' + '"' + str(data_dict["key_MO3-TEG-2_Step43"]) + '"/>' + '\n' +
#            '               </TestStep>' + '\n' +
#            '               <TestStep Name=' + '"' + teststep_dict["TestStep11"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step17" Units="sccm" Value=' + '"' + str(data_dict["key_MO4-TMI-2_Step17"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step43" Units="sccm" Value=' + '"' + str(data_dict["key_MO4-TMI-2_Step43"]) + '"/>' + '\n' +
#            '               </TestStep>' + '\n' +
#            '               <TestStep Name=' + '"' + teststep_dict["TestStep12"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step31" Units="sccm" Value=' + '"' + str(data_dict["key_MO5-TMI-3_Step31"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step39" Units="sccm" Value=' + '"' + str(data_dict["key_MO5-TMI-3_Step39"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step47" Units="sccm" Value=' + '"' + str(data_dict["key_MO5-TMI-3_Step47"]) + '"/>' + '\n' +
#            '               </TestStep>' + '\n' +
#            '               <TestStep Name=' + '"' + teststep_dict["TestStep13"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step43" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-1-20percent_Step43"]) + '"/>' + '\n' +
#            '               </TestStep>' + '\n' +
#            '               <TestStep Name=' + '"' + teststep_dict["TestStep14"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step9" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-21-5percent_Step9"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step13" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-21-5percent_Step13"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step17" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-21-5percent_Step17"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step21" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-21-5percent_Step21"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step25" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-21-5percent_Step25"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step26" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-21-5percent_Step26"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step35" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-21-5percent_Step35"]) + '"/>' + '\n' +
#            '               </TestStep>' + '\n' +
#            '               <TestStep Name=' + '"' + teststep_dict["TestStep15"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step9" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-2-50percent_Step9"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step13" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-2-50percent_Step13"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step17" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-2-50percent_Step17"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step21" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-2-50percent_Step21"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step25" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-2-50percent_Step25"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step26" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-2-50percent_Step26"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step31" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-2-50percent_Step31"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step35" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-2-50percent_Step35"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step39" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-2-50percent_Step39"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step47" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-2-50percent_Step47"]) + '"/>' + '\n' +
#            '               </TestStep>' + '\n' +
#            '               <TestStep Name=' + '"' + teststep_dict["TestStep16"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step26" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-2-01percent_Step26"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step31" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-2-01percent_Step31"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step35" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-2-01percent_Step35"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step39" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-2-01percent_Step39"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step43" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-2-01percent_Step43"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step47" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-2-01percent_Step47"]) + '"/>' + '\n' +
#            '               </TestStep>' + '\n' +
#            '               <TestStep Name=' + '"' + teststep_dict["TestStep17"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step9" Units="sccm" Value=' + '"' + str(data_dict["key_Si2H6-10ppm_Step9"]) + '"/>' + '\n' +
#            '               </TestStep>' + '\n' +
#            '               <TestStep Name=' + '"' + teststep_dict["TestStep18"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step9" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step9"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step13" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step13"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step17" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step17"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step21" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step21"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step25" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step25"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step26" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step26"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step31" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step31"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step35" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step35"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step39" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step39"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step43" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step43"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step47" Units="degree" Value=' + '"' + str(data_dict["key_GrowthTemperature_Step47"]) + '"/>' + '\n' +
#            '               </TestStep>' + '\n' +
#            '               <TestStep Name=' + '"' + teststep_dict["TestStep19"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
#            '                   <Data DataType="String" Name="Step9" Value=' + '"' + str(data_dict["key_LayerNo_Step9"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
#            '                   <Data DataType="String" Name="Step13" Value=' + '"' + str(data_dict["key_LayerNo_Step13"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
#            '                   <Data DataType="String" Name="Step17" Value=' + '"' + str(data_dict["key_LayerNo_Step17"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
#            '                   <Data DataType="String" Name="Step21" Value=' + '"' + str(data_dict["key_LayerNo_Step21"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
#            '                   <Data DataType="String" Name="Step25" Value=' + '"' + str(data_dict["key_LayerNo_Step25"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
#            '                   <Data DataType="String" Name="Step26" Value=' + '"' + str(data_dict["key_LayerNo_Step26"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
#            '                   <Data DataType="String" Name="Step31" Value=' + '"' + str(data_dict["key_LayerNo_Step31"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
#            '                   <Data DataType="String" Name="Step35" Value=' + '"' + str(data_dict["key_LayerNo_Step35"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
#            '                   <Data DataType="String" Name="Step39" Value=' + '"' + str(data_dict["key_LayerNo_Step39"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
#            '                   <Data DataType="String" Name="Step43" Value=' + '"' + str(data_dict["key_LayerNo_Step43"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
#            '                   <Data DataType="String" Name="Step47" Value=' + '"' + str(data_dict["key_LayerNo_Step47"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
#            '               </TestStep>' + '\n' +
#            '               <TestStep Name=' + '"' + teststep_dict["TestStep20"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
#            '                   <Data DataType="String" Name="Step9" Value=' + '"' + str(data_dict["key_Comment_Step9"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
#            '                   <Data DataType="String" Name="Step13" Value=' + '"' + str(data_dict["key_Comment_Step13"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
#            '                   <Data DataType="String" Name="Step17" Value=' + '"' + str(data_dict["key_Comment_Step17"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
#            '                   <Data DataType="String" Name="Step21" Value=' + '"' + str(data_dict["key_Comment_Step21"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
#            '                   <Data DataType="String" Name="Step25" Value=' + '"' + str(data_dict["key_Comment_Step25"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
#            '                   <Data DataType="String" Name="Step26" Value=' + '"' + str(data_dict["key_Comment_Step26"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
#            '                   <Data DataType="String" Name="Step31" Value=' + '"' + str(data_dict["key_Comment_Step31"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
#            '                   <Data DataType="String" Name="Step35" Value=' + '"' + str(data_dict["key_Comment_Step35"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
#            '                   <Data DataType="String" Name="Step39" Value=' + '"' + str(data_dict["key_Comment_Step39"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
#            '                   <Data DataType="String" Name="Step43" Value=' + '"' + str(data_dict["key_Comment_Step43"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
#            '                   <Data DataType="String" Name="Step47" Value=' + '"' + str(data_dict["key_Comment_Step47"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
#            '               </TestStep>' + '\n' +
#            '               <TestStep Name=' + '"' + teststep_dict["TestStep21"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step9" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step9"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step13" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step13"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step17" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step17"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step21" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step21"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step25" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step25"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step26" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step26"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step31" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step31"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step35" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step35"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step39" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step39"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step43" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step43"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step47" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Step_Step47"]) + '"/>' + '\n' +
#            '               </TestStep>' + '\n' +
#            '               <TestStep Name=' + '"' + teststep_dict["TestStep22"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step9" Units="cm-3" Value=' + '"' + str(data_dict["key_CarrierConcentration_Step9"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step26" Units="cm-3" Value=' + '"' + str(data_dict["key_CarrierConcentration_Step26"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step31" Units="cm-3" Value=' + '"' + str(data_dict["key_CarrierConcentration_Step31"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step35" Units="cm-3" Value=' + '"' + str(data_dict["key_CarrierConcentration_Step35"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step39" Units="cm-3" Value=' + '"' + str(data_dict["key_CarrierConcentration_Step39"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step43" Units="cm-3" Value=' + '"' + str(data_dict["key_CarrierConcentration_Step43"]) + '"/>' + '\n' +
#            '                   <Data DataType="Numeric" Name="Step47" Units="cm-3" Value=' + '"' + str(data_dict["key_CarrierConcentration_Step47"]) + '"/>' + '\n' +
#            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep23"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="F1" Units="percent" Value=' + '"' + str(data_dict["key_Piezocon_F1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="F1_Inverse" Units="percent" Value=' + '"' + str(data_dict["key_Piezocon_F1_Inverse"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep24"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="AsH3_5percent_BonbeDMZn" Units="percent" Value=' + '"' + str(data_dict["key_AsH3_BonbeConc_AsH3_5percent_BonbeDMZn"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep25"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="BallastN2" Units="slm" Value=' + '"' + str(data_dict["key_BallastN2_BallastN2"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep26"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="DMZn_ConcAfterReplacing" Units="percent" Value=' + '"' + str(data_dict["key_DMZn_Conc_DMZn_ConcAfterReplacing"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="DMZn_Conc" Units="percent" Value=' + '"' + str(data_dict["key_DMZn_Conc_DMZn_Conc"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Volatility_5percentOrMore" Units="percent" Value=' + '"' + str(data_dict["key_DMZn_Conc_Volatility_5percentOrMore"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep27"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="MO1-TEG" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO1-TEG"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO2-TMI" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO2-TMI"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO3-TEG" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO3-TEG"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO4-TMI" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO4-TMI"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO5-TMI" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO5-TMI"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO6-CBr4" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO6-CBr4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO7-Cp2Mg" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO7-Cp2Mg"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO8-TMA" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO8-TMA"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep28"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="STARTTIME_SORTED" Units="" Value=' + '"' + str(data_dict["key_STARTTIME_SORTED"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="SORTNUMBER" Units="" Value=' + '"' + str(data_dict["key_SORTNUMBER"]) + '"/>' + '\n' +
            '                   <Data DataType="String" Name="LotNumber_5" Value=' + '"' + str(data_dict["key_serial_number"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="LotNumber_9" Value=' + '"' + str(data_dict["key_LotNumber_9"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '\n'
            '               <TestEquipment>' + '\n' +
            '                   <Item DeviceName="SEM" DeviceSerialNumber="' + data_dict["key_TestEquipment_SEM"] + '"/>' + '\n' +
            '                   <Item DeviceName="PLmapper" DeviceSerialNumber="' + data_dict["key_TestEquipment_PLmapper"] + '"/>' + '\n' +
            '                   <Item DeviceName="XRD" DeviceSerialNumber="' + data_dict["key_TestEquipment_XRD"] + '"/>' + '\n' +
            '                   <Item DeviceName="MOCVD" DeviceSerialNumber="' + data_dict["key_TestEquipment_MOCVD"] + '"/>' + '\n' +
            '               </TestEquipment>' + '\n' +
            '\n'
            '               <ErrorData/>' + '\n' +
            '               <FailureData/>' + '\n' +
            '               <Configuration/>' + '\n' +
            '       </Result>' + '\n' +
            '</Results>'
            )
    f.close()


########## シートの判定からXML変換までの関数 ##########
def Data_Extract(filepath, SheetList):

    # ----- ログ書込：データ変換処理 -----
    Log.Log_Info(Log_File, 'Sub Start')

    wb = px.load_workbook(filepath, read_only=True, data_only=True)

    # ----- 処理するSheetsリストを昇順に並べる -----
    Sheets = sorted(SheetList)

    for Sheet_Name in Sheets:

        Sheet = wb[Sheet_Name]
        Initial = str(Sheet['M8'].value)[0]

        if Initial not in SerialNumber_list or not any(x in str(Sheet['W3'].value) for x in [
            '13B2-LD',
            '13B4-LD',
            '13B5-LD',
            '13B7-LD',
            '13B8-LD',
            '13E1-LD'
        ]):
            Log.Log_Error(Log_File, Sheet_Name + ' : ' + 'Not Covered\n')
            continue
            
        # ----- 空欄チェック -----
        if Get_Cells_Info(Sheet):
            Log.Log_Error(Log_File, "Blank Error\n")
            continue

        # ----- データの取得 -----
        data_dict = Open_Data_Sheet(Sheet, os.path.basename(filepath), Sheet_Name)
        if not data_dict: # part_number_groupが集計対象外だった(2025-06-23時点ではB5,B8のみ集計対象)
            continue

        # ----- oldファイルの実行時のみ、着工者が空欄であれば'-'に置き換える -----
        if data_dict["key_operator"] == "":
            if old_check:
                data_dict["key_operator"] = '-'
            else:
                Log.Log_Error(Log_File, Sheet_Name + ' : ' + 'Operator None\n')
                continue

        # ----- 日付フォーマット変換 -----
        if len(data_dict['key_start_date_time']) != 19 or '年' in data_dict['key_start_date_time']:
            Log.Log_Error(Log_File, data_dict["key_serial_number"] + ' : ' + "Date Error\n")
            continue

        # ----- ロット番号をキーとして品名が得られなかった -----
        if len(data_dict["key_part_number"]) == 0:
            Log.Log_Error(Log_File, data_dict["key_serial_number"] + ' : ' + "Part Number Error\n")
            continue

        # ----- STARTTIME_SORTEDの追加 -----

        # 日付をExcel時間に変換する
        date = datetime.strptime(str(data_dict["key_start_date_time"]).replace('T', ' ').replace('.', ':'), "%Y-%m-%d %H:%M:%S")
        date_excel_number = int(str(date - datetime(1899, 12, 30)).split()[0])

        # エピ番号の数値部だけを取得する
        Epi_Number = 0
        for i in data_dict["key_batch_number"]:
            try:
                if 0<=int(i)<=9:
                    Epi_Number = Epi_Number*10+int(i)
            except:
                pass

        # エピ番号を10^6で割って、excel時間に加算する
        date_excel_number += Epi_Number/10**6

        # data_dictに登録する
        data_dict["key_STARTTIME_SORTED"] = date_excel_number
        data_dict["key_SORTNUMBER"] = Epi_Number

        # ----- データ型の確認 -----
        result = Check.Data_Type(key_type, data_dict)
        if result == False:
            Log.Log_Error(Log_File, data_dict["key_serial_number"] + ' : ' + "Data Error\n")
            continue

        # ----- XMLファイルの作成 -----
        xml_file = 'Site=' + Site + ',ProductFamily=' + ProductFamily + ',Operation=' + Operation + \
                   ',Partnumber=' + data_dict["key_part_number"] + ',Serialnumber=' + data_dict["key_serial_number"] + \
                   ',Testdate=' + data_dict["key_start_date_time"].replace(':','.') + '.xml'

        Output_XML(xml_file, data_dict)
        Log.Log_Info(Log_File, data_dict["key_serial_number"] + ' : ' + "OK\n")

    wb.close()


########## Main処理 ##########
if __name__ == '__main__':

    # ----- ログ書込：Main処理の開始 -----
    Log.Log_Info(Log_File, 'Main Start')

    # ----- path内のフォルダ、ファイルを全部取得 -----
    all_files = os.listdir(Path)

    # ----- ログ書込：着工ファイル検索 -----
    Log.Log_Info(Log_File, 'File Search')

    # ----- ファイルパスの取得 -----
    array = []
    for filename in all_files:
        filepath = os.path.join(Path, filename)
        # Only process Excel files
        if filename.lower().endswith(('.xlsx', '.xlsm')) and "FM" in filename and "$" not in filename and os.path.isfile(filepath):
            dt = strftime("%Y-%m-%d %H:%M:%S", localtime(os.path.getctime(filepath)))
            array.append([filepath, dt])

    # ----- 着工ファイルが見つからなかったときはエラー処理 -----
    if len(array) == 0:
        Log.Log_Info(Log_File, 'Folder Error')
        sys.exit()

    # 最終更新日時順に並び替える
    array = sorted(array, key=lambda x: x[1])
    FileName = os.path.basename(array[0][0])
    Log.Log_Info(Log_File, FileName)

    # ----- 前回処理したエピ番号のNumber部分を取り出す -----
    Number = ""
    for i in Before_FileName:
        if "0" <= i <= "9":
            Number += i

    # ----- ファイルの切り替わりを確認 -----
    if Number not in FileName:

        # ----- ログ書込：フォルダ検索 -----
        Log.Log_Info(Log_File, 'Folder Serach')

        Old_file_path = MOCVD_OldFileSearch.F6(Number)

        if Old_file_path == -1:
            Log.Log_Info(Log_File, 'Old Folder Error')
            sys.exit()

        # ----- ログ書込：シート名の取得 -----
        Log.Log_Info(Log_File, 'OLD Get SheetName')

        # ----- 上記で指定したファイルのシート一覧を取得する -----
        wb = px.load_workbook(Old_file_path)
        Old_SheetName = wb.sheetnames
        wb.close()

        # ----- ログ書込：前Excelファイルのデータ取得 -----
        Log.Log_Info(Log_File, 'OLD Excel File Get Data')

        # ---- 過去ファイルの処理 -----
        Data_Extract(Old_file_path, Old_SheetName)

    # ----- ログ書込：Excelファイルのデータ取得 -----
    Log.Log_Info(Log_File, 'Excel File Get Data')

    # ----- arrayに格納されている全てのファイルの処理を行う -----
    for File_Path, _ in array:

        Log.Log_Info(Log_File, os.path.basename(File_Path))

        # ----- 対象ファイルを開き、シートの一覧を取得する -----
        wb = px.load_workbook(File_Path)
        SheetName = wb.sheetnames
        wb.close()

        Data_Extract(File_Path, SheetName)

    # ----- ログ書込：テキストファイルにシート名を上書きで書込する -----
    Log.Log_Info(Log_File, 'Write FileName')

    # ----- 先ほど処理を行ったファイル名の書き込み -----
    with open('F6_FileName_Format1.txt', 'w', encoding='utf-8') as textfile:
        textfile.write(FileName)

    # ----- 最終行を書き込んだファイルをGドライブに転送 -----
    # shutil.copy('F6_FileName_Format1.txt', 'T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開発/040_LD-EML/F6/13_ProgramUsedFile/')


########## ログ書込：プログラムの終了 ##########
Log.Log_Info(Log_File, 'Program End')