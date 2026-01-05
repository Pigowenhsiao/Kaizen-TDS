import openpyxl as px
import pandas as pd
import logging
import shutil
import pyodbc
import xlrd
import glob
import sys
import os

from datetime import datetime, timedelta, date
from time import strftime

# 自作関数の定義
sys.path.append('../../MyModule')
import SQL
import Log
import ExpandExp
import Convert_Date
import Row_Number_Func
import Check


########## 全体パラメータの定義 ##########
Site = '350'
ProductFamily = 'SAG FAB'
Operation = 'LD-EML_F6_Format2'
TestStation = 'LD-EML'
X = '999999'
Y = '999999'


########## Logフォルダ名の定義 ##########
Log_FolderName = str(date.today())

# ----- 格納するLogフォルダがなければ作成する -----
if not os.path.exists("../../Log/" + Log_FolderName):
    os.makedirs("../../Log/" + Log_FolderName)

# ----- ログ書き込み先ファイルパス -----
Log_File = '../../Log/' + Log_FolderName + '/040_LD-EML_F6_Format2.log'

# ----- ログ書込：プログラムの開始 -----
Log.Log_Info(Log_File, 'Program Start')


########## XML出力先ファイルパス ##########
Output_filepath = '//li.lumentuminc.net/data/SAG/TDS/Data/Files to Insert/XML/'
# Output_filepath = '../../XML/040_LD-EML/F6_Format2/'
#Output_filepath = 'C:/Users/hsi67063/Documents/TEMP/'  #for test


########## シート名定義 ##########
Data_sheet_name = "まとめ"


########## TestStepの定義 ##########
teststep_dict = {
    'TestStep1': 'Coordinate',
    'TestStep2': 'Mapper_Average',
    'TestStep3': 'Mapper_Adjust',
    'TestStep4': 'Center',
    'TestStep5': 'A',
    'TestStep6': 'B',
    'TestStep7': 'C',
    'TestStep8': 'D',
    'TestStep9': 'AB',
    'TestStep10': 'AC',
    'TestStep11': 'BC',
    'TestStep12': 'CD',
    'TestStep13': 'AA',
    'TestStep14': 'BB',
    'TestStep15': 'CC',
    'TestStep16': 'DD',
    'TestStep17': 'SORTED_DATA',
}


########## 取得した項目と型の対応表を定義 ##########
key_type = {
    "key_start_date_time": str,
    "key_part_number": str,
    "key_serial_number": str,
    "key_operator": str,
    "key_batch_number": str,
    "key_Mapper_Average_Mapper_PL_Lambda": float,
    "key_Mapper_Average_Mapper_PL_Lambda2": float, #2025/01/15 update
    "key_Mapper_Average_Mapper_PL_Intensity": float,
    "key_Mapper_Average_Mapper_PL_FWHM": float,
    "key_Mapper_Adjust_Mapper_TargetWavelength": float,
    "key_Mapper_Adjust_Mapper_Wavelength(1.3um)": float,
    "key_Mapper_Adjust_Mapper_CheckingAdjustValue": float,
    "key_Center_X": float,
    "key_Center_Y": float,
    "key_Center_Lambda": float,
    "key_Center_Intensity": float,
    "key_Center_FWHM": float,
    "key_Center_Tails": float,
    "key_Center_DeltaLambda": float,
    "key_A_X": float,
    "key_A_Y": float,
    "key_A_Lambda": float,
    "key_A_Intensity": float,
    "key_A_FWHM": float,
    "key_A_Tails": float,
    "key_A_DeltaLambda": float,
    "key_B_X": float,
    "key_B_Y": float,
    "key_B_Lambda": float,
    "key_B_Intensity": float,
    "key_B_FWHM": float,
    "key_B_Tails": float,
    "key_B_DeltaLambda": float,
    "key_C_X": float,
    "key_C_Y": float,
    "key_C_Lambda": float,
    "key_C_Intensity": float,
    "key_C_FWHM": float,
    "key_C_Tails": float,
    "key_C_DeltaLambda": float,
    "key_D_X": float,
    "key_D_Y": float,
    "key_D_Lambda": float,
    "key_D_Intensity": float,
    "key_D_FWHM": float,
    "key_D_Tails": float,
    "key_D_DeltaLambda": float,
    "key_AB_X": float,
    "key_AB_Y": float,
    "key_AB_Lambda": float,
    "key_AB_Intensity": float,
    "key_AB_FWHM": float,
    "key_AB_Tails": float,
    "key_AB_DeltaLambda": float,
    "key_AC_X": float,
    "key_AC_Y": float,
    "key_AC_Lambda": float,
    "key_AC_Intensity": float,
    "key_AC_FWHM": float,
    "key_AC_Tails": float,
    "key_AC_DeltaLambda": float,
    "key_BC_X": float,
    "key_BC_Y": float,
    "key_BC_Lambda": float,
    "key_BC_Intensity": float,
    "key_BC_FWHM": float,
    "key_BC_Tails": float,
    "key_BC_DeltaLambda": float,
    "key_CD_X": float,
    "key_CD_Y": float,
    "key_CD_Lambda": float,
    "key_CD_Intensity": float,
    "key_CD_FWHM": float,
    "key_CD_Tails": float,
    "key_CD_DeltaLambda": float,
    "key_AA_X": float,
    "key_AA_Y": float,
    "key_AA_Lambda": float,
    "key_AA_Intensity": float,
    "key_AA_FWHM": float,
    "key_AA_Tails": float,
    "key_AA_DeltaLambda": float,
    "key_BB_X": float,
    "key_BB_Y": float,
    "key_BB_Lambda": float,
    "key_BB_Intensity": float,
    "key_BB_FWHM": float,
    "key_BB_Tails": float,
    "key_BB_DeltaLambda": float,
    "key_CC_X": float,
    "key_CC_Y": float,
    "key_CC_Lambda": float,
    "key_CC_Intensity": float,
    "key_CC_FWHM": float,
    "key_CC_Tails": float,
    "key_CC_DeltaLambda": float,
    "key_DD_X": float,
    "key_DD_Y": float,
    "key_DD_Lambda": float,
    "key_DD_Intensity": float,
    "key_DD_FWHM": float,
    "key_DD_Tails": float,
    "key_DD_DeltaLambda": float,
    "key_Cmb_Date_Time": "datetime",
    "key_Cmb_Spectral_Range": str,
    "key_Cmb_Peak_Position": float,
    "key_Cmb_Peak_Intensity": float,
    "key_Cmb_FWHM(nm)": float,
    "key_Cmb_FWHM(meV)": float,
    "key_Cmb_SideBandHight": float,
    "key_Cmb_X": float,
    "key_Cmb_Y": float,
    "key_STARTTIME_SORTED": float,
    "key_SORTNUMBER": float,
    "key_LotNumber_9": str
}


########## 対象フォルダの全ファイルを格納 ##########
ALL_FILES = []


########## 取得しないフォルダ名 ##########
NotUsedDir = {'2017年', '2018年', '2019年','2020年','2021年','2022年','2023年', 'ヒメジ検討', '2015', '2016', '2017', '2018', '2018', '2019','未使用'}


########## 対象ロット番号のイニシャルを書込したファイルを取得する ##########
Log.Log_Info(Log_File, 'Get SerialNumber Initial List ')
#with open('T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開発/004_T2-EML/SerialNumber_Initial.txt', 'r', encoding='utf-8') as textfile:
with open('C:/Users/hsi67063/Downloads/SerialNumber_Initial.txt', 'r', encoding='utf-8') as textfile:

    SerialNumber_list = {s.strip() for s in textfile.readlines()}


########## 処理を行ったファイル名をテキストファイルから呼び出す ##########
with open('EndsFile_F6_Format2.txt', 'r', encoding='utf-8') as textfile:
    EndFiles = set(s.strip() for s in textfile.readlines())


########## 空欄チェック ##########
def Get_Cells_Info(file_path):
    
    # ----- ログ書込：空欄判定 -----
    Log.Log_Info(Log_File, "Blank Check")
    
    try:
    
        # ----- ファイルとシートを指定し開く -----
        if file_path.endswith('.xls'):
            return False
        wb = px.load_workbook(file_path, data_only=True)
        sheet = wb[Data_sheet_name]
    
        # ----- False -> 空欄がない -----
        is_cells_empty = False
        


        # ----- 空欄はNone表示となる -----
        if sheet.cell(5, 8).value is None or \
            sheet.cell(4, 8).value is None:
            is_cells_empty = True
    
        wb.close()
   
    except:
        
        # ----- ファイルとシートを指定し開く -----
        wb = px.load_workbook(file_path)
        sheet = wb[Data_sheet_name]
    
        # ----- False -> 空欄がない -----
        is_cells_empty = False
        

    
        # ----- 空欄はNone表示となる -----
        if sheet.cell(4, 8).value is None or \
            sheet.cell(5, 8).value is None:
            is_cells_empty = True
                    
    
        wb.close()
    

    return is_cells_empty

########## Excelからデータを取得 ##########
def Open_Data_Sheet(file_path):

    # ----- ログ書込：データ取得 -----
    Log.Log_Info(Log_File, 'Data Acquisition')

    # ----- 辞書の初期化 -----
    data_dict = dict()

    # ----- Excelファイルとシートの定義 -----
    wb = px.load_workbook(file_path, data_only=True)
    sheet = wb.worksheets[0]
    date_sheet = wb['データファイル']

    # ----- ロット番号を取得し、Primeにロット番号に対応する品種があるか調べる -----
    serial_number = sheet.cell(4, 8).value



    # ----- ロット番号が数値の時があるので、数値の時はエラーを出す -----
    if type(serial_number) is not str:
        return None, None

    # ----- Prime接続 -----
    conn, cursor = SQL.connSQL()
    if conn is None:
        Log.Log_Error(Log_File, serial_number + ' : ' + 'Connection with Prime Failed')
        sys.exit()

    # ----- 品名を取得 -----
    part_number, Nine_Serial_Number = SQL.selectSQL(cursor, serial_number)
    SQL.disconnSQL(conn, cursor)
    print(cursor, serial_number,part_number, Nine_Serial_Number)
    #input("Enter")
    # ----- [データファイル]シートのA5を抜き出す -----
    start_date = date_sheet.cell(5, 1).value

    # ----- Operatorが空欄であれば、'-'に置き換える -----
    Operator = str(sheet.cell(3, 5).value)
    if Operator == "None" or Operator == "":
        Operator = "-"

    # ----- PL-MAP -----
    PL_MAP = '1'
    if '#2' in str(file_path):
        PL_MAP = '2'

    # ----- データの取得 -----
    data_dict = {
        "key_start_date_time": start_date,
        "key_part_number": part_number,
        "key_serial_number": serial_number,
        "key_operator": Operator,
        "key_LotNumber_9": Nine_Serial_Number,
        "key_batch_number": sheet.cell(3, 8).value,
        "key_Mapper_Average_Mapper_PL_Lambda": sheet.cell(3, 14).value,
        "key_Mapper_Average_Mapper_PL_Lambda2": (
            sheet.cell(14, 6).value + sheet.cell(15, 6).value + sheet.cell(16, 6).value + sheet.cell(17, 6).value
        ) / 4,  # 2025/01/15 update
        "key_Mapper_Average_Mapper_PL_Intensity": sheet.cell(4, 14).value,
        "key_Mapper_Average_Mapper_PL_FWHM": sheet.cell(5, 14).value,
        "key_Mapper_Adjust_Mapper_TargetWavelength": sheet.cell(27, 5).value,
        "key_Mapper_Adjust_Mapper_Wavelength(1.3um)": sheet.cell(28, 5).value,
        "key_Mapper_Adjust_Mapper_CheckingAdjustValue": sheet.cell(29, 5).value,
        "key_Center_X": sheet.cell(9, 4).value,
        "key_Center_Y": sheet.cell(9, 5).value,
        "key_Center_Lambda": sheet.cell(9, 6).value,
        "key_Center_Intensity": sheet.cell(9, 7).value,
        "key_Center_FWHM": sheet.cell(9, 8).value,
        "key_Center_Tails": sheet.cell(9, 9).value,
        "key_Center_DeltaLambda": sheet.cell(9, 10).value,
        "key_A_X": sheet.cell(10, 4).value,
        "key_A_Y": sheet.cell(10, 5).value,
        "key_A_Lambda": sheet.cell(10, 6).value,
        "key_A_Intensity": sheet.cell(10, 7).value,
        "key_A_FWHM": sheet.cell(10, 8).value,
        "key_A_Tails": sheet.cell(10, 9).value,
        "key_A_DeltaLambda": sheet.cell(10, 10).value,
        "key_B_X": sheet.cell(11, 4).value,
        "key_B_Y": sheet.cell(11, 5).value,
        "key_B_Lambda": sheet.cell(11, 6).value,
        "key_B_Intensity": sheet.cell(11, 7).value,
        "key_B_FWHM": sheet.cell(11, 8).value,
        "key_B_Tails": sheet.cell(11, 9).value,
        "key_B_DeltaLambda": sheet.cell(11, 10).value,
        "key_C_X": sheet.cell(12, 4).value,
        "key_C_Y": sheet.cell(12, 5).value,
        "key_C_Lambda": sheet.cell(12, 6).value,
        "key_C_Intensity": sheet.cell(12, 7).value,
        "key_C_FWHM": sheet.cell(12, 8).value,
        "key_C_Tails": sheet.cell(12, 9).value,
        "key_C_DeltaLambda": sheet.cell(12, 10).value,
        "key_D_X": sheet.cell(13, 4).value,
        "key_D_Y": sheet.cell(13, 5).value,
        "key_D_Lambda": sheet.cell(13, 6).value,
        "key_D_Intensity": sheet.cell(13, 7).value,
        "key_D_FWHM": sheet.cell(13, 8).value,
        "key_D_Tails": sheet.cell(13, 9).value,
        "key_D_DeltaLambda": sheet.cell(13, 10).value,
        "key_AB_X": sheet.cell(14, 4).value,
        "key_AB_Y": sheet.cell(14, 5).value,
        "key_AB_Lambda": sheet.cell(14, 6).value,
        "key_AB_Intensity": sheet.cell(14, 7).value,
        "key_AB_FWHM": sheet.cell(14, 8).value,
        "key_AB_Tails": sheet.cell(14, 9).value,
        "key_AB_DeltaLambda": sheet.cell(14, 10).value,
        "key_AC_X": sheet.cell(15, 4).value,
        "key_AC_Y": sheet.cell(15, 5).value,
        "key_AC_Lambda": sheet.cell(15, 6).value,
        "key_AC_Intensity": sheet.cell(15, 7).value,
        "key_AC_FWHM": sheet.cell(15, 8).value,
        "key_AC_Tails": sheet.cell(15, 9).value,
        "key_AC_DeltaLambda": sheet.cell(15, 10).value,
        "key_BC_X": sheet.cell(16, 4).value,
        "key_BC_Y": sheet.cell(16, 5).value,
        "key_BC_Lambda": sheet.cell(16, 6).value,
        "key_BC_Intensity": sheet.cell(16, 7).value,
        "key_BC_FWHM": sheet.cell(16, 8).value,
        "key_BC_Tails": sheet.cell(16, 9).value,
        "key_BC_DeltaLambda": sheet.cell(16, 10).value,
        "key_CD_X": sheet.cell(17, 4).value,
        "key_CD_Y": sheet.cell(17, 5).value,
        "key_CD_Lambda": sheet.cell(17, 6).value,
        "key_CD_Intensity": sheet.cell(17, 7).value,
        "key_CD_FWHM": sheet.cell(17, 8).value,
        "key_CD_Tails": sheet.cell(17, 9).value,
        "key_CD_DeltaLambda": sheet.cell(17, 10).value,
        "key_AA_X": sheet.cell(18, 4).value,
        "key_AA_Y": sheet.cell(18, 5).value,
        "key_AA_Lambda": sheet.cell(18, 6).value,
        "key_AA_Intensity": sheet.cell(18, 7).value,
        "key_AA_FWHM": sheet.cell(18, 8).value,
        "key_AA_Tails": sheet.cell(18, 9).value,
        "key_AA_DeltaLambda": sheet.cell(18, 10).value,
        "key_BB_X": sheet.cell(19, 4).value,
        "key_BB_Y": sheet.cell(19, 5).value,
        "key_BB_Lambda": sheet.cell(19, 6).value,
        "key_BB_Intensity": sheet.cell(19, 7).value,
        "key_BB_FWHM": sheet.cell(19, 8).value,
        "key_BB_Tails": sheet.cell(19, 9).value,
        "key_BB_DeltaLambda": sheet.cell(19, 10).value,
        "key_CC_X": sheet.cell(20, 4).value,
        "key_CC_Y": sheet.cell(20, 5).value,
        "key_CC_Lambda": sheet.cell(20, 6).value,
        "key_CC_Intensity": sheet.cell(20, 7).value,
        "key_CC_FWHM": sheet.cell(20, 8).value,
        "key_CC_Tails": sheet.cell(20, 9).value,
        "key_CC_DeltaLambda": sheet.cell(20, 10).value,
        "key_DD_X": sheet.cell(21, 4).value,
        "key_DD_Y": sheet.cell(21, 5).value,
        "key_DD_Lambda": sheet.cell(21, 6).value,
        "key_DD_Intensity": sheet.cell(21, 7).value,
        "key_DD_FWHM": sheet.cell(21, 8).value,
        "key_DD_Tails": sheet.cell(21, 9).value,
        "key_DD_DeltaLambda": sheet.cell(21, 10).value,
        "key_PL_MAP": PL_MAP
    }

    wb.close()

    # ----- 空欄箇所はNoneとして取得される。Noneは文字列に変換できないため、空欄("")に置き換える -----
    for keys in data_dict:
        if data_dict[keys] is None or data_dict[keys] == "#DIV/0!" or (data_dict[keys] == '-' and keys != "key_operator"):
            data_dict[keys] = ""
        # ----- 指数表記を展開する -----
        if type(data_dict[keys]) is float and 'e' in str(data_dict[keys]) and keys != "key_start_date_time":
            data_dict[keys] = ExpandExp.Expand(data_dict[keys])

    # ----- データファイルシートからデータを取得する -----
    Data_file_sheet_array = []
    df = pd.read_excel(file_path, header=None, sheet_name='データファイル', usecols="A:I", skiprows=1)
    row, col = df.shape
    for i in range(row):
        tmp = []
        for j in range(col):
            tmp.append(str(df.iloc[i, j]))
        Data_file_sheet_array.append(tmp)

    # ----- 1行目のデータだけ、データ型確認用として辞書に追加する -----
    data_dict['key_Cmb_Date_Time'] = Data_file_sheet_array[0][0]
    data_dict['key_Cmb_Spectral_Range'] = Data_file_sheet_array[0][1]
    data_dict['key_Cmb_Peak_Position'] = Data_file_sheet_array[0][2]
    data_dict['key_Cmb_Peak_Intensity'] = Data_file_sheet_array[0][3]
    data_dict['key_Cmb_FWHM(nm)'] = Data_file_sheet_array[0][4]
    data_dict['key_Cmb_FWHM(meV)'] = Data_file_sheet_array[0][5]
    data_dict['key_Cmb_SideBandHight'] = Data_file_sheet_array[0][6]
    data_dict['key_Cmb_X'] = Data_file_sheet_array[0][7]
    data_dict['key_Cmb_Y'] = Data_file_sheet_array[0][8]
    data_dict['Lamnda_diff']= data_dict["key_Mapper_Average_Mapper_PL_Lambda"]-data_dict["key_Mapper_Adjust_Mapper_Wavelength(1.3um)"] #2025/01/15 add
    data_dict['Lamnda2_diff'] =  data_dict["key_Mapper_Average_Mapper_PL_Lambda2"]-data_dict["key_Center_Lambda"] #2025/01/15 add

    return data_dict, Data_file_sheet_array


########## XML変換 ##########
def Output_XML(xml_file, data_dict, Data_file_sheet_array):
    print("Output_XML")
    input("Enter")
    # ----- ログ書込：XML変換 -----
    Log.Log_Info(Log_File, 'Excel File To XML File Conversion')
    
    XML = '<?xml version="1.0" encoding="utf-8"?>' + '\n' + \
        '<Results xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema">' + '\n' + \
        '       <Result startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Result="Passed">' + '\n' + \
        '               <Header SerialNumber=' + '"' + data_dict["key_serial_number"] + '"' + ' PartNumber=' + '"' + data_dict["key_part_number"] + '"' + ' Operation=' + '"' + Operation + '"' + ' TestStation=' + '"' + TestStation + '"' + ' Operator=' + '"' + data_dict["key_operator"] + '"' + ' StartTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Site=' + '"' + Site + '"' + ' BatchNumber=' + '"' + data_dict["key_batch_number"] + '"' + ' LotNumber=' + '"' + data_dict["key_serial_number"] + '"/>' + '\n' + \
        '\n' + \
        '               <TestStep Name=' + '"' + teststep_dict["TestStep1"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
        '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + X + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + Y + '"/>' + '\n' + \
        '               </TestStep>' + '\n' + \
        '               <TestStep Name=' + '"' + teststep_dict["TestStep2"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
        '                   <Data DataType="Numeric" Name="Mapper_PL_Lambda" Units="nm" Value=' + '"' + str(data_dict["key_Mapper_Average_Mapper_PL_Lambda"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Mapper_PL_Intensity" Units="mV" Value=' + '"' + str(data_dict["key_Mapper_Average_Mapper_PL_Intensity"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Mapper_PL_FWHM" Units="meV" Value=' + '"' + str(data_dict["key_Mapper_Average_Mapper_PL_FWHM"]) + '"/>' + '\n' + \
        '               </TestStep>' + '\n' + \
        '               <TestStep Name=' + '"' + teststep_dict["TestStep3"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
        '                   <Data DataType="Numeric" Name="Mapper_TargetWavelength" Units="nm" Value=' + '"' + str(data_dict["key_Mapper_Adjust_Mapper_TargetWavelength"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Mapper_Wavelength(1.3um)" Units="nm" Value=' + '"' + str(data_dict["key_Mapper_Adjust_Mapper_Wavelength(1.3um)"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Mapper_CheckingAdjustValue" Units="nm" Value=' + '"' + str(data_dict["key_Mapper_Adjust_Mapper_CheckingAdjustValue"]) + '"/>' + '\n' + \
        '               </TestStep>' + '\n' + \
        '               <TestStep Name=' + '"' + teststep_dict["TestStep4"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
        '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_Center_X"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_Center_Y"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_Center_Lambda"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Lambda_Diff" Units="nm" Value=' + '"' + str(data_dict["Lamnda_diff"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Lambda2_Diff" Units="nm" Value=' + '"' + str(data_dict['Lamnda2_diff']) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_Center_Intensity"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_Center_FWHM"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_Center_Tails"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_Center_DeltaLambda"]) + '"/>' + '\n' + \
        '               </TestStep>' + '\n' + \
        '               <TestStep Name=' + '"' + teststep_dict["TestStep5"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
        '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_A_X"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_A_Y"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_A_Lambda"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_A_Intensity"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_A_FWHM"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_A_Tails"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_A_DeltaLambda"]) + '"/>' + '\n' + \
        '               </TestStep>' + '\n' + \
        '               <TestStep Name=' + '"' + teststep_dict["TestStep6"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
        '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_B_X"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_B_Y"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_B_Lambda"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_B_Intensity"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_B_FWHM"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_B_Tails"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_B_DeltaLambda"]) + '"/>' + '\n' + \
        '               </TestStep>' + '\n' + \
        '               <TestStep Name=' + '"' + teststep_dict["TestStep7"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
        '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_C_X"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_C_Y"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_C_Lambda"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_C_Intensity"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_C_FWHM"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_C_Tails"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_C_DeltaLambda"]) + '"/>' + '\n' + \
        '               </TestStep>' + '\n' + \
        '               <TestStep Name=' + '"' + teststep_dict["TestStep8"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
        '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_D_X"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_D_Y"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_D_Lambda"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_D_Intensity"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_D_FWHM"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_D_Tails"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_D_DeltaLambda"]) + '"/>' + '\n' + \
        '               </TestStep>' + '\n' + \
        '               <TestStep Name=' + '"' + teststep_dict["TestStep9"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
        '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_AB_X"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_AB_Y"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_AB_Lambda"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_AB_Intensity"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_AB_FWHM"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_AB_Tails"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_AB_DeltaLambda"]) + '"/>' + '\n' + \
        '               </TestStep>' + '\n' + \
        '               <TestStep Name=' + '"' + teststep_dict["TestStep10"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
        '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_AC_X"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_AC_Y"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_AC_Lambda"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_AC_Intensity"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_AC_FWHM"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_AC_Tails"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_AC_DeltaLambda"]) + '"/>' + '\n' + \
        '               </TestStep>' + '\n' + \
        '               <TestStep Name=' + '"' + teststep_dict["TestStep11"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
        '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_BC_X"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_BC_Y"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_BC_Lambda"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_BC_Intensity"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_BC_FWHM"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_BC_Tails"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_BC_DeltaLambda"]) + '"/>' + '\n' + \
        '               </TestStep>' + '\n' + \
        '               <TestStep Name=' + '"' + teststep_dict["TestStep12"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
        '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_CD_X"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_CD_Y"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_CD_Lambda"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_CD_Intensity"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_CD_FWHM"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_CD_Tails"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_CD_DeltaLambda"]) + '"/>' + '\n' + \
        '               </TestStep>' + '\n' + \
        '               <TestStep Name=' + '"' + teststep_dict["TestStep13"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
        '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_AA_X"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_AA_Y"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_AA_Lambda"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_AA_Intensity"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_AA_FWHM"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_AA_Tails"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_AA_DeltaLambda"]) + '"/>' + '\n' + \
        '               </TestStep>' + '\n' + \
        '               <TestStep Name=' + '"' + teststep_dict["TestStep14"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
        '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_BB_X"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_BB_Y"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_BB_Lambda"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_BB_Intensity"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_BB_FWHM"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_BB_Tails"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_BB_DeltaLambda"]) + '"/>' + '\n' + \
        '               </TestStep>' + '\n' + \
        '               <TestStep Name=' + '"' + teststep_dict["TestStep15"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
        '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_CC_X"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_CC_Y"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_CC_Lambda"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_CC_Intensity"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_CC_FWHM"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_CC_Tails"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_CC_DeltaLambda"]) + '"/>' + '\n' + \
        '               </TestStep>' + '\n' + \
        '               <TestStep Name=' + '"' + teststep_dict["TestStep16"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
        '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_DD_X"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_DD_Y"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_DD_Lambda"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_DD_Intensity"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_DD_FWHM"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_DD_Tails"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_DD_DeltaLambda"]) + '"/>' + '\n' + \
        '               </TestStep>' + '\n' + \
        '               <TestStep Name=' + '"' + teststep_dict["TestStep17"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
        '                   <Data DataType="Numeric" Name="STARTTIME_SORTED" Units="" Value=' + '"' + str(data_dict["key_STARTTIME_SORTED"]) + '"/>' + '\n' + \
        '                   <Data DataType="Numeric" Name="SORTNUMBER" Units="" Value=' + '"' + str(data_dict["key_SORTNUMBER"]) + '"/>' + '\n' + \
        '                   <Data DataType="String" Name="LotNumber_5" Value=' + '"' + str(data_dict["key_serial_number"]) + '"' + ' CompOperation="LOG"/>' + '\n' + \
        '                   <Data DataType="String" Name="LotNumber_9" Value=' + '"' + str(data_dict["key_LotNumber_9"]) + '"' + ' CompOperation="LOG"/>' + '\n' + \
        '               </TestStep>' + '\n'

    # Cmbの付与
    for i in range(len(Data_file_sheet_array)):
        TestStep = 'Cmb' + str(i + 1)
        XML += '               <TestStep Name=' + '"' + TestStep + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
               '                   <Data DataType="String" Name="Cmb_Date_Time" Value=' + '"' + str(Data_file_sheet_array[i][0]) + '"' + ' CompOperation="LOG"/>' + '\n' + \
               '                   <Data DataType="String" Name="Cmb_Spectral_Range" Value=' + '"' + str(Data_file_sheet_array[i][1]) + '"' + ' CompOperation="LOG"/>' + '\n' + \
               '                   <Data DataType="Numeric" Name="Cmb_Peak_Position" Units="nm" Value=' + '"' + str(Data_file_sheet_array[i][2]) + '"/>' + '\n' + \
               '                   <Data DataType="Numeric" Name="Cmb_Peak_Intensity" Units="counts" Value=' + '"' + str(Data_file_sheet_array[i][3]) + '"/>' + '\n' + \
               '                   <Data DataType="Numeric" Name="Cmb_FWHM(nm)" Units="nm" Value=' + '"' + str(Data_file_sheet_array[i][4]) + '"/>' + '\n' + \
               '                   <Data DataType="Numeric" Name="Cmb_FWHM(meV)" Units="meV" Value=' + '"' + str(Data_file_sheet_array[i][5]) + '"/>' + '\n' + \
               '                   <Data DataType="Numeric" Name="Cmb_SideBandHight" Units="counts" Value=' + '"' + str(Data_file_sheet_array[i][6]) + '"/>' + '\n' + \
               '                   <Data DataType="Numeric" Name="Cmb_X" Units="um" Value=' + '"' + str(Data_file_sheet_array[i][7]) + '"/>' + '\n' + \
               '                   <Data DataType="Numeric" Name="Cmb_Y" Units="um" Value=' + '"' + str(Data_file_sheet_array[i][8]) + '"/>' + '\n' + \
               '               </TestStep>' + '\n'
    
    XML += '\n' \
            '               <TestEquipment>' + '\n' + \
            '                   <Item DeviceName="PLmapper" DeviceSerialNumber="' + str(data_dict["key_PL_MAP"]) + '"/>' + '\n' + \
            '                   <Item DeviceName="MOCVD" DeviceSerialNumber="' + "6" + '"/>' + '\n' + \
            '               </TestEquipment>' + '\n' + \
            '\n' \
            '               <ErrorData/>' + '\n' + \
            '               <FailureData/>' + '\n' + \
            '               <Configuration/>' + '\n' + \
            '       </Result>' + '\n' + \
            '</Results>'

    
        
    f = open(Output_filepath + xml_file, 'w', encoding="utf-8")
            
    f.write(XML)
    f.close()


########## 指定パス内のすべてのファイルを取得 ##########
def ALL_FILE_FETCH(Path):
    global ALL_FILES

    for path, dir, file in os.walk(Path):
        # ----- dir の再帰でPathを生成しているため、省くフォルダについてはdirを書き換える必要がある -----
        dir[:] = [d for d in dir if d not in NotUsedDir]
        for f in file:
            if (f.endswith('.xlsx') or f.endswith('.xlsm') or f.endswith('.xls')) and "PL" in str(f):
                ALL_FILES.append(os.path.join(path, f))


########## main処理 ##########
if __name__ == '__main__':

    # ----- ログ書込：Main処理の開始 -----
    Log.Log_Info(Log_File, 'Main Start')

    # ----- 各フォルダ内の拡張子が[.xl*]のものをすべて取り出す -----
    #Path = ["Z:/PL-MAP/★PLﾏｯﾊﾟｰ判定/3ｲﾝﾁHTL13B2/F6LD/", "Z:/PL-MAP/★PLﾏｯﾊﾟｰ判定/3ｲﾝﾁHTL13B4/F6 LD", "Z:/PL-MAP/★PLﾏｯﾊﾟｰ判定/3ｲﾝﾁHTL13B5/F6 LD/", "Z:/PL-MAP#2/★PLﾏｯﾊﾟｰ判定/HL13B4/F6 LD/", "Z:/PL-MAP#2/★PLﾏｯﾊﾟｰ判定/HL13B5/F6 LD/"]
    Path = ["Z:/PL-MAP/★PLﾏｯﾊﾟｰ判定/3ｲﾝﾁHTL13B2/F6LD/",  "Z:/PL-MAP/★PLﾏｯﾊﾟｰ判定/3ｲﾝﾁHTL13B5/F6 LD/", "Z:/PL-MAP#2/★PLﾏｯﾊﾟｰ判定/HL13B4/F6 LD/", "Z:/PL-MAP#2/★PLﾏｯﾊﾟｰ判定/HL13B5/F6 LD/"]
    for p in Path:
        ALL_FILE_FETCH(p)

    # ----- 取得した全ファイルの処理 -----
    for FilePath in ALL_FILES:

        # ----- ファイルパスからファイル名を取り出し定義 -----
        File = os.path.basename(FilePath)

        # ----- 処理を行ったファイルであれば、処理を行わず次のファイルへ -----
        if FilePath in EndFiles or '~$' in File:
            continue

        Log.Log_Info(Log_File, File)

        # ----- 空欄判定 -----
        if Get_Cells_Info(FilePath):
            Log.Log_Error(Log_File, "Blank Error\n")
            print("path:", FilePath)
            continue
        
        # ----- 処理を行ったファイル名は、EndsFileに加え次から処理を行わないようにする -----
        EndFiles.add(FilePath)
        print(FilePath)

        # ----- データの取得 -----
        data_dict, Data_file_sheet_array = Open_Data_Sheet(FilePath)
        
        # ----- ロット番号エラー時、辞書は空で返ってくる -----
        if data_dict is None:
            Log.Log_Error(Log_File, "Lot Error\n")
            continue

        # ----- Primeにロット番号に対応する品名が入ってなければエラー処理を行う -----
        if data_dict['key_part_number'] == "":
            Log.Log_Error(Log_File, data_dict["key_serial_number"] + ' : ' + "Part Number Error\n")
            continue

        # ----- 作業者が空だった場合、'-'とする -----
        if data_dict['key_operator'] == "":
            data_dict["key_operator"] = "-"

        # ----- ログ書込：日付フォーマットの変換 -----
        Log.Log_Info(Log_File, 'Date Format Conversion')
        print(data_dict['key_start_date_time'])
        data_dict['key_start_date_time'] = Convert_Date.Edit_Date(data_dict['key_start_date_time']).replace('.', ':')
        print(data_dict['key_start_date_time'])
        
        # ----- 日付変換が失敗した(空欄で返ったきた)とき、エラー処理を行う -----
        if data_dict['key_start_date_time'] == "":
            Log.Log_Error(Log_File, data_dict["key_serial_number"] + ' : ' + "Date Error\n")
            continue

        # ----- STARTTIME_SORTEDの追加 -----

        # 日付をExcel時間に変換する
        date = datetime.strptime(str(data_dict["key_start_date_time"]).replace('T', ' ').replace('.', ':'), "%Y-%m-%d %H:%M:%S")
        date_excel_number = int(str(date - datetime(1899, 12, 30)).split()[0])

        # エピ番号の数値部だけを取得する
        Epi_Number = 0
        for i in data_dict["key_batch_number"]:
            try:
                if 0 <= int(i) <= 9:
                    Epi_Number = Epi_Number * 10 + int(i)
            except:
                pass

        # エピ番号を10^6で割って、excel時間に加算する
        date_excel_number += Epi_Number/10**6

        # data_dictに登録する
        data_dict["key_STARTTIME_SORTED"] = date_excel_number
        data_dict["key_SORTNUMBER"] = Epi_Number
        data_dict['key_start_date_time']=date


        # ----- データ型の確認 -----
        result = Check.Data_Type(key_type, data_dict)
        if result == False:
            Log.Log_Error(Log_File, data_dict["key_serial_number"] + ' : ' + "Data Error\n")
            continue

        # ----- XMLファイル名の作成 -----
        xml_file = 'Site=' + Site + ',ProductFamily=' + ProductFamily + ',Operation=' + Operation + \
                   ',Partnumber=' + data_dict["key_part_number"] + ',Serialnumber=' + data_dict["key_serial_number"] + \
                   ',Testdate=' + data_dict["key_start_date_time"].replace(':', '.') + '.xml'

        Output_XML(xml_file, data_dict, Data_file_sheet_array)
        Log.Log_Info(Log_File, data_dict["key_serial_number"] + ' : ' + "OK\n")


    # ----- 処理が完了したファイルをテキストファイルに書き込む -----
    EndFiles_list = sorted(list(EndFiles))
    EndFiles_str = "\n".join(EndFiles_list)
    with open('EndsFile_F6_Format2.txt', 'w', encoding='utf-8') as textfile:
        textfile.write(EndFiles_str)

    # ----- 最終行を書き込んだファイルをGドライブに転送 -----
    #shutil.copy('EndsFile_F6_Format2.txt', 'T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開発/040_LD-EML/F6/13_ProgramUsedFile/')


########## ログ書込：プログラムの終了 ##########
Log.Log_Info(Log_File, 'Program End')