# LD-EML_F6_Format2.py (based on LD-EML_F10_Format2.py) 2025-06-20 yoshida

import openpyxl as px
import pandas as pd
import logging
import shutil
import pyodbc
import xlrd
import glob
import sys
import os

from datetime import datetime, timedelta, date
from time import strftime

########## 自作関数の定義 ##########
sys.path.append('../../MyModule')
import SQL
import Log
import ExpandExp
import Convert_Date
import Row_Number_Func
import Check


########## 全体パラメータの定義 ##########
MOCVD_Equipment = 'F6'

Site = '350'
ProductFamily = 'SAG FAB'
Operation = 'LD-EML_' + MOCVD_Equipment + '_Format2'
TestStation = 'LD-EML'
X = '999999'
Y = '999999'


########## Logフォルダ名の定義 ##########
Log_FolderName = str(date.today())

# ----- 格納するLogフォルダがなければ作成する -----
if not os.path.exists("../../Log/" + Log_FolderName):
    os.makedirs("../../Log/" + Log_FolderName)

# ----- ログ書き込み先ファイルパス -----
Log_File = '../../Log/' + Log_FolderName + '/040_LD-EML_' + MOCVD_Equipment + '_Format2.log'

# ----- ログ書込：プログラムの開始 -----
Log.Log_Info(Log_File, 'Program Start ' + Operation)

# ------ 入力ファイルパス ------
input_filepath = [
    "Z:/PL-MAP/★PLﾏｯﾊﾟｰ判定/3ｲﾝﾁHTL13B5/F6 LD/",
    "Z:/PL-MAP#2/★PLﾏｯﾊﾟｰ判定/HL13B5/F6 LD/",
    "Z:/PL-MAP#3/★PLﾏｯﾊﾟｰ判定/HL13B5/F6LD/",
    "Z:/PL-MAP/★PLﾏｯﾊﾟｰ判定/3ｲﾝﾁHL13B8/F6LD/",
    "Z:/PL-MAP#2/★PLﾏｯﾊﾟｰ判定/HL13B8/F6LD/",
    "Z:/PL-MAP#3/★PLﾏｯﾊﾟｰ判定/HL13B8/F6LD/",
    "Z:/PL-MAP/★PLﾏｯﾊﾟｰ判定/3ｲﾝﾁHL13E1/F6LD/",
    "Z:/PL-MAP#2/★PLﾏｯﾊﾟｰ判定/HL13E1/F6LD/",
    "Z:/PL-MAP#3/★PLﾏｯﾊﾟｰ判定/HL13E1/F6LD/"

]
# input_filepath = ["../../InputFolder/"] # debug

########## XML出力先ファイルパス ##########
Output_filepath = '//li.lumentuminc.net/data/SAG/TDS/Data/Files to Insert/XML/'
# Output_filepath = '../../XML/' # debug


########## シート名定義 ##########
Data_sheet_name = "まとめ"


########## TestStepの定義 ##########
teststep_dict = {
    'TestStep1': 'Coordinate',
    'TestStep2': 'Mapper_Average',
    'TestStep3': 'Mapper_Adjust',
    'TestStep4': 'Center',
    'TestStep5': 'A',
    'TestStep6': 'B',
    'TestStep7': 'C',
    'TestStep8': 'D',
    'TestStep9': 'AB',
    'TestStep10': 'AC',
    'TestStep11': 'BC',
    'TestStep12': 'CD',
    'TestStep13': 'AA',
    'TestStep14': 'BB',
    'TestStep15': 'CC',
    'TestStep16': 'DD',
    'TestStep17': 'SORTED_DATA',
}


########## 取得した項目と型の対応表を定義 ##########
key_type = {
    "key_start_date_time": str,
    "key_part_number": str,
    "key_serial_number": str,
    "key_operator": str,
    "key_batch_number": str,
    "key_Mapper_Average_Mapper_PL_Lambda": float, 
    "key_Mapper_Average_Mapper_PL_Lambda2": float,
    "key_Mapper_Average_Mapper_PL_Intensity": float,
    "key_Mapper_Average_Mapper_PL_FWHM": float,
    "key_Mapper_Adjust_Mapper_TargetWavelength": float,
    "key_Mapper_Adjust_Mapper_Wavelength(1.3um)": float,
    "key_Mapper_Adjust_Mapper_CheckingAdjustValue": float,
    "key_Center_X": float,
    "key_Center_Y": float,
    "key_Center_Lambda": float,
    "key_Center_Intensity": float,
    "key_Center_FWHM": float,
    "key_Center_Tails": float,
    "key_Center_DeltaLambda": float,
    "key_A_X": float,
    "key_A_Y": float,
    "key_A_Lambda": float,
    "key_A_Intensity": float,
    "key_A_FWHM": float,
    "key_A_Tails": float,
    "key_A_DeltaLambda": float,
    "key_B_X": float,
    "key_B_Y": float,
    "key_B_Lambda": float,
    "key_B_Intensity": float,
    "key_B_FWHM": float,
    "key_B_Tails": float,
    "key_B_DeltaLambda": float,
    "key_C_X": float,
    "key_C_Y": float,
    "key_C_Lambda": float,
    "key_C_Intensity": float,
    "key_C_FWHM": float,
    "key_C_Tails": float,
    "key_C_DeltaLambda": float,
    "key_D_X": float,
    "key_D_Y": float,
    "key_D_Lambda": float,
    "key_D_Intensity": float,
    "key_D_FWHM": float,
    "key_D_Tails": float,
    "key_D_DeltaLambda": float,
    "key_AB_X": float,
    "key_AB_Y": float,
    "key_AB_Lambda": float,
    "key_AB_Intensity": float,
    "key_AB_FWHM": float,
    "key_AB_Tails": float,
    "key_AB_DeltaLambda": float,
    "key_AC_X": float,
    "key_AC_Y": float,
    "key_AC_Lambda": float,
    "key_AC_Intensity": float,
    "key_AC_FWHM": float,
    "key_AC_Tails": float,
    "key_AC_DeltaLambda": float,
    "key_BC_X": float,
    "key_BC_Y": float,
    "key_BC_Lambda": float,
    "key_BC_Intensity": float,
    "key_BC_FWHM": float,
    "key_BC_Tails": float,
    "key_BC_DeltaLambda": float,
    "key_CD_X": float,
    "key_CD_Y": float,
    "key_CD_Lambda": float,
    "key_CD_Intensity": float,
    "key_CD_FWHM": float,
    "key_CD_Tails": float,
    "key_CD_DeltaLambda": float,
    "key_AA_X": float,
    "key_AA_Y": float,
    "key_AA_Lambda": float,
    "key_AA_Intensity": float,
    "key_AA_FWHM": float,
    "key_AA_Tails": float,
    "key_AA_DeltaLambda": float,
    "key_BB_X": float,
    "key_BB_Y": float,
    "key_BB_Lambda": float,
    "key_BB_Intensity": float,
    "key_BB_FWHM": float,
    "key_BB_Tails": float,
    "key_BB_DeltaLambda": float,
    "key_CC_X": float,
    "key_CC_Y": float,
    "key_CC_Lambda": float,
    "key_CC_Intensity": float,
    "key_CC_FWHM": float,
    "key_CC_Tails": float,
    "key_CC_DeltaLambda": float,
    "key_DD_X": float,
    "key_DD_Y": float,
    "key_DD_Lambda": float,
    "key_DD_Intensity": float,
    "key_DD_FWHM": float,
    "key_DD_Tails": float,
    "key_DD_DeltaLambda": float,
    "key_Cmb_Date_Time": "datetime",
    "key_Cmb_Spectral_Range": str,
    "key_Cmb_Peak_Position": float,
    "key_Cmb_Peak_Intensity": float,
    "key_Cmb_FWHM(nm)": float,
    "key_Cmb_FWHM(meV)": float,
    "key_Cmb_SideBandHight": float,
    "key_Cmb_X": float,
    "key_Cmb_Y": float,
    "key_STARTTIME_SORTED": float,
    "key_SORTNUMBER" : float,
    "key_LotNumber_9": str
}


########## 対象フォルダの全ファイルを格納 ##########
ALL_FILES = []


########## 取得しないフォルダ名 ##########
#NotUsedDir = {'2017年', '2018年', '2019年', 'ヒメジ検討', '2015', '2016', '2017', '2018', '2019','2020','2021','2022','2023','未使用'}
NotUsedDir = {'2020','2021','2022','2023','未使用'} #2025/01/09 update the path

########## 対象ロット番号のイニシャルを書込したファイルを取得する ##########
Log.Log_Info(Log_File, 'Get SerialNumber Initial List ')
with open('../../SerialNumber_Initial.txt', 'r', encoding='utf-8') as textfile:
    SerialNumber_list = {s.strip() for s in textfile.readlines()}


########## 処理を行ったファイル名をテキストファイルから呼び出す ##########
with open('EndsFile_' + MOCVD_Equipment + '_Format2.txt', 'r', encoding='utf-8') as textfile:
    EndFiles = set(s.strip() for s in textfile.readlines())


# # get index offset for .xls or .xlsm(.xlsx)
# def get_index_offset(file_path):
#     _,ext = os.path.splitext(file_path)
#     if ext == '.xls':
#         return 0
#     elif ext == '.xlsm' or ext == '.xlsx':
#         return 1
#     else:
#         Log.Log_Error(Log_File, file_path + ' : ' + 'unknown extension')
#         return None


########## エピ番号、ロット番号の未記載チェック ##########
def missing_epinum_or_lotnum(file_path):
    Log.Log_Info(Log_File, "Check for missing of Epi No. and Lot No.")
    ext = os.path.splitext(file_path)[1].lower()
    is_cells_empty = False

    try:
        if ext == '.xls':
            wb = xlrd.open_workbook(file_path, on_demand=True)
            sheet = wb.sheet_by_name(Data_sheet_name)
            # xlrd uses 0-based indexing
            if sheet.cell(2, 7).value is None or sheet.cell(3, 7).value is None:
                is_cells_empty = True
            wb.release_resources()
        elif ext in ['.xlsx', '.xlsm']:
            wb = px.load_workbook(file_path, data_only=True)
            sheet = wb[Data_sheet_name]
            # openpyxl uses 1-based indexing
            if sheet.cell(row=3, column=8).value is None or sheet.cell(row=4, column=8).value is None:
                is_cells_empty = True
            wb.close()
        else:
            Log.Log_Error(Log_File, f"Unsupported file type: {file_path}")
            is_cells_empty = True
    except xlrd.biffh.XLRDError as e:
        Log.Log_Error(Log_File, f"xlrd error: {str(e)}")
        is_cells_empty = True
    except Exception as e:
        Log.Log_Error(Log_File, f"Excel read error: {str(e)}")
        is_cells_empty = True

    return is_cells_empty


########## Excelからデータを取得 ##########
def Open_Data_Sheet(file_path):
    Log.Log_Info(Log_File, 'Data Acquisition')

    # by using pd.read_excel(), both xls and xlsm(xlsx) can be supported.
    df = pd.read_excel(file_path, sheet_name=Data_sheet_name, header=None)

    # ----- ロット番号を取得し、Primeにロット番号に対応する品種があるか調べる -----
    serial_number = df.iloc[3, 7]

    # ----- ロット番号が数値の時があるので、数値の時はエラーを出す -----
    if type(serial_number) is not str:
        return None, None

    # ----- Prime接続 -----
    try:
        conn, cursor = SQL.connSQL()
        if conn is None:
            Log.Log_Error(Log_File, serial_number + ' : ' + 'Connection with Prime Failed')
            return None, None
        part_number, Nine_Serial_Number = SQL.selectSQL(cursor, serial_number)
        SQL.disconnSQL(conn, cursor)
    except Exception as e:
        Log.Log_Error(Log_File, serial_number + ' : ' + 'SQL Error: ' + str(e))
        return None, None

    # ----- [データファイル]シートのA5を抜き出す -----
    df_tmp=pd.read_excel(file_path, sheet_name='データファイル', header=None, nrows=5)
    start_date = df_tmp.iloc[4, 0]

    # ----- operatorが空欄であれば、'-'に置き換える -----
    operator = df.iloc[3, 5]
    if pd.isna(operator):
        operator = "-"
    else:
        operator = str(operator)

    # ----- PL-MAP -----
    if 'PL-MAP#2' in str(file_path):
        PL_MAP = 'PL-MAP2'
    elif 'PL-MAP#3' in str(file_path):
        PL_MAP = 'PL-MAP3'
    elif 'PL-MAP#4' in str(file_path):
        PL_MAP = 'PL-MAP4'
    elif 'PL-MAP#6' in str(file_path):
        PL_MAP = 'PL-MAP6'
    elif 'PL-MAP' in str(file_path):
        PL_MAP = 'PL-MAP1'
    else:
        PL_MAP = 'unknown'

    # ----- データの取得 -----
    data_dict = {
        "key_start_date_time": start_date,
        "key_part_number": part_number,
        "key_LotNumber_9": Nine_Serial_Number,
        "key_serial_number": serial_number,
        "key_operator": operator,
        "key_batch_number": df.iloc[2, 7],
        "key_Mapper_Average_Mapper_PL_Lambda": df.iloc[2, 13],
        "key_Mapper_Average_Mapper_PL_Lambda2": (df.iloc[13, 5]+df.iloc[14, 5]+df.iloc[15, 5]+df.iloc[16, 5])/4,
        "key_Mapper_Average_Mapper_PL_Intensity": df.iloc[3, 13],
        "key_Mapper_Average_Mapper_PL_FWHM": df.iloc[4, 13],
        "key_Mapper_Adjust_Mapper_TargetWavelength": df.iloc[26, 4],
        "key_Mapper_Adjust_Mapper_Wavelength(1.3um)": df.iloc[27, 4],
        "key_Mapper_Adjust_Mapper_CheckingAdjustValue": df.iloc[28, 4],
        "key_Center_X": df.iloc[8, 3],
        "key_Center_Y": df.iloc[8, 4],
        "key_Center_Lambda": df.iloc[8, 5],
        "key_Center_Intensity": df.iloc[8, 6],
        "key_Center_FWHM": df.iloc[8, 7],
        "key_Center_Tails": df.iloc[8, 8],
        "key_Center_DeltaLambda": df.iloc[8, 9],
        "key_A_X": df.iloc[9, 3],
        "key_A_Y": df.iloc[9, 4],
        "key_A_Lambda": df.iloc[9, 5],
        "key_A_Intensity": df.iloc[9, 6],
        "key_A_FWHM": df.iloc[9, 7],
        "key_A_Tails": df.iloc[9, 8],
        "key_A_DeltaLambda": df.iloc[9, 9],
        "key_B_X": df.iloc[10, 3],
        "key_B_Y": df.iloc[10, 4],
        "key_B_Lambda": df.iloc[10, 5],
        "key_B_Intensity": df.iloc[10, 6],
        "key_B_FWHM": df.iloc[10, 7],
        "key_B_Tails": df.iloc[10, 8],
        "key_B_DeltaLambda": df.iloc[10, 9],
        "key_C_X": df.iloc[11, 3],
        "key_C_Y": df.iloc[11, 4],
        "key_C_Lambda": df.iloc[11, 5],
        "key_C_Intensity": df.iloc[11, 6],
        "key_C_FWHM": df.iloc[11, 7],
        "key_C_Tails": df.iloc[11, 8],
        "key_C_DeltaLambda": df.iloc[11, 9],
        "key_D_X": df.iloc[12, 3],
        "key_D_Y": df.iloc[12, 4],
        "key_D_Lambda": df.iloc[12, 5],
        "key_D_Intensity": df.iloc[12, 6],
        "key_D_FWHM": df.iloc[12, 7],
        "key_D_Tails": df.iloc[12, 8],
        "key_D_DeltaLambda": df.iloc[12, 9],
        "key_AB_X": df.iloc[13, 3],
        "key_AB_Y": df.iloc[13, 4],
        "key_AB_Lambda": df.iloc[13, 5],
        "key_AB_Intensity": df.iloc[13, 6],
        "key_AB_FWHM": df.iloc[13, 7],
        "key_AB_Tails": df.iloc[13, 8],
        "key_AB_DeltaLambda": df.iloc[13, 9],
        "key_AC_X": df.iloc[14, 3],
        "key_AC_Y": df.iloc[14, 4],
        "key_AC_Lambda": df.iloc[14, 5],
        "key_AC_Intensity": df.iloc[14, 6],
        "key_AC_FWHM": df.iloc[14, 7],
        "key_AC_Tails": df.iloc[14, 8],
        "key_AC_DeltaLambda": df.iloc[14, 9],
        "key_BC_X": df.iloc[15, 3],
        "key_BC_Y": df.iloc[15, 4],
        "key_BC_Lambda": df.iloc[15, 5],
        "key_BC_Intensity": df.iloc[15, 6],
        "key_BC_FWHM": df.iloc[15, 7],
        "key_BC_Tails": df.iloc[15, 8],
        "key_BC_DeltaLambda": df.iloc[15, 9],
        "key_CD_X": df.iloc[16, 3],
        "key_CD_Y": df.iloc[16, 4],
        "key_CD_Lambda": df.iloc[16, 5],
        "key_CD_Intensity": df.iloc[16, 6],
        "key_CD_FWHM": df.iloc[16, 7],
        "key_CD_Tails": df.iloc[16, 8],
        "key_CD_DeltaLambda": df.iloc[16, 9],
        "key_AA_X": df.iloc[17, 3],
        "key_AA_Y": df.iloc[17, 4],
        "key_AA_Lambda": df.iloc[17, 5],
        "key_AA_Intensity": df.iloc[17, 6],
        "key_AA_FWHM": df.iloc[17, 7],
        "key_AA_Tails": df.iloc[17, 8],
        "key_AA_DeltaLambda": df.iloc[17, 9],
        "key_BB_X": df.iloc[18, 3],
        "key_BB_Y": df.iloc[18, 4],
        "key_BB_Lambda": df.iloc[18, 5],
        "key_BB_Intensity": df.iloc[18, 6],
        "key_BB_FWHM": df.iloc[18, 7],
        "key_BB_Tails": df.iloc[18, 8],
        "key_BB_DeltaLambda": df.iloc[18, 9],
        "key_CC_X": df.iloc[19, 3],
        "key_CC_Y": df.iloc[19, 4],
        "key_CC_Lambda": df.iloc[19, 5],
        "key_CC_Intensity": df.iloc[19, 6],
        "key_CC_FWHM": df.iloc[19, 7],
        "key_CC_Tails": df.iloc[19, 8],
        "key_CC_DeltaLambda": df.iloc[19, 9],
        "key_DD_X": df.iloc[20, 3],
        "key_DD_Y": df.iloc[20, 4],
        "key_DD_Lambda": df.iloc[20, 5],
        "key_DD_Intensity": df.iloc[20, 6],
        "key_DD_FWHM": df.iloc[20, 7],
        "key_DD_Tails": df.iloc[20, 8],
        "key_DD_DeltaLambda": df.iloc[20, 9],
        "key_PL_MAP": PL_MAP
    }
    
    # Temporary variable (ex:"F9" means cell F9 value)
    F9=df.iloc[8, 5]
    F14=df.iloc[13, 5]
    F15=df.iloc[14, 5]
    F16=df.iloc[15, 5]
    F17=df.iloc[16, 5]

    # ----- 空欄箇所はNoneとして取得される。Noneは文字列に変換できないため、空欄("")に置き換える -----
    for key in data_dict:
        if data_dict[key] is None or data_dict[key] == "#DIV/0!" or (data_dict[key] == '-' and key != "key_operator"):
            data_dict[key] = ""
        # ----- 指数表記を展開する -----
        # I can't understand...
        if type(data_dict[key]) is float and 'e' in str(data_dict[key]) and key != "key_start_date_time":
            data_dict[key] = ExpandExp.Expand(data_dict[key])

    # ----- データファイルシートからデータを取得する -----
    # above 'データファイルシート' (data file sheet) is the sheet which contains data of individual chips rather than whole wafer .
    Data_file_sheet_array = []
    df = pd.read_excel(file_path, header=None, sheet_name='データファイル', usecols="A:I", skiprows=1)
    row, col = df.shape
    for i in range(row):
        tmp = []
        for j in range(col):
            tmp.append(str(df.iloc[i, j]))
        Data_file_sheet_array.append(tmp)

    # ----- 1行目のデータだけ、データ型確認用として辞書に追加する -----
    # I don't know what 'Cmb' means.
    data_dict['key_Cmb_Date_Time'] = Data_file_sheet_array[0][0]
    data_dict['key_Cmb_Spectral_Range'] = Data_file_sheet_array[0][1]
    data_dict['key_Cmb_Peak_Position'] = Data_file_sheet_array[0][2]
    data_dict['key_Cmb_Peak_Intensity'] = Data_file_sheet_array[0][3]
    data_dict['key_Cmb_FWHM(nm)'] = Data_file_sheet_array[0][4]
    data_dict['key_Cmb_FWHM(meV)'] = Data_file_sheet_array[0][5]
    data_dict['key_Cmb_SideBandHight'] = Data_file_sheet_array[0][6]
    data_dict['key_Cmb_X'] = Data_file_sheet_array[0][7]
    data_dict['key_Cmb_Y'] = Data_file_sheet_array[0][8]
#    data_dict['Lambda_diff']= data_dict["key_Mapper_Average_Mapper_PL_Lambda"]-data_dict["key_Mapper_Adjust_Mapper_Wavelength(1.3um)"] #2025/01/15 add
    data_dict['Lambda_diff']= data_dict["key_Mapper_Average_Mapper_PL_Lambda"]-data_dict["key_Mapper_Adjust_Mapper_TargetWavelength"] #2025/02/07 modify E28 cell → E27 cell yoshida
#    data_dict['Lambda2_diff'] =  data_dict["key_Mapper_Average_Mapper_PL_Lambda2"]-data_dict["key_Center_Lambda"] #2025/01/15 add
    data_dict['Lambda2_diff'] =  max(abs(F14-F9),abs(F15-F9),abs(F16-F9),abs(F17-F9)) #2025/02/07 modify Lambda2_diff=max(|F14-F9|,,,,|F17-F9|) yoshida

    return data_dict, Data_file_sheet_array

########## XML変換 ##########
def Output_XML(xml_file, data_dict, Data_file_sheet_array):

    # ----- ログ書込：XML変換 -----
    Log.Log_Info(Log_File, 'Excel File To XML File Conversion')
    
    XML =   '<?xml version="1.0" encoding="utf-8"?>' + '\n' + \
            '<Results xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema">' + '\n' + \
            '       <Result startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Result="Passed">' + '\n' + \
            '               <Header SerialNumber=' + '"' + data_dict["key_serial_number"] + '"' + ' PartNumber=' + '"' + data_dict["key_part_number"] + '"' + ' Operation=' + '"' + Operation + '"' + ' TestStation=' + '"' + TestStation + '"' + ' Operator=' + '"' + data_dict["key_operator"] + '"' + ' StartTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Site=' + '"' + Site + '"' + ' BatchNumber=' + '"' + data_dict["key_batch_number"] + '"' + ' LotNumber=' + '"' + data_dict["key_serial_number"] + '"/>' + '\n' + \
            '\n' + \
            '               <TestStep Name=' + '"' + teststep_dict["TestStep1"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
            '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + X + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + Y + '"/>' + '\n' + \
            '               </TestStep>' + '\n' + \
            '               <TestStep Name=' + '"' + teststep_dict["TestStep2"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
            '                   <Data DataType="Numeric" Name="Mapper_PL_Lambda" Units="nm" Value=' + '"' + str(data_dict["key_Mapper_Average_Mapper_PL_Lambda"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Mapper_PL_Intensity" Units="mV" Value=' + '"' + str(data_dict["key_Mapper_Average_Mapper_PL_Intensity"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Mapper_PL_FWHM" Units="meV" Value=' + '"' + str(data_dict["key_Mapper_Average_Mapper_PL_FWHM"]) + '"/>' + '\n' + \
            '               </TestStep>' + '\n' + \
            '               <TestStep Name=' + '"' + teststep_dict["TestStep3"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
            '                   <Data DataType="Numeric" Name="Mapper_TargetWavelength" Units="nm" Value=' + '"' + str(data_dict["key_Mapper_Adjust_Mapper_TargetWavelength"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Mapper_Wavelength(1.3um)" Units="nm" Value=' + '"' + str(data_dict["key_Mapper_Adjust_Mapper_Wavelength(1.3um)"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Mapper_CheckingAdjustValue" Units="nm" Value=' + '"' + str(data_dict["key_Mapper_Adjust_Mapper_CheckingAdjustValue"]) + '"/>' + '\n' + \
            '               </TestStep>' + '\n' + \
            '               <TestStep Name=' + '"' + teststep_dict["TestStep4"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
            '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_Center_X"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_Center_Y"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_Center_Lambda"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Lambda_Diff" Units="nm" Value=' + '"' + str(data_dict["Lambda_diff"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Lambda2_Diff" Units="nm" Value=' + '"' + str(data_dict['Lambda2_diff']) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_Center_Intensity"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict['key_Center_FWHM']) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_Center_Tails"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_Center_DeltaLambda"]) + '"/>' + '\n' + \
            '               </TestStep>' + '\n' + \
            '               <TestStep Name=' + '"' + teststep_dict["TestStep5"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
            '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_A_X"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_A_Y"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_A_Lambda"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_A_Intensity"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_A_FWHM"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_A_Tails"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_A_DeltaLambda"]) + '"/>' + '\n' + \
            '               </TestStep>' + '\n' + \
            '               <TestStep Name=' + '"' + teststep_dict["TestStep6"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
            '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_B_X"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_B_Y"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_B_Lambda"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_B_Intensity"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_B_FWHM"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_B_Tails"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_B_DeltaLambda"]) + '"/>' + '\n' + \
            '               </TestStep>' + '\n' + \
            '               <TestStep Name=' + '"' + teststep_dict["TestStep7"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
            '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_C_X"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_C_Y"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_C_Lambda"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_C_Intensity"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_C_FWHM"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_C_Tails"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_C_DeltaLambda"]) + '"/>' + '\n' + \
            '               </TestStep>' + '\n' + \
            '               <TestStep Name=' + '"' + teststep_dict["TestStep8"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
            '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_D_X"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_D_Y"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_D_Lambda"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_D_Intensity"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_D_FWHM"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_D_Tails"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_D_DeltaLambda"]) + '"/>' + '\n' + \
            '               </TestStep>' + '\n' + \
            '               <TestStep Name=' + '"' + teststep_dict["TestStep9"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
            '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_AB_X"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_AB_Y"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_AB_Lambda"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_AB_Intensity"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_AB_FWHM"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_AB_Tails"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_AB_DeltaLambda"]) + '"/>' + '\n' + \
            '               </TestStep>' + '\n' + \
            '               <TestStep Name=' + '"' + teststep_dict["TestStep10"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
            '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_AC_X"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_AC_Y"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_AC_Lambda"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_AC_Intensity"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_AC_FWHM"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_AC_Tails"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_AC_DeltaLambda"]) + '"/>' + '\n' + \
            '               </TestStep>' + '\n' + \
            '               <TestStep Name=' + '"' + teststep_dict["TestStep11"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
            '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_BC_X"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_BC_Y"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_BC_Lambda"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_BC_Intensity"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_BC_FWHM"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_BC_Tails"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_BC_DeltaLambda"]) + '"/>' + '\n' + \
            '               </TestStep>' + '\n' + \
            '               <TestStep Name=' + '"' + teststep_dict["TestStep12"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
            '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_CD_X"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_CD_Y"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_CD_Lambda"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_CD_Intensity"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_CD_FWHM"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_CD_Tails"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_CD_DeltaLambda"]) + '"/>' + '\n' + \
            '               </TestStep>' + '\n' + \
            '               <TestStep Name=' + '"' + teststep_dict["TestStep13"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
            '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_AA_X"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_AA_Y"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_AA_Lambda"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_AA_Intensity"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_AA_FWHM"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_AA_Tails"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_AA_DeltaLambda"]) + '"/>' + '\n' + \
            '               </TestStep>' + '\n' + \
            '               <TestStep Name=' + '"' + teststep_dict["TestStep14"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
            '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_BB_X"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_BB_Y"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_BB_Lambda"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_BB_Intensity"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_BB_FWHM"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_BB_Tails"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_BB_DeltaLambda"]) + '"/>' + '\n' + \
            '               </TestStep>' + '\n' + \
            '               <TestStep Name=' + '"' + teststep_dict["TestStep15"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
            '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_CC_X"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_CC_Y"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_CC_Lambda"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_CC_Intensity"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_CC_FWHM"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_CC_Tails"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_CC_DeltaLambda"]) + '"/>' + '\n' + \
            '               </TestStep>' + '\n' + \
            '               <TestStep Name=' + '"' + teststep_dict["TestStep16"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
            '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + str(data_dict["key_DD_X"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + str(data_dict["key_DD_Y"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value=' + '"' + str(data_dict["key_DD_Lambda"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value=' + '"' + str(data_dict["key_DD_Intensity"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value=' + '"' + str(data_dict["key_DD_FWHM"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value=' + '"' + str(data_dict["key_DD_Tails"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value=' + '"' + str(data_dict["key_DD_DeltaLambda"]) + '"/>' + '\n' + \
            '               </TestStep>' + '\n' + \
            '               <TestStep Name=' + '"' + teststep_dict["TestStep17"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
            '                   <Data DataType="Numeric" Name="STARTTIME_SORTED" Units="" Value=' + '"' + str(data_dict["key_STARTTIME_SORTED"]) + '"/>' + '\n' + \
            '                   <Data DataType="Numeric" Name="SORTNUMBER" Units="" Value=' + '"' + str(data_dict["key_SORTNUMBER"]) + '"/>' + '\n' + \
            '                   <Data DataType="String" Name="LotNumber_5" Value=' + '"' + str(data_dict["key_serial_number"]) + '"' + ' CompOperation="LOG"/>' + '\n' + \
            '                   <Data DataType="String" Name="LotNumber_9" Value=' + '"' + str(data_dict["key_LotNumber_9"]) + '"' + ' CompOperation="LOG"/>' + '\n' + \
            '               </TestStep>' + '\n'

    # Cmbの付与
    # what's Cmb ...
    for i in range(len(Data_file_sheet_array)):
        TestStep = 'Cmb' + str(i + 1)
        XML += '               <TestStep Name=' + '"' + TestStep + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' + \
               '                   <Data DataType="String" Name="Cmb_Date_Time" Value=' + '"' + str(Data_file_sheet_array[i][0]) + '"' + ' CompOperation="LOG"/>' + '\n' + \
               '                   <Data DataType="String" Name="Cmb_Spectral_Range" Value=' + '"' + str(Data_file_sheet_array[i][1]) + '"' + ' CompOperation="LOG"/>' + '\n' + \
               '                   <Data DataType="Numeric" Name="Cmb_Peak_Position" Units="nm" Value=' + '"' + str(Data_file_sheet_array[i][2]) + '"/>' + '\n' + \
               '                   <Data DataType="Numeric" Name="Cmb_Peak_Intensity" Units="counts" Value=' + '"' + str(Data_file_sheet_array[i][3]) + '"/>' + '\n' + \
               '                   <Data DataType="Numeric" Name="Cmb_FWHM(nm)" Units="nm" Value=' + '"' + str(Data_file_sheet_array[i][4]) + '"/>' + '\n' + \
               '                   <Data DataType="Numeric" Name="Cmb_FWHM(meV)" Units="meV" Value=' + '"' + str(Data_file_sheet_array[i][5]) + '"/>' + '\n' + \
               '                   <Data DataType="Numeric" Name="Cmb_SideBandHight" Units="counts" Value=' + '"' + str(Data_file_sheet_array[i][6]) + '"/>' + '\n' + \
               '                   <Data DataType="Numeric" Name="Cmb_X" Units="um" Value=' + '"' + str(Data_file_sheet_array[i][7]) + '"/>' + '\n' + \
               '                   <Data DataType="Numeric" Name="Cmb_Y" Units="um" Value=' + '"' + str(Data_file_sheet_array[i][8]) + '"/>' + '\n' + \
               '               </TestStep>' + '\n'
    
    XML +=  '\n' \
            '               <TestEquipment>' + '\n' + \
            '                   <Item DeviceName="PLmapper" DeviceSerialNumber="' + "1" + '"/>' + '\n' + \
            '                   <Item DeviceName="MOCVD" DeviceSerialNumber="' + MOCVD_Equipment + '"/>' + '\n' + \
            '               </TestEquipment>' + '\n' + \
            '\n' \
            '               <ErrorData/>' + '\n' + \
            '               <FailureData/>' + '\n' + \
            '               <Configuration/>' + '\n' + \
            '       </Result>' + '\n' + \
            '</Results>' \

    
        
    f = open(Output_filepath + xml_file, 'w', encoding="utf-8")
            
    f.write(XML)

    f.close()


########## 指定パス内のすべてのファイルを取得 ##########
def ALL_FILE_FETCH(input_filepath):
    global ALL_FILES

    for path, dir, file in os.walk(input_filepath):
        # ----- dir の再帰でPathを生成しているため、省くフォルダについてはdirを書き換える必要がある -----
        dir[:] = [d for d in dir if d not in NotUsedDir]
        for f in file:
            if f.endswith('.xlsx') or f.endswith('.xlsm') or f.endswith('.xls'):
                ALL_FILES.append(os.path.join(path, f))



########## main処理 ##########
if __name__ == '__main__':

    # ----- ログ書込：Main処理の開始 -----
    Log.Log_Info(Log_File, 'Main Start')

    for p in input_filepath:
        ALL_FILE_FETCH(p)

    # ----- 取得した全ファイルの処理 -----
    for FilePath in ALL_FILES:

        # ----- ファイルパスからファイル名を取り出し定義 -----
        File = os.path.basename(FilePath)

        # ----- 処理を行ったファイルであれば、処理を行わず次のファイルへ -----
        if FilePath in EndFiles or '~$' in File:
            continue

        Log.Log_Info(Log_File, File)

        # ----- 空欄チェック -----
        if missing_epinum_or_lotnum(FilePath):
            Log.Log_Error(Log_File, "Missing EpiNum or LotNum\n")
            continue

        # ----- 処理を行ったファイル名は、EndsFileに加え次から処理を行わないようにする -----
        EndFiles.add(FilePath)

        # ----- データの取得 -----
        data_dict, Data_file_sheet_array = Open_Data_Sheet(FilePath)

        # ----- 辞書がNone → ロット番号にstring以外が入っていてreturnされた -----
        if data_dict is None:
            Log.Log_Error(Log_File, "Lot Error\n")
            continue

        # ----- Primeにロット番号に対応する品名が入ってなければエラー処理を行う -----
        if data_dict['key_part_number'] == "":
            Log.Log_Error(Log_File, data_dict["key_serial_number"] + ' : ' + "Part Number Error\n")
            continue

        # ----- ログ書込：日付フォーマットの変換 -----
        Log.Log_Info(Log_File, 'Date Format Conversion')
        data_dict['key_start_date_time'] = Convert_Date.Edit_Date(data_dict['key_start_date_time']).replace('.', ':')

        # ----- 日付変換が失敗した(空欄で返ったきた)とき、エラー処理を行う -----
        if data_dict['key_start_date_time'] == "":
            Log.Log_Error(Log_File, data_dict["key_serial_number"] + ' : ' + "Date Error\n")
            continue

        # ----- STARTTIME_SORTEDの追加 -----

        # 日付をExcel時間に変換する
        date = datetime.strptime(str(data_dict["key_start_date_time"]).replace('T', ' ').replace('.', ':'), "%Y-%m-%d %H:%M:%S")
        date_excel_number = int(str(date - datetime(1899, 12, 30)).split()[0])

        # エピ番号の数値部だけを取得する
        Epi_Number = 0
        for i in data_dict["key_batch_number"]:
            try:
                if 0<=int(i)<=9:
                    Epi_Number = Epi_Number*10+int(i)
            except:
                pass

        # エピ番号を10^6で割って、excel時間に加算する
        date_excel_number += Epi_Number/10**6

        # data_dictに登録する
        data_dict["key_STARTTIME_SORTED"] = date_excel_number
        data_dict["key_SORTNUMBER"] = Epi_Number

        # ----- データ型の確認 -----
        result = Check.Data_Type(key_type, data_dict)
        if result == False:
            Log.Log_Error(Log_File, data_dict["key_serial_number"] + ' : ' + "Data Error\n")
            continue

        # ----- XMLファイルの作成 -----
        xml_file = 'Site=' + Site + ',ProductFamily=' + ProductFamily + ',Operation=' + Operation + \
                   ',Partnumber=' + data_dict["key_part_number"] + ',Serialnumber=' + data_dict["key_serial_number"] + \
                   ',Testdate=' + data_dict["key_start_date_time"].replace(':', '.') + '.xml'

        Output_XML(xml_file, data_dict, Data_file_sheet_array)
        Log.Log_Info(Log_File, data_dict["key_serial_number"] + ' : ' + "OK\n")


    # ----- 処理が完了したファイルをテキストファイルに書き込む -----
    EndFiles_list = sorted(list(EndFiles))
    EndFiles_str = "\n".join(EndFiles_list)
    with open('EndsFile_' + MOCVD_Equipment + '_Format2.txt', 'w', encoding='utf-8') as textfile:
        textfile.write(EndFiles_str)



########## ログ書込：プログラムの終了 ##########
Log.Log_Info(Log_File, 'Program End')