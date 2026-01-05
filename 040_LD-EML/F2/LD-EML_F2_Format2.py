import openpyxl as px
import pandas as pd
import logging
import shutil
import pyodbc
import xlrd
import glob
import sys
import os
from datetime import datetime, timedelta, date
from time import strftime

# 自作関数の定義
sys.path.append('../../MyModule')
import SQL
import Log
import ExpandExp
import Convert_Date
import Row_Number_Func
import Check

# 全體參數的定義
Site = '350'
ProductFamily = 'SAG FAB'
Operation = 'LD-EML_F2_Format2'
TestStation = 'LD-EML'
X = '999999'
Y = '999999'

# Log資料夾名稱定義
Log_FolderName = str(date.today())
if not os.path.exists("../../Log/" + Log_FolderName):
    os.makedirs("../../Log/" + Log_FolderName)
Log_File = '../../Log/' + Log_FolderName + '/040_LD-EML_F2_Format2.log'
Log.Log_Info(Log_File, 'Program Start')

# XML輸出檔案路徑 (for test)
#Output_filepath = 'C:/Users/hsi67063/Documents/TEMP/'
Output_filepath = '//li.lumentuminc.net/data/SAG/TDS/Data/Files to Insert/XML/'

# 工作表名稱定義
Data_sheet_name = "まとめ"

# TestStep 的定義
teststep_dict = {
    'TestStep1': 'Coordinate',
    'TestStep2': 'Mapper_Average',
    'TestStep3': 'Mapper_Adjust',
    'TestStep4': 'Center',
    'TestStep5': 'A',
    'TestStep6': 'B',
    'TestStep7': 'C',
    'TestStep8': 'D',
    'TestStep9': 'AB',
    'TestStep10': 'AC',
    'TestStep11': 'BC',
    'TestStep12': 'CD',
    'TestStep13': 'AA',
    'TestStep14': 'BB',
    'TestStep15': 'CC',
    'TestStep16': 'DD',
    'TestStep17': 'SORTED_DATA',
}

# 取得項目與型態的對應表
key_type = {
    "key_start_date_time": str,
    "key_part_number": str,
    "key_serial_number": str,
    "key_operator": str,
    "key_batch_number": str,
    "key_Mapper_Average_Mapper_PL_Lambda": float,
    "key_Mapper_Average_Mapper_PL_Lambda2": float,
    "key_Mapper_Average_Mapper_PL_Intensity": float,
    "key_Mapper_Average_Mapper_PL_FWHM": float,
    "key_Mapper_Adjust_Mapper_TargetWavelength": float,
    "key_Mapper_Adjust_Mapper_Wavelength(1.3um)": float,
    "key_Mapper_Adjust_Mapper_CheckingAdjustValue": float,
    "key_Thickness_Thickness_Core" : float,
    "key_Center_X": float,
    "key_Center_Y": float,
    "key_Center_Lambda": float,
    "key_Center_Intensity": float,
    "key_Center_FWHM": float,
    "key_Center_Tails": float,
    "key_Center_DeltaLambda": float,
    "key_A_X": float,
    "key_A_Y": float,
    "key_A_Lambda": float,
    "key_A_Intensity": float,
    "key_A_FWHM": float,
    "key_A_Tails": float,
    "key_A_DeltaLambda": float,
    "key_B_X": float,
    "key_B_Y": float,
    "key_B_Lambda": float,
    "key_B_Intensity": float,
    "key_B_FWHM": float,
    "key_B_Tails": float,
    "key_B_DeltaLambda": float,
    "key_C_X": float,
    "key_C_Y": float,
    "key_C_Lambda": float,
    "key_C_Intensity": float,
    "key_C_FWHM": float,
    "key_C_Tails": float,
    "key_C_DeltaLambda": float,
    "key_D_X": float,
    "key_D_Y": float,
    "key_D_Lambda": float,
    "key_D_Intensity": float,
    "key_D_FWHM": float,
    "key_D_Tails": float,
    "key_D_DeltaLambda": float,
    "key_AB_X": float,
    "key_AB_Y": float,
    "key_AB_Lambda": float,
    "key_AB_Intensity": float,
    "key_AB_FWHM": float,
    "key_AB_Tails": float,
    "key_AB_DeltaLambda": float,
    "key_AC_X": float,
    "key_AC_Y": float,
    "key_AC_Lambda": float,
    "key_AC_Intensity": float,
    "key_AC_FWHM": float,
    "key_AC_Tails": float,
    "key_AC_DeltaLambda": float,
    "key_BC_X": float,
    "key_BC_Y": float,
    "key_BC_Lambda": float,
    "key_BC_Intensity": float,
    "key_BC_FWHM": float,
    "key_BC_Tails": float,
    "key_BC_DeltaLambda": float,
    "key_CD_X": float,
    "key_CD_Y": float,
    "key_CD_Lambda": float,
    "key_CD_Intensity": float,
    "key_CD_FWHM": float,
    "key_CD_Tails": float,
    "key_CD_DeltaLambda": float,
    "key_AA_X": float,
    "key_AA_Y": float,
    "key_AA_Lambda": float,
    "key_AA_Intensity": float,
    "key_AA_FWHM": float,
    "key_AA_Tails": float,
    "key_AA_DeltaLambda": float,
    "key_BB_X": float,
    "key_BB_Y": float,
    "key_BB_Lambda": float,
    "key_BB_Intensity": float,
    "key_BB_FWHM": float,
    "key_BB_Tails": float,
    "key_BB_DeltaLambda": float,
    "key_CC_X": float,
    "key_CC_Y": float,
    "key_CC_Lambda": float,
    "key_CC_Intensity": float,
    "key_CC_FWHM": float,
    "key_CC_Tails": float,
    "key_CC_DeltaLambda": float,
    "key_DD_X": float,
    "key_DD_Y": float,
    "key_DD_Lambda": float,
    "key_DD_Intensity": float,
    "key_DD_FWHM": float,
    "key_DD_Tails": float,
    "key_DD_DeltaLambda": float,
    "key_Cmb_Date_Time": "datetime",
    "key_Cmb_Spectral_Range": str,
    "key_Cmb_Peak_Position": float,
    "key_Cmb_Peak_Intensity": float,
    "key_Cmb_FWHM(nm)": float,
    "key_Cmb_FWHM(meV)": float,
    "key_Cmb_SideBandHight": float,
    "key_Cmb_X": float,
    "key_Cmb_Y": float,
    "key_STARTTIME_SORTED": float,
    "key_SORTNUMBER": float,
    "key_LotNumber_9": str
}

ALL_FILES = []
NotUsedDir = {'2017年', '2018年', '2019年', '2020年', '2021年', '2022年', '2023年', 'ヒメジ検討', '2015', '2016', '2017', '2018', '2019', '未使用', 'old'}

Log.Log_Info(Log_File, 'Get SerialNumber Initial List ')
try:
    # with open('T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開発/004_T2-EML/SerialNumber_Initial.txt', 'r', encoding='utf-8') as textfile:
    with open('../../SerialNumber_Initial.txt', 'r', encoding='utf-8') as textfile:
        SerialNumber_list = {s.strip() for s in textfile.readlines()}
except:
    with open('C:/Users/hsi67063/Downloads/SerialNumber_Initial.txt', 'r', encoding='utf-8') as textfile:    
        SerialNumber_list = {s.strip() for s in textfile.readlines()}

with open('EndsFile_F2_Format2.txt', 'r', encoding='utf-8') as textfile:
    EndFiles = set(s.strip() for s in textfile.readlines())

# 修改後的空欄檢查函數
def Get_Cells_Info(file_path):
    Log.Log_Info(Log_File, "Blank Check")
    ext = os.path.splitext(file_path)[1].lower()
    is_cells_empty = False

    try:
        if ext == '.xls':
            wb = xlrd.open_workbook(file_path, on_demand=True)
            sheet = wb.sheet_by_name(Data_sheet_name)
            if sheet.cell(2, 7).value is None or sheet.cell(3, 7).value is None:
                is_cells_empty = True
            wb.release_resources()
        elif ext in ['.xlsx', '.xlsm']:
            wb = px.load_workbook(file_path)
            sheet = wb[Data_sheet_name]
            if sheet.cell(2, 7).value is None or sheet.cell(3, 7).value is None:
                is_cells_empty = True
            wb.close()
        else:
            Log.Log_Error(Log_File, f"Unsupported file type: {file_path}")
            is_cells_empty = True
    except xlrd.biffh.XLRDError as e:
        Log.Log_Error(Log_File, f"xlrd error: {str(e)}")
        is_cells_empty = True
    except Exception as e:
        Log.Log_Error(Log_File, f"openpyxl error: {str(e)}")
        is_cells_empty = True

    return is_cells_empty

# 修改後的資料讀取函數
def Open_Data_Sheet(file_path):
    Log.Log_Info(Log_File, 'Data Acquisition')
    data_dict = dict()
    if file_path.lower().endswith('.xls'):
        wb = xlrd.open_workbook(file_path, on_demand=True)
        sheet = wb.sheet_by_name(Data_sheet_name)
        date_sheet = wb.sheet_by_name('データファイル')
        serial_number = sheet.cell(3, 7).value  # 第4列第8欄
        if type(serial_number) is not str:
            wb.release_resources()
            return None, None
        try:
            conn, cursor = SQL.connSQL()
            if conn is None:
                Log.Log_Error(Log_File, serial_number + ' : ' + 'Connection with Prime Failed')
                wb.release_resources()
                return None, None
            part_number, Nine_Serial_Number = SQL.selectSQL(cursor, serial_number)
            SQL.disconnSQL(conn, cursor)
        except Exception as e:
            Log.Log_Error(Log_File, serial_number + ' : ' + 'SQL Error: ' + str(e))
            wb.release_resources()
            return None, None
        start_date = date_sheet.cell(4, 0).value
        Operator = str(sheet.cell(3, 5).value)
        if Operator == "None" or Operator == "":
            Operator = "-"
        PL_MAP = '1'
        if '#2' in str(file_path):
            PL_MAP = '2'
        if '#3' in str(file_path):
            PL_MAP = '3' 
        if '#4' in str(file_path):
            PL_MAP = '4'  
        data_dict = {
            "key_start_date_time": start_date,
            "key_part_number": part_number,
            "key_serial_number": serial_number,
            "key_operator": Operator,
            "key_LotNumber_9": Nine_Serial_Number,
            "key_batch_number": sheet.cell(2, 7).value,
            "key_Mapper_Average_Mapper_PL_Lambda": sheet.cell(2, 13).value,
            "key_Mapper_Average_Mapper_PL_Lambda2": (sheet.cell(13, 5).value + sheet.cell(14, 5).value + sheet.cell(15, 5).value + sheet.cell(16, 5).value) / 4,
            "key_Mapper_Average_Mapper_PL_Intensity": sheet.cell(3, 13).value,
            "key_Mapper_Average_Mapper_PL_FWHM": sheet.cell(4, 13).value,
            "key_Mapper_Adjust_Mapper_TargetWavelength": sheet.cell(26, 4).value,
            "key_Mapper_Adjust_Mapper_Wavelength(1.3um)": sheet.cell(27, 4).value,
            "key_Mapper_Adjust_Mapper_CheckingAdjustValue": sheet.cell(28, 4).value,
            "key_Thickness_Thickness_Core": sheet.cell(76, 13).value,            
            "key_Center_X": sheet.cell(8, 3).value,
            "key_Center_Y": sheet.cell(8, 4).value,
            "key_Center_Lambda": sheet.cell(8, 5).value,
            "key_Center_Intensity": sheet.cell(8, 6).value,
            "key_Center_FWHM": sheet.cell(8, 7).value,
            "key_Center_Tails": sheet.cell(8, 8).value,
            "key_Center_DeltaLambda": sheet.cell(8, 9).value,
            "key_A_X": sheet.cell(9, 3).value,
            "key_A_Y": sheet.cell(9, 4).value,
            "key_A_Lambda": sheet.cell(9, 5).value,
            "key_A_Intensity": sheet.cell(9, 6).value,
            "key_A_FWHM": sheet.cell(9, 7).value,
            "key_A_Tails": sheet.cell(9, 8).value,
            "key_A_DeltaLambda": sheet.cell(9, 9).value,
            "key_B_X": sheet.cell(10, 3).value,
            "key_B_Y": sheet.cell(10, 4).value,
            "key_B_Lambda": sheet.cell(10, 5).value,
            "key_B_Intensity": sheet.cell(10, 6).value,
            "key_B_FWHM": sheet.cell(10, 7).value,
            "key_B_Tails": sheet.cell(10, 8).value,
            "key_B_DeltaLambda": sheet.cell(10, 9).value,
            "key_C_X": sheet.cell(11, 3).value,
            "key_C_Y": sheet.cell(11, 4).value,
            "key_C_Lambda": sheet.cell(11, 5).value,
            "key_C_Intensity": sheet.cell(11, 6).value,
            "key_C_FWHM": sheet.cell(11, 7).value,
            "key_C_Tails": sheet.cell(11, 8).value,
            "key_C_DeltaLambda": sheet.cell(11, 9).value,
            "key_D_X": sheet.cell(12, 3).value,
            "key_D_Y": sheet.cell(12, 4).value,
            "key_D_Lambda": sheet.cell(12, 5).value,
            "key_D_Intensity": sheet.cell(12, 6).value,
            "key_D_FWHM": sheet.cell(12, 7).value,
            "key_D_Tails": sheet.cell(12, 8).value,
            "key_D_DeltaLambda": sheet.cell(12, 9).value,
            "key_AB_X": sheet.cell(13, 3).value,
            "key_AB_Y": sheet.cell(13, 4).value,
            "key_AB_Lambda": sheet.cell(13, 5).value,
            "key_AB_Intensity": sheet.cell(13, 6).value,
            "key_AB_FWHM": sheet.cell(13, 7).value,
            "key_AB_Tails": sheet.cell(13, 8).value,
            "key_AB_DeltaLambda": sheet.cell(13, 9).value,
            "key_AC_X": sheet.cell(14, 3).value,
            "key_AC_Y": sheet.cell(14, 4).value,
            "key_AC_Lambda": sheet.cell(14, 5).value,
            "key_AC_Intensity": sheet.cell(14, 6).value,
            "key_AC_FWHM": sheet.cell(14, 7).value,
            "key_AC_Tails": sheet.cell(14, 8).value,
            "key_AC_DeltaLambda": sheet.cell(14, 9).value,
            "key_BC_X": sheet.cell(15, 3).value,
            "key_BC_Y": sheet.cell(15, 4).value,
            "key_BC_Lambda": sheet.cell(15, 5).value,
            "key_BC_Intensity": sheet.cell(15, 6).value,
            "key_BC_FWHM": sheet.cell(15, 7).value,
            "key_BC_Tails": sheet.cell(15, 8).value,
            "key_BC_DeltaLambda": sheet.cell(15, 9).value,
            "key_CD_X": sheet.cell(16, 3).value,
            "key_CD_Y": sheet.cell(16, 4).value,
            "key_CD_Lambda": sheet.cell(16, 5).value,
            "key_CD_Intensity": sheet.cell(16, 6).value,
            "key_CD_FWHM": sheet.cell(16, 7).value,
            "key_CD_Tails": sheet.cell(16, 8).value,
            "key_CD_DeltaLambda": sheet.cell(16, 9).value,
            "key_AA_X": sheet.cell(17, 3).value,
            "key_AA_Y": sheet.cell(17, 4).value,
            "key_AA_Lambda": sheet.cell(17, 5).value,
            "key_AA_Intensity": sheet.cell(17, 6).value,
            "key_AA_FWHM": sheet.cell(17, 7).value,
            "key_AA_Tails": sheet.cell(17, 8).value,
            "key_AA_DeltaLambda": sheet.cell(17, 9).value,
            "key_BB_X": sheet.cell(18, 3).value,
            "key_BB_Y": sheet.cell(18, 4).value,
            "key_BB_Lambda": sheet.cell(18, 5).value,
            "key_BB_Intensity": sheet.cell(18, 6).value,
            "key_BB_FWHM": sheet.cell(18, 7).value,
            "key_BB_Tails": sheet.cell(18, 8).value,
            "key_BB_DeltaLambda": sheet.cell(18, 9).value,
            "key_CC_X": sheet.cell(19, 3).value,
            "key_CC_Y": sheet.cell(19, 4).value,
            "key_CC_Lambda": sheet.cell(19, 5).value,
            "key_CC_Intensity": sheet.cell(19, 6).value,
            "key_CC_FWHM": sheet.cell(19, 7).value,
            "key_CC_Tails": sheet.cell(19, 8).value,
            "key_CC_DeltaLambda": sheet.cell(19, 9).value,
            "key_DD_X": sheet.cell(20, 3).value,
            "key_DD_Y": sheet.cell(20, 4).value,
            "key_DD_Lambda": sheet.cell(20, 5).value,
            "key_DD_Intensity": sheet.cell(20, 6).value,
            "key_DD_FWHM": sheet.cell(20, 7).value,
            "key_DD_Tails": sheet.cell(20, 8).value,
            "key_DD_DeltaLambda": sheet.cell(20, 9).value,
            "key_PL_MAP": PL_MAP
        }
        #Temporary variable (ex:"F9" means cell F9 value)
        F9=sheet.cell(8,5).value
        F14=sheet.cell(13,5).value
        F15=sheet.cell(14,5).value
        F16=sheet.cell(15,5).value
        F17=sheet.cell(16,5).value
        
       
        
        df = pd.read_excel(file_path, header=None, sheet_name='データファイル', usecols="A:I", skiprows=1)
        Data_file_sheet_array = []
        for i in range(len(df)):
            Data_file_sheet_array.append([str(x) for x in df.iloc[i].tolist()])

        data_dict['key_Cmb_Date_Time'] = Data_file_sheet_array[0][0]
        data_dict['key_Cmb_Spectral_Range'] = Data_file_sheet_array[0][1]
        data_dict['key_Cmb_Peak_Position'] = Data_file_sheet_array[0][2]
        data_dict['key_Cmb_Peak_Intensity'] = Data_file_sheet_array[0][3]
        data_dict['key_Cmb_FWHM(nm)'] = Data_file_sheet_array[0][4]
        data_dict['key_Cmb_FWHM(meV)'] = Data_file_sheet_array[0][5]
        data_dict['key_Cmb_SideBandHight'] = Data_file_sheet_array[0][6]
        data_dict['key_Cmb_X'] = Data_file_sheet_array[0][7]
        data_dict['key_Cmb_Y'] = Data_file_sheet_array[0][8]
       #data_dict['Lambda_diff'] = data_dict["key_Mapper_Average_Mapper_PL_Lambda"] - data_dict["key_Mapper_Adjust_Mapper_Wavelength(1.3um)"]
        data_dict['Lambda_diff']= data_dict["key_Mapper_Average_Mapper_PL_Lambda"]-data_dict["key_Mapper_Adjust_Mapper_TargetWavelength"]
        #data_dict['Lambda2_diff'] = data_dict["key_Mapper_Average_Mapper_PL_Lambda2"] - data_dict["key_Center_Lambda"]
        data_dict['Lambda2_diff'] =  max(abs(F14-F9),abs(F15-F9),abs(F16-F9),abs(F17-F9)) #2025/02/07 modify Lambda2_diff=max(|F14-F9|,,,,|F17-F9|) yoshida

        wb.release_resources()
    else:
        wb = px.load_workbook(file_path, data_only=True)
        sheet = wb[Data_sheet_name]
        date_sheet = wb['データファイル']
        serial_number = sheet.cell(row=4, column=8).value  # 第4列第8欄 (1-indexed)
        if type(serial_number) is not str:
            wb.close()
            return None, None
        try:
            conn, cursor = SQL.connSQL()
            if conn is None:
                Log.Log_Error(Log_File, serial_number + ' : ' + 'Connection with Prime Failed')
                wb.close()
                return None, None
            part_number, Nine_Serial_Number = SQL.selectSQL(cursor, serial_number)
            SQL.disconnSQL(conn, cursor)
        except Exception as e:
            Log.Log_Error(Log_File, serial_number + ' : ' + 'SQL Error: ' + str(e))
            wb.close()
            return None, None
        start_date = date_sheet.cell(row=5, column=1).value
        Operator = str(sheet.cell(row=4, column=6).value)
        if Operator == "None" or Operator == "":
            Operator = "-"
        PL_MAP = '1'
        if '#2' in str(file_path):
            PL_MAP = '2'
        if '#3' in str(file_path):
            PL_MAP = '3' 
        if '#4' in str(file_path):
            PL_MAP = '4'                       
        data_dict = {
            "key_start_date_time": start_date,
            "key_part_number": part_number,
            "key_serial_number": serial_number,
            "key_operator": Operator,
            "key_LotNumber_9": Nine_Serial_Number,
            "key_batch_number": sheet.cell(row=3, column=8).value,
            "key_Mapper_Average_Mapper_PL_Lambda": sheet.cell(row=3, column=14).value,
            "key_Mapper_Average_Mapper_PL_Lambda2": (sheet.cell(row=14, column=6).value + sheet.cell(row=15, column=6).value + sheet.cell(row=16, column=6).value + sheet.cell(row=17, column=6).value) / 4,
            "key_Mapper_Average_Mapper_PL_Intensity": sheet.cell(row=4, column=14).value,
            "key_Mapper_Average_Mapper_PL_FWHM": sheet.cell(row=5, column=14).value,
            "key_Mapper_Adjust_Mapper_TargetWavelength": sheet.cell(row=27, column=5).value,
            "key_Mapper_Adjust_Mapper_Wavelength(1.3um)": sheet.cell(row=28, column=5).value,
            "key_Mapper_Adjust_Mapper_CheckingAdjustValue": sheet.cell(row=29, column=5).value,
            "key_Thickness_Thickness_Core": sheet.cell(row=76, column=13).value,             
            "key_Center_X": sheet.cell(row=9, column=4).value,
            "key_Center_Y": sheet.cell(row=9, column=5).value,
            "key_Center_Lambda": sheet.cell(row=9, column=6).value,
            "key_Center_Intensity": sheet.cell(row=9, column=7).value,
            "key_Center_FWHM": sheet.cell(row=9, column=8).value,
            "key_Center_Tails": sheet.cell(row=9, column=9).value,
            "key_Center_DeltaLambda": sheet.cell(row=9, column=10).value,
            "key_A_X": sheet.cell(row=10, column=4).value,
            "key_A_Y": sheet.cell(row=10, column=5).value,
            "key_A_Lambda": sheet.cell(row=10, column=6).value,
            "key_A_Intensity": sheet.cell(row=10, column=7).value,
            "key_A_FWHM": sheet.cell(row=10, column=8).value,
            "key_A_Tails": sheet.cell(row=10, column=9).value,
            "key_A_DeltaLambda": sheet.cell(row=10, column=10).value,
            "key_B_X": sheet.cell(row=11, column=4).value,
            "key_B_Y": sheet.cell(row=11, column=5).value,
            "key_B_Lambda": sheet.cell(row=11, column=6).value,
            "key_B_Intensity": sheet.cell(row=11, column=7).value,
            "key_B_FWHM": sheet.cell(row=11, column=8).value,
            "key_B_Tails": sheet.cell(row=11, column=9).value,
            "key_B_DeltaLambda": sheet.cell(row=11, column=10).value,
            "key_C_X": sheet.cell(row=12, column=4).value,
            "key_C_Y": sheet.cell(row=12, column=5).value,
            "key_C_Lambda": sheet.cell(row=12, column=6).value,
            "key_C_Intensity": sheet.cell(row=12, column=7).value,
            "key_C_FWHM": sheet.cell(row=12, column=8).value,
            "key_C_Tails": sheet.cell(row=12, column=9).value,
            "key_C_DeltaLambda": sheet.cell(row=12, column=10).value,
            "key_D_X": sheet.cell(row=13, column=4).value,
            "key_D_Y": sheet.cell(row=13, column=5).value,
            "key_D_Lambda": sheet.cell(row=13, column=6).value,
            "key_D_Intensity": sheet.cell(row=13, column=7).value,
            "key_D_FWHM": sheet.cell(row=13, column=8).value,
            "key_D_Tails": sheet.cell(row=13, column=9).value,
            "key_D_DeltaLambda": sheet.cell(row=13, column=10).value,
            "key_AB_X": sheet.cell(row=14, column=4).value,
            "key_AB_Y": sheet.cell(row=14, column=5).value,
            "key_AB_Lambda": sheet.cell(row=14, column=6).value,
            "key_AB_Intensity": sheet.cell(row=14, column=7).value,
            "key_AB_FWHM": sheet.cell(row=14, column=8).value,
            "key_AB_Tails": sheet.cell(row=14, column=9).value,
            "key_AB_DeltaLambda": sheet.cell(row=14, column=10).value,
            "key_AC_X": sheet.cell(row=15, column=4).value,
            "key_AC_Y": sheet.cell(row=15, column=5).value,
            "key_AC_Lambda": sheet.cell(row=15, column=6).value,
            "key_AC_Intensity": sheet.cell(row=15, column=7).value,
            "key_AC_FWHM": sheet.cell(row=15, column=8).value,
            "key_AC_Tails": sheet.cell(row=15, column=9).value,
            "key_AC_DeltaLambda": sheet.cell(row=15, column=10).value,
            "key_BC_X": sheet.cell(row=16, column=4).value,
            "key_BC_Y": sheet.cell(row=16, column=5).value,
            "key_BC_Lambda": sheet.cell(row=16, column=6).value,
            "key_BC_Intensity": sheet.cell(row=16, column=7).value,
            "key_BC_FWHM": sheet.cell(row=16, column=8).value,
            "key_BC_Tails": sheet.cell(row=16, column=9).value,
            "key_BC_DeltaLambda": sheet.cell(row=16, column=10).value,
            "key_CD_X": sheet.cell(row=17, column=4).value,
            "key_CD_Y": sheet.cell(row=17, column=5).value,
            "key_CD_Lambda": sheet.cell(row=17, column=6).value,
            "key_CD_Intensity": sheet.cell(row=17, column=7).value,
            "key_CD_FWHM": sheet.cell(row=17, column=8).value,
            "key_CD_Tails": sheet.cell(row=17, column=9).value,
            "key_CD_DeltaLambda": sheet.cell(row=17, column=10).value,
            "key_AA_X": sheet.cell(row=18, column=4).value,
            "key_AA_Y": sheet.cell(row=18, column=5).value,
            "key_AA_Lambda": sheet.cell(row=18, column=6).value,
            "key_AA_Intensity": sheet.cell(row=18, column=7).value,
            "key_AA_FWHM": sheet.cell(row=18, column=8).value,
            "key_AA_Tails": sheet.cell(row=18, column=9).value,
            "key_AA_DeltaLambda": sheet.cell(row=18, column=10).value,
            "key_BB_X": sheet.cell(row=19, column=4).value,
            "key_BB_Y": sheet.cell(row=19, column=5).value,
            "key_BB_Lambda": sheet.cell(row=19, column=6).value,
            "key_BB_Intensity": sheet.cell(row=19, column=7).value,
            "key_BB_FWHM": sheet.cell(row=19, column=8).value,
            "key_BB_Tails": sheet.cell(row=19, column=9).value,
            "key_BB_DeltaLambda": sheet.cell(row=19, column=10).value,
            "key_CC_X": sheet.cell(row=20, column=4).value,
            "key_CC_Y": sheet.cell(row=20, column=5).value,
            "key_CC_Lambda": sheet.cell(row=20, column=6).value,
            "key_CC_Intensity": sheet.cell(row=20, column=7).value,
            "key_CC_FWHM": sheet.cell(row=20, column=8).value,
            "key_CC_Tails": sheet.cell(row=20, column=9).value,
            "key_CC_DeltaLambda": sheet.cell(row=20, column=10).value,
            "key_DD_X": sheet.cell(row=21, column=4).value,
            "key_DD_Y": sheet.cell(row=21, column=5).value,
            "key_DD_Lambda": sheet.cell(row=21, column=6).value,
            "key_DD_Intensity": sheet.cell(row=21, column=7).value,
            "key_DD_FWHM": sheet.cell(row=21, column=8).value,
            "key_DD_Tails": sheet.cell(row=21, column=9).value,
            "key_DD_DeltaLambda": sheet.cell(row=21, column=10).value,
            "key_PL_MAP": PL_MAP
        }
        
        #Temporary variable (ex:"F9" means cell F9 value)
        F9=sheet.cell(9,6).value
        F14=sheet.cell(14,6).value
        F15=sheet.cell(15,6).value
        F16=sheet.cell(16,6).value
        F17=sheet.cell(17,6).value
        
        
        df = pd.read_excel(file_path, header=None, sheet_name='データファイル', usecols="A:I", skiprows=1)
        Data_file_sheet_array = []
        for i in range(len(df)):
            Data_file_sheet_array.append([str(x) for x in df.iloc[i].tolist()])
        data_dict['key_Cmb_Date_Time'] = Data_file_sheet_array[0][0]
        data_dict['key_Cmb_Spectral_Range'] = Data_file_sheet_array[0][1]
        data_dict['key_Cmb_Peak_Position'] = Data_file_sheet_array[0][2]
        data_dict['key_Cmb_Peak_Intensity'] = Data_file_sheet_array[0][3]
        data_dict['key_Cmb_FWHM(nm)'] = Data_file_sheet_array[0][4]
        data_dict['key_Cmb_FWHM(meV)'] = Data_file_sheet_array[0][5]
        data_dict['key_Cmb_SideBandHight'] = Data_file_sheet_array[0][6]
        data_dict['key_Cmb_X'] = Data_file_sheet_array[0][7]
        data_dict['key_Cmb_Y'] = Data_file_sheet_array[0][8]
        #data_dict['Lambda_diff'] = data_dict["key_Mapper_Average_Mapper_PL_Lambda"] - data_dict["key_Mapper_Adjust_Mapper_Wavelength(1.3um)"]
        data_dict['Lambda_diff']= data_dict["key_Mapper_Average_Mapper_PL_Lambda"]-data_dict["key_Mapper_Adjust_Mapper_TargetWavelength"]
        #data_dict['Lambda2_diff'] = data_dict["key_Mapper_Average_Mapper_PL_Lambda2"] - data_dict["key_Center_Lambda"]
        data_dict['Lambda2_diff'] =  max(abs(F14-F9),abs(F15-F9),abs(F16-F9),abs(F17-F9)) #2025/02/07 modify Lambda2_diff=max(|F14-F9|,,,,|F17-F9|) yoshida
        try:
            if isinstance(data_dict['key_Cmb_Date_Time'], datetime):
                data_dict['key_Cmb_Date_Time'] = data_dict['key_Cmb_Date_Time'].strftime("%Y-%m-%d %H:%M:%S")
        except ValueError as e:
            print(data_dict['key_Cmb_Date_Time'])
            print(e)
        wb.close()
    return data_dict, Data_file_sheet_array

# XML轉換函數
def Output_XML(xml_file, data_dict, Data_file_sheet_array):
    
    Log.Log_Info(Log_File, 'Excel File To XML File Conversion')
    XML =  '<?xml version="1.0" encoding="utf-8"?>\n' + \
           '<Results xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema">\n' + \
           '       <Result startDateTime="' + str(data_dict["key_start_date_time"]) + '" Result="Passed">\n' + \
           '               <Header SerialNumber="' + str(data_dict["key_serial_number"]) + '" PartNumber="' + str(data_dict["key_part_number"]) + \
           '" Operation="' + Operation + '" TestStation="' + TestStation + '" Operator="' + str(data_dict["key_operator"]) + \
           '" StartTime="' + str(data_dict["key_start_date_time"]) + '" Site="' + Site + '" BatchNumber="' + str(data_dict["key_batch_number"]) + \
           '" LotNumber="' + str(data_dict["key_serial_number"]) + '"/>\n' + \
           '\n' + \
           '               <TestStep Name="' + teststep_dict["TestStep1"] + '" startDateTime="' + str(data_dict["key_start_date_time"]) + '" Status="Passed">\n' + \
           '                   <Data DataType="Numeric" Name="X" Units="um" Value="' + X + '"/>\n' + \
           '                   <Data DataType="Numeric" Name="Y" Units="um" Value="' + Y + '"/>\n' + \
           '               </TestStep>\n' + \
           '               <TestStep Name="' + teststep_dict["TestStep2"] + '" startDateTime="' + str(data_dict["key_start_date_time"]) + '" Status="Passed">\n' + \
           '                   <Data DataType="Numeric" Name="Mapper_PL_Lambda" Units="nm" Value="' + str(data_dict["key_Mapper_Average_Mapper_PL_Lambda"]) + '"/>\n' + \
           '                   <Data DataType="Numeric" Name="Mapper_PL_Intensity" Units="mV" Value="' + str(data_dict["key_Mapper_Average_Mapper_PL_Intensity"]) + '"/>\n' + \
           '                   <Data DataType="Numeric" Name="Mapper_PL_FWHM" Units="meV" Value="' + str(data_dict["key_Mapper_Average_Mapper_PL_FWHM"]) + '"/>\n' + \
           '                   <Data DataType="Numeric" Name="Thickness_Thickness_Core" Units="nm" Value="' + str(data_dict["key_Thickness_Thickness_Core"]) + '"/>\n' + \
           '               </TestStep>\n' + \
           '               <TestStep Name="' + teststep_dict["TestStep3"] + '" startDateTime="' + str(data_dict["key_start_date_time"]) + '" Status="Passed">\n' + \
           '                   <Data DataType="Numeric" Name="Mapper_TargetWavelength" Units="nm" Value="' + str(data_dict["key_Mapper_Adjust_Mapper_TargetWavelength"]) + '"/>\n' + \
           '                   <Data DataType="Numeric" Name="Mapper_Wavelength(1.3um)" Units="nm" Value="' + str(data_dict["key_Mapper_Adjust_Mapper_Wavelength(1.3um)"]) + '"/>\n' + \
           '                   <Data DataType="Numeric" Name="Mapper_CheckingAdjustValue" Units="nm" Value="' + str(data_dict["key_Mapper_Adjust_Mapper_CheckingAdjustValue"]) + '"/>\n' + \
           '               </TestStep>\n' + \
           '               <TestStep Name="' + teststep_dict["TestStep4"] + '" startDateTime="' + str(data_dict["key_start_date_time"]) + '" Status="Passed">\n' + \
           '                   <Data DataType="Numeric" Name="X" Units="um" Value="' + str(data_dict["key_Center_X"]) + '"/>\n' + \
           '                   <Data DataType="Numeric" Name="Y" Units="um" Value="' + str(data_dict["key_Center_Y"]) + '"/>\n' + \
           '                   <Data DataType="Numeric" Name="Lambda" Units="nm" Value="' + str(data_dict["key_Center_Lambda"]) + '"/>\n' + \
           '                   <Data DataType="Numeric" Name="Lambda_Diff" Units="nm" Value="' + str(data_dict["Lambda_diff"]) + '"/>\n' + \
           '                   <Data DataType="Numeric" Name="Lambda2_Diff" Units="nm" Value="' + str(data_dict['Lambda2_diff']) + '"/>\n' + \
           '                   <Data DataType="Numeric" Name="Intensity" Units="counts" Value="' + str(data_dict["key_Center_Intensity"]) + '"/>\n' + \
           '                   <Data DataType="Numeric" Name="FWHM" Units="meV" Value="' + str(data_dict["key_Center_FWHM"]) + '"/>\n' + \
           '                   <Data DataType="Numeric" Name="Tails" Units="AU" Value="' + str(data_dict["key_Center_Tails"]) + '"/>\n' + \
           '                   <Data DataType="Numeric" Name="DeltaLambda" Units="nm" Value="' + str(data_dict["key_Center_DeltaLambda"]) + '"/>\n' + \
           '               </TestStep>\n'
    # ※ 其他 TestStep（TestStep5 ~ TestStep17）的資料可依原邏輯補充
    # 加入 Cmb 的 TestStep
    for i in range(len(Data_file_sheet_array)):
        TestStep = 'Cmb' + str(i + 1)
        XML += '               <TestStep Name="' + TestStep + '" startDateTime="' + str(data_dict["key_start_date_time"]) + '" Status="Passed">\n' + \
               '                   <Data DataType="String" Name="Cmb_Date_Time" Value="' + str(Data_file_sheet_array[i][0]) + '" CompOperation="LOG"/>\n' + \
               '                   <Data DataType="String" Name="Cmb_Spectral_Range" Value="' + str(Data_file_sheet_array[i][1]) + '" CompOperation="LOG"/>\n' + \
               '                   <Data DataType="Numeric" Name="Cmb_Peak_Position" Units="nm" Value="' + str(Data_file_sheet_array[i][2]) + '"/>\n' + \
               '                   <Data DataType="Numeric" Name="Cmb_Peak_Intensity" Units="counts" Value="' + str(Data_file_sheet_array[i][3]) + '"/>\n' + \
               '                   <Data DataType="Numeric" Name="Cmb_FWHM(nm)" Units="nm" Value="' + str(Data_file_sheet_array[i][4]) + '"/>\n' + \
               '                   <Data DataType="Numeric" Name="Cmb_FWHM(meV)" Units="meV" Value="' + str(Data_file_sheet_array[i][5]) + '"/>\n' + \
               '                   <Data DataType="Numeric" Name="Cmb_SideBandHight" Units="counts" Value="' + str(Data_file_sheet_array[i][6]) + '"/>\n' + \
               '                   <Data DataType="Numeric" Name="Cmb_X" Units="um" Value="' + str(Data_file_sheet_array[i][7]) + '"/>\n' + \
               '                   <Data DataType="Numeric" Name="Cmb_Y" Units="um" Value="' + str(Data_file_sheet_array[i][8]) + '"/>\n' + \
               '               </TestStep>\n'
    XML += '               <TestEquipment>\n' + \
           '                   <Item DeviceName="PLmapper" DeviceSerialNumber="' + str(data_dict["key_PL_MAP"]) + '"/>\n' + \
           '                   <Item DeviceName="MOCVD" DeviceSerialNumber="2"/>\n' + \
           '               </TestEquipment>\n' + \
           '               <ErrorData/>\n' + \
           '               <FailureData/>\n' + \
           '               <Configuration/>\n' + \
           '       </Result>\n' + \
           '</Results>'
    with open(Output_filepath + xml_file, 'w', encoding="utf-8") as f:
        print(Output_filepath + xml_file)
        f.write(XML)

# 取得指定路徑中所有 Excel 檔案
def ALL_FILE_FETCH(Path):
    global ALL_FILES
    for path, dir, file in os.walk(Path):
        dir[:] = [d for d in dir if d not in NotUsedDir]
        for f in file:
            if f.endswith('.xlsx') or f.endswith('.xlsm') or f.endswith('.xls'):
                ALL_FILES.append(os.path.join(path, f))

if __name__ == '__main__':
    Log.Log_Info(Log_File, 'Main Start')
    # 指定需要處理的路徑
    Path = [ 
        "Z:/PL-MAP/★PLﾏｯﾊﾟｰ判定/3ｲﾝﾁHTL13B5/F2 LD/",
        "Z:/PL-MAP#2/★PLﾏｯﾊﾟｰ判定/HL13B5/F2 LD/",
        "Z:/PL-MAP/★PLﾏｯﾊﾟｰ判定/3ｲﾝﾁHL13E1/F2_LD",
        "Z:/PL-MAP#2/★PLﾏｯﾊﾟｰ判定/HL13E1/F2LD",
        "Z:/PL-MAP#3/★PLﾏｯﾊﾟｰ判定/HL13E1/F2LD"
    ]
    for p in Path:
        ALL_FILE_FETCH(p)
        
    for FilePath in ALL_FILES:
        file_mod_time = datetime.fromtimestamp(os.path.getmtime(FilePath))
        if (datetime.now() - file_mod_time).days > 30:
            Log.Log_Info(Log_File, f"File {FilePath} is older than 15 days. Skipping.")
            continue
        File = os.path.basename(FilePath)
        if FilePath in EndFiles or '~$' in File:
            print("the file is temp file or upload already skip it.")
            continue
        Log.Log_Info(Log_File, File)
        if Get_Cells_Info(FilePath):
            Log.Log_Error(Log_File, "Blank Error\n")
            continue
        EndFiles.add(FilePath)
        data_dict, Data_file_sheet_array = Open_Data_Sheet(FilePath)
        if data_dict is None:
            Log.Log_Error(Log_File, "Lot Error\n")
            continue
        if data_dict['key_part_number'] == "":
            Log.Log_Error(Log_File, str(data_dict["key_serial_number"]) + ' : ' + "Part Number Error\n")
            continue
        if data_dict['key_operator'] == "":
            data_dict["key_operator"] = "-"
        Log.Log_Info(Log_File, 'Date Format Conversion')
        data_dict['key_start_date_time'] = Convert_Date.Edit_Date(data_dict['key_start_date_time']).replace('.', ':')
        if data_dict['key_start_date_time'] == "":
            Log.Log_Error(Log_File, str(data_dict["key_serial_number"]) + ' : ' + "Date Error\n")
            continue
        date_obj = datetime.strptime(str(data_dict["key_start_date_time"]).replace('T', ' ').replace('.', ':'), "%Y-%m-%d %H:%M:%S")
        date_excel_number = int(str(date_obj - datetime(1899, 12, 30)).split()[0])
        Epi_Number = 0
        for i in data_dict["key_batch_number"]:
            try:
                if 0 <= int(i) <= 9:
                    Epi_Number = Epi_Number * 10 + int(i)
            except:
                pass
        date_excel_number += Epi_Number/10**6
        data_dict["key_STARTTIME_SORTED"] = date_excel_number
        data_dict["key_SORTNUMBER"] = Epi_Number

        result = Check.Data_Type(key_type, data_dict)
        if result == False:
            Log.Log_Error(Log_File, str(data_dict["key_serial_number"]) + ' : ' + "Data Error\n")
            continue
        xml_file = 'Site=' + Site + ',ProductFamily=' + ProductFamily + ',Operation=' + Operation + \
                   ',Partnumber=' + str(data_dict["key_part_number"]) + ',Serialnumber=' + str(data_dict["key_serial_number"]) + \
                   ',Testdate=' + data_dict["key_start_date_time"].replace(':', '.') + '.xml'
        Output_XML(xml_file, data_dict, Data_file_sheet_array)
        Log.Log_Info(Log_File, str(data_dict["key_serial_number"]) + ' : ' + "OK\n")
    EndFiles_list = sorted(list(EndFiles))
    EndFiles_str = "\n".join(EndFiles_list)
    with open('EndsFile_F2_Format2.txt', 'w', encoding='utf-8') as textfile:
        textfile.write(EndFiles_str)
    #shutil.copy('EndsFile_F2_Format2.txt', 'T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開發/040_LD-EML/F2/13_ProgramUsedFile/')
    Log.Log_Info(Log_File, 'Program End')
