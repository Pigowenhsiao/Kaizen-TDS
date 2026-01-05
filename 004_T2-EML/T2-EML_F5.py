import os
import sys
import glob
import openpyxl as px
import pprint
import shutil
import logging

from datetime import datetime, timedelta, date
from time import strftime, localtime


########## 自作関数の定義 ##########
sys.path.append('../MyModule')
import SQL
import Log
import MOCVD_OldFileSearch
import Check


########## 全体パラメータ定義 ##########
Site = '350'
ProductFamily = 'SAG FAB'
Operation = 'T2-EML'
Teststation = 'T2-EML'
X = '999999'
Y = '999999'


########## Logの設定 ##########
Log_FolderName = str(date.today())

# ----- 格納するLogフォルダがなければ作成する -----
if not os.path.exists("../Log/" + Log_FolderName):
    os.makedirs("../Log/" + Log_FolderName)

# ----- ログ書き込み先ファイルパス -----
Log_file = '../Log/' + Log_FolderName + '/004_T2-EML_F5.log'

# ----- ログ書込：プログラムの開始 -----
Log.Log_Info(Log_file, 'Program Start')


########## 処理ファイルのあるディレクトリ定義 ##########
Path = 'Z:/MOCVD/MOCVD過去プログラム/F5炉/'
# Path = 'C:/Users/hor78296/Desktop/F2炉/'


########## XML出力先ファイルパス ##########
Output_filepath = '//li.lumentuminc.net/data/SAG/TDS/Data/Files to Insert/XML/'
# Output_filepath = '../XML/004_T2-EML/'
#Output_filepath = 'C:/Users/hsi67063/Documents/TEMP/'  #for test


########## TestStepの定義 ##########
teststep_dict = {
    'TestStep1' : 'Coordinate',
    'TestStep2' : 'CarrierConcentration',
    'TestStep3' : 'Thickness',
    'TestStep4' : 'Dulation',
    'TestStep5' : 'MO1-TMI',
    'TestStep6' : 'MO2-TEG',
    'TestStep7' : 'MO3-TMI',
    'TestStep8' : 'MO4-TEG',
    'TestStep9' : 'MO5-TMI',
    'TestStep10' : 'AsH3-A-20percent',
    'TestStep11' : 'AsH3-5percent',
    'TestStep12' : 'PH3-A-50percent',
    'TestStep13' : 'DMZn-A-0.1percent',
    'TestStep14' : 'DMZn-B-0.1percent',
    'TestStep15' : 'Temperature',
    'TestStep16' : 'Piezocon',
    'TestStep17' : 'BallastN2',
    'TestStep18' : 'MO-Temperature',
    'TestStep19' : 'SORTED_DATA'
}


########## HeaderMiscの定義 ##########
HeaderMisc_dict = {
    'HeaderMisc1' : 'RecipeName-Macro',
    'HeaderMisc2' : 'RecipeName-Program',
    'HeaderMisc3' : 'RecipeName-Folder'
}


########## 取得した項目と型の対応表を定義 ##########
key_type = {
    "key_start_date_time" : str,
    "key_serial_number" : str,
    "key_part_number" : str,
    "key_operator" : str,
    "key_batch_number" : str,
    "key_HeaderMisc1" : str,
    "key_HeaderMisc2" : str,
    "key_HeaderMisc3" : str,
    "key_Clad_2" : float,
    "key_Contact_2" : float,
    "key_Minimum_2" : float,
    "key_Clad_3" : float,
    "key_Contact_3": float,
    "key_Step9_4" : float,
    "key_Step10_4" : float,
    "key_Step11_4" : float,
    "key_Step12_4" : float,
    "key_Step13_4" : float,
    "key_Step14_4" : float,
    "key_Step20_4" : float,
    "key_Step26_4" : float,
    "key_Step32_4" : float,
    "key_Step33_4" : float,
    "key_Step34_4" : float,
    "key_Step35_4" : float,
    "key_Step40_4" : float,
    "key_Step42_4" : float,
    "key_Step20_5" : float,
    "key_Step20_6" : float,
    "key_Step26_7" : float,
    "key_Step32_7" : float,
    "key_Step33_7" : float,
    "key_Step34_7" : float,
    "key_Step35_7" : float,
    "key_Step26_8" : float,
    "key_Step32_8" : float,
    "key_Step33_8" : float,
    "key_Step34_8" : float,
    "key_Step35_8" : float,
    "key_Step9_9" : float,
    "key_Step10_9" : float,
    "key_Step11_9" : float,
    "key_Step12_9" : float,
    "key_Step13_9" : float,
    "key_Step14_9" : float,
    "key_Step40_9" : float,
    "key_Step26_10" : float,
    "key_Step32_10" : float,
    "key_Step33_10" : float,
    "key_Step34_10" : float,
    "key_Step35_10" : float,
    "key_AsH3_11" : float,
    "key_Step9_12" : float,
    "key_Step10_12" : float,
    "key_Step11_12" : float,
    "key_Step12_12" : float,
    "key_Step13_12" : float,
    "key_Step14_12" : float,
    "key_Step20_12" : float,
    "key_Step26_12" : float,
    "key_Step40_12" : float,
    "key_Step42_12" : float,
    "key_Step13_13" : float,
    "key_Step14_13" : float,
    "key_Step20_13" : float,
    "key_Step26_13" : float,
    "key_Step32_13" : float,
    "key_Step33_13" : float,
    "key_Step34_13" : float,
    "key_Step35_13" : float,
    "key_Step9_14" : float,
    "key_Step10_14" : float,
    "key_Step11_14" : float,
    "key_Step12_14" : float,
    "key_Step9_15" : float,
    "key_Step10_15" : float,
    "key_Step11_15" : float,
    "key_Step12_15" : float,
    "key_Step13_15" : float,
    "key_Step14_15" : float,
    "key_Step20_15" : float,
    "key_Step26_15" : float,
    "key_Step32_15" : float,
    "key_Step33_15" : float,
    "key_Step34_15" : float,
    "key_Step35_15" : float,
    "key_Step40_15" : float,
    "key_Step42_15" : float,
    "key_F1_16" : float,
    "key_F2-Start_16" : float,
    "key_F2-Current_16" : float,
    "key_F2-T2CH_16" : float,
    "key_BallastN2_17" : float,
    "key_MO1_18" : float,
    "key_MO2_18" : float,
    "key_MO3_18" : float,
    "key_MO4_18" : float,
    "key_MO5_18" : float,
    "key_MO6_18" : float,
    "key_Blank_18" : float,
    "key_STARTTIME_SORTED": float,
    "key_SORTNUMBER" : float,
    "key_LotNumber_9" : str
}


########## 対象ロット番号のイニシャルを記載したファイルを取得する ##########
Log.Log_Info(Log_file, 'Get SerialNumber Initial List ')
#with open('T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開発/004_T2-EML/SerialNumber_Initial.txt', 'r', encoding='utf-8') as textfile:
with open('C:/Users/hsi67063/Downloads/SerialNumber_Initial.txt', 'r', encoding='utf-8') as textfile:  #for test
    SerialNumber_list = {s.strip() for s in textfile.readlines()}


########## 前回処理を行ったファイル名を取得する ##########
with open('F5_FileName.txt', 'r', encoding='utf-8') as textfile:
    Before_FileName = textfile.readline()
    print('before_filename:',Before_FileName)


########## 空欄チェック ##########
def Get_Cells_Info(Sheet):

    # ----- ログ記載：空欄判定 -----
    Log.Log_Info(Log_file, "Blank Check")

    # ----- False -> 空欄がない -----
    is_cells_empty = False

    # ----- 日付かロット番号かエピ番号かバラストN2流量が空欄ならば処理を行わない -----
    if  Sheet['I8'].value is None or Sheet['M8'].value is None or Sheet['Q7'].value is None or Sheet['AE40'].value is None:
        is_cells_empty = True

    return is_cells_empty


########## データの取得 ##########
def Open_Data_Sheet(Sheet, filepath, SheetName):

    # ----- ログ記載：データ取得 -----
    Log.Log_Info(Log_file, 'Data Acquisition')

    # ----- データを格納する辞書を作成 -----
    data_dict = dict()

    # ----- ロット番号の取得 -----
    serial_number = Sheet["M8"].value

    # ----- ロット番号が数値の時があるので、数値の時はエラーを出す -----
    if type(serial_number) is not str:
        return None

    # ----- SQLサーバと接続して、品名を抜き出す -----
    conn,cursor = SQL.connSQL()

    # ----- Prime接続できなかったときはエラーを出して終了 -----
    if conn is None:
        Log.Log_Error(Log_file, serial_number + ' : ' + 'Connection with Prime Failed')
        sys.exit()

    # ----- Primeから品名を取得 -----
    part_number, Nine_Serial_Number = SQL.selectSQL(cursor, serial_number)
    SQL.disconnSQL(conn, cursor)

    # ----- Polaron / SEM / MOCVD の装置Noを取得 -----
    Equipment1,Equipment2,Equipment3 = '1','1','5'  #F2->F5
    if '#2' in str(Sheet["J45"].value):
        Equipment1 = '2'
    if '#2' in str(Sheet['J50'].value):
        Equipment2 = '2'

    # ----- データの格納 -----
    data_dict = {
        "key_start_date_time" : str(Sheet["Q7"].value).replace(" ","T"),
        "key_serial_number" : serial_number,
        "key_LotNumber_9": Nine_Serial_Number,
        "key_part_number" : part_number,
        "key_operator" : Sheet["V8"].value,
        "key_batch_number" : Sheet["I8"].value,
        "key_HeaderMisc1" : Sheet["U3"].value,
        "key_HeaderMisc2" : Sheet["U4"].value,
        "key_HeaderMisc3" : Sheet["U5"].value,
        "key_Clad_2" : Sheet["M47"].value,
        "key_Contact_2" : Sheet["M46"].value,
        "key_Minimum_2" : Sheet["M48"].value,
        "key_Clad_3" : Sheet["M52"].value,
        "key_Contact_3": Sheet["M51"].value,
        "key_Step9_4" : Sheet["I21"].value,
        "key_Step10_4" : Sheet["I22"].value,
        "key_Step11_4" : Sheet["I23"].value,
        "key_Step12_4" : Sheet["I24"].value,
        "key_Step13_4" : Sheet["I25"].value,
        "key_Step14_4" : Sheet["I26"].value,
        "key_Step20_4" : Sheet["I27"].value,
        "key_Step26_4" : Sheet["I28"].value,
        "key_Step32_4" : Sheet["I29"].value,
        "key_Step33_4" : Sheet["I30"].value,
        "key_Step34_4" : Sheet["I31"].value,
        "key_Step35_4" : Sheet["I32"].value,
        "key_Step40_4" : Sheet["I33"].value,
        "key_Step42_4" : Sheet["I34"].value,
        "key_Step20_5" : Sheet["K27"].value,
        "key_Step20_6" : Sheet["L27"].value,
        "key_Step26_7" : Sheet["M28"].value,
        "key_Step32_7" : Sheet["M29"].value,
        "key_Step33_7" : Sheet["M30"].value,
        "key_Step34_7" : Sheet["M31"].value,
        "key_Step35_7" : Sheet["M32"].value,
        "key_Step26_8" : Sheet["N28"].value,
        "key_Step32_8" : Sheet["N29"].value,
        "key_Step33_8" : Sheet["N30"].value,
        "key_Step34_8" : Sheet["N31"].value,
        "key_Step35_8" : Sheet["N32"].value,
        "key_Step9_9" : Sheet["O21"].value,
        "key_Step10_9" : Sheet["O22"].value,
        "key_Step11_9" : Sheet["O23"].value,
        "key_Step12_9" : Sheet["O24"].value,
        "key_Step13_9" : Sheet["O25"].value,
        "key_Step14_9" : Sheet["O26"].value,
        "key_Step40_9" : Sheet["O33"].value,
        "key_Step26_10" : Sheet["R28"].value,
        "key_Step32_10" : Sheet["R29"].value,
        "key_Step33_10" : Sheet["R30"].value,
        "key_Step34_10" : Sheet["R31"].value,
        "key_Step35_10" : Sheet["R32"].value,
        "key_AsH3_11" : Sheet["T27"].value,
        "key_Step9_12" : Sheet["U21"].value,
        "key_Step10_12" : Sheet["U22"].value,
        "key_Step11_12" : Sheet["U23"].value,
        "key_Step12_12" : Sheet["U24"].value,
        "key_Step13_12" : Sheet["U25"].value,
        "key_Step14_12" : Sheet["U26"].value,
        "key_Step20_12" : Sheet["U27"].value,
        "key_Step26_12" : Sheet["U28"].value,
        "key_Step40_12" : Sheet["U33"].value,
        "key_Step42_12" : Sheet["U34"].value,
        "key_Step13_13" : Sheet["X25"].value,
        "key_Step14_13" : Sheet["X26"].value,
        "key_Step20_13" : Sheet["X27"].value,
        "key_Step26_13" : Sheet["X28"].value,
        "key_Step32_13" : Sheet["X29"].value,
        "key_Step33_13" : Sheet["X30"].value,
        "key_Step34_13" : Sheet["X31"].value,
        "key_Step35_13" : Sheet["X32"].value,
        "key_Step9_14" : Sheet["Y21"].value,
        "key_Step10_14" : Sheet["Y22"].value,
        "key_Step11_14" : Sheet["Y23"].value,
        "key_Step12_14" : Sheet["Y24"].value,
        "key_Step9_15" : Sheet["Z21"].value,
        "key_Step10_15" : Sheet["Z22"].value,
        "key_Step11_15" : Sheet["Z23"].value,
        "key_Step12_15" : Sheet["Z24"].value,
        "key_Step13_15" : Sheet["Z25"].value,
        "key_Step14_15" : Sheet["Z26"].value,
        "key_Step20_15" : Sheet["Z27"].value,
        "key_Step26_15" : Sheet["Z28"].value,
        "key_Step32_15" : Sheet["Z29"].value,
        "key_Step33_15" : Sheet["Z30"].value,
        "key_Step34_15" : Sheet["Z31"].value,
        "key_Step35_15" : Sheet["Z32"].value,
        "key_Step40_15" : Sheet["Z33"].value,
        "key_Step42_15" : Sheet["Z34"].value,
        "key_F1_16" : Sheet["AE38"].value,
        "key_F2-Start_16" : Sheet["AE44"].value,
        "key_F2-Current_16" : Sheet["AE45"].value,
        "key_F2-T2CH_16" : Sheet["AE46"].value,
        "key_BallastN2_17" : Sheet["AE41"].value,
        "key_MO1_18" : Sheet["K43"].value,
        "key_MO2_18" : Sheet["L43"].value,
        "key_MO3_18" : Sheet["M43"].value,
        "key_MO4_18" : Sheet["N43"].value,
        "key_MO5_18" : Sheet["O43"].value,
        "key_MO6_18" : Sheet["P43"].value,
        "key_Blank_18" : Sheet["Q43"].value,
        "key_Equipment1" : Equipment1,
        "key_Equipment2" : Equipment2,
        "key_Equipment3" : Equipment3
    }

    # ----- 空欄箇所はNoneとして取得される。Noneは文字列に変換できないため、空欄("")に置き換える -----
    for keys in data_dict:
        if data_dict[keys] is None or data_dict[keys] == "None" or data_dict[keys] == '-':
            data_dict[keys] = ""
        # ----- 指数表記箇所はint型に変換する -----
        if type(data_dict[keys]) is float and 'e' in str(data_dict[keys]) and keys != "key_start_date_time":
            data_dict[keys] = int(float(data_dict[keys]))

    return data_dict


########## XMLファイルに変換 ##########
def Output_XML(xml_file, data_dict):

    # ----- ログ記載：XML変換 -----
    Log.Log_Info(Log_file, 'Excel File To XML File Conversion')

    f = open(Output_filepath + xml_file, 'w', encoding="utf-8")

    f.write('<?xml version="1.0" encoding="utf-8"?>' + '\n' +
            '<Results xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema">' + '\n' +
            '       <Result startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Result="Passed">' + '\n' +
            '               <Header SerialNumber=' + '"' + data_dict["key_serial_number"] + '"' + ' PartNumber=' + '"' + data_dict["key_part_number"] + '"' + ' Operation=' + '"' + Operation + '"' + ' TestStation=' + '"' + Operation + '"' + ' Operator=' + '"' + data_dict["key_operator"] + '"' + ' StartTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Site=' + '"' + Site + '"' + ' BatchNumber=' + '"' + data_dict["key_batch_number"] + '"' + ' LotNumber=' + '"' + data_dict["key_serial_number"] + '"/>' + '\n' +
            '               <HeaderMisc>' + '\n' +
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc1"] + '">' + data_dict["key_HeaderMisc1"] + '</Item>' + '\n'
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc2"] + '">' + data_dict["key_HeaderMisc2"] + '</Item>' + '\n'
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc3"] + '">' + data_dict["key_HeaderMisc3"] + '</Item>' + '\n'
            '               </HeaderMisc>' + '\n' +
            '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep1"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + X + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + Y + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep2"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="CarrierConcentration_Clad" Units="cm-3" Value=' + '"' + str(data_dict["key_Clad_2"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="CarrierConcentration_Contact" Units="cm-3" Value=' + '"' + str(data_dict["key_Contact_2"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="CarrierConcentration_Minimum" Units="cm-3" Value=' + '"' + str(data_dict["key_Minimum_2"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep3"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Thickness_Clad" Units="nm" Value=' + '"' + str(data_dict["key_Clad_3"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Thickness_Contact" Units="nm" Value=' + '"' + str(data_dict["key_Contact_3"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep4"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step9" Units="min" Value=' + '"' + str(data_dict["key_Step9_4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sec" Value=' + '"' + str(data_dict["key_Step10_4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step11" Units="min" Value=' + '"' + str(data_dict["key_Step11_4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12" Units="min" Value=' + '"' + str(data_dict["key_Step12_4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="min" Value=' + '"' + str(data_dict["key_Step13_4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="min" Value=' + '"' + str(data_dict["key_Step14_4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step20" Units="sec" Value=' + '"' + str(data_dict["key_Step20_4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step26" Units="sec" Value=' + '"' + str(data_dict["key_Step26_4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step32" Units="sec" Value=' + '"' + str(data_dict["key_Step32_4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step33" Units="sec" Value=' + '"' + str(data_dict["key_Step33_4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step34" Units="min" Value=' + '"' + str(data_dict["key_Step34_4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step35" Units="min" Value=' + '"' + str(data_dict["key_Step35_4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step40" Units="min" Value=' + '"' + str(data_dict["key_Step40_4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step42" Units="sec" Value=' + '"' + str(data_dict["key_Step42_4"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep5"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step20" Units="sccm" Value=' + '"' + str(data_dict["key_Step20_5"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep6"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step20" Units="sccm" Value=' + '"' + str(data_dict["key_Step20_6"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep7"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step26" Units="sccm" Value=' + '"' + str(data_dict["key_Step26_7"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step32" Units="sccm" Value=' + '"' + str(data_dict["key_Step32_7"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step33" Units="sccm" Value=' + '"' + str(data_dict["key_Step33_7"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step34" Units="sccm" Value=' + '"' + str(data_dict["key_Step34_7"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step35" Units="sccm" Value=' + '"' + str(data_dict["key_Step35_7"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep8"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step26" Units="sccm" Value=' + '"' + str(data_dict["key_Step26_8"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step32" Units="sccm" Value=' + '"' + str(data_dict["key_Step32_8"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step33" Units="sccm" Value=' + '"' + str(data_dict["key_Step33_8"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step34" Units="sccm" Value=' + '"' + str(data_dict["key_Step34_8"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step35" Units="sccm" Value=' + '"' + str(data_dict["key_Step35_8"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep9"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step9" Units="sccm" Value=' + '"' + str(data_dict["key_Step9_9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sccm" Value=' + '"' + str(data_dict["key_Step10_9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step11" Units="sccm" Value=' + '"' + str(data_dict["key_Step11_9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12" Units="sccm" Value=' + '"' + str(data_dict["key_Step12_9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="sccm" Value=' + '"' + str(data_dict["key_Step13_9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="sccm" Value=' + '"' + str(data_dict["key_Step14_9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step40" Units="sccm" Value=' + '"' + str(data_dict["key_Step40_9"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep10"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step26" Units="sccm" Value=' + '"' + str(data_dict["key_Step26_10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step32" Units="sccm" Value=' + '"' + str(data_dict["key_Step32_10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step33" Units="sccm" Value=' + '"' + str(data_dict["key_Step33_10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step34" Units="sccm" Value=' + '"' + str(data_dict["key_Step34_10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step35" Units="sccm" Value=' + '"' + str(data_dict["key_Step35_10"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep11"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="AsH3-5percent-Step20" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3_11"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep12"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step9" Units="sccm" Value=' + '"' + str(data_dict["key_Step9_12"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sccm" Value=' + '"' + str(data_dict["key_Step10_12"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step11" Units="sccm" Value=' + '"' + str(data_dict["key_Step11_12"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12" Units="sccm" Value=' + '"' + str(data_dict["key_Step12_12"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="sccm" Value=' + '"' + str(data_dict["key_Step13_12"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="sccm" Value=' + '"' + str(data_dict["key_Step14_12"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step20" Units="sccm" Value=' + '"' + str(data_dict["key_Step20_12"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step26" Units="sccm" Value=' + '"' + str(data_dict["key_Step26_12"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step40" Units="sccm" Value=' + '"' + str(data_dict["key_Step40_12"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step42" Units="sccm" Value=' + '"' + str(data_dict["key_Step42_12"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep13"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="sccm" Value=' + '"' + str(data_dict["key_Step13_13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="sccm" Value=' + '"' + str(data_dict["key_Step14_13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step20" Units="sccm" Value=' + '"' + str(data_dict["key_Step20_13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step26" Units="sccm" Value=' + '"' + str(data_dict["key_Step26_13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step32" Units="sccm" Value=' + '"' + str(data_dict["key_Step32_13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step33" Units="sccm" Value=' + '"' + str(data_dict["key_Step33_13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step34" Units="sccm" Value=' + '"' + str(data_dict["key_Step34_13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step35" Units="sccm" Value=' + '"' + str(data_dict["key_Step35_13"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep14"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step9" Units="sccm" Value=' + '"' + str(data_dict["key_Step9_14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sccm" Value=' + '"' + str(data_dict["key_Step10_14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step11" Units="sccm" Value=' + '"' + str(data_dict["key_Step11_14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12" Units="sccm" Value=' + '"' + str(data_dict["key_Step12_14"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep15"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step9" Units="degree" Value=' + '"' + str(data_dict["key_Step9_15"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="degree" Value=' + '"' + str(data_dict["key_Step10_15"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step11" Units="degree" Value=' + '"' + str(data_dict["key_Step11_15"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12" Units="degree" Value=' + '"' + str(data_dict["key_Step12_15"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="degree" Value=' + '"' + str(data_dict["key_Step13_15"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="degree" Value=' + '"' + str(data_dict["key_Step14_15"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step20" Units="degree" Value=' + '"' + str(data_dict["key_Step20_15"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step26" Units="degree" Value=' + '"' + str(data_dict["key_Step26_15"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step32" Units="degree" Value=' + '"' + str(data_dict["key_Step32_15"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step33" Units="degree" Value=' + '"' + str(data_dict["key_Step33_15"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step34" Units="degree" Value=' + '"' + str(data_dict["key_Step34_15"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step35" Units="degree" Value=' + '"' + str(data_dict["key_Step35_15"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step40" Units="degree" Value=' + '"' + str(data_dict["key_Step40_15"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step42" Units="degree" Value=' + '"' + str(data_dict["key_Step42_15"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep16"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="F1" Units="percent" Value=' + '"' + str(data_dict["key_F1_16"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="F2-Start" Units="percent" Value=' + '"' + str(data_dict["key_F2-Start_16"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="F2-Current" Units="percent" Value=' + '"' + str(data_dict["key_F2-Current_16"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="F2-T2CH" Units="percent" Value=' + '"' + str(data_dict["key_F2-T2CH_16"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep17"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="BallastN2" Units="slm" Value=' + '"' + str(data_dict["key_BallastN2_17"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep18"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="MO1-TMI" Units="degree" Value=' + '"' + str(data_dict["key_MO1_18"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO2-TEG" Units="degree" Value=' + '"' + str(data_dict["key_MO2_18"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO3-TMI" Units="degree" Value=' + '"' + str(data_dict["key_MO3_18"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO4-TEG" Units="degree" Value=' + '"' + str(data_dict["key_MO4_18"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO5-TMI" Units="degree" Value=' + '"' + str(data_dict["key_MO5_18"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO6-TEG" Units="degree" Value=' + '"' + str(data_dict["key_MO6_18"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Blank" Units="degree" Value=' + '"' + str(data_dict["key_Blank_18"]).replace('ー', '-') + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep19"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="STARTTIME_SORTED" Units="" Value=' + '"' + str(data_dict["key_STARTTIME_SORTED"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="SORTNUMBER" Units="" Value=' + '"' + str(data_dict["key_SORTNUMBER"]) + '"/>' + '\n' +
            '                   <Data DataType="String" Name="LotNumber_5" Value=' + '"' + str(data_dict["key_serial_number"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="LotNumber_9" Value=' + '"' + str(data_dict["key_LotNumber_9"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '\n'
            '               <TestEquipment>' + '\n' +
            '                   <Item DeviceName="Polaron" DeviceSerialNumber="' + data_dict["key_Equipment1"] + '"/>' + '\n' +
            '                   <Item DeviceName="SEM" DeviceSerialNumber="' + data_dict["key_Equipment2"] + '"/>' + '\n' +
            '                   <Item DeviceName="MOCVD" DeviceSerialNumber="' + data_dict["key_Equipment3"] + '"/>' + '\n' +
            '               </TestEquipment>' + '\n' +
            '\n'
            '               <ErrorData/>' + '\n' +
            '               <FailureData/>' + '\n' +
            '               <Configuration/>' + '\n' +
            '       </Result>' + '\n' +
            '</Results>'
            )
    f.close()


########## シートの判定からXML変換まで処理 ##########
def Data_Extract(filepath, SheetList, old_check):

    # ----- ログ記載：データ変換処理 -----
    Log.Log_Info(Log_file, 'Sub Program Main\n')

    wb = px.load_workbook(filepath, read_only = True, data_only = True)

    # ----- 取得したシートの処理 -----
    for Sheet_Name in SheetList[::-1]:

        Sheet = wb[Sheet_Name]
        Initial = str(Sheet['M8'].value)[0]
        
        # ----- シートが処理対象シートかどうか確認 -----
        if '13B2-LD' not in str(Sheet['U3'].value) or Initial not in SerialNumber_list: # need to check the recipe
            Log.Log_Error(Log_file, Sheet_Name + ' : ' + 'Not Covered\n')
            continue

        # ----- 空欄チェック -----
        if Get_Cells_Info(Sheet):
            Log.Log_Error(Log_file, "Blank Error\n")
            continue

        # ----- データの取得 -----
        data_dict = Open_Data_Sheet(Sheet, os.path.basename(filepath), Sheet_Name)

        # ----- oldファイルの実行時のみ、着工者が空欄であれば'-'に置き換える -----
        if data_dict["key_operator"] == "":
            if old_check:
                data_dict["key_operator"] = '-'
            else:
                Log.Log_Error(Log_file, Sheet_Name + ' : ' + 'Operator None\n')
                continue

        # ----- 辞書が空で返ってきたらエラーとして処理 -----
        if data_dict is None:
            Log.Log_Error(Log_file, "Lot No Error\n")
            continue

        # ----- 日付フォーマット変換 -----
        if len(data_dict['key_start_date_time']) != 19 or '年' in data_dict['key_start_date_time']:
            Log.Log_Error(Log_file, data_dict["key_serial_number"] + ' : ' + "Date Error\n")
            continue

        # ----- STARTTIME_SORTEDの追加 -----

        # 日付をExcel時間に変換する
        date = datetime.strptime(str(data_dict["key_start_date_time"]).replace('T', ' ').replace('.', ':'), "%Y-%m-%d %H:%M:%S")
        date_excel_number = int(str(date - datetime(1899, 12, 30)).split()[0])

        # ----- 品名チェック -----
        if len(data_dict["key_part_number"]) == 0:
            Log.Log_Error(Log_file, data_dict["key_serial_number"] + ' : ' + "Part Number Error\n")
            continue

        # ----- エピ番号の数値部だけを取得する -----
        Epi_Number = 0
        for i in data_dict["key_batch_number"]:
            try:
                if 0<=int(i)<=9:
                    Epi_Number = Epi_Number*10+int(i)
            except:
                pass

        # エピ番号を10^6で割って、excel時間に加算する
        date_excel_number += Epi_Number/10**6

        # data_dictに登録する
        data_dict["key_STARTTIME_SORTED"] = date_excel_number
        data_dict["key_SORTNUMBER"] = Epi_Number

        # ----- データ型の確認 -----
        result = Check.Data_Type(key_type, data_dict)
        if result == False:
            Log.Log_Error(Log_file, data_dict["key_serial_number"] + ' : ' + "Data Error\n")
            continue

        # ----- XMLファイル名定義 -----
        xml_file = 'Site=' + Site + ',ProductFamily=' + ProductFamily + ',Operation=' + Operation + \
                   ',Partnumber=' + data_dict["key_part_number"] + ',Serialnumber=' + data_dict["key_serial_number"] + \
                   ',Testdate=' + data_dict["key_start_date_time"].replace(':','.') + '.xml'

        Output_XML(xml_file, data_dict)
        Log.Log_Info(Log_file, data_dict["key_serial_number"] + ' : ' + "OK\n")
    
    wb.close()


########## Main処理 ##########
if __name__ == "__main__":

    # ----- ログ記載：Main処理の開始 -----
    Log.Log_Info(Log_file, 'Main Start')

    # ----- path内のフォルダ、ファイルを全部取得 -----
    all_files = os.listdir(Path)

    # ----- ログ記載：ファイル検索 -----
    Log.Log_Info(Log_file, 'File Search')

    # ----- 先頭にFHが付いているファイル名と最終更新日時(sec)を格納 -----
    array = []
    for filename in all_files:        
        filepath = os.path.join(Path,filename)
        if "FB" in str(filename) and "$" not in str(filename) and os.path.isfile(filepath): #update F5 tool Lot Name
            dt = strftime("%Y-%m-%d %H:%M:%S", localtime(os.path.getctime(filepath)))
            array.append([filepath, dt])
    # ----- FHファイルが見つからなかった -----
    if len(array)==0:
        Log.Log_Info(Log_file, 'Folder Error')
        sys.exit()

    # ----- 最終更新日時順に並び替える -----
    array = sorted(array, key=lambda x:x[1])
    FileName = os.path.basename(array[0][0])
    Log.Log_Info(Log_file, FileName)

    # ----- 前回処理したエピ番号のNumber部分を取り出す -----
    Number = ""
    for i in Before_FileName:
        if "0" <= i <= "9":
            Number += i

    print(Number, Number[0:4],FileName)
    
    # ----- ファイルの切り替わりを確認 -----
    if Number[0:4] not in FileName:

        # ----- ログ記載：FHフォルダ検索 -----
        Log.Log_Info(Log_file, 'Folder Serach')

        # ----- 過去フォルダからNumberが含まれたファイル名を取り出す -----
        Old_File_Path = MOCVD_OldFileSearch.F5(Number)
        if Old_File_Path == -1:
            Log.Log_Info(Log_file, 'Old Folder Error')
            sys.exit()

        # ----- ログ記載：シート名の取得 -----
        Log.Log_Info(Log_file, 'OLD Get SheetName')

        #　----- 上記で指定したファイルのシート一覧を取得する -----
        wb = px.load_workbook(Old_File_Path)
        Old_SheetName = wb.sheetnames
        wb.close()

        # ----- ログ記載：前Excelファイルのデータ取得 -----
        Log.Log_Info(Log_file, 'OLD Excel File Get Data')

        # ----- 過去ファイルの処理 -----
        Data_Extract(Old_File_Path, Old_SheetName, 1)

    # ----- ログ書込：Excelファイルのデータ取得 -----
    Log.Log_Info(Log_file, 'Excel File Get Data')

    # ----- arrayに格納されている全てのファイルの処理を行う -----
    for file_path, _ in array:

        Log.Log_Info(Log_file, os.path.basename(file_path))

        # ----- 対象ファイルを開き、シートの一覧を取得する -----
        wb = px.load_workbook(file_path)
        SheetName = wb.sheetnames
        wb.close()

        # ----- 全シートの処理を行う -----
        Data_Extract(file_path, SheetName, 0)

    # ----- ログ書込：テキストファイルにファイル名を上書きで書込する -----
    Log.Log_Info(Log_file, 'Write FileName')

    # ----- 先ほど処理を行ったファイル名の書き込み -----
    with open('F5_FileName.txt', 'w', encoding='utf-8') as textfile:
        textfile.write(FileName)

    # ----- 最終行を書き込んだファイルをGドライブに転送 -----
    shutil.copy("F5_FileName.txt", 'T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開発/004_T2-EML/13_ProgramUsedFile/')
    

########## ログ記載：プログラムの終了 ##########
Log.Log_Info(Log_file, 'Program End')