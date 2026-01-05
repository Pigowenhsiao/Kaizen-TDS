import os
import sys
import glob
import openpyxl as px
import shutil
import logging

from datetime import datetime, timedelta, date
from time import strftime, localtime


########## 自作関数の定義 ##########
sys.path.append('../MyModule')
import SQL
import Log
import MOCVD_OldFileSearch
import Check


########## 全体パラメータ定義 ##########
Site = '350'
ProductFamily = 'SAG FAB'
Operation = 'T2-EML'
Teststation = 'T2-EML'
X = '999999'
Y = '999999'


########## Logの設定 ##########
Log_FolderName = str(date.today())

# ----- 格納するLogフォルダがなければ作成する -----
if not os.path.exists("../Log/" + Log_FolderName):
    os.makedirs("../Log/" + Log_FolderName)

# ----- ログ書き込み先ファイルパス -----
Log_file = '../Log/' + Log_FolderName + '/004_T2-EML_F6.log'

# ----- ログ書込：プログラムの開始 -----
Log.Log_Info(Log_file, 'Program Start')


########## 処理ファイルのあるディレクトリ定義 ##########
Path = 'Z:/MOCVD/MOCVD過去プログラム/F6炉/'
# Path = 'C:/Users/hor78296/Desktop/F6炉/'


########## XML出力先ファイルパス ##########
Output_filepath = '//li.lumentuminc.net/data/SAG/TDS/Data/Files to Insert/XML/'
# Output_filepath = '../XML/004_T2-EML/'
#Output_filepath = 'C:/Users/hsi67063/Downloads/処理済みフォルダ/'

########## TestStepの定義 ##########
teststep_dict = {
    'TestStep1' : 'Coordinate',
    'TestStep2' : 'XRayDiffraction',
    'TestStep3' : 'CarrierConcentration',
    'TestStep4' : 'Thickness',
    'TestStep5' : 'Particles',
    'TestStep6' : 'Dulation',
    'TestStep7' : 'MO1-TEG',
    'TestStep8' : 'MO2-TMI',
    'TestStep9' : 'MO3-TEG',
    'TestStep10' : 'MO4-TMI',
    'TestStep11' : 'MO5-TMI',
    'TestStep12' : 'AsH3-1-20percent',
    'TestStep13' : 'AsH3-21-5percent',
    'TestStep14' : 'AsH3-3-20percent',
    'TestStep15' : 'PH3-1-50percent',
    'TestStep16' : 'PH3-2-50percent',
    'TestStep17' : 'DMZn-1-0.1percent',
    'TestStep18' : 'DMZn-2-0.1percent',
    'TestStep19' : 'Temperature',
    'TestStep20' : 'Piezocon',
    'TestStep21' : 'BallastN2',
    'TestStep22' : 'MO-Temperature',
    'TestStep23' : 'SORTED_DATA',
}


########## HeaderMiscの定義 ##########
HeaderMisc_dict = {
    'HeaderMisc1' : 'RecipeName-Macro',
    'HeaderMisc2' : 'RecipeName-Program',
    'HeaderMisc3' : 'RecipeName-Folder'
}


########## 取得した項目と型の対応表を定義 ##########
key_type = {
    "key_start_date_time" : str,
    "key_serial_number" : str,
    "key_part_number" : str,
    "key_batch_number" : str,
    "key_HeaderMisc1" : str,
    "key_HeaderMisc2" : str,
    "key_HeaderMisc3" : str,
    "key_XRayDiffraction_Strain" : float,
    "key_CarrierConcentration_Clad" : float,
    "key_CarrierConcentration_Contact" : float,
    "key_CarrierConcentration_Minimum" : float,
    "key_Thickness_Clad" : float,
    "key_Thickness_Contact": float,
    "key_Particles" : float,
    "key_Dulation_Step9" : float,
    "key_Dulation_Step10" : float,
    "key_Dulation_Step11" : float,
    "key_Dulation_Step12" : float,
    "key_Dulation_Step13" : float,
    "key_Dulation_Step14" : float,
    "key_Dulation_Step18" : float,
    "key_Dulation_Step22" : float,
    "key_Dulation_Step26" : float,
    "key_Dulation_Step27" : float,
    "key_Dulation_Step28" : float,
    "key_Dulation_Step29" : float,
    "key_Dulation_Step30" : float,
    "key_Dulation_Step31" : float,
    "key_Dulation_Step35" : float,
    "key_Dulation_Step37" : float,
    "key_MO1-TEG_Step18" : float,
    "key_MO2-TMI_Step18" : float,
    "key_MO3-TEG_Step22" : float,
    "key_MO3-TEG_Step26" : float,
    "key_MO3-TEG_Step27" : float,
    "key_MO3-TEG_Step28" : float,
    "key_MO3-TEG_Step29" : float,
    "key_MO3-TEG_Step30" : float,
    "key_MO3-TEG_Step31" : float,
    "key_MO4-TMI_Step22" : float,
    "key_MO4-TMI_Step26" : float,
    "key_MO4-TMI_Step27" : float,
    "key_MO4-TMI_Step28" : float,
    "key_MO4-TMI_Step29" : float,
    "key_MO4-TMI_Step30" : float,
    "key_MO4-TMI_Step31" : float,
    "key_MO5-TMI_Step9" : float,
    "key_MO5-TMI_Step10" : float,
    "key_MO5-TMI_Step11" : float,
    "key_MO5-TMI_Step12" : float,
    "key_MO5-TMI_Step13" : float,
    "key_MO5-TMI_Step14" : float,
    "key_MO5-TMI_Step35" : float,
    "key_AsH3-1-20percent_Step26" : float,
    "key_AsH3-1-20percent_Step27" : float,
    "key_AsH3-1-20percent_Step28" : float,
    "key_AsH3-1-20percent_Step29" : float,
    "key_AsH3-1-20percent_Step30" : float,
    "key_AsH3-1-20percent_Step31" : float,
    "key_AsH3-21-5percent_Step18" : float,
    "key_AsH3-3-20percent_Step22" : float,
    "key_PH3-1-50percent_Step9" : float,
    "key_PH3-1-50percent_Step10" : float,
    "key_PH3-1-50percent_Step11" : float,
    "key_PH3-1-50percent_Step12" : float,
    "key_PH3-1-50percent_Step13" : float,
    "key_PH3-1-50percent_Step14" : float,
    "key_PH3-1-50percent_Step35" : float,
    "key_PH3-1-50percent_Step37" : float,
    "key_PH3-2-50percent_Step18" : float,
    "key_PH3-2-50percent_Step22" : float,
    "key_DMZn-1-0.1percent_Step9" : float,
    "key_DMZn-1-0.1percent_Step10" : float,
    "key_DMZn-1-0.1percent_Step11" : float,
    "key_DMZn-1-0.1percent_Step12" : float,
    "key_DMZn-2-0.1percent_Step13" : float,
    "key_DMZn-2-0.1percent_Step14" : float,
    "key_DMZn-2-0.1percent_Step18" : float,
    "key_DMZn-2-0.1percent_Step22" : float,
    "key_DMZn-2-0.1percent_Step26" : float,
    "key_DMZn-2-0.1percent_Step27" : float,
    "key_DMZn-2-0.1percent_Step28" : float,
    "key_DMZn-2-0.1percent_Step29" : float,
    "key_DMZn-2-0.1percent_Step30" : float,
    "key_DMZn-2-0.1percent_Step31" : float,
    "key_Temperature_Step9" : float,
    "key_Temperature_Step10" : float,
    "key_Temperature_Step11" : float,
    "key_Temperature_Step12" : float,
    "key_Temperature_Step13" : float,
    "key_Temperature_Step14" : float,
    "key_Temperature_Step18" : float,
    "key_Temperature_Step22" : float,
    "key_Temperature_Step26" : float,
    "key_Temperature_Step27" : float,
    "key_Temperature_Step28" : float,
    "key_Temperature_Step29" : float,
    "key_Temperature_Step30" : float,
    "key_Temperature_Step31" : float,
    "key_Temperature_Step35" : float,
    "key_Temperature_Step37" : float,
    "key_Piezocon_F1" : float,
    "key_Piezocon_F1_Inverse" : float,
    "key_BallastN2_BallastN2" : float,
    "key_MO-Temperature_MO1-TEG" : float,
    "key_MO-Temperature_MO2-TMI" : float,
    "key_MO-Temperature_MO3-TEG" : float,
    "key_MO-Temperature_MO4-TMI" : float,
    "key_MO-Temperature_MO5-TMI" : float,
    "key_MO-Temperature_MO6-CBr4" : float,
    "key_MO-Temperature_MO7-Cp2Mg" : float,
    "key_MO-Temperature_MO8-TMA" : float,
    "key_STARTTIME_SORTED": float,
    "key_SORTNUMBER" : float,
    "key_LotNumber_9" : str
}


########## 対象ロット番号のイニシャルを記載したファイルを取得する ##########
Log.Log_Info(Log_file, 'Get SerialNumber Initial List ')
with open('T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開発/004_T2-EML/SerialNumber_Initial.txt', 'r', encoding='utf-8') as textfile:
#with open('C:/Users/hsi67063/Downloads/SerialNumber_Initial.txt', 'r', encoding='utf-8') as textfile:
    SerialNumber_list = {s.strip() for s in textfile.readlines()}


########## 前回処理を行ったファイル名を取得する ##########
with open('F6_FileName.txt', 'r', encoding='utf-8') as textfile:
    Before_FileName = textfile.readline()


########## 空欄チェック ##########
def Get_Cells_Info(Sheet):

    # ----- ログ記載：空欄判定 -----
    Log.Log_Info(Log_file, "Blank Check")

    # ----- False -> 空欄がない -----
    is_cells_empty = False

    # ----- 日付かエピ番号かバラストN2流量が空欄ならば処理を行わない -----
    if  Sheet['I8'].value is None or Sheet['R7'].value is None or Sheet['AG44'].value is None:
            is_cells_empty = True

    return is_cells_empty


########## データの取得 ##########
def Open_Data_Sheet(Sheet, filepath, SheetName):

    # ----- ログ記載：データ取得 -----
    Log.Log_Info(Log_file, 'Data Acquisition')

    # ----- データを格納する辞書を作成 -----
    data_dict = dict()

    # ----- ロット番号の取得 -----
    serial_number = Sheet["M8"].value

    # ----- ロット番号が数値の時があるので、数値の時はエラーを出す -----
    if type(serial_number) is not str:
        return None

    # ----- SQLサーバと接続して、品名を抜き出す -----
    conn,cursor = SQL.connSQL()

    # ----- Prime接続できなかったときはNoneが返ってくる -----
    if conn is None:
        Log.Log_Error(Log_file, serial_number + ' : ' + 'Connection with Prime Failed')
        sys.exit()

    # ----- Primeから品名を取得 -----
    part_number, Nine_Serial_Number = SQL.selectSQL(cursor, serial_number)
    SQL.disconnSQL(conn, cursor)

    # ----- Polaron / SEM / MOCVD の装置Noを取得 -----
    Equipment1,Equipment2,Equipment3, Equipment4 = '1','1','1','6'
    if '#2' in str(Sheet["J52"].value):
        Equipment1 = '2'
    if '#2' in str(Sheet['J57'].value):
        Equipment2 = '2'
    if '#2' in str(Sheet['J48'].value):
        Equipment3 = '2'

    # ----- データの格納 -----
    data_dict = {
        "key_start_date_time" : str(Sheet["R7"].value).replace(" ","T"),
        "key_serial_number" : serial_number,
        "key_LotNumber_9": Nine_Serial_Number,
        "key_part_number" : part_number,
        "key_batch_number" : Sheet["I8"].value,
        "key_HeaderMisc1" : Sheet["W3"].value,
        "key_HeaderMisc2" : Sheet["W4"].value,
        "key_HeaderMisc3" : Sheet["W5"].value,
        "key_XRayDiffraction_Strain" : Sheet["M50"].value,
        "key_CarrierConcentration_Clad" : Sheet["M54"].value,
        "key_CarrierConcentration_Contact" : Sheet["M53"].value,
        "key_CarrierConcentration_Minimum" : Sheet["M55"].value,
        "key_Thickness_Clad" : Sheet["M59"].value,
        "key_Thickness_Contact": Sheet["M58"].value,
        "key_Particles" : Sheet["M63"].value,
        "key_Dulation_Step9" : Sheet["I22"].value,
        "key_Dulation_Step10" : Sheet["I23"].value,
        "key_Dulation_Step11" : Sheet["I24"].value,
        "key_Dulation_Step12" : Sheet["I25"].value,
        "key_Dulation_Step13" : Sheet["I26"].value,
        "key_Dulation_Step14" : Sheet["I27"].value,
        "key_Dulation_Step18" : Sheet["I28"].value,
        "key_Dulation_Step22" : Sheet["I29"].value,
        "key_Dulation_Step26" : Sheet["I30"].value,
        "key_Dulation_Step27" : Sheet["I31"].value,
        "key_Dulation_Step28" : Sheet["I32"].value,
        "key_Dulation_Step29" : Sheet["I33"].value,
        "key_Dulation_Step30" : Sheet["I34"].value,
        "key_Dulation_Step31" : Sheet["I35"].value,
        "key_Dulation_Step35" : Sheet["I36"].value,
        "key_Dulation_Step37" : Sheet["I37"].value,
        "key_MO1-TEG_Step18" : Sheet['K28'].value,
        "key_MO2-TMI_Step18" : Sheet['L28'].value,
        "key_MO3-TEG_Step22" : Sheet['M29'].value,
        "key_MO3-TEG_Step26" : Sheet['M30'].value,
        "key_MO3-TEG_Step27" : Sheet['M31'].value,
        "key_MO3-TEG_Step28" : Sheet['M32'].value,
        "key_MO3-TEG_Step29" : Sheet['M33'].value,
        "key_MO3-TEG_Step30" : Sheet['M34'].value,
        "key_MO3-TEG_Step31" : Sheet['M35'].value,
        "key_MO4-TMI_Step22" : Sheet['N29'].value,
        "key_MO4-TMI_Step26" : Sheet['N30'].value,
        "key_MO4-TMI_Step27" : Sheet['N31'].value,
        "key_MO4-TMI_Step28" : Sheet['N32'].value,
        "key_MO4-TMI_Step29" : Sheet['N33'].value,
        "key_MO4-TMI_Step30" : Sheet['N34'].value,
        "key_MO4-TMI_Step31" : Sheet['N35'].value,
        "key_MO5-TMI_Step9" : Sheet['O22'].value,
        "key_MO5-TMI_Step10" : Sheet['O23'].value,
        "key_MO5-TMI_Step11" : Sheet['O24'].value,
        "key_MO5-TMI_Step12" : Sheet['O25'].value,
        "key_MO5-TMI_Step13" : Sheet['O26'].value,
        "key_MO5-TMI_Step14" : Sheet['O27'].value,
        "key_MO5-TMI_Step35" : Sheet['O36'].value,
        "key_AsH3-1-20percent_Step26" : Sheet["S30"].value,
        "key_AsH3-1-20percent_Step27" : Sheet["S31"].value,
        "key_AsH3-1-20percent_Step28" : Sheet["S32"].value,
        "key_AsH3-1-20percent_Step29" : Sheet["S33"].value,
        "key_AsH3-1-20percent_Step30" : Sheet["S34"].value,
        "key_AsH3-1-20percent_Step31" : Sheet["S35"].value,
        "key_AsH3-21-5percent_Step18" : Sheet["T28"].value,
        "key_AsH3-3-20percent_Step22" : Sheet["V29"].value,
        "key_PH3-1-50percent_Step9" : Sheet["W22"].value,
        "key_PH3-1-50percent_Step10" : Sheet["W23"].value,
        "key_PH3-1-50percent_Step11" : Sheet["W24"].value,
        "key_PH3-1-50percent_Step12" : Sheet["W25"].value,
        "key_PH3-1-50percent_Step13" : Sheet["W26"].value,
        "key_PH3-1-50percent_Step14" : Sheet["W27"].value,
        "key_PH3-1-50percent_Step35" : Sheet["W36"].value,
        "key_PH3-1-50percent_Step37" : Sheet["W37"].value,
        "key_PH3-2-50percent_Step18" : Sheet["X28"].value,
        "key_PH3-2-50percent_Step22" : Sheet["X29"].value,
        "key_DMZn-1-0.1percent_Step9" : Sheet["Y22"].value,
        "key_DMZn-1-0.1percent_Step10" : Sheet["Y23"].value,
        "key_DMZn-1-0.1percent_Step11" : Sheet["Y24"].value,
        "key_DMZn-1-0.1percent_Step12" : Sheet["Y25"].value,
        "key_DMZn-2-0.1percent_Step13" : Sheet["Z26"].value,
        "key_DMZn-2-0.1percent_Step14" : Sheet["Z27"].value,
        "key_DMZn-2-0.1percent_Step18" : Sheet["Z28"].value,
        "key_DMZn-2-0.1percent_Step22" : Sheet["Z29"].value,
        "key_DMZn-2-0.1percent_Step26" : Sheet["Z30"].value,
        "key_DMZn-2-0.1percent_Step27" : Sheet["Z31"].value,
        "key_DMZn-2-0.1percent_Step28" : Sheet["Z32"].value,
        "key_DMZn-2-0.1percent_Step29" : Sheet["Z33"].value,
        "key_DMZn-2-0.1percent_Step30" : Sheet["Z34"].value,
        "key_DMZn-2-0.1percent_Step31" : Sheet["Z35"].value,
        "key_Temperature_Step9" : Sheet["AB22"].value,
        "key_Temperature_Step10" : Sheet["AB23"].value,
        "key_Temperature_Step11" : Sheet["AB24"].value,
        "key_Temperature_Step12" : Sheet["AB25"].value,
        "key_Temperature_Step13" : Sheet["AB26"].value,
        "key_Temperature_Step14" : Sheet["AB27"].value,
        "key_Temperature_Step18" : Sheet["AB28"].value,
        "key_Temperature_Step22" : Sheet["AB29"].value,
        "key_Temperature_Step26" : Sheet["AB30"].value,
        "key_Temperature_Step27" : Sheet["AB31"].value,
        "key_Temperature_Step28" : Sheet["AB32"].value,
        "key_Temperature_Step29" : Sheet["AB33"].value,
        "key_Temperature_Step30" : Sheet["AB34"].value,
        "key_Temperature_Step31" : Sheet["AB35"].value,
        "key_Temperature_Step35" : Sheet["AB36"].value,
        "key_Temperature_Step37" : Sheet["AB37"].value,
        "key_Piezocon_F1" : Sheet["AG41"].value,
        "key_Piezocon_F1_Inverse" : Sheet["AG42"].value,
        "key_BallastN2_BallastN2" : Sheet["AG44"].value,
        "key_MO-Temperature_MO1-TEG" : Sheet["K46"].value,
        "key_MO-Temperature_MO2-TMI" : Sheet["L46"].value,
        "key_MO-Temperature_MO3-TEG" : Sheet["M46"].value,
        "key_MO-Temperature_MO4-TMI" : Sheet["N46"].value,
        "key_MO-Temperature_MO5-TMI" : Sheet["O46"].value,
        "key_MO-Temperature_MO6-CBr4" : Sheet["P46"].value,
        "key_MO-Temperature_MO7-Cp2Mg" : Sheet["Q46"].value,
        "key_MO-Temperature_MO8-TMA" : Sheet["R46"].value,
        "key_Equipment1" : Equipment1,
        "key_Equipment2" : Equipment2,
        "key_Equipment3" : Equipment3,
        "key_Equipment4" : Equipment4
    }

    # ----- 空欄箇所はNoneとして取得される。Noneは文字列に変換できないため、空欄("")に置き換える -----
    for keys in data_dict:
        if data_dict[keys] is None or data_dict[keys] == "None" or data_dict[keys] == '-':
            data_dict[keys] = ""
        # ----- 指数表記箇所はint型に変換する -----
        if type(data_dict[keys]) is float and 'e' in str(data_dict[keys]) and keys != "key_start_date_time":
            data_dict[keys] = int(float(data_dict[keys]))

    return data_dict


########## XMLファイルに変換 ##########
def Output_XML(xml_file, data_dict):

    # ----- ログ記載：XML変換 -----
    Log.Log_Info(Log_file, 'Excel File To XML File Conversion')

    f = open(Output_filepath + xml_file, 'w', encoding="utf-8")

    f.write('<?xml version="1.0" encoding="utf-8"?>' + '\n' +
            '<Results xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema">' + '\n' +
            '       <Result startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Result="Passed">' + '\n' +
            '               <Header SerialNumber=' + '"' + data_dict["key_serial_number"] + '"' + ' PartNumber=' + '"' + data_dict["key_part_number"] + '"' + ' Operation=' + '"' + Operation + '"' + ' TestStation=' + '"' + Operation + '"' + ' Operator=' + '"' + '-' + '"' + ' StartTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Site=' + '"' + Site + '"' + ' BatchNumber=' + '"' + data_dict["key_batch_number"] + '"' + ' LotNumber=' + '"' + data_dict["key_serial_number"] + '"/>' + '\n' +
            '               <HeaderMisc>' + '\n' +
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc1"] + '">' + data_dict["key_HeaderMisc1"] + '</Item>' + '\n'
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc2"] + '">' + data_dict["key_HeaderMisc2"] + '</Item>' + '\n'
            '                   <Item Description=' + '"' + HeaderMisc_dict["HeaderMisc3"] + '">' + data_dict["key_HeaderMisc3"] + '</Item>' + '\n'
            '               </HeaderMisc>' + '\n' +
            '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep1"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="X" Units="um" Value=' + '"' + X + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Y" Units="um" Value=' + '"' + Y + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep2"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="XRayDiffraction_Strain" Units="percent" Value=' + '"' + str(data_dict["key_XRayDiffraction_Strain"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep3"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="CarrierConcentration_Clad" Units="cm-3" Value=' + '"' + str(data_dict["key_CarrierConcentration_Clad"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="CarrierConcentration_Contact" Units="cm-3" Value=' + '"' + str(data_dict["key_CarrierConcentration_Contact"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="CarrierConcentration_Minimum" Units="cm-3" Value=' + '"' + str(data_dict["key_CarrierConcentration_Minimum"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep4"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Thickness_Clad" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Clad"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Thickness_Contact" Units="nm" Value=' + '"' + str(data_dict["key_Thickness_Contact"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep5"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Particles" Units="pieces" Value=' + '"' + str(data_dict["key_Particles"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep6"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step9" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step11" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step11"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step12"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step18" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step18"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step22" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step22"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step26" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step26"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step27" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step27"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step28" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step28"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step29" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step29"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step30" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step30"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step31" Units="min" Value=' + '"' + str(data_dict["key_Dulation_Step31"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step35" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step35"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step37" Units="sec" Value=' + '"' + str(data_dict["key_Dulation_Step37"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep7"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step18" Units="sccm" Value=' + '"' + str(data_dict["key_MO1-TEG_Step18"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep8"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step18" Units="sccm" Value=' + '"' + str(data_dict["key_MO2-TMI_Step18"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep9"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step22" Units="sccm" Value=' + '"' + str(data_dict["key_MO3-TEG_Step22"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step26" Units="sccm" Value=' + '"' + str(data_dict["key_MO3-TEG_Step26"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step27" Units="sccm" Value=' + '"' + str(data_dict["key_MO3-TEG_Step27"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step28" Units="sccm" Value=' + '"' + str(data_dict["key_MO3-TEG_Step28"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step29" Units="sccm" Value=' + '"' + str(data_dict["key_MO3-TEG_Step29"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step30" Units="sccm" Value=' + '"' + str(data_dict["key_MO3-TEG_Step30"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step31" Units="sccm" Value=' + '"' + str(data_dict["key_MO3-TEG_Step31"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep10"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step22" Units="sccm" Value=' + '"' + str(data_dict["key_MO4-TMI_Step22"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step26" Units="sccm" Value=' + '"' + str(data_dict["key_MO4-TMI_Step26"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step27" Units="sccm" Value=' + '"' + str(data_dict["key_MO4-TMI_Step27"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step28" Units="sccm" Value=' + '"' + str(data_dict["key_MO4-TMI_Step28"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step29" Units="sccm" Value=' + '"' + str(data_dict["key_MO4-TMI_Step29"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step30" Units="sccm" Value=' + '"' + str(data_dict["key_MO4-TMI_Step30"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step31" Units="sccm" Value=' + '"' + str(data_dict["key_MO4-TMI_Step31"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep11"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step9" Units="sccm" Value=' + '"' + str(data_dict["key_MO5-TMI_Step9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sccm" Value=' + '"' + str(data_dict["key_MO5-TMI_Step10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step11" Units="sccm" Value=' + '"' + str(data_dict["key_MO5-TMI_Step11"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12" Units="sccm" Value=' + '"' + str(data_dict["key_MO5-TMI_Step12"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="sccm" Value=' + '"' + str(data_dict["key_MO5-TMI_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="sccm" Value=' + '"' + str(data_dict["key_MO5-TMI_Step14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step35" Units="sccm" Value=' + '"' + str(data_dict["key_MO5-TMI_Step35"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep12"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step26" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-1-20percent_Step26"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step27" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-1-20percent_Step27"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step28" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-1-20percent_Step28"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step29" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-1-20percent_Step29"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step30" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-1-20percent_Step30"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step31" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-1-20percent_Step31"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep13"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step18" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-21-5percent_Step18"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep14"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step22" Units="sccm" Value=' + '"' + str(data_dict["key_AsH3-3-20percent_Step22"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep15"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step9" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-1-50percent_Step9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-1-50percent_Step10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step11" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-1-50percent_Step11"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-1-50percent_Step12"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-1-50percent_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-1-50percent_Step14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step35" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-1-50percent_Step35"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step37" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-1-50percent_Step37"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep16"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step18" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-2-50percent_Step18"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step22" Units="sccm" Value=' + '"' + str(data_dict["key_PH3-2-50percent_Step22"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep17"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step9" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-1-0.1percent_Step9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-1-0.1percent_Step10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step11" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-1-0.1percent_Step11"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-1-0.1percent_Step12"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep18"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-2-0.1percent_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-2-0.1percent_Step14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step18" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-2-0.1percent_Step18"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step22" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-2-0.1percent_Step22"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step26" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-2-0.1percent_Step26"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step27" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-2-0.1percent_Step27"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step28" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-2-0.1percent_Step28"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step29" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-2-0.1percent_Step29"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step30" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-2-0.1percent_Step30"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step31" Units="sccm" Value=' + '"' + str(data_dict["key_DMZn-2-0.1percent_Step31"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep19"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="Step9" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step9"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step10" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step10"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step11" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step11"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step12" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step12"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step13" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step13"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step14" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step14"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step18" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step18"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step22" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step22"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step26" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step26"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step27" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step27"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step28" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step28"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step29" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step29"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step30" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step30"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step31" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step31"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step35" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step35"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="Step37" Units="degree" Value=' + '"' + str(data_dict["key_Temperature_Step37"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep20"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="F1" Units="percent" Value=' + '"' + str(data_dict["key_Piezocon_F1"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="F1_Inverse" Units="percent" Value=' + '"' + str(data_dict["key_Piezocon_F1_Inverse"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep21"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="BallastN2" Units="slm" Value=' + '"' + str(data_dict["key_BallastN2_BallastN2"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep22"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="MO1-TEG" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO1-TEG"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO2-TMI" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO2-TMI"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO3-TEG" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO3-TEG"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO4-TMI" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO4-TMI"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO5-TMI" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO5-TMI"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO6-CBr4" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO6-CBr4"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO7-Cp2Mg" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO7-Cp2Mg"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="MO8-TMA" Units="degree" Value=' + '"' + str(data_dict["key_MO-Temperature_MO8-TMA"]) + '"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '               <TestStep Name=' + '"' + teststep_dict["TestStep23"] + '"' + ' startDateTime=' + '"' + data_dict["key_start_date_time"] + '"' + ' Status="Passed">' + '\n' +
            '                   <Data DataType="Numeric" Name="STARTTIME_SORTED" Units="" Value=' + '"' + str(data_dict["key_STARTTIME_SORTED"]) + '"/>' + '\n' +
            '                   <Data DataType="Numeric" Name="SORTNUMBER" Units="" Value=' + '"' + str(data_dict["key_SORTNUMBER"]) + '"/>' + '\n' +
            '                   <Data DataType="String" Name="LotNumber_5" Value=' + '"' + str(data_dict["key_serial_number"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '                   <Data DataType="String" Name="LotNumber_9" Value=' + '"' + str(data_dict["key_LotNumber_9"]) + '"' + ' CompOperation="LOG"/>' + '\n' +
            '               </TestStep>' + '\n' +
            '\n'
            '               <TestEquipment>' + '\n' +
            '                   <Item DeviceName="Polaron" DeviceSerialNumber="' + data_dict["key_Equipment1"] + '"/>' + '\n' +
            '                   <Item DeviceName="SEM" DeviceSerialNumber="' + data_dict["key_Equipment2"] + '"/>' + '\n' +
            '                   <Item DeviceName="XRD" DeviceSerialNumber="' + data_dict["key_Equipment3"] + '"/>' + '\n' +
            '                   <Item DeviceName="MOCVD" DeviceSerialNumber="' + data_dict["key_Equipment4"] + '"/>' + '\n' +
            '               </TestEquipment>' + '\n' +
            '\n'
            '               <ErrorData/>' + '\n' +
            '               <FailureData/>' + '\n' +
            '               <Configuration/>' + '\n' +
            '       </Result>' + '\n' +
            '</Results>'
            )
    f.close()


########## シートの判定からXML変換まで処理 ##########
def Data_Extract(filepath, SheetList):

    # ----- ログ記載：データ変換処理 -----
    Log.Log_Info(Log_file, 'Sub Program Main\n')

    wb = px.load_workbook(filepath, read_only = True, data_only = True)

    # ----- 取得したシートの処理 -----
    for Sheet_Name in SheetList:

        Sheet = wb[Sheet_Name]
        Initial = str(Sheet['M8'].value)[0]

        # ----- シートが処理対象シートかどうか確認 -----
        if 'HTL13B2-HIMEJI-LD.exe' not in str(Sheet['W3'].value) or Initial not in SerialNumber_list:
            Log.Log_Error(Log_file, Sheet_Name + ' : ' + 'Not Covered\n')
            continue

        # ----- 空欄チェック -----
        if Get_Cells_Info(Sheet):
            Log.Log_Error(Log_file, "Blank Error\n")
            continue

        # ----- データの取得 -----
        data_dict = Open_Data_Sheet(Sheet, os.path.basename(filepath), Sheet_Name)

        # ----- 辞書が空で返ってきたらエラーとして処理 -----
        if data_dict is None:
            Log.Log_Error(Log_file, "Lot No Error\n")
            continue

        # ----- 日付フォーマット変換 -----
        if len(data_dict['key_start_date_time']) != 19 or '年' in data_dict['key_start_date_time']:
            Log.Log_Error(Log_file, data_dict["key_serial_number"] + ' : ' + "Date Error\n")
            continue

        # ----- 品名チェック -----
        if len(data_dict["key_part_number"]) == 0:
            Log.Log_Error(Log_file, data_dict["key_serial_number"] + ' : ' + "Part Number Error\n")
            continue

        # ----- STARTTIME_SORTEDの追加 -----

        # 日付をExcel時間に変換する
        date = datetime.strptime(str(data_dict["key_start_date_time"]).replace('T', ' ').replace('.', ':'), "%Y-%m-%d %H:%M:%S")
        date_excel_number = int(str(date - datetime(1899, 12, 30)).split()[0])

        # エピ番号の数値部だけを取得する
        Epi_Number = 0
        for i in data_dict["key_batch_number"]:
            try:
                if 0<=int(i)<=9:
                    Epi_Number = Epi_Number*10+int(i)
            except:
                pass

        # エピ番号を10^6で割って、excel時間に加算する
        date_excel_number += Epi_Number/10**6

        # data_dictに登録する
        data_dict["key_STARTTIME_SORTED"] = date_excel_number
        data_dict["key_SORTNUMBER"] = Epi_Number


        # ----- データ型の確認 -----
        result = Check.Data_Type(key_type, data_dict)
        if result == False:
            Log.Log_Error(Log_file, data_dict["key_Serial_number"] + ' : ' + "Data Error\n")
            continue

        # ----- XMLファイル名定義 -----
        xml_file = 'Site=' + Site + ',ProductFamily=' + ProductFamily + ',Operation=' + Operation + \
                   ',Partnumber=' + data_dict["key_part_number"] + ',Serialnumber=' + data_dict["key_serial_number"] + \
                   ',Testdate=' + data_dict["key_start_date_time"].replace(':','.') + '.xml'

        Output_XML(xml_file, data_dict)
        Log.Log_Info(Log_file, data_dict["key_serial_number"] + ' : ' + "OK\n")

    wb.close()


########## Main処理 ##########
if __name__ == "__main__":

    # ----- ログ記載：Main処理の開始 -----
    Log.Log_Info(Log_file, 'Main Start')

    # ----- path内のフォルダ、ファイルを全部取得 -----
    all_files = os.listdir(Path)

    # ----- ログ記載：ファイル検索 -----
    Log.Log_Info(Log_file, 'File Search')

    # ----- FMが含まれているファイル名と最終更新日時(sec)を格納 -----
    array = []
    for filename in all_files:
        filepath = os.path.join(Path, filename)
        if "FM" in str(filename) and "$" not in str(filename) and os.path.isfile(filepath):
            dt = strftime("%Y-%m-%d %H:%M:%S", localtime(os.path.getctime(filepath)))
            array.append([filepath, dt])

    # ----- FMファイルが見つからなかった -----
    if len(array)==0:
        Log.Log_Info(Log_file, 'Folder Error')
        sys.exit()

    # ----- 最終更新日時順に並び替える -----
    array = sorted(array, key=lambda x:x[1])
    FileName = os.path.basename(array[0][0])
    Log.Log_Info(Log_file, FileName)

    # ----- 前回処理したエピ番号のNumber部分を取り出す -----
    Number = ""
    for i in Before_FileName:
        if "0" <= i <= "9":
            Number += i

    # ----- ファイルの切り替わりを確認 -----
    if Number not in FileName:

        # ----- ログ記載：FMフォルダ検索 -----
        Log.Log_Info(Log_file, 'Folder Serach')

        # ----- 過去フォルダからNumberが含まれたファイル名を取り出す -----
        Old_File_Path = MOCVD_OldFileSearch.F6(Number)
        if Old_File_Path == -1:
            Log.Log_Info(Log_file, 'Old Folder Error')
            sys.exit()

        # ----- ログ記載：シート名の取得 -----
        Log.Log_Info(Log_file, 'OLD Get SheetName')

        #　----- 上記で指定したファイルのシート一覧を取得する -----
        wb = px.load_workbook(Old_File_Path)
        Old_SheetName = wb.sheetnames
        wb.close()

        # ----- ログ記載：前Excelファイルのデータ取得 -----
        Log.Log_Info(Log_file, 'OLD Excel File Get Data')
        Log.Log_Info(Log_file, Old_File_Path)

        # ----- 切り替わり前のファイルの処理 -----
        Data_Extract(Old_File_Path, Old_SheetName)

    # ----- ログ書込：Excelファイルのデータ取得 -----
    Log.Log_Info(Log_file, 'Excel File Get Data')

    # ----- arrayに格納されている全てのファイルの処理を行う -----
    for file_path, _ in array:

        Log.Log_Info(Log_file, os.path.basename(file_path))

        # ----- 対象ファイルを開き、シートの一覧を取得する -----
        wb = px.load_workbook(file_path)
        SheetName = wb.sheetnames
        wb.close()

        # ----- 全シートの処理を行う -----
        Data_Extract(file_path, SheetName)

    # ----- ログ書込：テキストファイルにファイル名を上書きで書込する -----
    Log.Log_Info(Log_file, 'Write FileName')

    # ----- 先ほど処理を行ったファイル名の書き込み -----
    with open('F6_FileName.txt', 'w', encoding='utf-8') as textfile:
        textfile.write(FileName)

    # ----- 最終行を書き込んだファイルをGドライブに転送 -----
    shutil.copy("F6_FileName.txt", 'T:/04_プロセス関係/10_共通/91_KAIZEN-TDS/01_開発/004_T2-EML/13_ProgramUsedFile/')


########## ログ記載：プログラムの終了 ##########
Log.Log_Info(Log_file, 'Program End')